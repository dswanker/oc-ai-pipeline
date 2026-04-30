"""
Sanitize XLSForms before passing to pyxform for Build Preview rendering.

pyxform is stricter than OpenClinica's own validator. Two known cases
where OC accepts a form but pyxform rejects it:

  1. OC-8 phantom end group — OC requires:
         begin repeat NAME
         end group          ← no matching begin group (phantom)
         end repeat
     This is stripped via stack-based context detection. Name-based
     detection ('PHANTOM' in name) is kept as a fallback for forms built
     by older edc-builder versions that still had the name set.

  2. Legacy *_PHANTOM named rows — older edc-builder versions emitted
     explicitly-named phantom rows. Strip those too.
"""
import io
import openpyxl


def sanitize_xlsform_bytes(src_bytes: bytes) -> bytes:
    """Take XLSForm .xlsx bytes, return cleaned .xlsx bytes safe for pyxform."""
    wb = openpyxl.load_workbook(io.BytesIO(src_bytes))
    if 'survey' not in wb.sheetnames:
        out = io.BytesIO(); wb.save(out); return out.getvalue()

    ws = wb['survey']
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        out = io.BytesIO(); wb.save(out); return out.getvalue()

    headers = [c.value for c in rows[0]]
    type_idx = headers.index('type') if 'type' in headers else None
    name_idx = headers.index('name') if 'name' in headers else None

    if type_idx is None:
        out = io.BytesIO(); wb.save(out); return out.getvalue()

    rows_to_delete = set()

    # ── Pass 1: stack-based OC-8 phantom detection ─────────────────────────
    # An 'end group' is a phantom when the innermost open block is a repeat,
    # not a group. This is independent of the row's name.
    stack = []  # 'group' or 'repeat'
    for i, row in enumerate(rows[1:], start=1):   # i = 0-based data index
        t = str(row[type_idx].value or '').strip().lower() if type_idx < len(row) else ''
        if t == 'begin group':
            stack.append('group')
        elif t == 'begin repeat':
            stack.append('repeat')
        elif t == 'end group':
            if stack and stack[-1] == 'repeat':
                rows_to_delete.add(i + 1)  # +1 → 1-based sheet row (row 1 = header)
            elif stack:
                stack.pop()
        elif t == 'end repeat':
            if stack and stack[-1] == 'repeat':
                stack.pop()

    # ── Pass 2: legacy name-based PHANTOM detection (older edc-builder) ────
    if name_idx is not None:
        for i, row in enumerate(rows[1:], start=1):
            name = str(row[name_idx].value or '')
            type_v = str(row[type_idx].value or '').strip()
            if ('PHANTOM' in name and
                    type_v in ('end_group', 'end group', 'begin_group', 'begin group')):
                rows_to_delete.add(i + 1)

    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def get_form_settings_bytes(xlsform_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(xlsform_bytes), read_only=True)
    if 'settings' not in wb.sheetnames:
        return {}
    s = wb['settings']
    rows = list(s.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {}
    return dict(zip(rows[0], rows[1]))

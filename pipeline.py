"""
pipeline.py — oc-ai-pipeline orchestration

Architecture
────────────
  call_claude()  → JSON text  (analysis, fast, no code execution)
  run_*_*()      → real binary files via LOCAL scripts in skills/*/scripts/

File generation happens via local python imports from skills/*/scripts/
rather than the Anthropic Skills API sandbox (which previously failed to
reliably return file_ids for file retrieval). Every chain imports its
skill's scripts directly and runs them in a thread pool executor.

Flow (fresh run):
  1. call_claude           : protocol PDF  → Study Spec JSON
  2. run_study_spec_files  : JSON          → Study Spec PDF + XLSX
  3. call_claude           : JSON          → Protocol Summary JSON
  4. run_protocol_summary_pdf : JSON       → Protocol Summary PDF
  5. run_pricing_quote     : JSON          → Quote PDFs + XLSXs
  6. run_edc_build         : JSON          → EDC Build ZIP
  7. run_dvs_xlsx          : JSON + ZIP    → DVS XLSX
  8. create_oc_study       : JSON          → OC study + design board

Human-in-the-loop paths:
  A. Edited Study Spec XLSX uploaded  → skip steps 1-2, run 3-8
  B. Edited Build ZIP uploaded        → skip steps 1-6, run 7 only
  C. Edited DVS uploaded              → translate changes → rebuild ZIP + DVS
  D. Edited Quote XLSX uploaded       → DEPRECATED, logs a message + skips
  E. Edited SOE CSV uploaded          → update SOE in OpenClinica (not impl)
"""

import asyncio, io, json, os, sys, tempfile, time, traceback, zipfile, datetime as _dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from monday_client import (get_item, download_file, upload_file, set_status,
                            append_log, set_text, set_link, download_column_file,
                            list_column_filenames, COL, BOARD_ID)
from auth_manager import AuthManager
from claude_client  import call_claude, extract_json, run_skill
from migration_pipeline import run_migration as run_edc_migration
from trainer_integration import (
    run_protocol_analysis_quick,
    retrieve_examples,
    create_pending_row,
    format_examples_block,
    trainer_enabled,
)
from prompts        import (
    EDC_STRUCTURE_PROMPT, PRICING_SUMMARY_PROMPT,
    DVS_TRANSLATE_PROMPT,
)
from uat_loader import run_uat_loader, UAT_STATUS

SKILLS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'skills')

STATUS = {
    "not_started":            "Not Started",
    "paused_for_auth":        "Paused for Authentication",
    "analysis_running":       "Analysis Running",
    "analysis_complete":      "Analysis Complete",
    "build_pricing_running":  "Build + Pricing Running",
    "build_complete":         "Build Complete",
    "pricing_complete":       "Pricing Complete",
    "dvs_running":            "DVS Running",
    "dvs_complete":           "DVS Complete — Awaiting Review",
    "creating_oc_study":      "Creating OC Study",
    "all_complete":           "All Complete",
    "build_preview_running":  "Build Preview Running",
    "failed":                 "Failed",
    # Design change intake statuses (added 2026-06-02)
    "change_intake_running":  "Change Intake Running",
    "change_intake_complete": "Change Intake Complete",
    "change_intake_failed":   "Change Intake Failed",
}
# Trainer retrieval — number of similar past pairs to request.
# Phase 1 starts at 3; raise after observing prompt length & quality.
TRAINER_K = 3

# ── Helpers ───────────────────────────────────────────────────────────────────

def _find(files: dict, *patterns) -> bytes | None:
    """Return bytes for first filename ending with any pattern."""
    for pat in patterns:
        for name, data in files.items():
            if name.lower().endswith(pat.lower()):
                return data
    return None


async def _noop_bytes():
    """Awaitable that returns None — used as a placeholder in asyncio.gather."""
    return None


def _vendor_slug_from_display_name(display_name):
    """Translate a monday `source_edc_system` display name (e.g. "REDCap")
    to a conventions/vendors/ slug (e.g. "redcap") via the existing
    VENDOR_CONVENTION_FILES dict in migration/odm_to_spec.py.

    Returns None for unknown / empty input — caller treats that as
    non-migration build (cascade vendor bucket is skipped).

    odm_to_spec.py lives in migration/ but uses bare-name imports
    (`from odm_reader import ...`), so we add migration/ to sys.path
    on first use — same pattern as migration_pipeline.py.

    >>> _vendor_slug_from_display_name("REDCap")
    'redcap'
    >>> _vendor_slug_from_display_name("Castor EDC")
    'castor'
    >>> _vendor_slug_from_display_name("UnknownEDC") is None
    True
    >>> _vendor_slug_from_display_name("") is None
    True
    >>> _vendor_slug_from_display_name(None) is None
    True
    """
    if not display_name:
        return None
    import os as _os, sys as _sys
    _mig_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "migration")
    if _mig_dir not in _sys.path:
        _sys.path.insert(0, _mig_dir)
    try:
        from odm_to_spec import VENDOR_CONVENTION_FILES
    except ImportError:
        return None
    filename = VENDOR_CONVENTION_FILES.get(display_name)
    if not filename:
        return None
    return filename[:-3] if filename.endswith(".md") else filename


# ── Customer Convention Questions (CQ) ─────────────────────────────────────────
# Customers can supply convention preferences via columns on the AI Hub board
# whose titles start with "CQ " (preferred, full question becomes the key) or
# "CQ_" (legacy underscore form). The pipeline reads ALL such columns
# dynamically — adding a new question to the board requires zero code changes,
# the next pipeline run picks it up automatically.
#
# These customer answers are injected into the EDC structure prompt as part of
# extra_parts, so Claude sees them when generating the Study Spec JSON. The
# Study Spec then flows through the rest of the pipeline (build, DVS, etc.),
# so conventions injected here propagate to all downstream stages.

CQ_PREFIX_NEW    = "CQ "    # human-readable: "CQ How do you want X?"
CQ_PREFIX_LEGACY = "CQ_"    # short identifier: "CQ_How_Do_You_Want_X"


def _strip_cq_prefix(title: str) -> str:
    """Remove the CQ prefix and normalize the question text for use as a key."""
    if title.startswith(CQ_PREFIX_NEW):
        return title[len(CQ_PREFIX_NEW):].strip()
    if title.startswith(CQ_PREFIX_LEGACY):
        return title[len(CQ_PREFIX_LEGACY):].replace("_", " ").strip()
    return title


def _extract_customer_conventions(cols: dict) -> dict:
    """
    Extract customer convention answers from the cols dict (column_id -> column).
    Recognizes columns whose title starts with 'CQ ' or 'CQ_'. Empty answers
    are skipped. Returns dict[question_text -> answer_text].
    """
    out = {}
    for col_id, col in cols.items():
        title = (col.get("title") or "").strip()
        is_cq = (title.startswith(CQ_PREFIX_NEW)
                 or title.startswith(CQ_PREFIX_LEGACY)
                 or col_id.startswith(CQ_PREFIX_LEGACY))
        if not is_cq:
            continue
        answer = (col.get("text") or "").strip()
        if not answer:
            continue
        out[_strip_cq_prefix(title)] = answer
    return out


def _build_customer_conventions_block(conventions: dict) -> str:
    """Format customer conventions as a prompt-ready text block. Empty when no answers."""
    if not conventions:
        return ""
    lines = [
        "Customer Convention Preferences (apply these when generating the Study Spec):",
    ]
    for question, answer in conventions.items():
        lines.append(f"  - Q: {question}")
        lines.append(f"    A: {answer}")
    return "\n".join(lines)


def _xl_header_row(ws, headers, bg="1B3A6B", fg="FFFFFF"):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(name="Arial", bold=True, color=fg, size=10)
    aln  = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font, cell.fill, cell.alignment = font, fill, aln

def _xl_data_row(ws, values, bold=False):
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = Font(name="Arial", bold=bold, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

def _xl_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── XLSForm ZIP builder (local, from JSON) ────────────────────────────────────

def _xlsform_zip(build_json):
    """Convert EDC Build JSON into a ZIP of XLSForm xlsx files. Returns bytes.

    Uses build_single_xlsform from the edc-builder skill scripts so every
    output form carries the OC form template (reference tabs + dropdown).
    Falls back to building from scratch if the skill scripts are unavailable.
    """
    # Standard survey columns — used to detect extra_cols from JSON data
    _SURVEY_COLS = {
        "type", "name", "label", "bind::oc:itemgroup", "hint", "appearance",
        "bind::oc:briefdescription", "bind::oc:description", "relevant",
        "required", "required_message", "constraint", "constraint_message",
        "default", "calculation", "trigger", "readonly", "image",
        "repeat_count", "bind::oc:external"
    }

    # Try to import the edc-builder script so we get the template + dropdown
    _add_scripts("edc-builder")
    try:
        from build_xlsforms import build_single_xlsform as _build_single
        _use_skill = True
    except ImportError:
        _build_single = None
        _use_skill = False
        print("_xlsform_zip: build_xlsforms not available — using scratch builder",
              flush=True)

    forms   = build_json.get("forms", {})
    zip_buf = io.BytesIO()

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, form_data in forms.items():

            # ── CSV pass-through (timepoint, lab ranges, checklist) ────────
            if filename.endswith('.csv'):
                if isinstance(form_data, str):
                    zf.writestr(filename, form_data)
                elif isinstance(form_data, list):
                    import csv as _csv
                    cbuf = io.StringIO()
                    if form_data:
                        writer = _csv.DictWriter(cbuf, fieldnames=form_data[0].keys())
                        writer.writeheader()
                        writer.writerows(form_data)
                    zf.writestr(filename, cbuf.getvalue())
                continue

            survey   = form_data.get("survey", [])
            choices  = form_data.get("choices", [])
            settings = form_data.get("settings", {}) or {}

            # Derive form_id from settings or strip extension from filename
            form_id = (settings.get("form_id")
                       or os.path.splitext(filename)[0])

            if _use_skill:
                # ── Template-based path via build_single_xlsform ──────────
                # Detect any extra columns beyond the standard 20
                extra_cols = []
                for row in survey:
                    for k in row:
                        if k not in _SURVEY_COLS and k not in extra_cols:
                            extra_cols.append(k)

                skill_form_data = {
                    "form_id":    form_id,
                    "form_title": settings.get("form_title", form_id),
                    "settings":   settings,
                    "survey":     survey,
                    "choices":    choices,
                    "extra_cols": extra_cols,
                }
                build_log = {"placeholder_applied": [], "build_errors": []}
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
                    tmp_path = tf.name
                try:
                    _build_single(skill_form_data, tmp_path, build_log)
                    with open(tmp_path, "rb") as f:
                        zf.writestr(filename, f.read())
                finally:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass

            else:
                # ── Scratch fallback (no template, no reference tabs) ──────
                wb   = Workbook()
                ws_s = wb.active
                ws_s.title = "survey"
                if survey:
                    hdrs = list(survey[0].keys())
                    _xl_header_row(ws_s, hdrs)
                    for row in survey:
                        _xl_data_row(ws_s, [row.get(h, "") for h in hdrs])

                ws_c = wb.create_sheet("choices")
                if choices:
                    hdrs = list(choices[0].keys())
                    _xl_header_row(ws_c, hdrs)
                    for row in choices:
                        _xl_data_row(ws_c, [row.get(h, "") for h in hdrs])

                ws_t = wb.create_sheet("settings")
                if settings:
                    _xl_header_row(ws_t, list(settings.keys()))
                    _xl_data_row(ws_t, list(settings.values()))

                xbuf = io.BytesIO()
                wb.save(xbuf)
                zf.writestr(filename, xbuf.getvalue())

        checklist = build_json.get("study_checklist")
        if checklist and isinstance(checklist, list) and checklist:
            import csv as _csv
            cbuf = io.StringIO()
            writer = _csv.DictWriter(cbuf, fieldnames=checklist[0].keys())
            writer.writeheader()
            writer.writerows(checklist)
            zf.writestr("study_checklist.csv", cbuf.getvalue())

    zip_buf.seek(0)
    return zip_buf.getvalue()


def _convert_to_pdf(file_bytes: bytes, filename: str) -> bytes:
    """
    Convert a document file to PDF bytes for Claude ingestion.
    Supports: .docx, .doc (via LibreOffice)
    Returns PDF bytes, or empty bytes if conversion fails.
    """
    import subprocess, tempfile, shutil, os
    ext = (filename.rsplit('.', 1)[-1] if '.' in filename else '').lower()
    if ext not in ('docx', 'doc'):
        return b''
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            src_path = os.path.join(tmpdir, f'protocol.{ext}')
            with open(src_path, 'wb') as f:
                f.write(file_bytes)
            result = subprocess.run(
                ['libreoffice', '--headless', '--convert-to', 'pdf',
                 '--outdir', tmpdir, src_path],
                capture_output=True, timeout=90, text=True
            )
            pdf_path = os.path.join(tmpdir, f'protocol.pdf')
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    pdf = f.read()
                print(f"Converted {filename} → PDF ({len(pdf):,} bytes)", flush=True)
                return pdf
            else:
                print(f"LibreOffice conversion failed for {filename}: "
                      f"{result.stderr[:200]}", flush=True)
                return b''
    except Exception as e:
        print(f"_convert_to_pdf error: {e}", flush=True)
        return b''


def _extract_docx_as_text(file_bytes: bytes) -> str:
    """
    Extract plain text from a .docx file using python-docx.
    Fallback when LibreOffice is unavailable.
    """
    try:
        import io
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        print(f"_extract_docx_as_text error: {e}", flush=True)
        return ''


def _google_doc_export_url(link: str) -> str:
    """
    Convert a Google Docs / Drive share link into a PDF-export URL.

    Handles the three common share-link shapes:
      - https://docs.google.com/document/d/<ID>/edit?usp=sharing
      - https://docs.google.com/document/d/<ID>/view
      - https://drive.google.com/file/d/<ID>/view?usp=sharing

    Returns the public ``?export=pdf`` URL when the document ID can be
    extracted; otherwise returns "" so callers can skip cleanly. Only
    works for documents whose share setting is "Anyone with the link"
    — protected docs will 401/403 and the caller treats that as empty.
    """
    import re as _re
    m = _re.search(r"/d/([A-Za-z0-9_-]{20,})", link or "")
    if not m:
        return ""
    doc_id = m.group(1)
    if "drive.google.com" in link:
        return f"https://drive.google.com/uc?export=download&id={doc_id}"
    return f"https://docs.google.com/document/d/{doc_id}/export?format=pdf"


def _detect_oc_standard_type(file_bytes):
    """
    Detect whether the file in the oc_standard column is an ODM XML or a
    ZIP of XLSForms. Returns 'ODM_XML', 'XLSFORM_ZIP', or 'UNKNOWN'.
    """
    if not file_bytes:
        return 'UNKNOWN'
    # ZIP magic bytes: PK\x03\x04
    if file_bytes[:4] == b'PK\x03\x04':
        return 'XLSFORM_ZIP'
    # XML: starts with BOM or <?xml or <ODM
    head = file_bytes[:200].lstrip()
    if (head.startswith(b'<?xml') or head.startswith(b'<ODM') or
            head.startswith(b'\xef\xbb\xbf<?xml')):
        return 'ODM_XML'
    # Try decoding as text and check for ODM signature
    try:
        text = file_bytes[:500].decode('utf-8', errors='ignore')
        if '<ODM' in text or 'xmlns:odm' in text.lower():
            return 'ODM_XML'
    except Exception:
        pass
    return 'UNKNOWN'


def _read_zip_xlsforms(zip_bytes):
    """Read a ZIP of XLSForm xlsx files. Returns forms dict."""
    import openpyxl
    forms = {}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if not name.endswith('.xlsx') or name.startswith('__'):
                continue
            src = openpyxl.load_workbook(io.BytesIO(zf.read(name)))
            form_data = {}
            for sheet_name in ['survey', 'choices', 'settings']:
                if sheet_name in src.sheetnames:
                    ws   = src[sheet_name]
                    rows = list(ws.values)
                    if not rows:
                        form_data[sheet_name] = [] if sheet_name != 'settings' else {}
                        continue
                    headers = [str(h).strip() if h else '' for h in rows[0]]
                    if sheet_name == 'settings':
                        form_data[sheet_name] = dict(zip(headers, [
                            str(v) if v is not None else '' for v in rows[1]
                        ])) if len(rows) > 1 else {}
                    else:
                        form_data[sheet_name] = [
                            {h: (str(v) if v is not None else '')
                             for h, v in zip(headers, row)}
                            for row in rows[1:]
                            if any(v is not None for v in row)
                        ]
                else:
                    form_data[sheet_name] = [] if sheet_name != 'settings' else {}
            forms[os.path.basename(name)] = form_data
    print(f"Read {len(forms)} XLSForm(s) from ZIP", flush=True)
    return {"forms": forms}


def _dvs_xlsx_to_text(dvs_bytes):
    """Extract DVS XLSX as structured text for Claude to read."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(dvs_bytes))
    lines = []
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.values)
        if not rows:
            continue
        lines.append(f"\n=== Sheet: {sheet_name} ===")
        headers = [str(h) if h else '' for h in rows[0]]
        lines.append('\t'.join(headers))
        for row in rows[1:]:
            if any(v is not None for v in row):
                lines.append('\t'.join(str(v) if v is not None else '' for v in row))
    return '\n'.join(lines)


# ── Pricing model — run scripts locally ───────────────────────────────────────

def _add_scripts(skill_name):
    path = os.path.join(SKILLS_DIR, skill_name, "scripts")
    if path not in sys.path:
        sys.path.insert(0, path)


def run_pricing_quote(pricing_summary_dict,
                      additional_sub_disc=0.0, additional_svc_disc=0.0,
                      edc_structure=None):
    """Run pricing-quote scripts locally. Returns dict of file bytes.

    Args:
      pricing_summary_dict: the Protocol Summary JSON (primary pricing input)
      additional_sub_disc:  user-entered subscription discount (decimal,
                            e.g., 0.10 for 10% off). Applied to all module
                            totals (monthly_fee × duration) after volume +
                            platform discounts.
      additional_svc_disc:  user-entered services discount (decimal). Applied
                            to the build_fee (ps_hours + pm_hours + contingency).
      edc_structure:        Study Spec JSON for enriching flag comments.
    """
    _add_scripts("pricing-quote")
    from pricing_engine      import calculate_quote
    from generate_quote_pdf  import build_quote_pdfs
    from generate_quote_xlsx import build_quote_xlsx

    quote    = calculate_quote(pricing_summary_dict,
                               additional_sub_disc=additional_sub_disc,
                               additional_svc_disc=additional_svc_disc,
                               edc_structure=edc_structure)
    protocol = quote["study_meta"].get("protocol_number", "STUDY")

    with tempfile.TemporaryDirectory() as tmp:
        paths = {
            "internal_pdf":  os.path.join(tmp, f"{protocol}_Quote_Internal.pdf"),
            "client_pdf":    os.path.join(tmp, f"{protocol}_Quote_Client.pdf"),
            "internal_xlsx": os.path.join(tmp, f"{protocol}_Quote_Internal.xlsx"),
            "client_xlsx":   os.path.join(tmp, f"{protocol}_Quote_Client.xlsx"),
        }
        build_quote_pdfs(quote, paths["internal_pdf"], paths["client_pdf"])
        build_quote_xlsx(quote, paths["internal_xlsx"], paths["client_xlsx"])
        return {k: open(v, "rb").read() for k, v in paths.items()}


# ── Local runners for Study Spec, Protocol Summary, EDC Build, DVS ──────────
# These mirror test_skills_locally.py — imported directly from the skills
# folder's scripts/ directory and run in a thread pool executor.

def run_study_spec_files(struct_json, customer_subdomain="", migration_source=None):
    """Generate Study Spec PDF + XLSX locally. Returns {'pdf': bytes, 'xlsx': bytes}."""
    _add_scripts("protocol-analysis")
    from generate_study_spec_pdf  import build_edc_pdf
    from generate_study_spec_xlsx import build_edc_xlsx

    # Compute conventions_applied metrics from the forms data per
    try:
        from conventions_engine import apply_conventions
        study_meta = struct_json.get("study_meta", {})
        study_id = study_meta.get("protocol_number", "UNKNOWN")
        
        apply_conventions(
            spec=struct_json,
            study_id=study_id,
            customer_subdomain=customer_subdomain,
            migration_source=migration_source,
        )
        
        applied_list = struct_json.get("study_meta", {}).get("conventions_engine_applied", [])
        print(f"conventions_engine: applied {len(applied_list)} conventions", flush=True)
    except Exception as ex:
        print(f"conventions_engine FAILED — continuing without conventions: {ex}", flush=True)
        import traceback
        traceback.print_exc()
    protocol = (struct_json.get("study_meta", {}).get("protocol_number")
                or "STUDY")
    with tempfile.TemporaryDirectory() as tmp:
        pdf_path  = os.path.join(tmp, f"{protocol}_Study_Specification.pdf")
        xlsx_path = os.path.join(tmp, f"{protocol}_Study_Specification.xlsx")
        build_edc_pdf(struct_json, pdf_path)
        build_edc_xlsx(struct_json, xlsx_path)
        return {
            "pdf":  open(pdf_path, "rb").read(),
            "xlsx": open(xlsx_path, "rb").read(),
        }


def run_protocol_summary_pdf(pricing_json, struct_json=None):
    """Generate Protocol Summary PDF locally. Returns bytes.

    If struct_json is provided, the PDF includes a Study Event Schedule
    sub-table in Section 3 (Timepoint Label | Arm | Forms Assigned —
    Event OID omitted for client-facing simplicity).
    """
    _add_scripts("protocol-analysis")
    from generate_protocol_summary_pdf import build_pricing_pdf

    protocol = (pricing_json.get("study_meta", {}).get("protocol_number")
                or "STUDY")
    with tempfile.TemporaryDirectory() as tmp:
        pdf_path = os.path.join(tmp, f"{protocol}_Protocol_Summary.pdf")
        build_pricing_pdf(pricing_json, pdf_path, struct_json=struct_json)
        return open(pdf_path, "rb").read()


def run_edc_build(struct_json):
    """Build EDC ZIP locally. Returns (zip_bytes, build_log, forms_json)."""
    _add_scripts("edc-builder")
    from build_xlsforms  import build_all_xlsforms, write_timepoint_csv, write_labranges_csv, write_calendar_artifacts
    from build_checklist import build_checklist_pdf, build_checklist_xlsx
    from build_package   import build_package

    protocol = (struct_json.get("study_meta", {}).get("protocol_number")
                or "STUDY")
    build_log = {
        'forms_built':         [],
        'forms_skipped':       [],
        'build_errors':        [],
        'build_warnings':      [],
        'placeholder_applied': [],
        'oid_placeholders':    [],
    }

    with tempfile.TemporaryDirectory() as tmp:
        forms_dir     = os.path.join(tmp, 'forms')
        csv_dir       = os.path.join(tmp, 'csv')
        checklist_dir = os.path.join(tmp, 'checklist')
        package_dir   = os.path.join(tmp, 'package')
        for d in (forms_dir, csv_dir, checklist_dir, package_dir):
            os.makedirs(d, exist_ok=True)

        build_all_xlsforms(struct_json, forms_dir, build_log)
        write_timepoint_csv(struct_json.get('timepoint_csv', {}),
                            os.path.join(csv_dir, f'{protocol}_tpt.csv'),
                            build_log)
        write_labranges_csv(struct_json.get('labranges_csv', {}),
                            os.path.join(csv_dir, f'{protocol}_labranges.csv'),
                            build_log)
        write_calendar_artifacts(struct_json, csv_dir, build_log)
        build_checklist_pdf(struct_json, build_log,
                            os.path.join(checklist_dir,
                                         f'{protocol}_Build_Checklist.pdf'))
        build_checklist_xlsx(struct_json, build_log,
                             os.path.join(checklist_dir,
                                          f'{protocol}_Build_Checklist.xlsx'))

        zip_path = build_package(struct_json, build_log,
                                 forms_dir, csv_dir, checklist_dir, package_dir)
        zip_bytes = open(zip_path, "rb").read()

        # Also build the forms_json view used by DVS (survey rows per form)
        forms_json = {"forms": {}}
        for fname in sorted(os.listdir(forms_dir)):
            if fname.lower().endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(os.path.join(forms_dir, fname),
                                            read_only=True, data_only=True)
                survey_rows = []
                if 'survey' in wb.sheetnames:
                    ws = wb['survey']
                    rows = list(ws.iter_rows(values_only=True))
                    if rows:
                        headers = [str(h or '').strip() for h in rows[0]]
                        for r in rows[1:]:
                            row_dict = {headers[i]: r[i] for i in range(len(headers))
                                        if i < len(r) and r[i] is not None}
                            if row_dict:
                                survey_rows.append(row_dict)
                choice_rows = []
                if 'choices' in wb.sheetnames:
                    ws_c = wb['choices']
                    c_rows = list(ws_c.iter_rows(values_only=True))
                    if c_rows:
                        c_hdrs = [str(h or '').strip() for h in c_rows[0]]
                        for r in c_rows[1:]:
                            rd = {c_hdrs[i]: r[i] for i in range(len(c_hdrs))
                                  if i < len(r) and r[i] is not None}
                            if rd:
                                choice_rows.append(rd)
                forms_json["forms"][fname] = {"survey": survey_rows, "choices": choice_rows}
        return zip_bytes, build_log, forms_json


def run_dvs_xlsx(struct_json, forms_json):
    """Build DVS XLSX locally. Returns bytes or None if builder not available.

    The DVS is a MECHANICAL MIRROR of what's actually in the XLSForms —
    not an invention of new checks from the protocol. Every constraint,
    required flag, calculation, and relevant expression in the forms
    produces corresponding rows in DVS_OC4, Protocol_Extraction,
    Query_Text_Library, and UAT_Cases. Humans add missing checks by
    editing the DVS and re-uploading (DVS_TRANSLATE_PROMPT picks up the
    diff and injects it back into the forms)."""
    _add_scripts("dvs-specification")
    try:
        from generate_dvs import build_dvs
        from extract_dvs_from_forms import extract_dvs_data
    except ImportError as e:
        print(f"DVS builder not available locally: {e}", flush=True)
        return None

    protocol = (struct_json.get("study_meta", {}).get("protocol_number")
                or "STUDY")

    # Mechanical extraction — walks every survey row in every form and
    # emits the 4 DVS content arrays in the shape build_dvs expects.
    dvs_data = extract_dvs_data(struct_json, forms_json)
    print(f"DVS extraction — {len(dvs_data['dvs_oc4'])} checks, "
          f"{len(dvs_data['query_text_library'])} unique messages, "
          f"{len(dvs_data['uat_cases'])} UAT cases", flush=True)

    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = os.path.join(tmp, f"{protocol}_DVS.xlsx")
        build_dvs(dvs_data, xlsx_path)
        return open(xlsx_path, "rb").read()

def _extract_scheduling_block(struct_json):
    """Second-pass targeted extraction of the scheduling block.

    Called when protocol-analysis produces a struct_json without a scheduling
    array. Uses the already-extracted timepoint_csv rows as the sole input —
    no protocol PDF needed. Much cheaper and more reliable than including
    scheduling in the main protocol-analysis call.

    Returns struct_json with scheduling populated (mutates in place and returns).
    Fails silently — calendaring falls back gracefully if this doesn't fire.
    """
    import anthropic as _anthropic

    existing = struct_json.get("scheduling")
    if existing:
        print("[scheduling-pass] scheduling block already present — skipping second pass", flush=True)
        return struct_json

    tpt_rows = struct_json.get("timepoint_csv", {}).get("rows", [])
    if not tpt_rows:
        print("[scheduling-pass] no timepoint rows — cannot extract scheduling", flush=True)
        return struct_json

    protocol_number = struct_json.get("study_meta", {}).get("protocol_number", "STUDY")
    event_list = [
        {
            "event_oid":       r.get("event", ""),
            "timepoint_label": r.get("timepoint", ""),
            "arm":             r.get("arm", "BOTH"),
        }
        for r in tpt_rows if r.get("event")
    ]

    prompt = f"""You are extracting scheduling data for a clinical trial study.

Study: {protocol_number}

These events were already extracted from the protocol Schedule of Events:
{json.dumps(event_list, indent=2)}

For each event return a JSON array entry with:
- event_oid: (copy from input, unchanged)
- anchor_event_oid: the event OID this visit anchors to. null for the index event (first scheduled visit, typically SE_SCREENING or SE_BASELINE). For all other events, set to whichever event they are measured from.
- offset_target_days: integer days from the anchor. "Day 30" → 30. "Week 4" → 28. "Day 2-6" → 4 (midpoint). null if unclear.
- window_lower_days: lower window bound in days (negative = before target). null if not specified.
- window_upper_days: upper window bound in days. null if not specified.
- repeating: true if this event recurs within the study, false otherwise.
- arm: copy from input exactly.
- conditional_trigger: free text if event is triggered by a clinical event, null if purely calendar-scheduled.

Rules:
1. The first event (lowest visit number) is the index event: anchor_event_oid null, offset_target_days 0.
2. SE_UNSCHEDULED / SE_UNSCH: anchor null, offset null, repeating true, conditional_trigger "Unscheduled visit — triggered by clinical need".
3. SE_COMMON, SE_EOT, SE_EOS, SE_FOLLOWUP: anchor null unless the protocol explicitly links them to another event.
4. Day N labels anchor to the index event unless the protocol specifies otherwise.
5. Use null for any field you cannot determine. Never guess.

Respond with ONLY a valid JSON array — no markdown, no explanation, no preamble.

Example:
[
  {{"event_oid":"SE_SCREENING","anchor_event_oid":null,"offset_target_days":0,"window_lower_days":null,"window_upper_days":null,"repeating":false,"arm":"BOTH","conditional_trigger":null}},
  {{"event_oid":"SE_DAY_30","anchor_event_oid":"SE_SCREENING","offset_target_days":30,"window_lower_days":-3,"window_upper_days":3,"repeating":false,"arm":"BOTH","conditional_trigger":null}}
]"""

    try:
        client = _anthropic.Anthropic()
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        # Strip markdown fences if present
        if text.startswith("```"):
            parts = text.split("```")
            text = parts[1] if len(parts) > 1 else text
            if text.startswith("json"):
                text = text[4:].strip()
        scheduling = json.loads(text)
        struct_json["scheduling"] = scheduling
        print(f"[scheduling-pass] Extracted {len(scheduling)} scheduling entries for {protocol_number}", flush=True)
    except Exception as _exc:
        print(f"[scheduling-pass] Failed ({type(_exc).__name__}: {_exc}) — calendaring will use fallback", flush=True)

    return struct_json


def run_calendaring_rules(struct_json, forms_json):
    """Generate calendaring rules zip. Returns bytes or None if skill not available.

    Tier 1: mechanical Event Action rules from the scheduling block.
    Falls back to timepoint_csv rows with NEEDS_REVIEW flag when scheduling absent.
    Pipeline wiring:
      _want() trigger : dropdown_mm2nc7d4 label ID 7 = "Calendaring Rules"
      output column   : file_mm3te0de  (Calendaring Output)
      update input    : file_mm3tgqeg  (Calendaring Rules Update Input)
    """
    _add_scripts("calendaring-rules")
    try:
        from extract_calendar_rules import extract_calendar_rules
        from validate_rules import validate_rules
        from generate_rule_artifacts import generate_rule_artifacts
    except ImportError as e:
        print(f"Calendaring rules skill not available: {e}", flush=True)
        return None

    build_log = {"build_warnings": []}

    rule_data = extract_calendar_rules(struct_json, forms_json)
    rule_data = validate_rules(rule_data)

    summary = rule_data.get("validation_summary", {})
    print(
        f"Calendaring — {len(rule_data['rules'])} rules, "
        f"{summary.get('passed', 0)} passed, {summary.get('failed', 0)} failed",
        flush=True,
    )
    for w in rule_data.get("warnings", []):
        print(f"  WARNING: {w}", flush=True)

    with tempfile.TemporaryDirectory() as tmp:
        zip_bytes = generate_rule_artifacts(rule_data, tmp, build_log)
        for w in build_log.get("build_warnings", []):
            print(f"  BUILD WARNING: {w}", flush=True)
        return zip_bytes


# ── OpenClinica Study Service API ─────────────────────────────────────────────

async def _get_oc_token(subdomain, is_production=False):
    """
    Get OC auth token via user-service password grant.

    Note: `is_production` is retained for logging/monday-column compatibility
    but no longer affects the URL — all OC traffic goes to openclinica.io
    (the single customer-facing environment).
    """
    import httpx
    username = os.environ.get("OC_API_USERNAME", "").strip()
    password = os.environ.get("OC_API_PASSWORD", "").strip()
    if not username or not password:
        raise ValueError("OC_API_USERNAME or OC_API_PASSWORD not set")
    url = f"https://{subdomain}.build.openclinica.io/user-service/api/oauth/token"
    print(f"Getting OC auth token from {url}", flush=True)
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post(url,
                         headers={"Content-Type": "application/json"},
                         json={"username": username, "password": password})
    if r.status_code != 200:
        raise RuntimeError(f"OC auth failed {r.status_code}: {r.text[:200]}")
    return r.text.strip()


async def _check_study_exists(subdomain, token, protocol_num, is_production=False):
    import httpx
    url = f"https://{subdomain}.build.openclinica.io/study-service/api/studies"
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.get(url,
                        headers={"Authorization": f"Bearer {token}",
                                 "Content-Type": "application/json"},
                        params={"archived": "false", "size": 500})
    if r.status_code != 200:
        return None
    uid = protocol_num[:30].lower()
    for s in r.json():
        if s.get("uniqueIdentifier", "").lower() == uid:
            return s.get("uuid")
    return None


def _build_board_json(struct_json):
    """
    Build a board.json payload for the OpenClinica Study Designer
    from the Study Specification JSON.

    board.json structure:
      lists = Events (one per timepoint row)
      cards = Forms (one per form per event it is assigned to)

    Uses Meteor-style 17-char random IDs generated from the OIDs
    so the import is deterministic and repeatable.

    Deduplicates events by `event` OID before building lists — the Claude
    extraction sometimes emits duplicate rows when protocols contain multiple
    overlapping SOE tables (e.g., detailed injection schedule + summary
    weekly schedule). We keep the FIRST occurrence of each event OID; later
    duplicates are dropped with a log line.
    """
    import hashlib

    def _meteor_id(seed):
        """Generate a stable 17-char alphanumeric ID from a seed string."""
        chars = "23456789ABCDEFGHJKLMNPQRSTWXYZabcdefghijkmnopqrstuvwxyz"
        h = int(hashlib.md5(seed.encode()).hexdigest(), 16)
        result = []
        for _ in range(17):
            result.append(chars[h % len(chars)])
            h //= len(chars)
        return ''.join(result)

    def _form_ocoid(form_id):
        """Board cards reference forms by their OpenClinica-stored OID, which
        carries the F_ prefix OC adds internally (e.g. 'F_AE'). The XLSForm
        settings form_id stays bare per OC contract — only the board card
        formOcoid is prefixed so it matches OC's stored OID. Idempotent."""
        fid = str(form_id or "").strip()
        if not fid:
            return fid
        return fid if fid.upper().startswith("F_") else f"F_{fid}"

    raw_timepoint_rows = struct_json.get("timepoint_csv", {}).get("rows", [])
    forms              = struct_json.get("forms", [])

    # ── Deduplicate events by OID (preserve first-seen order) ─────────────
    seen_oids      = set()
    timepoint_rows = []
    dropped        = []
    for row in raw_timepoint_rows:
        oid = row.get("event", "")
        if not oid:
            continue
        if oid in seen_oids:
            dropped.append(oid)
            continue
        seen_oids.add(oid)
        timepoint_rows.append(row)
    if dropped:
        print(f"_build_board_json: dropped {len(dropped)} duplicate "
              f"event row(s): {sorted(set(dropped))}", flush=True)

    # Build event list (lists)
    lists = []
    event_id_map = {}   # event_oid → meteor _id
    for i, row in enumerate(timepoint_rows):
        event_oid   = row.get("event", f"SE_EVENT{i+1}")
        label       = row.get("timepoint", event_oid)
        # Seed Meteor ID with (oid, index) so any future dedup leak still
        # produces unique IDs rather than silently colliding.
        meteor_id   = _meteor_id(f"{event_oid}|{i}")
        event_id_map[event_oid] = meteor_id

        # Determine if repeating — common events don't have visit windows
        is_repeating = "UNSCH" in event_oid.upper() or "COMMON" in event_oid.upper()
        event_type   = "Common" if is_repeating else "Visit-Based"

        lists.append({
            "_id":         meteor_id,
            "title":       label,
            "sort":        i,
            "eventOcoid":  event_oid,
            "isRepeating": is_repeating,
            "type":        event_type,
        })

    # Build form cards
    cards = []
    card_sort = {}          # event_oid → current sort index
    original_card_id = {}  # form_id → first meteor card _id (for _parentId)

    for form in forms:
        form_id      = form.get("form_id", "")
        if form_id.upper().startswith("F_"):
            print(f"[board-build] WARNING: form_id {form_id!r} has F_ "
                  f"prefix — skipping this form card. Fix the form_id "
                  f"in the study spec JSON.", flush=True)
            continue
        form_title   = form.get("form_title", form_id)
        visits       = form.get("visits_assigned", [])
        first_card   = True

        # Dedup the visits_assigned list too — same root cause can produce
        # a form assigned to the same event twice.
        visits = list(dict.fromkeys(visits))  # preserves order

        for event_oid in visits:
            if event_oid not in event_id_map:
                continue
            list_id  = event_id_map[event_oid]
            sort_idx = card_sort.get(event_oid, 0)
            card_sort[event_oid] = sort_idx + 1

            # Generate stable card ID from form+event+sort to guarantee
            # uniqueness even if some form/event combo somehow repeats.
            card_id  = _meteor_id(f"{form_id}|{event_oid}|{sort_idx}")

            # Repeating flag for THIS card's event (same rule as the lists
            # loop above) — drives the per-card required/SDV properties.
            _card_repeating = ("UNSCH" in event_oid.upper()
                               or "COMMON" in event_oid.upper())
            card = {
                "_id":      card_id,
                # ── Two-step title strategy ───────────────────────────────
                # OC derives the form OID from the card title at import
                # time.  We want F_ICF, not F_INFORMED_CONSENT, so we
                # send the bare form_id ("ICF") as the title during
                # importStudy.  After import, _rename_board_card_titles()
                # updates every card's title to the human-readable
                # form_title ("Informed Consent").  display_title carries
                # that value through the board JSON for the rename pass.
                "title":         form_id,
                "display_title": form_title,
                "listId":   list_id,
                "formOcoid": _form_ocoid(form_id),
                "sort":     sort_idx,
                "required": not _card_repeating,   # Visit-Based True; Common/Unsch False
                "sdv":      "required_item_level",  # Item-Level SDV for all cards
                "itemLevelSdv": True,               # boolean flag on all OC cards
            }

            # First occurrence is the original; subsequent ones reference it
            if first_card:
                original_card_id[form_id] = card_id
                first_card = False
            else:
                card["_parentId"] = original_card_id[form_id]

            cards.append(card)

    return {"labels": [], "lists": lists, "cards": cards}


async def _get_board_id(subdomain, study_uuid, is_production, token=None):
    """
    Get the Study Designer board ID for a newly created study.
    The board ID is embedded in the currentBoardUrl returned by the study-service.
    URL format: https://{subdomain}.design.openclinica.io/b/{boardId}/...
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    base_url = f"https://{subdomain}.build.openclinica.io"
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.get(
            f"{base_url}/study-service/api/studies/{study_uuid}",
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
        )
    if r.status_code != 200:
        raise RuntimeError(f"Could not fetch study details: {r.status_code} {r.text[:200]}")
    data          = r.json()
    board_url     = data.get("currentBoardUrl", "")
    # Extract board ID from URL: .../b/{boardId}/...
    if "/b/" in board_url:
        parts    = board_url.split("/b/")
        board_id = parts[1].split("/")[0]
        print(f"Board ID: {board_id}", flush=True)
        return board_id
    raise RuntimeError(f"Could not extract board ID from URL: {board_url}")


async def _get_board_card_ids(
    subdomain: str, board_id: str,
    is_production: bool = False, token: str = None,
) -> set | None:
    """GET /api/boards/{board_id} and return set of non-archived card _ids.

    Used by create_oc_study to give the publisher a precise filter of
    "cards belonging to the current run" — avoids the publisher walking
    every stale card left in the DOM by prior imports.

    Returns:
        Set of Meteor card _ids (strings like "7a3JP37ytrJ9RN4vF") for
        cards where archived=False; None on API failure (caller should
        fall back to no filter rather than blocking the publish).
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = f"https://{subdomain}.design.openclinica.io/api/boards/{board_id}"
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(url, headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            })
    except Exception as e:
        print(f"[board-card-ids] GET request failed: {e}", flush=True)
        return None
    if r.status_code != 200:
        print(f"[board-card-ids] GET /api/boards/{board_id} returned "
              f"{r.status_code} — skipping filter", flush=True)
        return None
    try:
        data = r.json()
    except Exception as e:
        print(f"[board-card-ids] response not JSON: {e}", flush=True)
        return None
    cards = data.get("cards") or []
    card_ids = {c.get("_id") for c in cards
                if c.get("_id") and not c.get("archived")}
    print(f"[board-card-ids] {len(card_ids)} non-archived card _ids "
          f"({len(cards)} total cards)", flush=True)
    return card_ids if card_ids else None


async def _count_board_cards(
    subdomain: str, board_id: str,
    is_production: bool = False, token: str = None,
) -> int | None:
    """GET /api/boards/{board_id} and return the non-archived card COUNT.

    Unlike _get_board_card_ids (which returns None for an empty board, to
    signal "no filter"), this distinguishes 0 (genuinely empty board) from
    None (API/parse failure). The fast-rerun path uses it to decide whether
    to force a full reimport: an empty board means the prior import never
    populated it, so skipping reimport would leave nothing to publish.
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = f"https://{subdomain}.design.openclinica.io/api/boards/{board_id}"
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(url, headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            })
    except Exception as e:
        print(f"[board-card-count] GET request failed: {e}", flush=True)
        return None
    if r.status_code != 200:
        print(f"[board-card-count] GET /api/boards/{board_id} returned "
              f"{r.status_code}", flush=True)
        return None
    try:
        data = r.json()
    except Exception as e:
        print(f"[board-card-count] response not JSON: {e}", flush=True)
        return None
    cards = data.get("cards") or []
    n = sum(1 for c in cards if not c.get("archived"))
    print(f"[board-card-count] {n} non-archived cards on board {board_id}",
          flush=True)
    return n


async def _get_board_structure(
    subdomain: str, board_id: str,
    is_production: bool = False, token: str = None,
) -> dict:
    """GET /api/boards/{board_id} and return the board's current structure.

    Returns a dict with two keys:
        events: {title: list_id}              — non-archived lists (events)
        cards:  {formOcoid: {"_id","listId"}} — non-archived cards by form OID

    Used to make board import idempotent: a board that already has events is
    reconciled (card property updates only) rather than cleared + reimported,
    which would duplicate SOE events. On API/parse failure returns empty
    dicts so the caller treats it as a fresh board.
    """
    import httpx
    empty = {"events": {}, "cards": {}}
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = f"https://{subdomain}.design.openclinica.io/api/boards/{board_id}"
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(url, headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            })
    except Exception as e:
        print(f"[board-structure] GET request failed: {e}", flush=True)
        return empty
    if r.status_code != 200:
        print(f"[board-structure] GET /api/boards/{board_id} returned "
              f"{r.status_code}", flush=True)
        return empty
    try:
        data = r.json()
    except Exception as e:
        print(f"[board-structure] response not JSON: {e}", flush=True)
        return empty
    events = {}
    for lst in (data.get("lists") or []):
        if lst.get("archived"):
            continue
        title, lid = lst.get("title"), lst.get("_id")
        if title and lid:
            events[title] = lid
    cards = {}
    for c in (data.get("cards") or []):
        if c.get("archived"):
            continue
        foid, cid = c.get("formOcoid"), c.get("_id")
        if foid and cid:
            cards[foid] = {"_id": cid, "listId": c.get("listId")}
    print(f"[board-structure] {len(events)} events, {len(cards)} cards "
          f"on board {board_id}", flush=True)
    return {"events": events, "cards": cards}


async def _reconcile_board_cards(subdomain, board_id, board_json,
                                 existing_structure, session_path,
                                 is_production, token=None):
    """Apply card-property updates (required / sdv / itemLevelSdv) to cards
    that ALREADY exist on the board, via the Meteor client-side Cards.update
    DDP call — the same mechanism oc_form_publisher uses to set _version.

    Only touches cards whose formOcoid is already present in
    existing_structure["cards"]. Missing cards are logged as warnings, not
    created (the card-creation DDP method is unknown). Never clears or
    reimports, so it cannot duplicate SOE events.
    """
    from playwright.async_api import async_playwright

    # Build {existing_card_id: {required, sdv, itemLevelSdv}} from board_json,
    # matched to live cards by formOcoid. Dedupe by formOcoid (multiple cards
    # of the same form share an OID; existing_structure keys by OID too).
    existing_cards = (existing_structure or {}).get("cards", {}) or {}
    updates, missing, seen = {}, [], set()
    # _build_board_json emits cards at the top level; some board payloads
    # nest them under lists[].cards. Support both: prefer the nested form,
    # fall back to the top-level "cards" key so reconcile never iterates an
    # empty list against the current builder output.
    all_cards = [c for lst in board_json.get("lists", []) for c in lst.get("cards", [])]
    all_cards = all_cards or board_json.get("cards", [])
    for card in all_cards:
        foid = card.get("formOcoid")
        if not foid or foid in seen:
            continue
        seen.add(foid)
        match = existing_cards.get(foid)
        if match and match.get("_id"):
            updates[match["_id"]] = {
                "required":     card.get("required"),
                "sdv":          card.get("sdv"),
                "itemLevelSdv": card.get("itemLevelSdv"),
            }
        else:
            missing.append(foid)

    if not updates:
        print(f"[board-reconcile] updated 0 existing cards, {len(missing)} "
              f"cards not found (manual addition required)", flush=True)
        return

    ok = 0
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            ctx  = await browser.new_context(storage_state=session_path)
            page = await ctx.new_page()
            board_url = f"https://{subdomain}.design.openclinica.io/b/{board_id}"
            await page.goto(board_url, wait_until="networkidle", timeout=60000)
            # Wait for the Meteor client to connect so client-initiated
            # Cards.update calls reach the server (board DOM need not render —
            # the update is a server-side DDP method).
            try:
                await page.wait_for_function(
                    "() => typeof Meteor !== 'undefined' "
                    "&& Meteor.status().connected",
                    timeout=20000)
            except Exception:
                print("[board-reconcile] Meteor client not confirmed "
                      "connected — attempting updates anyway", flush=True)
            results = await page.evaluate(
                """
                async (updates) => {
                    const out = {};
                    for (const [cardId, set] of Object.entries(updates)) {
                        try {
                            Cards.update(cardId, {$set: set});
                            out[cardId] = {ok: true};
                        } catch(e) {
                            out[cardId] = {ok: false, error: e.toString()};
                        }
                    }
                    return out;
                }
                """,
                updates,
            )
            ok = sum(1 for v in results.values() if v.get("ok"))
            for cid, v in results.items():
                if not v.get("ok"):
                    print(f"[board-reconcile] Cards.update {cid} failed: "
                          f"{v.get('error')}", flush=True)
        finally:
            await browser.close()

    print(f"[board-reconcile] updated {ok} existing cards, {len(missing)} "
          f"cards not found (manual addition required)", flush=True)


async def _check_board_form_versions(
    subdomain: str, board_id: str,
    is_production: bool = False, token: str = None,
) -> tuple[bool, list]:
    """GET /api/boards/{board_id}, return (all_ok, missing_forms).

    Used by publish_to_test as a pre-flight before calling the publish
    API — that API returns 400 with a list of missing-version forms,
    and catching it here gives the operator a cleaner per-form signal
    on the monday row instead of a raw 400 trace.

    Returns:
        (True, []) if every non-archived form has at least one card with
        a non-empty versions array (i.e. an uploaded XLSForm).
        (False, [(form_name, formOcoid), ...]) listing forms whose cards
        all have empty versions arrays.
        Fail-open: if the API call itself errors, returns (True, []) so
        the caller still attempts publish (and OC surfaces the real
        error if any).
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = f"https://{subdomain}.design.openclinica.io/api/boards/{board_id}"
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(url, headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            })
    except Exception as e:
        print(f"[publish-preflight] GET request failed: {e} — "
              f"skipping pre-flight", flush=True)
        return True, []
    if r.status_code != 200:
        print(f"[publish-preflight] GET /api/boards/{board_id} "
              f"returned {r.status_code} — skipping pre-flight",
              flush=True)
        return True, []
    try:
        data = r.json()
    except Exception as e:
        print(f"[publish-preflight] response not JSON: {e}", flush=True)
        return True, []

    cards = data.get("cards") or []
    # Group non-archived cards by formOcoid; mark form as having a
    # version if ANY of its cards has a non-empty versions array (form
    # definitions are shared across cards of the same form).
    by_form: dict = {}
    for card in cards:
        if card.get("archived"):
            continue
        oid = card.get("formOcoid")
        if not oid:
            continue
        info = by_form.setdefault(oid, {
            "name": card.get("title") or oid,
            "has_version": False,
        })
        if card.get("versions"):
            info["has_version"] = True

    missing = [(info["name"], oid)
               for oid, info in by_form.items()
               if not info["has_version"]]
    return (not missing), missing


async def _get_board_form_oids(
    subdomain: str, board_id: str,
    is_production: bool = False, token: str = None,
) -> set | None:
    """GET /api/boards/{board_id} and return set of non-archived form OIDs.

    Companion to _check_board_form_versions: that helper asks "do forms
    ON the board have versions?"; this one asks "WHICH forms are on the
    board at all?" so the caller can compare against a spec to detect
    forms that should be on the board but never made it (incomplete
    import). Form OIDs are uppercased to match the spec convention.

    Returns set of uppercased formOcoid strings, or None on API failure
    (caller should skip the missing-from-board check rather than block).
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = f"https://{subdomain}.design.openclinica.io/api/boards/{board_id}"
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(url, headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            })
    except Exception as e:
        print(f"[board-form-oids] GET failed: {e}", flush=True)
        return None
    if r.status_code != 200:
        print(f"[board-form-oids] GET /api/boards/{board_id} returned "
              f"{r.status_code}", flush=True)
        return None
    try:
        data = r.json()
    except Exception as e:
        print(f"[board-form-oids] response not JSON: {e}", flush=True)
        return None
    oids = set()
    for card in (data.get("cards") or []):
        if card.get("archived"):
            continue
        oid = card.get("formOcoid")
        if oid:
            oids.add(oid.upper())
    print(f"[board-form-oids] {len(oids)} unique non-archived form OIDs",
          flush=True)
    return oids


# ── Pipeline upload record (per-item storage for conflict detection) ──────────
# Tracks which OC form-version IDs the pipeline created, so on the next run we
# can detect any OC version that we didn't upload → indicates a human edit in
# OC4 Designer (CONFLICT). Persisted to a Railway volume so it survives deploys.

UPLOAD_RECORDS_DIR = "/data/pipeline_upload_records"


def _read_upload_record(item_id: str) -> dict:
    """Read the publisher's per-item upload record from disk.

    Returns {} on missing file or parse error (fail-safe — treats the
    item as first-ever run, prompting full uploads with no conflict
    claims).
    """
    if not item_id:
        return {}
    path = os.path.join(UPLOAD_RECORDS_DIR, f"{item_id}.json")
    try:
        with open(path, "r") as f:
            return json.load(f) or {}
    except FileNotFoundError:
        return {}
    except Exception as e:
        print(f"[upload-record] read failed for {item_id} ({e}); "
              f"treating as empty", flush=True)
        return {}


def _write_upload_record(item_id: str, record: dict) -> None:
    """Persist the publisher's per-item upload record to disk.

    Creates parent directory if missing. Failures are logged but not
    raised — record loss isn't fatal; next run just treats as empty
    record and re-uploads (with possible false-positive conflicts if
    OC has stale versions).
    """
    if not item_id:
        return
    try:
        os.makedirs(UPLOAD_RECORDS_DIR, exist_ok=True)
        path = os.path.join(UPLOAD_RECORDS_DIR, f"{item_id}.json")
        with open(path, "w") as f:
            json.dump(record, f, indent=2)
        print(f"[upload-record] wrote {path}", flush=True)
    except Exception as e:
        print(f"[upload-record] write failed for {item_id} ({e}); "
              f"next run will see stale state", flush=True)


async def _fetch_oc_versions_by_oid(
    subdomain: str, board_id: str,
    is_production: bool = False, token: str = None,
) -> dict:
    """GET /api/boards/{board_id}, return {formOcoid: set[int]}.

    For each non-archived form OID, returns the union of version IDs
    across all of its cards. Cards of the same form share the version
    definitions in practice but we union defensively.

    Returns {} on API failure (caller decides whether to proceed —
    typically by skipping conflict detection rather than blocking).
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = f"https://{subdomain}.design.openclinica.io/api/boards/{board_id}"
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(url, headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            })
    except Exception as e:
        print(f"[oc-versions] GET failed: {e}", flush=True)
        return {}
    if r.status_code != 200:
        print(f"[oc-versions] GET returned {r.status_code}", flush=True)
        return {}
    try:
        data = r.json()
    except Exception as e:
        print(f"[oc-versions] response not JSON: {e}", flush=True)
        return {}

    by_oid: dict = {}
    for card in (data.get("cards") or []):
        if card.get("archived"):
            continue
        oid = (card.get("formOcoid") or "").upper()
        if not oid:
            continue
        version_ids = by_oid.setdefault(oid, set())
        for v in (card.get("versions") or []):
            vid = v.get("id")
            if isinstance(vid, int):
                version_ids.add(vid)
    print(f"[oc-versions] {len(by_oid)} non-archived form OIDs "
          f"({sum(len(v) for v in by_oid.values())} total versions)",
          flush=True)
    return by_oid


async def _clear_board(board_url: str, session_path: str) -> None:
    """Archive all lists on the OC designer board via a Meteor method call.

    Clears the board before reimporting so cards don't accumulate.
    Discovered via DevTools: Meteor.call
        'updateArchiveStatusOfMedicalCodingLayoutItems'
            (boardId, listId, [], '0', true)
    archives one list; iterate over Lists collection client-side to hit
    each list on the board.

    Args:
        board_url:    Full slug URL (e.g. .../b/{boardId}/{slug}); the
                      slug-less form just redirects to the studies list
                      and won't load any minicards.
        session_path: Path to a saved Playwright storage_state JSON for
                      the OC SSO session.
    """
    from playwright.async_api import async_playwright
    board_id = board_url.split("/b/")[1].split("/")[0]

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            ctx  = await browser.new_context(storage_state=session_path)
            page = await ctx.new_page()
            await page.goto(board_url, wait_until="networkidle",
                            timeout=60000)
            # Wait for Meteor client to load + first minicard to render.
            try:
                await page.wait_for_selector('.js-minicard', timeout=20000)
            except Exception:
                print("[board-clear] no minicards found — board may "
                      "already be empty", flush=True)
                return

            # Read list IDs from the Meteor client-side collection.
            # If the collection is named differently or not yet loaded,
            # the try/catch returns [] and we cleanly skip.
            list_ids = await page.evaluate(f"""
                () => {{
                    try {{
                        return Lists.find(
                            {{boardId: '{board_id}', archived: false}}
                        ).fetch().map(l => l._id);
                    }} catch(e) {{
                        return [];
                    }}
                }}
            """)
            print(f"[board-clear] {len(list_ids)} lists to archive",
                  flush=True)

            # Archive each list. Per-list round-trip is slower than a
            # batched call but matches the discovered API surface.
            for list_id in list_ids:
                try:
                    await page.evaluate(f"""
                        () => new Promise((resolve) => {{
                            Meteor.call(
                                'updateArchiveStatusOfMedicalCodingLayoutItems',
                                '{board_id}', '{list_id}', [], '0', true,
                                (err) => resolve(err ? String(err) : 'ok')
                            );
                        }})
                    """)
                except Exception as _e:
                    print(f"[board-clear] archive {list_id} failed: "
                          f"{_e}", flush=True)

            # Wait for server to process all archives. Poll the Meteor
            # Lists collection until non-archived count hits 0, or
            # timeout after 30s. The fixed 3s wait we used before
            # raced the server and let _import_board reimport while
            # archives were still propagating (cards accumulated).
            print(f"[board-clear] waiting for {len(list_ids)} archives "
                  f"to propagate...", flush=True)
            deadline = 30  # seconds
            for _i in range(deadline):
                await page.wait_for_timeout(1000)
                remaining = await page.evaluate(f"""
                    () => {{
                        try {{
                            return Lists.find(
                                {{boardId: '{board_id}', archived: false}}
                            ).count();
                        }} catch(e) {{ return -1; }}
                    }}
                """)
                if remaining == 0:
                    print(f"[board-clear] all lists confirmed archived "
                          f"after {_i+1}s", flush=True)
                    break
                if remaining == -1:
                    print(f"[board-clear] Lists collection unavailable, "
                          f"waiting full {deadline}s", flush=True)
                    await page.wait_for_timeout((deadline - _i - 1) * 1000)
                    break
            else:
                print(f"[board-clear] timeout waiting for archives — "
                      f"{remaining} lists still active", flush=True)

            print(f"[board-clear] archived {len(list_ids)} lists — "
                  f"board cleared", flush=True)
        finally:
            await browser.close()


async def _rename_board_card_titles(subdomain, board_id, board_json,
                                     session_path, is_production):
    """
    After importStudy, update each card's title from the bare form_id
    (e.g. "ICF") to the human-readable display_title (e.g. "Informed
    Consent").

    importStudy uses the card title to derive the form OID, so we send
    form_id as title to get F_ICF.  This pass renames the cards to the
    display title so they look correct in the OC Designer UI.

    Only renames cards where display_title differs from title.
    Non-fatal — a failure here is logged but does not abort the run.
    """
    from playwright.async_api import async_playwright

    # Build a map of card_id → display_title for cards that need renaming
    rename_map = {}
    for card in board_json.get("cards", []):
        dt = card.get("display_title", "")
        t  = card.get("title", "")
        cid = card.get("_id", "")
        if dt and t and dt != t and cid:
            rename_map[cid] = dt

    if not rename_map:
        print("[board-rename] no cards need renaming", flush=True)
        return

    # Only rename the first card per form_id — sibling cards (_parentId
    # present) share the same form definition and OC propagates the title
    # from the parent; renaming them individually is unnecessary.
    seen_titles: set = set()
    filtered_rename: dict = {}
    for card in board_json.get("cards", []):
        cid = card.get("_id", "")
        if cid not in rename_map:
            continue
        if card.get("_parentId"):
            continue   # sibling — skip
        dt = rename_map[cid]
        if dt in seen_titles:
            continue
        seen_titles.add(dt)
        filtered_rename[cid] = dt

    print(f"[board-rename] renaming {len(filtered_rename)} card titles "
          f"via Meteor DDP...", flush=True)

    designer_url = f"https://{subdomain}.design.openclinica.io"
    board_url    = f"{designer_url}/b/{board_id}"

    import json as _json
    session_data = {}
    try:
        with open(session_path) as _sf:
            session_data = _json.load(_sf)
    except Exception as _se:
        print(f"[board-rename] could not read session: {_se} — skipping",
              flush=True)
        return

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        try:
            ctx = await browser.new_context()
            cookies = session_data.get("cookies", [])
            if cookies:
                await ctx.add_cookies(cookies)
            page = await ctx.new_page()
            await page.goto(board_url, wait_until="networkidle",
                            timeout=30000)
            # Wait for Meteor to be available — the board JS bundle loads
            # asynchronously and Meteor is not defined until it does.
            try:
                await page.wait_for_function(
                    "() => typeof Meteor !== 'undefined'",
                    timeout=20000,
                )
                await page.wait_for_timeout(1000)
            except Exception as _mwe:
                print(f"[board-rename] Meteor not available after 20s: "
                      f"{_mwe} — skipping rename", flush=True)
                return

            renamed = 0
            for card_id, display_title in filtered_rename.items():
                try:
                    await page.evaluate(
                        """([cid, newTitle]) => {
                            Meteor.call(
                                '/cards/update',
                                {_id: cid},
                                {$set: {
                                    title: newTitle,
                                    dateLastActivity: {$date: Date.now()}
                                }},
                                {}
                            );
                        }""",
                        [card_id, display_title]
                    )
                    renamed += 1
                except Exception as _re:
                    print(f"[board-rename] failed for card {card_id}: "
                          f"{_re}", flush=True)

            await page.wait_for_timeout(1000)
            print(f"[board-rename] renamed {renamed}/{len(filtered_rename)} "
                  f"cards", flush=True)
        finally:
            await browser.close()


async def _import_board(subdomain, board_id, board_json, is_production, token=None):
    """
    Import the board.json into the study designer.
    POST {designer_url}/api/importStudy/{boardId}
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    designer_url = f"https://{subdomain}.design.openclinica.io"
    endpoint    = f"{designer_url}/api/importStudy/{board_id}"

    print(f"Importing board to: {endpoint}", flush=True)
    async with httpx.AsyncClient(timeout=60) as c:
        r = await c.post(
            endpoint,
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
            json=board_json,
        )
    print(f"Board import: {r.status_code} {r.text[:200]}", flush=True)
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Board import failed {r.status_code}: {r.text[:300]}")
    # Extract the board URL from the response. The design-service import
    # endpoint returns the board-id+slug URL (e.g. /b/wQyCTnJFKjyGMQ9d9/
    # a-biomarker-study-in-...) which is what the OC designer actually
    # resolves to — the UUID-based URL the caller could derive on its
    # own redirects to the studies list instead of loading the board.
    # Field name unverified in the design-service docs; try the likely
    # candidates and let the next run's logs tell us which one fires.
    board_url = ""
    try:
        body = r.json()
        if isinstance(body, dict):
            board_url = (body.get("currentBoardUrl")
                         or body.get("boardUrl")
                         or body.get("url")
                         or body.get("studyBoardUrl")
                         or "")
        elif isinstance(body, str) and body.strip():
            # Design-service import returns a bare JSON string (the
            # URL) — not an object. This is the actually-observed shape.
            board_url = body.strip()
        # The import response includes the hostname already
        # (e.g. "cust1.design.openclinica.io/b/..."), so we only need
        # to prepend the scheme. Building base + sep + path produces
        # a double-hostname URL.
        if board_url and not board_url.startswith("http"):
            board_url = f"https://{board_url}"
    except Exception as e:
        print(f"Board import: response JSON parse failed ({e}); "
              f"caller will fall back to UUID-based study_url.", flush=True)
    print(f"Board import: extracted board_url={board_url!r}", flush=True)

    # importStudy returns 200 as soon as the payload is accepted, but the
    # design service creates the cards asynchronously — an immediate board
    # read has been observed returning only a fraction of the expected cards
    # (e.g. 10 of 121). Poll until the count reaches the expected total or
    # stops climbing, so the caller (and the publish-preflight) see the full
    # board. We only WARN on a shortfall — never raise and never auto-
    # re-import: re-importing a non-empty board duplicates the cards that DID
    # land, and the downstream "MISSING FROM BOARD" preflight already blocks
    # publish for an incomplete board.
    _expected = len(board_json.get("cards") or [])
    if _expected:
        _deadline, _interval, _stable_needed = 120, 5, 4
        _waited, _last, _stable, _n = 0, -1, 0, None
        while _waited <= _deadline:
            _n = await _count_board_cards(
                subdomain, board_id, is_production, token=token)
            if _n is not None and _n >= _expected:
                print(f"[board-import] cards settled: {_n}/{_expected} "
                      f"after ~{_waited}s", flush=True)
                break
            # Detect a stalled (partial) import: a non-zero count that hasn't
            # changed across several reads means the server is done and
            # created fewer cards than we sent. A count of 0 keeps waiting
            # (server may not have started) up to the deadline.
            if _n is not None and _n > 0:
                if _n == _last:
                    _stable += 1
                else:
                    _stable, _last = 1, _n
                if _stable >= _stable_needed:
                    print(f"[board-import] WARNING: card count stalled at "
                          f"{_n}/{_expected} (unchanged for {_stable} reads, "
                          f"~{_waited}s) — import appears PARTIAL. Missing "
                          f"forms will fail the publish-preflight; clear the "
                          f"board and re-import to retry.", flush=True)
                    break
            await asyncio.sleep(_interval)
            _waited += _interval
        else:
            print(f"[board-import] WARNING: card count {_n}/{_expected} after "
                  f"{_deadline}s timeout — board may be incomplete; missing "
                  f"forms will fail the publish-preflight.", flush=True)

    return board_url


async def create_oc_study(subdomain, struct_json, is_production=False,
                          edc_zip_url=None, oc_email=None, item_id=None,
                          fast_rerun=False):
    """
    Create a study in OpenClinica and import the Study Design Board (SOE).

    Steps:
    1. Create study shell via study-service API
       (skips if study already exists)
    2. Build board.json from struct_json (events + forms)
    3. Get the board ID from the newly created study
    4. Import board.json via study designer API
    """
    import httpx
    token    = await _get_oc_token(subdomain, is_production=is_production)
    headers  = {"Authorization": f"Bearer {token}",
                 "Content-Type": "application/json"}
    base_url = f"https://{subdomain}.build.openclinica.io"
    meta     = struct_json.get("study_meta", {})
    protocol_num = meta.get("protocol_number", "STUDY")

    # ── Step 1: Create or find the study ──────────────────────────────────────
    # Only reuse an existing study if the Monday UUID column is populated —
    # a blank UUID column means the operator intentionally reset for a fresh
    # run (cleanup procedure). Skipping _check_study_exists in that case
    # prevents reuse of a study created by a previously-killed pipeline run
    # whose UUID was never written back to Monday before the kill.
    _monday_uuid = ""
    if item_id:
        try:
            _item_check = await get_item(item_id)
            _cols_check = {c["id"]: c for c in (_item_check.get("column_values") or [])}
            _monday_uuid = (_cols_check.get(COL["study_uuid"], {}).get("text") or "").strip()
        except Exception as _ue:
            print(f"[study-create] could not read Monday UUID column ({_ue}); "
                  f"will check OC for existing study", flush=True)

    if _monday_uuid:
        # Monday has a UUID from a prior run — check OC to confirm it still exists
        existing_uuid = await _check_study_exists(subdomain, token, protocol_num,
                                                   is_production=is_production)
    else:
        # Monday UUID is blank — operator reset for a fresh run, always create new
        print("[study-create] Monday UUID column is blank — creating fresh study "
              "(skipping existence check)", flush=True)
        existing_uuid = None

    if existing_uuid:
        print(f"Study already exists (uuid: {existing_uuid}) — skipping creation.", flush=True)
        study_uuid = existing_uuid
    else:
        type_map  = {"interventional": "INTERVENTIONAL", "observational": "OBSERVATIONAL"}
        phase_map = {"phase i": "PHASEI", "phase 1": "PHASEI",
                     "phase ii": "PHASEII", "phase 2": "PHASEII",
                     "phase iii": "PHASEIII", "phase 3": "PHASEIII",
                     "phase iv": "PHASEIV", "phase 4": "PHASEIV"}
        today      = _dt.date.today().isoformat()
        dur_months = int(meta.get("total_study_duration_months", 24) or 24)
        end_date   = (_dt.date.today().replace(
                       year=_dt.date.today().year + dur_months // 12)).isoformat()

        payload = {
            "name":               protocol_num,
            "description":        meta.get("study_title",
                                   meta.get("description",
                                            f"{protocol_num} — {meta.get('indication', '')}")),
            "uniqueIdentifier":   protocol_num[:30],
            "type":               type_map.get(str(meta.get("type","")).lower(),
                                               "INTERVENTIONAL"),
            "phase":              phase_map.get(str(meta.get("study_phase","")).lower().strip(),
                                               "OTHER_NON_IND"),
            "expectedStartDate":  today,
            "expectedEndDate":    end_date,
            "expectedEnrollment": int(meta.get("expected_enrollment", 0) or 0),
            "collectSex":         True,
            "collectDateOfBirth": "ONLY_THE_YEAR",
            "collectPersonId":    "ALWAYS",
        }

        async with httpx.AsyncClient(timeout=60) as c:
            r = await c.post(f"{base_url}/study-service/api/studies",
                             headers=headers, json=payload)
        print(f"OC Study API: {r.status_code} {r.text[:300]}", flush=True)
        if r.status_code not in (200, 201):
            raise RuntimeError(f"OC Study API returned {r.status_code}: {r.text[:300]}")
        study_uuid = r.json().get("uuid", "")
        if not study_uuid:
            raise RuntimeError("Study created but no UUID returned")

    designer_url = f"https://{subdomain}.design.openclinica.io"
    study_url    = f"{designer_url}/b/{study_uuid}"

    # ── Step 2: Build board.json from struct_json ──────────────────────────────
    print("Building board.json from Study Specification...", flush=True)
    board_json = _build_board_json(struct_json)
    print(f"Board: {len(board_json['lists'])} events, "
          f"{len(board_json['cards'])} form cards", flush=True)
    print(f"[board-json] card properties set: required=True for "
          f"Visit-Based events, sdv=required_item_level for all cards",
          flush=True)

    # ── Steps 3 + 4: Board import via design service ──────────────────────────
    # NOTE: /api/importStudy/{boardId} is a CLONE-INTO-EMPTY operation. It
    # returns HTTP 500 {"error":"Copy error"} if the target board already
    # contains events/forms. When the pipeline creates a brand-new study,
    # the board is always empty so this succeeds. When re-running against
    # an existing study (study-already-exists branch above), the board
    # probably has content from the prior run and import will fail —
    # we handle that case gracefully with Option A (skip + log).
    board_imported = False
    board_error    = None
    try:
        board_id = await _get_board_id(subdomain, study_uuid, is_production, token=token)
        # Force a full reimport if a fast-rerun would otherwise skip import
        # but the board is actually EMPTY (0 cards) — e.g. a freshly-created
        # study/board, or one whose prior import failed. Skipping import on
        # an empty board leaves nothing to publish and trips the
        # publish-preflight "MISSING FROM BOARD" for every form. Only
        # downgrade on a CONFIRMED-empty board (count == 0); on a count
        # failure (None) keep fast_rerun as-is (safe default — clone-into-
        # empty on a populated board would duplicate cards).
        if fast_rerun:
            _board_card_n = await _count_board_cards(
                subdomain, board_id, is_production, token=token)
            if _board_card_n == 0:
                print("[board-import] fast-rerun requested but board has 0 "
                      "cards — forcing full reimport", flush=True)
                fast_rerun = False
        if fast_rerun:
            # Skip the import — board already exists from the prior
            # full run. _import_board is a CLONE-INTO-EMPTY op; running
            # it again on a populated board accumulates duplicate form
            # cards (observed: 131 cards vs 69 expected after re-runs).
            print("[board-import] fast-rerun — skipping board reimport, "
                  "using existing board", flush=True)
            board_imported = True
            # Try to read the full board URL (with slug) from the monday
            # row — it was written there by the prior full run. The
            # board-id-only short URL routes to the designer but the
            # Playwright publisher needs the slug form to actually load
            # the form cards. Fall back to short URL if not available.
            _short_url = f"{designer_url}/b/{board_id}"
            try:
                _item = await get_item(item_id) if item_id else None
                _saved_url = ""
                if _item:
                    for _cv in (_item.get("column_values") or []):
                        if _cv.get("id") == COL["oc_study_url"]:
                            _saved_url = (_cv.get("text") or "").strip()
                            break
                if _saved_url and "/b/" in _saved_url:
                    study_url = _saved_url
                    print(f"[board-import] fast-rerun — using saved "
                          f"board URL: {study_url}", flush=True)
                else:
                    study_url = _short_url
                    print(f"[board-import] fast-rerun — no saved URL, "
                          f"using short URL: {study_url}", flush=True)
            except Exception as _ue:
                study_url = _short_url
                print(f"[board-import] fast-rerun — URL read failed "
                      f"({_ue}), using short URL: {study_url}", flush=True)
            # Propagate SDV/required property changes to existing cards on
            # every fast-rerun, so board_json property edits always reach the
            # live board even when the reimport is skipped.
            _session_path = (f"/data/browser_sessions/{oc_email}.json"
                             if oc_email else "")
            if _session_path and os.path.exists(_session_path):
                try:
                    _existing = await _get_board_structure(
                        subdomain, board_id, is_production, token=token)
                    await _reconcile_board_cards(
                        subdomain, board_id, board_json, _existing,
                        _session_path, is_production, token=token)
                except Exception as _rce:
                    print(f"[board-reconcile] fast-rerun reconcile failed "
                          f"(non-fatal): {_rce}", flush=True)
            else:
                print("[board-reconcile] fast-rerun — skipped, no usable "
                      "session", flush=True)
        else:
            existing_structure = await _get_board_structure(
                subdomain, board_id, is_production, token=token)
        if not fast_rerun and existing_structure["events"]:
            print("[board-import] board has existing events — skipping "
                  "clear+reimport to prevent SOE duplication", flush=True)
            board_imported = True
            # Set study_url from the monday saved URL (same logic as the
            # fast-rerun path) — the slug form is needed downstream.
            _short_url = f"{designer_url}/b/{board_id}"
            try:
                _item = await get_item(item_id) if item_id else None
                _saved_url = ""
                if _item:
                    for _cv in (_item.get("column_values") or []):
                        if _cv.get("id") == COL["oc_study_url"]:
                            _saved_url = (_cv.get("text") or "").strip()
                            break
                if _saved_url and "/b/" in _saved_url:
                    study_url = _saved_url
                    print(f"[board-import] reconcile — using saved "
                          f"board URL: {study_url}", flush=True)
                else:
                    study_url = _short_url
                    print(f"[board-import] reconcile — no saved URL, "
                          f"using short URL: {study_url}", flush=True)
            except Exception as _ue:
                study_url = _short_url
                print(f"[board-import] reconcile — URL read failed "
                      f"({_ue}), using short URL: {study_url}", flush=True)
            # Apply required/sdv/itemLevelSdv updates to existing cards.
            _session_path = (f"/data/browser_sessions/{oc_email}.json"
                             if oc_email else "")
            if _session_path and os.path.exists(_session_path):
                try:
                    await _reconcile_board_cards(
                        subdomain, board_id, board_json,
                        existing_structure, _session_path,
                        is_production, token=token)
                except Exception as _rce:
                    print(f"[board-reconcile] failed (non-fatal): {_rce}",
                          flush=True)
            else:
                print("[board-reconcile] skipped — no usable session",
                      flush=True)
        elif not fast_rerun:
            # Clear the board via the Meteor archiveList method before
            # reimporting. _import_board is CLONE-INTO-EMPTY — calling
            # it on a non-empty board appends cards rather than
            # replacing them. The clear requires the slug-form URL
            # (slug-less URLs redirect to the studies list and load no
            # minicards) — read it from monday where the prior run
            # saved it. Skipped (no-op) on a fresh study or when no
            # session is available.
            _board_clear_url = ""
            if item_id:
                try:
                    _ci = await get_item(item_id)
                    for _cv in (_ci.get("column_values") or []):
                        if _cv.get("id") == COL["oc_study_url"]:
                            _board_clear_url = (_cv.get("text") or "").strip()
                            break
                except Exception as _re:
                    print(f"[board-clear] could not read prior board URL: "
                          f"{_re}", flush=True)
            _session_path = (f"/data/browser_sessions/{oc_email}.json"
                             if oc_email else "")
            if (_session_path and os.path.exists(_session_path)
                    and _board_clear_url and "/b/" in _board_clear_url):
                try:
                    await _clear_board(_board_clear_url, _session_path)
                except Exception as _ce:
                    print(f"[board-clear] clear failed (non-fatal): "
                          f"{_ce}", flush=True)
            else:
                print(f"[board-clear] skipped — fresh study or no usable "
                      f"session (url={_board_clear_url!r}, "
                      f"session={bool(_session_path and os.path.exists(_session_path))})",
                      flush=True)
            imported_board_url = await _import_board(
                subdomain, board_id, board_json, is_production, token=token)
            print("Study design board imported successfully.", flush=True)
            board_imported = True

            # ── Two-step title rename ─────────────────────────────────────
            # importStudy used bare form_id as card title to get clean OIDs
            # (F_ICF not F_INFORMED_CONSENT).  Now rename cards to their
            # human-readable display_title via Meteor DDP.
            _rename_session = (f"/data/browser_sessions/{oc_email}.json"
                               if oc_email else "")
            if _rename_session and os.path.exists(_rename_session):
                try:
                    await _rename_board_card_titles(
                        subdomain, board_id, board_json,
                        _rename_session, is_production)
                except Exception as _rne:
                    print(f"[board-rename] non-fatal error: {_rne}",
                          flush=True)
            else:
                print("[board-rename] skipped — no usable session",
                      flush=True)
            # Prefer the import response's board-id+slug URL — the
            # Playwright form-upload flow needs THAT to render the designer.
            # The UUID-based study_url we built above just redirects to the
            # studies list. Fall back to study_url unchanged if the response
            # didn't include a usable URL.
            if imported_board_url:
                study_url = imported_board_url
                # Persist the full slug URL to monday so fast-rerun can
                # read it back without needing to re-import the board.
                # Done here (not just at the run_pipeline call site) so
                # the URL is saved ASAP after import, surviving any
                # downstream crash before control returns to the caller.
                if item_id:
                    try:
                        await set_text(item_id, COL["oc_study_url"],
                                       study_url)
                    except Exception as _wu:
                        print(f"[board-import] failed to save board "
                              f"URL: {_wu}", flush=True)
    except Exception as e:
        board_error = str(e)
        # Classify the failure so the user gets an actionable message
        err_lower = board_error.lower()
        if "copy error" in err_lower or "500" in board_error:
            print(f"Board import failed — target board is not empty.",
                  flush=True)
            print(f"  The OpenClinica design service's importStudy endpoint "
                  f"only works on empty boards. To re-import, open the design "
                  f"board in the UI, delete existing events/forms, then re-run "
                  f"this pipeline.", flush=True)
            board_error = ("Board already has content — manual cleanup required "
                           "before re-import. See the design board URL above.")
        elif "401" in board_error or "unauthorized" in err_lower:
            print(f"Board import failed — authentication rejected (401).",
                  flush=True)
            print(f"  Check OC_API_USERNAME / OC_API_PASSWORD Railway env vars "
                  f"match a valid OpenClinica user account.", flush=True)
        else:
            print(f"Board import failed — {board_error}", flush=True)
            print(f"  Unexpected error — check Railway logs for full traceback.",
                  flush=True)

    # ── Step 5: Upload XLSForm files to create form versions ─────────────────
    # Required before publish_to_test() will succeed; OC errors with
    # "No form version defined" if forms have no published version.
    # Uses Playwright headless Chrome via oc_form_publisher (no REST API
    # exists for form-version upload). Skipped if (a) board import failed
    # — no point publishing forms with no design board — or (b) no EDC
    # ZIP URL was provided (caller didn't fetch it from Monday).
    forms_publish = None
    if board_imported and edc_zip_url and oc_email:
        # Auth-check was moved up to run_pipeline so we can bail before
        # spending 2-3 min on Claude analysis. By the time we get here
        # the session has been validated to exist (or oc_email is empty
        # and we're skipping form upload entirely below).
        try:
            from oc_form_publisher import publish_forms_to_openclinica
            # Fetch the set of card _ids for THIS run via REST. Lets the
            # publisher filter out stale .js-minicard elements left over
            # from prior imports — without this it walks every card on the
            # board (~220 vs ~70 actually-current) and crashes the browser.
            _allowed_card_ids = await _get_board_card_ids(
                subdomain, board_id,
                is_production=is_production, token=token,
            )

            # ── Pre-flight conflict detection ────────────────────────────
            # Compare OC's current version IDs against our stored record
            # of "version IDs the pipeline created". Any OC version we
            # didn't create = human edited in OC4 Designer → CONFLICT.
            # The publisher skips upload for conflict OIDs to avoid
            # overwriting the human's work.
            _upload_record = _read_upload_record(item_id) if item_id else {}
            _oc_versions_before = await _fetch_oc_versions_by_oid(
                subdomain, board_id,
                is_production=is_production, token=token,
            )
            _stored_forms = _upload_record.get("forms") or {}
            _conflict_oids: set = set()
            if not _stored_forms:
                # First-ever run for this item — no baseline to compare
                # against, so we can't distinguish "pipeline uploaded
                # this" from "human uploaded this". Skip conflict
                # detection; the post-publish record-write below
                # establishes the baseline for future runs.
                print(f"[conflict-detect] No stored upload record for "
                      f"item {item_id} — skipping conflict detection "
                      f"on first run, establishing baseline", flush=True)
            else:
                # Conflict semantics: only flag when the pipeline HAS
                # a record for this OID and OC has at least one version
                # ID outside that record. When the pipeline has no
                # record for an OID at all (fresh upload record, admin
                # cleared it, etc.), treat OC's existing version as
                # unmanaged — upload fresh rather than flagging.
                # Previously the absent-record case auto-flagged every
                # OC version as a conflict, which made the conflict
                # detector over-aggressive across re-runs and after
                # record resets.
                for _oid, _oc_vids in _oc_versions_before.items():
                    if _oid not in _stored_forms:
                        continue
                    _stored_vids = set(_stored_forms[_oid]
                                       .get("pipeline_version_ids", []))
                    if _oc_vids - _stored_vids:
                        _conflict_oids.add(_oid)
                if _conflict_oids:
                    print(f"[conflict-detect] {len(_conflict_oids)} "
                          f"form OID(s) flagged as conflicts (OC has "
                          f"versions the pipeline didn't create): "
                          f"{sorted(_conflict_oids)}", flush=True)

            print(f"Uploading XLSForm files to {study_url} via Playwright "
                  f"(SSO as {oc_email})...", flush=True)
            if item_id:
                try:
                    await set_status(item_id, COL["pipeline_status"],
                                     "Uploading Forms")
                    await append_log(item_id,
                        f"Form upload starting — "
                        f"{len(_allowed_card_ids)} cards / "
                        f"{len(set(_conflict_oids))} conflicts skipped")
                except Exception:
                    pass
            forms_publish = await publish_forms_to_openclinica(
                study_url=study_url,
                edc_zip_url=edc_zip_url,
                auth_token=token,
                user_email=oc_email,
                allowed_card_ids=_allowed_card_ids,
                conflict_oids=_conflict_oids if _conflict_oids else None,
                item_id=item_id,
            )
            print(f"Form publish: {forms_publish.forms_uploaded}/"
                  f"{forms_publish.forms_total} uploaded; "
                  f"errors={len(forms_publish.errors)}  "
                  f"conflicts={len(forms_publish.conflicts)}", flush=True)
            for err in forms_publish.errors[:5]:
                print(f"  form-upload error: {err}", flush=True)
            for _conf in forms_publish.conflicts[:5]:
                print(f"  conflict: {_conf}", flush=True)
            # Log upload result to Monday
            if item_id:
                try:
                    _up = forms_publish.forms_uploaded
                    _tot = forms_publish.forms_total
                    _errs = len(forms_publish.errors)
                    _confs = len(forms_publish.conflicts)
                    _msg = (f"Form-version upload: {_up}/{_tot} succeeded"
                            + (f"; {_confs} conflict(s) skipped" if _confs else "")
                            + (f"; {_errs} error(s)" if _errs else ""))
                    await append_log(item_id, _msg)
                    if _errs > 0 and _up == 0:
                        await append_log(item_id,
                            f"Form upload FAILED: "
                            + "; ".join(forms_publish.errors[:3]))
                except Exception:
                    pass

            # ── Post-publish: update stored upload record ────────────────
            # For each OID the publisher uploaded, OVERWRITE the stored
            # set with OC's current version IDs (pipeline now "owns" all
            # current versions of that OID). For conflict OIDs we DO NOT
            # update the stored set — the conflict persists on the next
            # run until the human resolves it.
            if item_id and forms_publish.uploaded_oids:
                try:
                    # OC's REST API has propagation lag — a fetch
                    # immediately after the publisher returns
                    # observed only ~12 of 23 forms' versions on the
                    # last run, so the stored record ended up with
                    # empty pipeline_version_ids for most OIDs and
                    # the next run over-flagged conflicts. 12s gives
                    # OC time to surface the newly-uploaded versions
                    # before we snapshot.
                    print(f"[upload-record] waiting 12s for OC version "
                          f"propagation before snapshot...", flush=True)
                    await asyncio.sleep(12)
                    _oc_versions_after = await _fetch_oc_versions_by_oid(
                        subdomain, board_id,
                        is_production=is_production, token=token,
                    )
                    _stored_forms = _upload_record.setdefault("forms", {})
                    _now = _dt.datetime.utcnow().strftime(
                        "%Y-%m-%dT%H:%M:%SZ")
                    for _oid in forms_publish.uploaded_oids:
                        _current = sorted(
                            _oc_versions_after.get(_oid.upper(), set()))
                        _stored_forms[_oid.upper()] = {
                            "pipeline_version_ids": _current,
                            "last_uploaded_at": _now,
                        }
                    _upload_record["last_updated"] = _now
                    _upload_record["item_id"] = str(item_id)
                    # Save a hash of the Study Spec JSON so future runs can
                    # detect whether the spec changed and skip re-upload if not.
                    if struct_json:
                        import hashlib as _hashlib
                        _spec_bytes = json.dumps(struct_json, sort_keys=True,
                                                 ensure_ascii=False).encode()
                        _upload_record["spec_hash"] = (
                            _hashlib.sha256(_spec_bytes).hexdigest()[:16])
                    _write_upload_record(item_id, _upload_record)
                except Exception as _ue:
                    print(f"[upload-record] post-publish update failed: "
                          f"{_ue}", flush=True)

            # ── Surface conflicts on the monday row ──────────────────────
            if item_id and forms_publish.conflicts:
                _confs = forms_publish.conflicts[:10]
                _more = (f" (+{len(forms_publish.conflicts) - 10} more)"
                         if len(forms_publish.conflicts) > 10 else "")
                try:
                    await append_log(item_id,
                        f"⚠️ Manual edit conflicts detected — pipeline "
                        f"did not overwrite: "
                        f"{', '.join(_confs)}{_more}. Review in OC4 "
                        f"Designer and re-run or manually resolve.")
                except Exception as _le:
                    print(f"[conflict-log] append_log failed: {_le}",
                          flush=True)
        except Exception as e:
            # Don't fail study creation if form publish errors; caller
            # surfaces the partial state via the return dict.
            print(f"Form publish raised: {type(e).__name__}: {e}", flush=True)
            print(traceback.format_exc(), flush=True)
            forms_publish = None
    elif not edc_zip_url:
        print(f"Form publish skipped — no edc_zip_url passed to "
              f"create_oc_study.", flush=True)
    elif not oc_email:
        print(f"Form publish skipped — no OpenClinica Email "
              f"(COL[oc_email] / emailothn6i3m) set on this monday row. "
              f"User must populate that column with their OC SSO email "
              f"before form upload can run; study + design board still "
              f"created successfully.", flush=True)

    # Return a dict so callers can surface both the URL and the import state
    return {
        "study_url":      study_url,
        "study_uuid":     study_uuid,    # written to COL["study_uuid"] by caller; used by publish_to_test
        "board_imported": board_imported,
        "board_error":    board_error,
        # FormPublishResult dict or None if skipped/raised. Caller can
        # log via append_log(item_id, ...) — we don't have item_id here.
        "forms_publish":  forms_publish.to_dict() if forms_publish else None,
    }


# ── Publish-to-Test workflow (button webhook → main.py → publish_to_test) ────
#
# Triggered when the user clicks the "Publish to Test" button on a row in
# Monday. main.py's /webhook/monday dispatches button-click events here.
#
# Flow:
#   1. read oc_subdomain + study_uuid from monday   (study_uuid is written
#      to COL["study_uuid"] by create_oc_study during the main pipeline)
#   2. GET /api/studies/{uuid}/study-environments → (env_uuid, oid)
#   3. POST /api/studies/{uuid}/study-versions with studyEnvironmentUuid
#   4. update Study OID + Published Status columns
#
# Status transitions on the Published Status column:
#   Publishing → Published   (happy path)
#   Publishing → Failed      (any exception)
#
# publish_to_test() catches all exceptions and writes them to ai_run_log +
# Published Status. It never raises — the webhook should always return.


async def _get_study_environment_uuid(
    subdomain,
    study_uuid,
    env_name="Test",
    is_production=False,
    token=None,
):
    """GET /api/studies/{study_uuid}/study-environments and return
    (env_uuid, oid) for the entry whose environmentName matches env_name
    (case-insensitive).

    Returns a 2-tuple so the caller can populate both the studyEnvironmentUuid
    input to /study-versions AND the Study OID monday column from one call.

    Raises RuntimeError if env_name isn't found, with the list of envs
    that WERE returned — actionable for the operator.
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = (f"https://{subdomain}.build.openclinica.io"
           f"/study-service/api/studies/{study_uuid}/study-environments")
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.get(url, headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/json",
        })
    if r.status_code != 200:
        raise RuntimeError(
            f"GET /study-environments failed: {r.status_code} {r.text[:300]}"
        )
    envs = r.json()
    if not isinstance(envs, list):
        raise RuntimeError(
            f"GET /study-environments returned non-list: "
            f"{type(envs).__name__} — {str(envs)[:200]}"
        )

    target = env_name.strip().lower()
    for env in envs:
        if (env.get("environmentName") or "").strip().lower() == target:
            env_uuid = env.get("uuid")
            oid      = env.get("oid") or ""
            if not env_uuid:
                raise RuntimeError(
                    f"Environment {env_name!r} found but has no uuid field. "
                    f"Full entry: {env}"
                )
            return env_uuid, oid

    available = [e.get("environmentName", "?") for e in envs]
    raise RuntimeError(
        f"No environment named {env_name!r} found for study {study_uuid}. "
        f"Available: {available}"
    )


async def _publish_study_version(
    subdomain,
    study_uuid,
    study_environment_uuid,
    version_name=None,
    description=None,
    is_production=False,
    token=None,
):
    """POST /api/studies/{study_uuid}/study-versions.

    Body:  {"studyEnvironmentUuid": ..., "versionName": ..., "description": ...}
    Returns: the parsed StudyVersion response dict.

    Raises RuntimeError on non-2xx with the response body in the message
    so the caller can write a meaningful Published Status note.
    """
    import httpx
    if token is None:
        token = await _get_oc_token(subdomain, is_production=is_production)
    url = (f"https://{subdomain}.build.openclinica.io"
           f"/study-service/api/studies/{study_uuid}/study-versions")
    body = {"studyEnvironmentUuid": study_environment_uuid}
    if version_name:
        body["versionName"] = version_name
    if description:
        body["description"] = description

    # Retry on the "No form version defined" 400 — OC's REST API can take
    # tens of seconds to surface form versions uploaded moments earlier,
    # so the publish call races against propagation. Any other 400 (or
    # other non-2xx) raises immediately — those are real failures.
    max_retries = 3
    retries_used = 0
    while True:
        async with httpx.AsyncClient(timeout=180) as c:
            r = await c.post(url, headers={
                "Authorization": f"Bearer {token}",
                "Content-Type":  "application/json",
            }, json=body)
        print(f"OC Publish API: {r.status_code} {r.text[:300]}", flush=True)
        if r.status_code in (200, 201):
            return r.json()
        if (r.status_code == 400
                and "No form version defined" in r.text
                and retries_used < max_retries):
            retries_used += 1
            print(f"[publish] version propagation retry "
                  f"{retries_used}/{max_retries} — waiting 30s...",
                  flush=True)
            await asyncio.sleep(30)
            continue
        raise RuntimeError(
            f"Publish failed: {r.status_code} {r.text[:300]}"
        )


async def _activate_test_environment(subdomain, study_uuid):
    """Activate the TEST environment after publish by PUTting the full
    StudyEnvironmentDTO with status="AVAILABLE".

    Steps:
      1. GET /study-service/api/studies/{study_uuid}/study-environments
         to find the env where environmentType == "TEST" or
         environmentName == "Test".
      2. Set its status to "AVAILABLE".
      3. PUT /study-service/api/study-environments with the full modified
         object as JSON body (UUID is carried inside the body, not the URL).

    Returns the PUT response JSON. Raises RuntimeError on any non-2xx.
    Uses Bearer token via _get_oc_token (same pattern as publish_to_test).
    """
    import httpx
    token = await _get_oc_token(subdomain, is_production=False)
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
    }
    base = f"https://{subdomain}.build.openclinica.io/study-service"

    # ── Step 1: GET the environments and locate the TEST one ──────────────
    get_url = f"{base}/api/studies/{study_uuid}/study-environments"
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.get(get_url, headers=headers)
    if r.status_code != 200:
        raise RuntimeError(
            f"Activate TEST: GET environments failed "
            f"{r.status_code}: {r.text[:300]}"
        )
    envs = r.json() or []
    test_env = None
    for env in envs:
        env_type = (env.get("environmentType") or "").upper()
        env_name = (env.get("environmentName") or "").upper()
        if env_type == "TEST" or env_name == "TEST":
            test_env = env
            break
    if test_env is None:
        names = [e.get("environmentName") for e in envs]
        raise RuntimeError(
            f"Activate TEST: no TEST environment found for study "
            f"{study_uuid}. Got: {names}"
        )

    # ── Step 2: Mutate status on the full object ──────────────────────────
    test_env["status"] = "AVAILABLE"

    # ── Step 3: PUT the modified object back ──────────────────────────────
    put_url = f"{base}/api/study-environments"
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.put(put_url, headers=headers, json=test_env)
    if r.status_code not in (200, 201, 204):
        raise RuntimeError(
            f"Activate TEST: PUT failed {r.status_code}: {r.text[:300]}"
        )
    try:
        return r.json()
    except Exception:
        return {"status_code": r.status_code, "body": r.text[:300]}


async def publish_calendaring_rules(subdomain, study_uuid, cal_zip_bytes, is_production=False):
    """POST validated calendaring rules to the OC4 rule-service.

    Idempotent — GETs existing rules first and skips any whose name already
    exists. Returns a summary dict with keys: uploaded, skipped, failed, errors.

    Endpoint: https://{subdomain}.build.openclinica.io/rule-service/api/studies/{study_uuid}/rules
    Auth:     Bearer token from _get_oc_token (same as publish_to_test)
    Body:     Raw rule JSON (exactly what generate_rule_artifacts writes to rules/*.json)
    """
    import zipfile, io as _io
    import httpx

    base_url = (
        f"https://{subdomain}.build.openclinica.io"
        f"/rule-service/api/studies/{study_uuid}/rules"
    )
    token = await _get_oc_token(subdomain, is_production=is_production)
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
    }

    # Step 1: GET existing rules to build idempotency set
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.get(base_url, headers=headers)
    if r.status_code != 200:
        raise RuntimeError(
            f"Failed to fetch existing rules: {r.status_code} {r.text[:200]}"
        )
    existing_rules = {rule["name"]: rule["uuid"] for rule in r.json()}
    print(f"[cal-publish] {len(existing_rules)} rules already on study", flush=True)

    # Step 2: Extract rule JSONs from zip
    try:
        zf = zipfile.ZipFile(_io.BytesIO(cal_zip_bytes))
        rule_files = [n for n in zf.namelist() if n.startswith("rules/") and n.endswith(".json")]
    except Exception as exc:
        raise RuntimeError(f"Could not read calendaring zip: {exc}")

    uploaded, skipped, failed, errors = 0, 0, 0, []

    for rule_file in sorted(rule_files):
        rule_json = json.loads(zf.read(rule_file).decode("utf-8"))
        rule_name = rule_json.get("name", rule_file)

        if rule_name in existing_rules:
            # Rule exists — update it via PUT
            rule_uuid = existing_rules[rule_name]
            try:
                async with httpx.AsyncClient(timeout=30) as c:
                    r = await c.put(
                        f"{base_url}/{rule_uuid}",
                        headers=headers,
                        json=rule_json,
                    )
                if r.status_code in (200, 201):
                    print(f"[cal-publish] UPDATE {rule_name} uuid={rule_uuid}", flush=True)
                    uploaded += 1
                else:
                    print(f"[cal-publish] FAIL UPDATE {rule_name}: {r.status_code} {r.text[:200]}", flush=True)
                    failed += 1
                    errors.append(f"{rule_name} (update): {r.status_code} {r.text[:200]}")
            except Exception as exc:
                print(f"[cal-publish] ERROR UPDATE {rule_name}: {exc}", flush=True)
                failed += 1
                errors.append(f"{rule_name} (update): {exc}")
            continue

        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.post(
                    base_url,
                    params={"newEpochOrCalendar": "false"},
                    headers=headers,
                    json=rule_json,
                )
            if r.status_code in (200, 201):
                rule_uuid = r.json().get("uuid", "?")
                print(f"[cal-publish] OK {rule_name} uuid={rule_uuid}", flush=True)
                uploaded += 1
                existing_rules[rule_name] = r.json().get("uuid", "")
            else:
                print(f"[cal-publish] FAIL {rule_name}: {r.status_code} {r.text[:200]}", flush=True)
                failed += 1
                errors.append(f"{rule_name}: {r.status_code} {r.text[:200]}")
        except Exception as exc:
            print(f"[cal-publish] ERROR {rule_name}: {exc}", flush=True)
            failed += 1
            errors.append(f"{rule_name}: {exc}")

    return {"uploaded": uploaded, "skipped": skipped, "failed": failed, "errors": errors}


async def publish_to_test(item_id, uploaded_oids=None):
    """Entry point invoked by main.py's safe_run_publish background task.

    See the section comment above for the full flow. This function never
    raises — every failure path is captured into Published Status="Failed"
    and append_log() so the operator sees what went wrong on the monday row.

    Args:
        item_id: monday item id (string-coerced).
        uploaded_oids: Optional iterable of form OIDs that were just
            uploaded by the publisher in THIS session. Used to suppress
            false "missing version" reports in the pre-flight when OC's
            REST API hasn't yet propagated the newly-uploaded versions.
            Pass None (default) when calling from a context where no
            recent upload happened (e.g. webhook button click much
            later) — propagation delay is irrelevant by then.
    """
    item_id = str(item_id)
    # Safe defaults — may be overwritten inside the try block. Used in the
    # except block's ReadTimeout verify path; if they were never set the
    # verify attempt is skipped and we fall through to marking Failed.
    oc_subdomain = ""
    study_uuid   = ""

    try:
        # 1. Show we're working on it
        await set_status(item_id, COL["published_status"], "Publishing")
        await append_log(item_id, "Publish to Test: starting")

        # 2. Read inputs from monday
        item = await get_item(item_id)
        cols = {cv["id"]: cv for cv in (item.get("column_values") or [])}
        oc_subdomain = (cols.get(COL["oc_subdomain"], {}).get("text") or "").strip()
        study_uuid   = (cols.get(COL["study_uuid"],   {}).get("text") or "").strip()

        if not oc_subdomain:
            raise RuntimeError(
                "OC Subdomain is empty. Set it on this row before clicking "
                "Publish to Test.")
        if not study_uuid:
            raise RuntimeError(
                "Study UUID is empty. Run the main pipeline (Send to AI) "
                "first to create the study in OpenClinica — that step "
                "writes the UUID to this column.")

        await append_log(item_id, f"Publish to Test: study_uuid={study_uuid}")

        # 4. Resolve env uuid + oid via /study-environments
        env_uuid, oid = await _get_study_environment_uuid(
            oc_subdomain, study_uuid,
            env_name="Test", is_production=False,
        )
        await append_log(item_id,
            f"Publish to Test: env_uuid={env_uuid}  oid={oid!r}")

        # 5. Persist the Study OID column (cheap bonus from the same call)
        if oid:
            await set_text(item_id, COL["study_oid"], oid)

        # 5b. Pre-flight checks — verify board is ready for publish.
        # Two distinct failure modes the OC publish API would otherwise
        # report as a raw 400; catching them here gives the operator a
        # cleaner per-form signal on the monday row:
        #   (1) MISSING FROM BOARD — form is in the study spec but no
        #       card exists for it on the board (incomplete import).
        #       Detected by comparing spec form_ids against board.
        #   (2) MISSING VERSION   — form has a card on the board but no
        #       XLSForm was uploaded for it (form-publish never ran or
        #       partially failed). Detected via the existing helper.
        # Both block publish; both get reported in the AI Run Log.
        try:
            _board_id = await _get_board_id(
                oc_subdomain, study_uuid, is_production=False)

            # Expected OIDs (from spec) — best-effort. If the spec
            # download or parse fails we just skip the missing-from-
            # board check, falling back to the existing board-only
            # version check.
            _expected_oid_to_name: dict = {}
            try:
                _spec_bytes = await download_column_file(
                    item_id, COL["spec_json"])
                if _spec_bytes:
                    _spec = json.loads(_spec_bytes.decode("utf-8"))
                    for _f in (_spec.get("forms") or []):
                        _foid = (_f.get("form_id") or "").strip()
                        if not _foid:
                            continue
                        # Match _build_board_json's F_-prefix filter:
                        # those forms are intentionally never imported,
                        # so reporting them as missing would be noise.
                        if _foid.upper().startswith("F_"):
                            continue
                        _expected_oid_to_name[_foid.upper()] = (
                            _f.get("form_title") or _foid)
            except Exception as _se:
                print(f"[publish-preflight] could not read spec_json "
                      f"({_se}) — falling back to board-only check",
                      flush=True)
                _expected_oid_to_name = {}

            # Missing-from-board check (only if we have expected OIDs).
            _missing_from_board: list = []
            if _expected_oid_to_name:
                _board_oids = await _get_board_form_oids(
                    oc_subdomain, _board_id, is_production=False)
                if _board_oids is not None:
                    # Board cards carry the F_-prefixed OID (e.g. F_SLEEP)
                    # that OC stores; the spec uses the bare OID (SLEEP).
                    # Strip a leading F_ (case-insensitive) from BOTH sides
                    # before diffing so the prefix alone never reads as
                    # "missing from board".
                    _strip_f = lambda o: o[2:] if o.upper().startswith("F_") else o
                    _board_bare = {_strip_f(_o) for _o in _board_oids}
                    _missing_from_board = sorted(
                        {_strip_f(_o) for _o in _expected_oid_to_name}
                        - _board_bare)

            # Missing-version check (existing helper, unchanged sig).
            _all_versions_ok, _missing_versions = (
                await _check_board_form_versions(
                    oc_subdomain, _board_id, is_production=False))

            # Trust just-uploaded OIDs: if the caller told us which
            # forms the publisher uploaded in this session, suppress
            # any "missing version" reports for those — OC's REST API
            # has propagation delay and may not yet show the new
            # versions even though the publisher confirmed upload.
            if uploaded_oids:
                _trusted = {oid.upper() for oid in uploaded_oids}
                _before = len(_missing_versions)
                _missing_versions = [
                    (fn, foid) for fn, foid in _missing_versions
                    if (foid or "").upper() not in _trusted
                ]
                _suppressed = _before - len(_missing_versions)
                if _suppressed:
                    print(f"[publish-preflight] Suppressed "
                          f"{_suppressed} missing-version report(s) "
                          f"for OIDs uploaded this session "
                          f"(REST API propagation delay)",
                          flush=True)

            # Trust just-uploaded OIDs for missing-from-board check:
            # When getForm runs, it rewrites the card's formOcoid from the
            # spec's short stem (e.g. AE, DM) to the form-service OID
            # (e.g. F_ADVERSEEVENT, F_DEMOGRAPHICS). The board then reports
            # long OIDs; the spec still has short stems — so every uploaded
            # form falsely appears "missing from board" when compared by name.
            # Since the publisher confirms each upload via DDP before adding
            # to uploaded_oids, we trust it entirely: if any forms were
            # uploaded this session, suppress ALL missing-from-board reports.
            # Substring matching (startswith) cannot cover cases like
            # AE → ADVERSEEVENT where neither is a prefix of the other.
            if uploaded_oids and _missing_from_board:
                _before = len(_missing_from_board)
                _missing_from_board = []
                print(f"[publish-preflight] Suppressed "
                      f"{_before} missing-from-board report(s) — "
                      f"forms confirmed uploaded via DDP this session "
                      f"(form-service OID rewrite)", flush=True)

            # Emit per-form log lines for each category.
            for _foid in _missing_from_board:
                _fname = _expected_oid_to_name.get(_foid, _foid)
                print(f"[publish-preflight] MISSING FROM BOARD: "
                      f"{_fname} (OID={_foid})", flush=True)
            for _fname, _foid in _missing_versions:
                print(f"[publish-preflight] MISSING VERSION: "
                      f"{_fname} (OID={_foid})", flush=True)

            # If either category has findings, skip the publish.
            if _missing_from_board or _missing_versions:
                _parts: list = []
                if _missing_from_board:
                    _from_board = ", ".join(_missing_from_board[:10])
                    if len(_missing_from_board) > 10:
                        _from_board += (
                            f" (+{len(_missing_from_board) - 10} more)")
                    _parts.append(f"missing from board: {_from_board}")
                if _missing_versions:
                    _v_oids = [oid for _, oid in _missing_versions]
                    _versions = ", ".join(_v_oids[:10])
                    if len(_v_oids) > 10:
                        _versions += f" (+{len(_v_oids) - 10} more)"
                    _parts.append(f"missing versions: {_versions}")
                _log_msg = "Publish skipped — " + " | ".join(_parts)
                print(f"[publish-preflight] {_log_msg}", flush=True)
                await set_status(item_id, COL["published_status"],
                                 "Failed")
                await append_log(item_id, _log_msg)
                return
        except Exception as _pe:
            # Pre-flight itself failed — don't block publish; if the
            # real issue surfaces in the API call we'll handle it there.
            print(f"[publish-preflight] check failed ({_pe}); "
                  f"proceeding with publish anyway", flush=True)

        # 6. Publish
        ts = _dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        version_name = f"UAT-{ts}"
        response = await _publish_study_version(
            oc_subdomain, study_uuid, env_uuid,
            version_name=version_name,
            is_production=False,
        )
        await append_log(item_id,
            f"Publish to Test: success — version={version_name}  "
            f"response_keys={list(response.keys())[:6]}")

        # 7. Mark as published
        await set_status(item_id, COL["published_status"], "Published")

    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        print(f"PUBLISH_TO_TEST FAILED for item {item_id}: {err}", flush=True)
        print(traceback.format_exc(), flush=True)
        # Best-effort error reporting; do NOT raise — webhook must return.
        try:
            # ReadTimeout means OC accepted the request but didn't respond
            # in time — it may have published successfully in the background.
            # Wait 30s then verify OC's actual study status before marking
            # Failed, so a slow-but-successful publish isn't mis-reported.
            import httpx as _httpx
            _verified_failed = True
            if isinstance(e, _httpx.ReadTimeout) and oc_subdomain and study_uuid:
                print("[publish] ReadTimeout — waiting 30s then verifying "
                      "OC study status before marking Failed...", flush=True)
                await asyncio.sleep(30)
                try:
                    _token = await _get_oc_token(oc_subdomain)
                    _ver_url = (f"https://{oc_subdomain}.build.openclinica.io"
                                f"/study-service/api/studies/{study_uuid}")
                    async with _httpx.AsyncClient(timeout=30) as _c:
                        _r = await _c.get(_ver_url, headers={
                            "Authorization": f"Bearer {_token}"})
                    if _r.status_code == 200:
                        _sdata = _r.json()
                        _penv = (_sdata.get("publishedEnvironmentType") or
                                 _sdata.get("publishedEnvType") or "")
                        if _penv and _penv != "NOT_PUBLISHED":
                            print(f"[publish] OC confirms study IS published "
                                  f"({_penv}) — overriding ReadTimeout failure",
                                  flush=True)
                            _verified_failed = False
                            await set_status(item_id,
                                             COL["published_status"],
                                             "Published")
                            await append_log(item_id,
                                "Publish to Test: succeeded (confirmed via "
                                "OC study API after ReadTimeout).")
                except Exception as _ve:
                    print(f"[publish] OC verify-after-timeout failed: "
                          f"{_ve} — marking Failed", flush=True)
            if _verified_failed:
                await set_status(item_id, COL["published_status"], "Failed")
                await append_log(item_id, f"Publish to Test FAILED: {err}")
        except Exception as inner:
            print(f"PUBLISH_TO_TEST status-update fallback also failed: "
                  f"{inner}", flush=True)


# ── Load-DVS-UAT-Data workflow (checkbox webhook → main.py → load_dvs_uat_data)
#
# Reads the DVS XLSX off the monday row, parses it via the oc-uat-runner
# package (uat_runner.parsers.dvs_parser), generates CDISC ODM 1.3.2 XML
# (uat_runner.generators.odm_generator), and POSTs it to OpenClinica via
# uat_runner.api.oc_client.OpenClinicaClient.import_odm_data.
#
# The OpenClinicaClient is synchronous (built on `requests`); we wrap its
# network call in asyncio.to_thread so we don't block the event loop.
#
# Fail-fast: if the study isn't published (study_oid empty), surface a
# friendly "click Publish to Test first" error instead of polling. The
# user can re-check the box after publish completes.

async def load_dvs_uat_data(item_id):
    """Entry point invoked by main.py's safe_run_load_dvs_uat_data task.

    FULL UAT WORKFLOW:
    1. Validate inputs (study published + DVS available)
    2. Create UAT site (or reuse existing if site_oid already set)
    3. Create test participants (UAT-001, UAT-002, etc.)
    4. Parse DVS and generate ODM XML
    5. Import ODM data into OpenClinica
    6. Retrieve clinical data for validation
    7. Generate reports:
       - Validation Traceability Matrix (XLSX)
       - Validation Summary Report (PDF)
       - Updated DVS with results (XLSX)
    8. Upload reports to Monday.com

    Never raises — all failures are captured into append_log() so the
    operator sees them on the monday row.
    """
    item_id = str(item_id)

    try:
        await append_log(item_id, "Load DVS UAT Data: starting")

        # 1. Read inputs from monday
        item = await get_item(item_id)
        cols = {cv["id"]: cv for cv in (item.get("column_values") or [])}
        oc_subdomain = (cols.get(COL["oc_subdomain"], {}).get("text") or "").strip()
        study_uuid   = (cols.get(COL["study_uuid"],   {}).get("text") or "").strip()
        study_oid    = (cols.get(COL["study_oid"],    {}).get("text") or "").strip()

        if not oc_subdomain:
            raise RuntimeError(
                "OC Subdomain is empty. Set it on this row before loading DVS UAT data.")
        if not study_uuid:
            raise RuntimeError(
                "Study UUID is empty. Run the main pipeline (Send to AI) first "
                "to create the study in OpenClinica — that step writes the UUID "
                "to this column.")
        if not study_oid:
            raise RuntimeError(
                "Study OID is empty — study is not published yet. Click "
                "'Publish to Test' first and wait for it to complete, then "
                "re-check the 'Load DVS UAT Data' box.")

        await append_log(item_id,
            f"Load DVS UAT Data: study_uuid={study_uuid}  study_oid={study_oid}")

        # 2. Fetch the DVS XLSX — prefer the pipeline-generated dvs_output,
        # fall back to a human-uploaded dvs_input. download_column_file
        # returns None for an empty column.
        dvs_source = None
        dvs_bytes  = await download_column_file(item_id, COL["dvs_output"])
        if dvs_bytes:
            dvs_source = "dvs_output"
        else:
            dvs_bytes = await download_column_file(item_id, COL["dvs_input"])
            if dvs_bytes:
                dvs_source = "dvs_input"

        if not dvs_bytes:
            raise RuntimeError(
                "No DVS XLSX found on this row. Either run the main pipeline "
                "to generate one (dvs_output column) or upload one manually to "
                "the DVS Input column, then re-check this box.")

        await append_log(item_id,
            f"Load DVS UAT Data: fetched DVS from {dvs_source}  "
            f"size={len(dvs_bytes)} bytes")

        # 3. Parse DVS → generate ODM XML → POST to OC
        # Lazy-import uat_runner so a missing/broken install surfaces as
        # a clean monday-log error rather than a module-load crash.
        from uat_runner.parsers.dvs_parser       import parse_dvs
        from uat_runner.generators.odm_generator import generate_odm
        from uat_runner.api.oc_client            import OpenClinicaClient

        # parse_dvs takes a file path (not bytes), so round-trip dvs_bytes
        # through a tempfile. delete=False because we need to read it after
        # closing the writer; explicit os.unlink in the finally below.
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False) as tmp:
                tmp.write(dvs_bytes)
                tmp_path = tmp.name

            test_cases = parse_dvs(tmp_path)
            await append_log(item_id,
                f"Load DVS UAT Data: parsed {len(test_cases)} test cases from DVS")

            odm_xml = generate_odm(test_cases, study_oid=study_oid)
            await append_log(item_id,
                f"Load DVS UAT Data: generated ODM XML ({len(odm_xml)} bytes)")

            # Reuse the existing helper for OC auth — same token issuer
            # publish_to_test uses, so the Test environment is consistent.
            token = await _get_oc_token(oc_subdomain, is_production=False)
            client = OpenClinicaClient(
                base_url=f"https://{oc_subdomain}.build.openclinica.io",
                auth_token=token,
            )
            # OpenClinicaClient uses synchronous `requests` — wrap the
            # network call so the event loop stays responsive. The import
            # endpoint is keyed by study_oid (NOT study_uuid).
            await asyncio.to_thread(
                client.import_odm_data, study_oid, odm_xml)

            await append_log(item_id,
                f"Load DVS UAT Data: imported {len(test_cases)} test cases "
                f"successfully")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

        await append_log(item_id, "Load DVS UAT Data: complete")

    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        print(f"LOAD_DVS_UAT_DATA FAILED for item {item_id}: {err}", flush=True)
        print(traceback.format_exc(), flush=True)
        try:
            await append_log(item_id, f"Load DVS UAT Data FAILED: {err}")
        except Exception as inner:
            print(f"LOAD_DVS_UAT_DATA log fallback failed: {inner}", flush=True)


# ── OC-9 backstop: Common Visit for cross-visit forms ────────────────────────

def _enforce_common_visit(struct_json):
    """RULE OC-9 backstop. Ensure SE_COMMON exists and AE/CM/DV/AESAE
    forms only live there. Runs after Claude returns the Study Spec JSON
    and fixes the structure deterministically if Claude forgot.

    Idempotent — safe to call multiple times.
    """
    COMMON_FORMS = {"AE", "CM", "DV", "AESAE"}

    if not isinstance(struct_json, dict):
        return struct_json

    forms = struct_json.get("forms", [])
    if not isinstance(forms, list):
        return struct_json

    # Events may be stored under "events" or "visits" depending on version.
    # We canonicalize to "events" but tolerate both.
    events = struct_json.get("events")
    if events is None:
        events = struct_json.get("visits", [])
    if not isinstance(events, list):
        events = []

    # Only create SE_COMMON if at least one of the common forms is actually
    # in scope for this protocol. If none are scoped, skip entirely.
    common_forms_in_study = [
        f for f in forms
        if isinstance(f, dict) and f.get("form_id") in COMMON_FORMS
    ]
    if not common_forms_in_study:
        return struct_json

    # Find enrollment event to anchor SE_COMMON's availability window
    enrollment_oid = None
    for ev in events:
        if not isinstance(ev, dict):
            continue
        oid = str(ev.get("event_oid", "")).upper()
        if "ENRL" in oid or "RAND" in oid or "ENROLL" in oid:
            enrollment_oid = ev.get("event_oid")
            break

    # Ensure SE_COMMON exists in events
    has_se_common = any(
        isinstance(ev, dict) and ev.get("event_oid") == "SE_COMMON"
        for ev in events
    )
    if not has_se_common:
        events.append({
            "event_oid":       "SE_COMMON",
            "event_title":     "Common Visit",
            "event_type":      "common",
            "is_repeating":    True,
            "available_after": enrollment_oid or "",
        })
        struct_json["events"] = events

    # Force visits_assigned=["SE_COMMON"] on each common form
    fixed_count = 0
    for f in common_forms_in_study:
        if f.get("visits_assigned") != ["SE_COMMON"]:
            f["visits_assigned"] = ["SE_COMMON"]
            fixed_count += 1

    if fixed_count:
        print(f"OC-9 backstop: reassigned {fixed_count} form(s) to SE_COMMON",
              flush=True)

    return struct_json


def _backfill_migration_fields(spec):
    """Add schedule_of_events + per-form migration lifecycle fields if
    missing. Idempotent — safe to call on every spec load."""
    if not isinstance(spec, dict):
        return spec

    # Top-level schedule_of_events
    if "schedule_of_events" not in spec:
        # Derive target-side visit_mappings from timepoint_csv
        tpt_rows = (spec.get("timepoint_csv") or {}).get("rows") or []
        visit_mappings = []
        seen_events = set()
        for row in tpt_rows:
            ev = row.get("event")
            if ev and ev not in seen_events:
                seen_events.add(ev)
                visit_mappings.append({
                    "source_oid": None,
                    "source_name": None,
                    "target_oid": ev,
                    "target_name": row.get("timepoint", ""),
                    "action": "pending",
                    "notes": "",
                })

        # Derive form_placements as a flat list (one row per form/visit)
        form_placements = []
        for f in spec.get("forms") or []:
            for v in f.get("visits_assigned") or []:
                form_placements.append({
                    "target_visit_oid": v,
                    "form_id": f.get("form_id", ""),
                    "required": True,
                    "repeating": bool(f.get("has_repeating_group")),
                    "notes": "",
                })

        # Arm mappings from study_meta.arms
        arms = (spec.get("study_meta") or {}).get("arms") or []
        arm_mappings = [
            {"source_arm": None, "target_arm": a.get("arm_code", ""), "action": "pending"}
            for a in arms
        ]

        spec["schedule_of_events"] = {
            "migration_status": "draft",
            "approved_by": "",
            "approved_at": "",
            "visit_mappings": visit_mappings,
            "form_placements": form_placements,
            "arm_mappings": arm_mappings,
        }

    # Top-level study_settings — separate from SOE per UX design
    if "study_settings" not in spec:
        # Migrate subject_id_rule from old SOE-nested location if present
        legacy_rule = (spec.get("schedule_of_events") or {}).pop("subject_id_rule", None)
        spec["study_settings"] = {
            "migration_status": "draft",
            "approved_by": "",
            "approved_at": "",
            "subject_id_rule": legacy_rule or {
                "mode": "passthrough",
                "template": "",
                "pattern": "",
                "replacement": "",
            },
        }
    else:
        # study_settings exists, but if SOE still has subject_id_rule from
        # an in-flight spec, fold it in (preferring study_settings if both)
        soe = spec.get("schedule_of_events") or {}
        if "subject_id_rule" in soe:
            spec["study_settings"].setdefault("subject_id_rule", soe.pop("subject_id_rule"))

    # Per-form lifecycle fields
    for f in spec.get("forms") or []:
        f.setdefault("migration_status", "draft")
        f.setdefault("approved_by", "")
        f.setdefault("approved_at", "")
        f.setdefault("rejected_reason", "")

    return spec


def _sanitize_form_titles(spec):
    """Strip characters from form_title that break OC's form-service.

    Idempotent — safe to call on every spec load.

    Root cause (proven CRS-135, 2026-06-02): a '+' anywhere in a form's
    display title silently DEADLOCKS OC's version-attach (the upload spins
    forever, no version is created, and the version-less form record then
    poisons every subsequent add/upload/publish in the study). This is the
    sole reason SLEEP ("Sleep Quality (NRS + PROMIS 8a)") failed for weeks
    while every other form built cleanly. Controlled test: same form/file
    with the '+' removed attaches a version on the first try.

    Fix at the SOURCE (spec creation), not the edc-builder: the builder
    renders whatever form_title the spec provides, so the spec must never
    emit a hostile character. We replace '+' with ' and ' (preserving
    meaning, e.g. "NRS + PROMIS" -> "NRS and PROMIS"), collapse the
    resulting double-spaces, and trim. Other URL/grammar-hostile chars can
    be added here later if OC reveals more offenders.
    """
    if not isinstance(spec, dict):
        return spec

    def _clean(title: str) -> str:
        if not isinstance(title, str) or "+" not in title:
            return title
        # "A + B" -> "A and B"; bare "A+B" -> "A and B" too.
        cleaned = title.replace("+", " and ")
        # Collapse any double spaces introduced by the replacement.
        while "  " in cleaned:
            cleaned = cleaned.replace("  ", " ")
        return cleaned.strip()

    for f in spec.get("forms") or []:
        if isinstance(f, dict) and "form_title" in f:
            orig = f.get("form_title")
            new = _clean(orig)
            if new != orig:
                print(f"[spec-sanitize] form_title contained '+', rewrote: "
                      f"{orig!r} -> {new!r} (OC form-service deadlocks on '+')",
                      flush=True)
                f["form_title"] = new

    return spec


# ── Session pre-flight (validates a saved Playwright session in ~15s) ──────────

async def _validate_oc_session(subdomain: str, session_path: str) -> bool:
    """SSO session preflight — temporarily disabled (always returns True).

    Previous approaches failed: the .design cookie check was false-valid;
    the Keycloak userinfo check was false-stale; the Playwright probe
    correctly detects a dead session but also rejects freshly-captured
    sessions on Railway because the headless browser can't replay a
    storage_state captured on a different machine/IP through OC's SSO.

    The publisher's own _authenticate_via_sso already handles a dead
    session correctly (deletes the stale file, errors out). The missing
    piece is re-prompting the user for auth when the publisher fails —
    that fix belongs in the publisher's failure path, not here.

    TODO: implement re-auth prompt in publisher's SSO failure handler,
    then re-enable a meaningful preflight check.
    """
    print(f"[session-preflight] check skipped (fail-open) — "
          f"publisher will validate SSO directly", flush=True)
    return True


# ── Main pipeline ──────────────────────────────────────────────────────────────

async def _session_keepalive(session_path: str, subdomain: str, interval_s: int = 60):
    """Background task: pings OC designer every interval_s seconds using the saved
    session cookies to prevent Keycloak from expiring the token mid-run.
    Runs until cancelled. Failure-open: errors are logged but never raised."""
    import httpx
    while True:
        await asyncio.sleep(interval_s)
        try:
            cookies = {}
            _iss = ""
            try:
                import json as _json
                import base64 as _b64
                with open(session_path) as _f:
                    _state = _json.load(_f)
                for c in _state.get("cookies", []):
                    cookies[c["name"]] = c["value"]
                # Pull the SSO token's issuer (realm base) so we can hit the
                # right Keycloak silent-auth endpoint. Token lives in
                # storage_state localStorage under jhi-authenticationtoken
                # (JSON-quoted). Decode its `iss` claim.
                _tok = ""
                for _o in _state.get("origins", []):
                    for _ls in _o.get("localStorage", []):
                        if _ls.get("name") in ("jhi-authenticationtoken",
                                               "jhi-idtoken") and not _tok:
                            _tok = (_ls.get("value") or "").strip().strip('"')
                if _tok and _tok.count(".") >= 2:
                    _pl = _tok.split(".")[1]
                    _pl += "=" * (-len(_pl) % 4)
                    _claims = _json.loads(_b64.urlsafe_b64decode(_pl))
                    _iss = (_claims.get("iss") or "").rstrip("/")
            except Exception:
                pass

            # Ping 1 — design app. Keeps the .design cookie warm (used by the
            # REST preflight / board reads). Does NOT touch Keycloak.
            url = f"https://{subdomain}.design.openclinica.io/api/boards"
            async with httpx.AsyncClient(timeout=10, follow_redirects=False) as client:
                await client.get(url, cookies=cookies)
            print(f"[session-keepalive] ping sent to {subdomain}.design.openclinica.io",
                  flush=True)

            # Ping 2 — Keycloak silent-auth (prompt=none). THIS is what keeps
            # the SSO session alive: the Playwright publisher renews via
            # implicit-flow silent re-auth, which depends on the Keycloak SSO
            # session cookie. That session has a short idle timeout (~20 min)
            # independent of the 24h token exp; the ~9-min build was idling it
            # out, so the publisher landed on the Keycloak login page. Hitting
            # /auth?prompt=none with the saved cookies resets the idle timer.
            # We follow redirects and log the landing so the run log SHOWS
            # whether the session is alive (lands on signin-callback) or dead
            # (lands on the login page) — self-validating, not blind.
            if _iss:
                try:
                    import time as _t
                    _sa = (f"{_iss}/protocol/openid-connect/auth"
                           f"?client_id=studymanager"
                           f"&redirect_uri=https://{subdomain}.build.openclinica.io"
                           f"/signin-callback.html"
                           f"&response_type=id_token%20token"
                           f"&scope=openid%20profile"
                           f"&prompt=none"
                           f"&state=ka{int(_t.time())}"
                           f"&nonce=ka{int(_t.time())}")
                    async with httpx.AsyncClient(
                            timeout=10, follow_redirects=True) as _kc:
                        _r = await _kc.get(_sa, cookies=cookies)
                    _final = str(_r.url)
                    if "signin-callback" in _final or "access_token" in _final:
                        _state_word = "ALIVE (renewed)"
                    elif "openid-connect/auth" in _final or "login" in _final.lower():
                        _state_word = "DEAD (would need re-auth)"
                    else:
                        _state_word = f"unknown (landed {_final[:80]})"
                    print(f"[session-keepalive] keycloak silent-auth: "
                          f"{_state_word} status={_r.status_code}", flush=True)
                except Exception as _kce:
                    print(f"[session-keepalive] keycloak ping error "
                          f"(non-fatal): {_kce}", flush=True)
            else:
                print("[session-keepalive] keycloak ping skipped — no token "
                      "issuer found in session", flush=True)

            # Ping 3 — EU clinical host. Keeps the session cookie warm on the
            # customer's clinical instance (e.g. cust1.eu.openclinica.io).
            # Required so UAT ODM import cookies are still valid at end of run.
            try:
                import csv as _csv
                from pathlib import Path as _Path
                _csv_path = (_Path(__file__).parent
                             / "references" / "customer_uuids.csv")
                _bridge = None
                if _csv_path.exists():
                    with open(_csv_path, newline="") as _csvf:
                        for _row in _csv.DictReader(_csvf):
                            if _row.get("subdomain","").lower() == subdomain.lower():
                                _bridge = _row.get("bridge_url","").rstrip("/")
                                break
                if _bridge:
                    _eu_ping_url = f"{_bridge}/MainMenu"
                    async with httpx.AsyncClient(
                            timeout=10, follow_redirects=False) as _eu:
                        _eu_r = await _eu.get(_eu_ping_url, cookies=cookies)
                    print(f"[session-keepalive] EU clinical ping: "
                          f"{_eu_ping_url} status={_eu_r.status_code}", flush=True)
            except Exception as _eue:
                print(f"[session-keepalive] EU clinical ping error "
                      f"(non-fatal): {_eue}", flush=True)

            # Ping 4 — build app My Studies. The publisher authenticates by
            # navigating to cust1.build.openclinica.io/#/account-study.  The
            # build app has its own Angular JWT that can idle out independently
            # of the Keycloak SSO session — hitting the API with the saved
            # cookies keeps that JWT active so the publisher doesn't land on
            # the Keycloak login page after a long analysis + build phase.
            try:
                _build_ping_url = (
                    f"https://{subdomain}.build.openclinica.io"
                    f"/study-service/api/studies"
                )
                async with httpx.AsyncClient(
                        timeout=10,
                        follow_redirects=False,
                        cookies=cookies) as _bp:
                    _bp_r = await _bp.get(_build_ping_url)
                print(f"[session-keepalive] build app ping: "
                      f"{_build_ping_url} status={_bp_r.status_code}",
                      flush=True)
            except Exception as _bpe:
                print(f"[session-keepalive] build app ping error "
                      f"(non-fatal): {_bpe}", flush=True)
        except asyncio.CancelledError:
            raise
        except Exception as e:
            print(f"[session-keepalive] ping error (non-fatal): {e}", flush=True)


async def run_pipeline(item_id):
    try:
        # ── 0. Fetch item from monday.com ─────────────────────────────────────
        item         = await get_item(item_id)
        cols         = {c["id"]: c for c in item["column_values"]}
        protocol_num = cols.get(COL["protocol_number"], {}).get("text", "STUDY")
        oc_subdomain = cols.get(COL["oc_subdomain"],    {}).get("text", "").strip()
        # Per-user OC SSO login for form upload (Playwright/storage_state).
        # Required for create_oc_study's Playwright form-publish step; if
        # empty, that step is skipped with a clear log.
        oc_email     = cols.get(COL["oc_email"],        {}).get("text", "").strip()

        # Fetch library filenames for both columns so we can inject them into
        # the Study Spec JSON's study_meta.library_files_provided (overrides
        # whatever Claude emits). This gives humans a clear record of which
        # library inputs were used.
        try:
            crf_lib_names    = await list_column_filenames(item_id, COL["crf_library"])
            oc_std_lib_names = await list_column_filenames(item_id, COL["oc_standard"])
        except Exception as e:
            print(f"Warning: could not fetch library filenames: {e}", flush=True)
            crf_lib_names, oc_std_lib_names = [], []
        library_files_provided = crf_lib_names + oc_std_lib_names
        print(f"Library files: CRF={crf_lib_names} | OC4={oc_std_lib_names}", flush=True)

        _now    = _dt.datetime.utcnow()
        version = f"V{_now.strftime('%m%d')}.{_now.strftime('%H%M')}"
        print(f"Protocol: {protocol_num} | Version: {version}", flush=True)

        # ── UAT-only shortcut ────────────────────────────────────────────────
        # If Study UUID is already populated AND Load UAT checkbox is checked
        # AND Publish to Test checkbox is NOT checked → skip all build stages
        # and run the UAT loader directly against the existing published study.
        # This is useful when only test data needs to be reloaded (e.g. DVS
        # was regenerated) without rebuilding or republishing the study.

        # ── Auth check must happen BEFORE any early exits ─────────────────────
        # Any checkbox that touches OC requires a valid Playwright session.
        def _bool_col_early(col_id: str) -> bool:
            try:
                return bool(json.loads(
                    cols.get(col_id, {}).get("value") or "{}"
                ).get("checked", False))
            except Exception:
                return False

        _oc_email_early = (cols.get(COL["oc_email"], {}).get("text") or "").strip()
        _needs_session_early = (
            _bool_col_early(COL["create_study"])
            or _bool_col_early("boolean_mm3g2vzf")
            or _bool_col_early("boolean_mm3z1xy8")
            or _bool_col_early("boolean_mm3gxe49")
        )
        if _needs_session_early and _oc_email_early:
            from auth_manager import AuthManager as _AM
            _am = _AM()
            _session_ok = False
            if _am.session_exists(_oc_email_early):
                _sess_path = str(_am.get_session_path(_oc_email_early))
                _age_s = time.time() - os.path.getmtime(_sess_path)
                if _age_s < 120:
                    _session_ok = True  # grace period — just created
                else:
                    from oc_form_publisher import probe_sso_session as _probe
                    _oc_sub_early = (cols.get(COL["oc_subdomain"], {}).get("text") or "").strip()
                    _session_ok = await _probe(_oc_sub_early, _sess_path)
                    if not _session_ok:
                        try:
                            os.remove(_sess_path)
                        except OSError:
                            pass
            if not _am.session_exists(_oc_email_early):
                _oc_sub_for_link = (cols.get(COL.get("oc_subdomain","text_mm3aa7cx"), {}).get("text") or "").strip()
                # Derive clinical host from CSV (region-aware)
                _clinical_host = ""
                if _oc_sub_for_link:
                    from uat_loader import _pages_base as _pb
                    try:
                        _bridge = _pb(_oc_sub_for_link)
                        from urllib.parse import urlparse as _up2
                        _clinical_host = _up2(_bridge).netloc
                    except Exception:
                        _clinical_host = f"{_oc_sub_for_link}.eu.openclinica.io"
                _auth_link = _am.generate_auth_link(
                    _oc_email_early,
                    "https://oc-ai-pipeline-production.up.railway.app",
                    context="uat",
                )
                # Include clinical host in link so instructions page can show it
                if _clinical_host:
                    from urllib.parse import urlencode as _ue
                    _auth_link += "&clinical_host=" + _clinical_host
                await append_log(item_id,
                    f"⚠️ Authentication Required\n\n"
                    f"Click here to authenticate your OpenClinica account:\n"
                    f"{_auth_link}\n\n"
                    f"After authentication, trigger 'Send to AI' again to continue.")
                await set_link(item_id, COL["oc_auth_link"], _auth_link,
                               text="Authenticate OpenClinica")
                await set_status(item_id, COL["pipeline_status"],
                                 "Paused for Authentication")
                try:
                    await set_status(item_id, COL["ai_trigger"],
                                     "Do not Send To AI Yet")
                except Exception:
                    pass
                print(f"[auth-early] Auth required for {_oc_email_early} — pausing.",
                      flush=True)
                return

        _existing_uuid_early = (cols.get(COL["study_uuid"], {})
                                .get("text") or "").strip()
        _load_uat_early = (cols.get(COL["load_dvs_uat_data"], {})
                           .get("text") or "").strip() == "v"
        _publish_early  = (cols.get(COL["publish_to_test"], {})
                           .get("text") or "").strip() == "v"
        if _existing_uuid_early and _load_uat_early and not _publish_early:
            print(f"[uat-only] Study UUID={_existing_uuid_early!r} already "
                  f"exists and Publish is not checked — running UAT loader "
                  f"directly, skipping all build stages.", flush=True)
            await set_status(item_id, COL["pipeline_status"],
                             UAT_STATUS["loading"])
            try:
                _uat_result = await run_uat_loader(item_id)
                if _uat_result["success"]:
                    await asyncio.gather(
                        set_status(item_id, COL["pipeline_status"],
                                   STATUS["all_complete"]),
                        append_log(item_id,
                                   f"[UAT-only] UAT load succeeded. "
                                   f"Site: {_uat_result['site_oid']}. "
                                   f"Participants: "
                                   f"{len(_uat_result['participants_created'])}."),
                    )
                    print(f"[uat-only] UAT load succeeded → All Complete",
                          flush=True)
                else:
                    _errs = "; ".join(_uat_result["errors"])
                    # If auth required, status is already set by uat_loader
                    # (Paused for Authentication + trigger reset) — don't overwrite
                    _is_auth = any("Authentication required" in e
                                   for e in _uat_result["errors"])
                    if not _is_auth:
                        await asyncio.gather(
                            set_status(item_id, COL["pipeline_status"],
                                       UAT_STATUS["failed"]),
                            append_log(item_id,
                                       f"[UAT-only] UAT load FAILED: {_errs}"),
                        )
                    else:
                        await append_log(item_id,
                                         f"[UAT-only] UAT load FAILED: {_errs}")
                    print(f"[uat-only] UAT load failed: {_errs}", flush=True)
            except Exception as _ue:
                print(f"[uat-only] UAT loader crashed: {_ue}", flush=True)
                await asyncio.gather(
                    set_status(item_id, COL["pipeline_status"],
                               UAT_STATUS["failed"]),
                    append_log(item_id,
                               f"[UAT-only] UAT loader ERROR: {_ue}"),
                )
            return  # ← exit run_pipeline; nothing else to do

        def _pct(col_key):
            raw = cols.get(COL[col_key], {}).get("text", "").strip()
            try:
                return float(raw) / 100.0 if raw else 0.0
            except ValueError:
                return 0.0

        additional_sub_disc = _pct("subscription_discount")
        additional_svc_disc = _pct("services_discount")

        output_raw        = cols.get(COL["output_requested"], {}).get("text", "") or ""
        output_selections = {s.strip().lower() for s in output_raw.split(",") if s.strip()}
        run_all = len(output_selections) == 0
        # Fast-rerun mode: a user opt-in (gated below on `not run_all`)
        # signal that struct_json comes from a non-Claude source — either
        # an edited human input file OR the prior run's spec_json on
        # monday (Path R, defined further down). When True, chains A/B/E
        # (spec PDF/XLSX, protocol summary, build preview) are skipped
        # so a re-run focuses on the user's actual iteration target
        # (typically Chain C build or Chain D form upload).
        fast_rerun = False
        _FAST_RERUN_SKIP = {
            "protocol specification",  # Chain A
            "protocol summary",        # Chain B (summary PDF)
            "price quote",             # Chain B (quote)
            "build preview",           # Chain E
        }
        def _want(label):
            if fast_rerun and label.lower() in _FAST_RERUN_SKIP:
                return False
            return run_all or label.lower() in output_selections
        print(f"Output requested: {output_raw!r} | run_all={run_all}", flush=True)

        create_study_val = cols.get(COL["create_study"], {}).get("value")
        try:
            parsed = json.loads(create_study_val or "{}")
            create_study = bool(parsed.get("checked", False)) if isinstance(parsed, dict) else bool(parsed)
        except Exception:
            create_study = False

        oc_production_val = cols.get(COL["oc_production"], {}).get("value")
        try:
            parsed = json.loads(oc_production_val or "{}")
            oc_production = bool(parsed.get("checked", False)) if isinstance(parsed, dict) else bool(parsed)
        except Exception:
            oc_production = False

        # Trainer integration is gated on TRAINER_URL being set in env
        # (see trainer_integration.trainer_enabled). No per-row Monday
        # checkbox — successful pipeline runs always feed the trainer
        # when the service is wired up; dedup on the trainer side
        # prevents duplicate corpus rows.
        trainer_on = trainer_enabled()

        print(f"Create OC Study: {create_study} | Subdomain: {oc_subdomain} | "
              f"Production: {oc_production} | Trainer enabled: {trainer_on}",
              flush=True)

        # ── Early OAuth check (saves chains A-E on first-time auth) ──────────
        # Auth check already handled above (before early exits) — see
        # _needs_session_early block. Session path set there for use below.

        # ── 1. Check for human-uploaded inputs (parallel downloads) ──────────
        (edited_spec_xlsx,
         edited_build_zip,
         edited_dvs_xlsx,
         edited_quote_xlsx,
         edited_soe_csv,
         source_edc_export_bytes) = await asyncio.gather(
            download_column_file(item_id, COL["edited_spec_input"]),
            download_column_file(item_id, COL["build_input"]),
            download_column_file(item_id, COL["dvs_input"]),
            download_column_file(item_id, COL["quote_input"]),
            download_column_file(item_id, COL["soe_input"]),
            download_column_file(item_id, COL["source_edc_export"]),
        )

        print(f"Human inputs — spec:{edited_spec_xlsx is not None} "
              f"build:{edited_build_zip is not None} dvs:{edited_dvs_xlsx is not None} "
              f"quote:{edited_quote_xlsx is not None} soe:{edited_soe_csv is not None} "
              f"edc_export:{source_edc_export_bytes is not None}",
              flush=True)

        # Fast-rerun trigger #1 — human-uploaded spec or build files.
        # When the user supplies an edited Study Spec XLSX or an edited
        # EDC build ZIP, they're iterating on Chain C/D outputs; the
        # Claude-driven chains A/B/E would just regenerate already-good
        # documents. Gated on `not run_all` so a run with no output
        # selections (the "full run" default) still produces everything.
        # See _want() above for what gets skipped.
        if (not run_all) and (edited_spec_xlsx or edited_build_zip):
            fast_rerun = True
            print(f"[fast-rerun] Human input detected — chains A/B/E "
                  f"will be skipped (spec={bool(edited_spec_xlsx)}, "
                  f"build={bool(edited_build_zip)})", flush=True)

        # ── Path D: Edited Quote XLSX → regenerate Quote PDFs ─────────────────
        # DEPRECATED: there's no local script to regenerate Quote PDFs from
        # an edited XLSX. To refresh the PDFs, edit the Protocol Summary JSON
        # or Study Spec and re-run the full pipeline.
        if edited_quote_xlsx:
            await append_log(item_id,
                "Edited Quote XLSX detected. Automatic PDF regeneration from edited "
                "XLSX is not currently supported — to regenerate PDFs, edit the "
                "underlying Protocol Summary JSON or Study Spec and re-run the "
                "full pipeline."
            )
            print("Path D: edited Quote XLSX detected — regeneration not supported, "
                  "skipping without changes.", flush=True)
            await set_status(item_id, COL["pipeline_status"], STATUS["all_complete"])
            await append_log(item_id, "Pipeline complete (Path D — no-op).")
            return

        # ── Path E: Edited SOE CSV → update OpenClinica ───────────────────────
        if edited_soe_csv and oc_subdomain:
            await append_log(item_id, "Edited SOE CSV detected — updating OpenClinica.")
            print("Path E: updating SOE in OpenClinica...", flush=True)
            # TODO: implement SOE update API call when OC API supports it
            await append_log(item_id, "SOE update in OpenClinica — not yet implemented.")
            await set_status(item_id, COL["pipeline_status"], STATUS["all_complete"])
            return

        # ── Download inputs in parallel ───────────────────────────────────────
        async def _get_protocol_doc():
            """
            Download the protocol document from the protocol column ONLY.

            Reads exclusively from COL["protocol"] (no whole-item asset scan).
            Supports:
              - PDF: returned as-is.
              - Word (.docx / .doc): converted to PDF via LibreOffice when
                possible; otherwise extracted as text and returned with the
                ``%%DOCX_TEXT%%`` marker so callers can pass it as
                ``extra_text`` instead of ``pdf_bytes``.
              - Google Doc / Drive link (URL in the column rather than an
                uploaded file): exported as PDF.

            Returns bytes on success, or None when the column is empty / the
            content can't be resolved. Callers must tolerate None.
            """
            if edited_spec_xlsx:
                return None  # Skip — Path A already gives us struct via XLSX

            col_entry = cols.get(COL["protocol"], {})
            raw_value = col_entry.get("value")

            # 1. Uploaded-file case (asset on the column).
            if raw_value:
                try:
                    parsed = json.loads(raw_value)
                except (ValueError, TypeError):
                    parsed = {}
                files = parsed.get("files") if isinstance(parsed, dict) else None
                if files:
                    body = await download_column_file(item_id, COL["protocol"])
                    if not body:
                        return None
                    fname = (files[-1].get("name") or "").lower()

                    # PDF magic or .pdf extension → pass through.
                    if body.startswith(b"%PDF-") or fname.endswith(".pdf"):
                        return body

                    # ZIP magic (.docx/.xlsx OOXML) or .docx/.doc extension
                    # → try LibreOffice, fall back to extracted text.
                    if body.startswith(b"PK\x03\x04") or \
                       fname.endswith((".docx", ".doc")):
                        print(f"Converting Word doc: {fname or '<unnamed>'}",
                              flush=True)
                        pdf = await asyncio.get_event_loop().run_in_executor(
                            None,
                            lambda: _convert_to_pdf(body, fname or "protocol.docx"),
                        )
                        if pdf:
                            return pdf
                        text = await asyncio.get_event_loop().run_in_executor(
                            None, lambda: _extract_docx_as_text(body),
                        )
                        if text:
                            return b"%%DOCX_TEXT%%" + text.encode("utf-8")
                        return None

                    # Anything else (no recognisable header) — accept the
                    # bytes verbatim and let Claude decide.
                    return body

            # 2. Link case — Monday file columns can also carry a URL
            # (e.g. Google Doc / Drive share link). Detect a docs/drive URL
            # in the column's `.text` and fetch the PDF export.
            link_text = (col_entry.get("text") or "").strip()
            if link_text and ("docs.google.com" in link_text or
                              "drive.google.com" in link_text):
                export_url = _google_doc_export_url(link_text)
                if export_url:
                    print(f"Exporting Google Doc as PDF: {link_text[:80]}",
                          flush=True)
                    return await download_file(export_url) or None

            return None

        # Use download_column_file for CRF / OC standard so we go through the
        # asset → public_url (S3) path that doesn't need session auth. The
        # previous code used the column's `.text` URL which is a Monday
        # `protected_static/...` link that returns HTTP 406 to bearer-token
        # requests, silently dropping the customer's library inputs.
        protocol_bytes, crf_pdf, oc_zip = await asyncio.gather(
            _get_protocol_doc(),
            download_column_file(item_id, COL["crf_library"]),
            download_column_file(item_id, COL["oc_standard"]),
        )
        _proto_desc = (
            f"{len(protocol_bytes):,} bytes PDF" if protocol_bytes and
            not protocol_bytes.startswith(b"%%DOCX_TEXT%%")
            else f"{len(protocol_bytes) - 14:,} chars text (Word doc)"
            if protocol_bytes else "0 bytes"
        )
        print(f"Protocol: {_proto_desc} | "
              f"CRF: {len(crf_pdf) if crf_pdf else 0} | "
              f"OC ZIP: {len(oc_zip) if oc_zip else 0}", flush=True)

        # ── Determine if analysis/chains are needed ───────────────────────────
        needs_analysis = (
            _want("protocol specification") or _want("protocol summary")
            or _want("price quote") or _want("study build zip")
            or (create_study and oc_subdomain)
        )

        # ── Steps 1-2: Study Specification ────────────────────────────────────
        struct_json = None

        # Holds NOTES_FOR_AI content extracted from the edited XLSX — injected
        # into the prompt so Claude understands what the reviewer changed and why.
        reviewer_notes_block = ""

        if edited_spec_xlsx:
            # Path A: User uploaded edited Study Spec XLSX
            await append_log(item_id, "Edited Study Specification XLSX detected.")
            print("Path A: reading edited Study Spec XLSX...", flush=True)
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(edited_spec_xlsx))

            # ── Extract NOTES_FOR_AI from all survey/choices sheets ───────────
            notes_collected = []
            for sheet_name in wb.sheetnames:
                if not (sheet_name.endswith('_survey') or sheet_name.endswith('_choices')):
                    continue
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                # Find ACTION col (A=0) and NOTES_FOR_AI col (B=1) — fixed positions
                # but verify by reading header row (row index 2 = row 3 in 1-indexed)
                header_row_idx = None
                for ri, row in enumerate(rows):
                    if row and str(row[0] or '').strip().upper() == 'ACTION':
                        header_row_idx = ri
                        break
                if header_row_idx is None:
                    continue
                notes_col_idx = 1   # column B = index 1
                action_col_idx = 0  # column A = index 0
                for row in rows[header_row_idx + 1:]:
                    if not row or len(row) <= notes_col_idx:
                        continue
                    action = str(row[action_col_idx] or '').strip().upper()
                    note   = str(row[notes_col_idx] or '').strip()
                    if note and note.lower() not in ('notes_for_ai', 'notes for ai'):
                        field_name = str(row[2] or '').strip() if len(row) > 2 else ''
                        prefix = f"[{sheet_name}][{action or 'MODIFIED'}]"
                        if field_name:
                            prefix += f" {field_name}:"
                        notes_collected.append(f"{prefix} {note}")

            if notes_collected:
                reviewer_notes_block = (
                    "\n\nREVIEWER NOTES FROM EDITED STUDY SPECIFICATION:\n"
                    + "\n".join(f"  • {n}" for n in notes_collected)
                    + "\nApply these notes when interpreting ACTION=DELETE/ADD rows "
                    "and when regenerating any affected forms."
                )
                print(f"Path A: extracted {len(notes_collected)} reviewer notes "
                      f"from edited XLSX", flush=True)

            # ── Try to extract embedded JSON ──────────────────────────────────
            for sheet_name in wb.sheetnames:
                if 'json' in sheet_name.lower() or 'spec' in sheet_name.lower():
                    ws = wb[sheet_name]
                    raw = '\n'.join(str(cell.value or '') for row in ws.iter_rows() for cell in row)
                    try:
                        struct_json = extract_json(
                            raw,
                            expected_keys=["study_meta", "forms"],
                        )
                        print("Extracted JSON from edited Study Spec XLSX.", flush=True)
                        # OC-9 backstop: apply to edited-XLSX path as well
                        struct_json = _enforce_common_visit(struct_json)
                        struct_json = _backfill_migration_fields(struct_json)
                        struct_json = _sanitize_form_titles(struct_json)
                        # ── Conventions engine pass + three-way conflict detection (Phase C.4) ──
                        # Path X.1 is the edited-XLSX path. Three snapshots make TRUE
                        # conflict detection possible:
                        #   baseline        = previous spec_json from monday (what user downloaded)
                        #   user_edit       = struct_json after XLSX parse (what user uploaded)
                        #   post_convention = struct_json after apply_conventions (what engine did)
                        # True conflict: engine touched a path the user ALSO touched.
                        # First run (no baseline) or download failure → falls back to
                        # Phase C.2 two-way behavior (all engine mutations reported).
                        # No false-negative suppression at any stage; degraded modes
                        # only inflate the conflict list, never silence it.
                        try:
                            import copy as _copy
                            from conventions_engine import apply_conventions, diff as _conv_diff, attribution as _conv_attr
                            _study_id = (struct_json.get("study_meta") or {}).get("protocol_number") or protocol_num
                            # TODO(B.1b follow-up): This path does not currently extract monday's
                            # source_edc_system column (dropdown_mm382w7d). Vendor conventions
                            # apply only on migration path (Path M). If non-migration builds need
                            # vendor conventions in future, extract the column at build entry and
                            # thread it through as migration_source here.

                            # Phase C.4: try to fetch the prior spec_json from monday as baseline.
                            # Failures (first run for this study, monday hiccup, malformed JSON)
                            # → None, which silently degrades to Phase C.2 two-way diff.
                            _baseline_spec = None
                            try:
                                _baseline_bytes = await download_column_file(item_id, COL["spec_json"])
                                _baseline_spec = json.loads(_baseline_bytes.decode("utf-8"))
                            except Exception as _be:
                                print(f"conventions_engine: no baseline spec_json available "
                                      f"(first run or download failed: {_be}); falling back to "
                                      f"Phase C.2 two-way diff", flush=True)

                            # Snapshot user's edit BEFORE the engine mutates anything.
                            _user_edit_spec = _copy.deepcopy(struct_json)

                            # User changes = diff(baseline, user_edit). Empty list when no baseline.
                            _user_changes = (
                                _conv_diff.deep_diff(_baseline_spec, _user_edit_spec)
                                if _baseline_spec is not None else []
                            )
                            _user_change_paths = {r["field_path"] for r in _user_changes}

                            apply_conventions(struct_json, study_id=_study_id,
                                              customer_subdomain=oc_subdomain)

                            # Engine changes = diff(user_edit, post_convention). Attribute to conventions.
                            _engine_changes = _conv_diff.deep_diff(_user_edit_spec, struct_json)
                            _applied_log = (struct_json.get("study_meta") or {}).get(
                                "conventions_engine_applied", [])
                            _engine_changes_attributed = _conv_attr.attribute_changes(
                                _engine_changes, _applied_log)

                            # True conflicts: filter engine changes to paths the user also touched.
                            # When no baseline → fallback to all engine mutations (Phase C.2 schema:
                            # 4-key rows lacking baseline_value, renderers handle the missing key).
                            if _baseline_spec is not None:
                                _conflicts = _conv_diff.filter_to_user_intersected(
                                    _engine_changes_attributed, _user_change_paths, _baseline_spec,
                                )
                            else:
                                _conflicts = _engine_changes_attributed

                            _sm = struct_json.setdefault("study_meta", {})
                            _sm["user_changes"] = _user_changes
                            _sm["convention_conflicts"] = _conflicts

                            print(f"conventions_engine: {len(_user_changes)} user change(s), "
                                  f"{len(_engine_changes_attributed)} engine mutation(s), "
                                  f"{len(_conflicts)} true conflict(s) on user-edited spec "
                                  f"(Path X.1, {'three-way' if _baseline_spec is not None else 'two-way fallback'})",
                                  flush=True)
                        except Exception as _ce:
                            print(f"conventions_engine FAILED — continuing without conventions: {_ce}",
                                  flush=True)
                            try:
                                await append_log(item_id, f"Conventions engine error (build continues): {_ce}")
                            except Exception:
                                pass
                        break
                    except ValueError:
                        pass
            if struct_json is None:
                await append_log(item_id, "Could not extract JSON from edited XLSX — running fresh analysis.")

        # ── Path M: Source EDC Export (migration) ─────────────────────────────
        # Routing trigger: the AI Study Hub board's path-label column
        # `label__1` must read "Migration". Migration is authoritative —
        # we no longer fall back to "did a file get uploaded" because
        # operators were occasionally attaching ODMs to non-migration
        # rows and getting silently routed into Path M. Both label AND
        # file are required; label without file is a hard error.
        _migration_post_build_state: dict | None = None
        _study_path_label = (cols.get(COL["study_path_label"]) or {}).get("text", "") or ""
        _is_migration_label = _study_path_label.strip().lower() == "migration"
        if struct_json is None and _is_migration_label:
            if not source_edc_export_bytes:
                msg = ("Migration label set but no file on Source EDC "
                       "Export column — upload an ODM XML (or ZIP) and "
                       "re-trigger.")
                print(f"Path M FAIL: {msg}", flush=True)
                await append_log(item_id, msg)
                await set_status(item_id, COL["pipeline_status"], STATUS["failed"])
                return
            # protocol_bytes may be a real PDF, the b"%%DOCX_TEXT%%"-marked
            # text fallback from Word docs, or b""/None. run_migration
            # accepts all three: it routes truthy bytes into enrichment
            # mode and unwraps the DOCX_TEXT marker internally.
            _proto_for_migration = protocol_bytes if protocol_bytes else None
            _enrichment = bool(_proto_for_migration)
            _mode = ("ODM+Protocol enrichment mode (AI-assisted)"
                     if _enrichment else "ODM-only mode")
            await append_log(item_id, f"Source EDC Export detected — running migration path ({_mode}).")
            print(f"Path M: running EDC migration from Source EDC Export — {_mode}", flush=True)
            await set_status(item_id, COL["pipeline_status"], STATUS["analysis_running"])
            mig_result = await run_edc_migration(
                item_id,
                raw_bytes=source_edc_export_bytes,
                protocol_bytes=_proto_for_migration,
            )
            if mig_result["status"] != "ok":
                msg = f"Migration {mig_result['status']}: {mig_result['summary']}"
                print(f"Path M FAIL: {msg}", flush=True)
                await append_log(item_id, msg)
                await set_status(item_id, COL["pipeline_status"], STATUS["failed"])
                return
            # Stash what the post-build gap-analysis hook will need so we
            # don't have to re-parse the ODM later in run_pipeline.
            _migration_post_build_state = {
                "odm_metadata":       mig_result.get("odm_metadata"),
                "source_system":      mig_result.get("source_system"),
                "source_odm_bytes":   mig_result.get("source_odm_bytes"),
                "source_odm_filename": mig_result.get("source_odm_filename"),
            }
            # On success, fetch the freshly-uploaded Study Spec JSON so
            # downstream stages see the same in-memory struct.
            spec_bytes = await download_column_file(item_id, COL["spec_json"])
            struct_json = json.loads(spec_bytes.decode("utf-8"))
            struct_json = _enforce_common_visit(struct_json)
            struct_json = _backfill_migration_fields(struct_json)
            struct_json = _sanitize_form_titles(struct_json)
            # ── Conventions engine pass (no-op until conventions/ store is populated) ─
            try:
                from conventions_engine import apply_conventions
                _study_id = (struct_json.get("study_meta") or {}).get("protocol_number") or protocol_num
                _vendor_slug = _vendor_slug_from_display_name(mig_result.get("source_system"))
                apply_conventions(struct_json, study_id=_study_id,
                                  customer_subdomain=oc_subdomain,
                                  migration_source=_vendor_slug)
            except Exception as _ce:
                print(f"conventions_engine FAILED — continuing without conventions: {_ce}",
                      flush=True)
                try:
                    await append_log(item_id, f"Conventions engine error (build continues): {_ce}")
                except Exception:
                    pass
            print(f"Path M: struct_json loaded — "
                  f"{len(struct_json.get('forms', []))} forms, "
                  f"source={mig_result.get('source_system')}", flush=True)

        # ── Guard: ODM XML present but label is not "Migration" ───────────────
        # A Source EDC Export file in file_mm386dte without label__1="Migration"
        # means the operator most likely forgot to set the label on a duplication
        # or is testing with a migration file on a non-migration row.  Letting
        # the pipeline fall through silently to Protocol Analysis here is what
        # caused the VAX1001-copy → PrTK05 data-corruption incident.  Fail hard.
        if struct_json is None and source_edc_export_bytes and not _is_migration_label:
            _odm_msg = (
                "FAILED: Source EDC Export file found but Study Path Label is "
                f"'{_study_path_label or '(empty)'}', not 'Migration'. "
                "Set the label column to 'Migration' and re-trigger."
            )
            print(f"Path M GUARD (label mismatch): {_odm_msg}", flush=True)
            await append_log(item_id, _odm_msg)
            await set_status(item_id, COL["pipeline_status"], STATUS["failed"])
            return

        # ── Read AI Instructions from edited XLSX (if present) ──────────────
        ai_instructions_block = ""
        if edited_spec_xlsx:
            try:
                import openpyxl as _opx
                _wb = _opx.load_workbook(io.BytesIO(edited_spec_xlsx), data_only=True)
                if "AI_INSTRUCTIONS" in _wb.sheetnames:
                    _ws = _wb["AI_INSTRUCTIONS"]
                    _rows = list(_ws.iter_rows(values_only=True))
                    _study_instrs  = []
                    _form_instrs   = []
                    _in_s1 = _in_s2 = False
                    for _row in _rows:
                        if not _row or not any(v for v in _row if v):
                            continue
                        _cell0 = str(_row[0] or "").strip().upper()
                        if "SECTION 1" in _cell0:
                            _in_s1, _in_s2 = True, False; continue
                        if "SECTION 2" in _cell0:
                            _in_s1, _in_s2 = False, True; continue
                        if "SECTION 3" in _cell0 or "VERSION HISTORY" in _cell0:
                            _in_s1 = _in_s2 = False; continue
                        if _cell0 in ("PRIORITY", "FORM OID", "VERSION"):
                            continue   # skip header rows
                        if _in_s1:
                            _instr = str(_row[1] or "").strip()
                            _pri   = str(_row[0] or "").strip()
                            if _instr and _instr.lower() != "instruction":
                                _study_instrs.append(
                                    f"[{_pri.upper() or 'MEDIUM'}] {_instr}")
                        elif _in_s2:
                            _foid  = str(_row[0] or "").strip()
                            _instr = str(_row[1] or "").strip()
                            if _foid and _instr and _instr.lower() != "instruction":
                                _form_instrs.append(f"  {_foid}: {_instr}")

                    parts = []
                    if _study_instrs:
                        parts.append("STUDY-LEVEL INSTRUCTIONS FROM HUMAN REVIEWER "
                                     "(apply to the entire study):\n" +
                                     "\n".join(f"  • {i}" for i in _study_instrs))
                    if _form_instrs:
                        parts.append("FORM-SPECIFIC INSTRUCTIONS FROM HUMAN REVIEWER:\n" +
                                     "\n".join(_form_instrs))
                    if parts:
                        ai_instructions_block = (
                            "\n\n══════════════════════════════════════════\n"
                            "AI INSTRUCTIONS — HIGHEST PRIORITY — APPLY BEFORE ALL OTHER INPUTS\n"
                            "══════════════════════════════════════════\n" +
                            "\n\n".join(parts) +
                            "\n══════════════════════════════════════════\n"
                        )
                        n_si = len(_study_instrs)
                        n_fi = len(_form_instrs)
                        print(f"AI Instructions: {n_si} study-level, "
                              f"{n_fi} form-specific", flush=True)
                        await append_log(item_id,
                            f"AI Instructions read: {n_si} study-level, "
                            f"{n_fi} form-specific.")
            except Exception as _ai_exc:
                print(f"AI Instructions read error: {_ai_exc}", flush=True)

        # ── Path R: re-use existing Study Spec JSON from monday ──────────────
        # Fast-rerun trigger #2. If struct_json is still None at this point
        # (no Path A edit, no Path M migration) AND the user did not request
        # a full run, try to download the spec_json that's already on the
        # row from a previous successful run. Saves ~7 min of Claude
        # extraction when the user is only iterating on Chains C/D.
        # Falls back to full Claude extraction silently on any error.
        if struct_json is None and not run_all:
            try:
                _existing_spec = await download_column_file(
                    item_id, COL["spec_json"])
                print(f"[fast-rerun-probe] spec_json column returned: "
                      f"{len(_existing_spec) if _existing_spec else 'None'} bytes",
                      flush=True)
                if _existing_spec:
                    struct_json = json.loads(_existing_spec.decode("utf-8"))
                    struct_json = _enforce_common_visit(struct_json)
                    struct_json = _backfill_migration_fields(struct_json)
                    struct_json = _sanitize_form_titles(struct_json)
                    fast_rerun = True
                    print(f"[fast-rerun] Using existing Study Spec JSON "
                          f"from monday ({len(_existing_spec)} bytes) — "
                          f"skipping Claude extraction", flush=True)
                    await append_log(item_id,
                        f"[fast-rerun] Re-using existing Study Spec "
                        f"JSON from monday — chains A/B/E skipped.")
            except Exception as _r_exc:
                print(f"[fast-rerun] Could not load existing spec_json "
                      f"({type(_r_exc).__name__}: {_r_exc}); falling "
                      f"back to full Claude extraction", flush=True)
                struct_json = None  # ensure fallback to Claude

        # Fresh analysis if needed and not already populated
        # B7: will hold the optional JSON-upload coroutine, to be awaited
        # concurrently with the chains at the main asyncio.gather below.
        # It's set to a coroutine only when we fresh-extract struct_json.
        spec_json_upload_task = None

        # ── Guard: no spec JSON and no protocol bytes ─────────────────────────
        # If struct_json is still None after Path A / Path M / Path R, and there
        # are no protocol bytes to give Claude, Protocol Analysis will run on
        # empty input.  Empty-input extraction can silently consume in-process
        # Railway worker state from a concurrent run (as in the VAX1001-copy
        # incident) or produce a spec for the wrong study entirely.  Fail hard
        # with a clear operator message instead.
        if struct_json is None and needs_analysis and not (protocol_bytes and len(protocol_bytes) > 1024):
            _no_proto_msg = (
                "FAILED: Protocol Analysis cannot run — no protocol document found "
                "(Protocol column is empty) and no valid Study Spec JSON is on this "
                "row to fast-rerun from.  Upload a protocol PDF or DOCX to the "
                "Protocol column and re-trigger."
            )
            print(f"Protocol Analysis blocked — no protocol bytes: {_no_proto_msg}", flush=True)
            await append_log(item_id, _no_proto_msg)
            await set_status(item_id, COL["pipeline_status"], STATUS["failed"])
            return

        if struct_json is None and needs_analysis:
            await set_status(item_id, COL["pipeline_status"], STATUS["analysis_running"])
            await append_log(item_id, "Protocol Analysis started.")

            extra_parts = []
            if ai_instructions_block:
                extra_parts.insert(0, ai_instructions_block.strip())
            if reviewer_notes_block:
                extra_parts.append(reviewer_notes_block.strip())
            # Customer Convention Questions (CQ_* / 'CQ ' columns).
            # Read dynamically from the cols dict — any column with a CQ prefix
            # is picked up automatically. New questions require zero code change.
            customer_conventions = _extract_customer_conventions(cols)
            if customer_conventions:
                cq_block = _build_customer_conventions_block(customer_conventions)
                extra_parts.append(cq_block)
                print(f"Customer conventions: {len(customer_conventions)} answer(s) provided",
                      flush=True)
                await append_log(item_id,
                    f"Customer conventions: {len(customer_conventions)} answer(s) "
                    f"injected into Study Spec generation.")
            else:
                print("Customer conventions: none provided", flush=True)
            if oc_zip:
                oc_file_type = _detect_oc_standard_type(oc_zip)
                if oc_file_type == 'ODM_XML':
                    extra_parts.append(
                        "Customer OpenClinica Study ODM XML attached — use as Priority 1 "
                        "(most authoritative source; reflects what customer has built in OC). "
                        "Extract: (a) StudyEventDef structure to understand the customer's "
                        "schedule-of-events pattern — note any StudyEventDef with "
                        "Type='Common' which indicates the customer uses a Common event for "
                        "non-visit-dependent forms like AE and CM rather than per-module "
                        "events; (b) FormDef/ItemGroupDef/ItemDef to use as form templates; "
                        "(c) CodeLists as choice list baselines; "
                        "(d) existing OIDs as the naming convention to follow."
                    )
                else:
                    extra_parts.append(
                        "Customer OC4 XLSForm Standards (ZIP) attached — use as Priority 1 "
                        "(most authoritative source; reflects what customer has built in OC)."
                    )
            if crf_pdf:
                extra_parts.append(
                    "Customer CRF Library (PDF) attached — use as Priority 2 "
                    "(fallback for any forms not found in the Priority 1 source)."
                )
            # ─── Trainer retrieval: fetch similar past pairs as few-shot examples ──
            # Gated on TRAINER_URL presence (trainer_enabled). When the trainer
            # is wired up, retrieval runs on every pipeline pass; otherwise it
            # is silently skipped (no noisy "not reachable" warnings on local
            # dev runs without a trainer).
            if not trainer_on:
                print("Step 0: Trainer retrieval — SKIPPED (TRAINER_URL not set)",
                      flush=True)
            else:
                try:
                    print("Step 0: Trainer retrieval — quick protocol analysis...", flush=True)
                    quick_analysis = await run_protocol_analysis_quick(protocol_bytes or b"")
                    if quick_analysis:
                        print(f"Step 0: Trainer retrieval — fetching examples (k={TRAINER_K})...",
                              flush=True)
                        matches = await retrieve_examples(
                            quick_analysis, k=TRAINER_K, reserve_same_sponsor=True,
                        )
                        if matches:
                            block = format_examples_block(
                                matches,
                                sponsor_hint=quick_analysis.get("sponsor"),
                                reserve_same_sponsor=True,
                            )
                            if block:
                                extra_parts.append(block)
                                await append_log(item_id,
                                    f"Trainer retrieval: {len(matches)} similar past "
                                    f"build(s) injected as examples.")
                except Exception as _trainer_exc:  # noqa: BLE001
                    print(f"Trainer retrieval failed: {_trainer_exc} — continuing without examples",
                          flush=True)

            print("Step 1: Claude extracting Study Spec JSON...", flush=True)
            # Start session keepalive — pings OC designer every 60s while Claude
            # works (7-10 min) so the Keycloak token doesn't expire before upload.
            _keepalive_task = None
            _session_file = str(AuthManager().get_session_path(oc_email)) if (
                create_study and oc_email) else None
            if _session_file and os.path.exists(_session_file):
                _keepalive_task = asyncio.create_task(
                    _session_keepalive(_session_file, oc_subdomain, interval_s=60))
                print(f"[session-keepalive] started for {oc_subdomain}", flush=True)
            # Handle DOCX-as-text fallback: protocol arrived as text, not PDF
            _docx_text_marker = b"%%DOCX_TEXT%%"
            if protocol_bytes and protocol_bytes.startswith(_docx_text_marker):
                docx_text = protocol_bytes[len(_docx_text_marker):].decode("utf-8",
                                                                            errors="replace")
                print(f"Protocol is Word text ({len(docx_text):,} chars) — "
                      f"passing as extra_text", flush=True)
                _pdf_arg   = None
                _text_args = [docx_text] + (extra_parts or [])
            else:
                _pdf_arg   = protocol_bytes or None
                _text_args = extra_parts or []

            struct_text = await call_claude(
                EDC_STRUCTURE_PROMPT,
                pdf_bytes  = _pdf_arg,
                extra_text = "\n".join(_text_args) if _text_args else None,
            )
            try:
                struct_json = extract_json(
                    struct_text,
                    expected_keys=["study_meta", "forms"],
                )
                # Normalize — handle list at top level
                if isinstance(struct_json, list):
                    struct_json = {"study_meta": {"protocol_number": protocol_num},
                                   "forms": struct_json, "review_flags": {}}
                # Normalize — ensure forms is a list of dicts, not strings
                forms = struct_json.get("forms", [])
                if forms and isinstance(forms[0], str):
                    struct_json["forms"] = [{"form_id": f, "form_title": f} for f in forms]
                print(f"Study Spec JSON extracted — "
                      f"{len(struct_json.get('forms', []))} forms, "
                      f"keys: {list(struct_json.keys())}", flush=True)
                if _keepalive_task and not _keepalive_task.done():
                    _keepalive_task.cancel()
                    print("[session-keepalive] cancelled after spec extraction",
                          flush=True)
            except ValueError:
                struct_json = {"study_meta": {"protocol_number": protocol_num},
                               "forms": [], "review_flags": {}}
                print("Warning: Study Spec JSON not valid — using empty fallback", flush=True)

            # OC-9 backstop: ensure SE_COMMON exists and AE/CM/DV/AESAE
            # forms live only there. Deterministic fix-up if Claude missed it.
            struct_json = _enforce_common_visit(struct_json)
            struct_json = _backfill_migration_fields(struct_json)
            struct_json = _sanitize_form_titles(struct_json)
            # ── Conventions engine pass (no-op until conventions/ store is populated) ─
            try:
                from conventions_engine import apply_conventions
                _study_id = (struct_json.get("study_meta") or {}).get("protocol_number") or protocol_num
                # TODO(B.1b follow-up): This path does not currently extract monday's
                # source_edc_system column (dropdown_mm382w7d). Vendor conventions
                # apply only on migration path (Path M). If non-migration builds need
                # vendor conventions in future, extract the column at build entry and
                # thread it through as migration_source here.
                apply_conventions(struct_json, study_id=_study_id,
                                  customer_subdomain=oc_subdomain)
            except Exception as _ce:
                print(f"conventions_engine FAILED — continuing without conventions: {_ce}",
                      flush=True)
                try:
                    await append_log(item_id, f"Conventions engine error (build continues): {_ce}")
                except Exception:
                    pass

            # Inject library filenames from monday columns into study_meta —
            # overrides whatever Claude may have guessed for
            # library_files_provided. This ensures the Study Spec PDF shows
            # the actual files that were uploaded to the item.
            if library_files_provided:
                sm = struct_json.setdefault("study_meta", {})
                sm["library_files_provided"] = library_files_provided

            # B7: JSON upload is prepared as a coroutine but NOT awaited here —
            # it runs concurrently with chains A-D via the main gather() below.
            # This block only executes when we freshly extracted struct_json
            # (guaranteed by the outer `if struct_json is None` guard).
            spec_json_upload_task = upload_file(
                item_id, COL["spec_json"],
                f"{protocol_num}_Study_Specification_{version}.json",
                json.dumps(struct_json, indent=2).encode()
            )

            # ── Trainer: create pending row on trainer board ────────────────
            # Fires unconditionally on every successful pipeline run when the
            # trainer is wired up (TRAINER_URL set). Best-effort — any failure
            # is logged but does not block the pipeline. The corpus dedup
            # logic in /pending-row prevents duplicates. The new row sits in
            # "Awaiting Build Completion" status until a human uploads the
            # final form definitions.
            if trainer_on and protocol_bytes:
                try:
                    sponsor_hint = (struct_json.get("study_meta", {})
                                    .get("sponsor")
                                    or struct_json.get("study_meta", {})
                                    .get("sponsor_name"))
                    print(f"[trainer] creating pending row: name={protocol_num!r} "
                          f"sponsor={sponsor_hint!r}", flush=True)
                    new_trainer_item_id = await create_pending_row(
                        protocol_bytes,
                        name=protocol_num,
                        protocol_filename=f"{protocol_num}.pdf",
                        sponsor_client=sponsor_hint,
                        source_pipeline_item=str(item_id),
                        protocol_number=protocol_num,  # Required for deduplication
                    )
                    if new_trainer_item_id:
                        await append_log(
                            item_id,
                            f"Trainer pending row created: item_id={new_trainer_item_id}",
                        )
                except Exception as _trainer_exc:  # noqa: BLE001
                    print(f"[trainer] create_pending_row failed: {_trainer_exc} "
                          f"— continuing without trainer row", flush=True)

        # ── Mapping review UI deep-link ─────────────────────────────────────
        # Populate COL["mapping_review_url"] once we have a Study Spec JSON,
        # regardless of which path produced it (Path B protocol-PDF or Path M
        # migration). Gated on MAPPING_UI_URL env — when unset (local dev,
        # mapping-ui not yet deployed) the column write is skipped silently.
        # Best-effort; failure does not block the chains.
        if struct_json:
            mapping_ui_base = os.environ.get("MAPPING_UI_URL", "").strip().rstrip("/")
            if mapping_ui_base:
                review_url = f"{mapping_ui_base}/?item={item_id}&board={BOARD_ID}"
                try:
                    await set_link(item_id, COL["mapping_review_url"], review_url,
                                   text="Open Mapping Review")
                    await append_log(item_id,
                        f"Mapping review URL written: {review_url}")
                except Exception as _link_exc:  # noqa: BLE001
                    print(f"mapping_review_url write failed (non-fatal): "
                          f"{_link_exc}", flush=True)

        # Scheduling-block second pass: if protocol-analysis did not emit a
        # scheduling array (common when the main call is large), extract it
        # now from the already-resolved timepoint_csv rows. Cheap, focused,
        # and mutates struct_json in place before any chain consumes it.
        if struct_json and not struct_json.get("scheduling"):
            struct_json = _extract_scheduling_block(struct_json)

        # ── Launch parallel chains if struct_json is available ────────────────
        if struct_json and needs_analysis:
            await set_status(item_id, COL["pipeline_status"], STATUS["build_pricing_running"])
            await append_log(item_id, "Chains A (spec files), B (summary+quote), C (build+DVS), D (OC study) starting in parallel.")

            # Shared state for parallel chains
            pricing_json = {"study_meta": {"protocol_number": protocol_num}}

            # ── Chain A: Study Spec files ──────────────────────────────────────
            async def chain_a():
                if not _want("protocol specification"):
                    return
                print("Chain A: Generating Study Spec PDF + XLSX (local)...", flush=True)
                try:
                    loop = asyncio.get_event_loop()
                    spec_files = await loop.run_in_executor(
                        None, lambda: run_study_spec_files(struct_json, oc_subdomain, None)
                    )
                    await asyncio.gather(
                        upload_file(item_id, COL["spec_pdf"],
                            f"{protocol_num}_Study_Specification_{version}.pdf",
                            spec_files["pdf"]),
                        upload_file(item_id, COL["spec_xlsx"],
                            f"{protocol_num}_Study_Specification_{version}.xlsx",
                            spec_files["xlsx"]),
                    )
                    print(f"Chain A complete — pdf:{len(spec_files['pdf'])} bytes "
                          f"xlsx:{len(spec_files['xlsx'])} bytes", flush=True)
                except Exception as e:
                    print(f"Chain A error: {e}", flush=True)
                    traceback.print_exc()
                    await append_log(item_id, f"Study Spec file generation error: {e}")

            # ── Chain B: Protocol Summary JSON → PDF + Quote ───────────────────
            async def chain_b():
                nonlocal pricing_json
                want_summary = _want("protocol summary")
                want_quote   = _want("price quote")
                if not want_summary and not want_quote:
                    return

                # Both branches need pricing_json — extract it once
                print("Chain B: Claude extracting Protocol Summary JSON...", flush=True)
                # Slim struct_json for Protocol Summary — keep keys that use
                # the actual Study Spec field names (form_id, form_title,
                # cdash_domain). Previously slimming to "name"/"domain" lost
                # all form identity → empty Protocol Summary → empty Quote.
                # Strip convention-engine debug payloads out of study_meta —
                # these are sized for engine audit (≈3 MB / 750K tokens for
                # CRS-136) and would blow the Claude 1M context cap on the
                # Chain B Protocol Summary call. Chain B only needs the
                # protocol-metadata fields. See pipeline failure 2026-05-18
                # at 19:19:25 UTC (1.22M-token prompt rejected).
                _META_BLOAT_KEYS = {
                    "conventions_prompt_block",
                    "conventions_engine_applied",
                    "customer_vendor_conflicts",
                }
                _study_meta_full = struct_json.get("study_meta", {})
                struct_slim = {
                    "study_meta":    {k: v for k, v in _study_meta_full.items()
                                      if k not in _META_BLOAT_KEYS},
                    "timepoint_csv": struct_json.get("timepoint_csv", {}),
                    "review_flags":  struct_json.get("review_flags", {}),
                    "forms": [
                        {"form_id":         f.get("form_id", ""),
                         "form_title":      f.get("form_title", ""),
                         "cdash_domain":    f.get("cdash_domain", ""),
                         "form_category":   f.get("form_category", ""),
                         "complexity":      f.get("complexity", ""),
                         "visits_assigned": f.get("visits_assigned", []),
                         "reuse_count":     f.get("reuse_count", 1),
                         "is_epro":         f.get("is_epro", False),
                         "has_repeating_group": f.get("has_repeating_group", False),
                         "arm_applicability":   f.get("arm_applicability", "ALL")}
                        for f in struct_json.get("forms", [])
                        if isinstance(f, dict)
                    ],
                }
                pricing_text = await call_claude(
                    PRICING_SUMMARY_PROMPT,
                    extra_text="Study Specification JSON:\n" + json.dumps(struct_slim),
                    max_tokens=64000,  # B4: Chain B may also produce large output
                )
                pricing_json_valid = False
                try:
                    pricing_json = extract_json(
                        pricing_text,
                        expected_keys=["study_meta", "patient_population",
                                       "visit_summary", "crf_summary"],
                    )
                    if isinstance(pricing_json, list):
                        print("Warning: Protocol Summary was a list, not a dict", flush=True)
                        pricing_json = {"study_meta": {"protocol_number": protocol_num}}
                    else:
                        # B5: Validate required keys actually came through.
                        # Fewer than 3 of the 4 expected keys means extraction
                        # gave us a fragment, not a complete summary.
                        expected = {"study_meta", "patient_population",
                                    "visit_summary", "crf_summary"}
                        present  = expected & set(pricing_json.keys())
                        if len(present) >= 3:
                            pricing_json_valid = True
                            print(f"Protocol Summary JSON extracted — "
                                  f"keys: {list(pricing_json.keys())}", flush=True)
                        else:
                            print(f"Warning: Protocol Summary missing expected keys. "
                                  f"Got: {sorted(present)}. Skipping Chain B work.", flush=True)
                except ValueError:
                    print("Warning: Protocol Summary JSON not valid — "
                          "skipping Chain B work.", flush=True)

                # B1+B5: Bail out of Chain B entirely if pricing_json is unusable.
                # Prevents uploading empty/garbage PDFs to monday.com.
                if not pricing_json_valid:
                    await append_log(item_id,
                        "Chain B skipped — Protocol Summary JSON invalid or incomplete. "
                        "Study Spec JSON is still available; re-run may recover.")
                    return

                # Steps 4 + 5 in parallel: Protocol Summary PDF + Pricing Quote
                async def gen_ps_pdf():
                    if not want_summary:
                        return
                    print("Chain B: Generating Protocol Summary PDF (local)...", flush=True)
                    try:
                        loop = asyncio.get_event_loop()
                        ps_pdf = await loop.run_in_executor(
                            None, lambda: run_protocol_summary_pdf(pricing_json, struct_json)
                        )
                        uploads = [upload_file(item_id, COL["pricing_summary"],
                            f"{protocol_num}_Protocol_Summary_{version}.json",
                            json.dumps(pricing_json, indent=2).encode())]
                        if ps_pdf:
                            uploads.append(upload_file(item_id, COL["pricing_summary"],
                                f"{protocol_num}_Protocol_Summary_{version}.pdf", ps_pdf))
                        await asyncio.gather(*uploads)
                        print(f"Protocol Summary PDF: {len(ps_pdf) if ps_pdf else 0} bytes", flush=True)
                    except Exception as e:
                        print(f"Protocol Summary PDF error: {e}", flush=True)
                        traceback.print_exc()
                        await append_log(item_id, f"Protocol Summary PDF error: {e}")
                        raise   # B6: surface failure to chain_b

                async def gen_quote():
                    if not want_quote:
                        return
                    print("Chain B: Generating Price Quote (local scripts)...", flush=True)
                    try:
                        loop = asyncio.get_event_loop()
                        qf = await loop.run_in_executor(
                            None,
                            lambda: run_pricing_quote(
                                pricing_json,
                                additional_sub_disc=additional_sub_disc,
                                additional_svc_disc=additional_svc_disc,
                                edc_structure=struct_json,
                            )
                        )
                        await asyncio.gather(
                            upload_file(item_id, COL["pricing_quote"],
                                f"{protocol_num}_Quote_Internal_{version}.pdf",  qf["internal_pdf"]),
                            upload_file(item_id, COL["pricing_quote"],
                                f"{protocol_num}_Quote_Client_{version}.pdf",    qf["client_pdf"]),
                            upload_file(item_id, COL["pricing_quote"],
                                f"{protocol_num}_Quote_Internal_{version}.xlsx", qf["internal_xlsx"]),
                            upload_file(item_id, COL["pricing_quote"],
                                f"{protocol_num}_Quote_Client_{version}.xlsx",   qf["client_xlsx"]),
                        )
                        await append_log(item_id, "Price Quote complete — 4 files uploaded.")
                    except Exception as e:
                        print(f"Price Quote error: {e}", flush=True)
                        traceback.print_exc()
                        await append_log(item_id, f"Price Quote error: {e}")
                        raise   # B6: surface failure to chain_b

                # B6: use return_exceptions so gen_ps_pdf failure doesn't cancel
                # gen_quote (or vice versa). If either raised, chain_b re-raises
                # the first exception so chain-level tracking sees the failure.
                sub_results = await asyncio.gather(gen_ps_pdf(), gen_quote(),
                                                    return_exceptions=True)
                sub_errors = [r for r in sub_results if isinstance(r, Exception)]
                if sub_errors:
                    await append_log(item_id,
                        f"Chain B finished with {len(sub_errors)} subtask failure(s).")
                    print(f"Chain B had {len(sub_errors)} failure(s); re-raising first.",
                          flush=True)
                    raise sub_errors[0]
                await append_log(item_id, "Chain B complete.")
                print("Chain B complete.", flush=True)

            # ── Chain C: EDC Build → DVS ──────────────────────────────────────
            build_zip_holder  = [None]   # mutable container for async closure
            build_json_holder = [{"forms": {}}]
            # Chain D writes forms_publish (dict from create_oc_study) here so
            # the post-pipeline auto-trigger can read uploaded_oids without
            # locals() — chain_d's locals are not visible from run_pipeline.
            forms_publish_holder = [None]
            # Event set when chain_c's build is done (or if chain_c is skipped).
            # Chain E waits on this instead of re-triggering a duplicate build
            # when both "study build zip" and "build preview" are selected.
            edc_build_event = asyncio.Event()

            async def chain_c():
                if not _want("study build zip"):
                    edc_build_event.set()   # not running — unblock chain_e
                    return
                print("Chain C: EDC Build starting...", flush=True)
                try:
                    await _run_edc_and_dvs(suppress_uploads=False)
                finally:
                    edc_build_event.set()   # always unblock chain_e
                print("Chain C complete.", flush=True)

            async def _run_edc_and_dvs(suppress_uploads=False):

                if edited_build_zip:
                    # Path B: User uploaded edited XLSForm ZIP
                    print("Path B: using user-uploaded XLSForm ZIP...", flush=True)
                    await append_log(item_id, "Using user-uploaded XLSForm ZIP.")
                    try:
                        build_json_holder[0] = _read_zip_xlsforms(edited_build_zip)
                        build_zip_holder[0]  = edited_build_zip
                    except Exception as e:
                        print(f"Error reading build ZIP: {e}", flush=True)
                        await append_log(item_id, f"Error reading build ZIP: {e}")

                elif edited_dvs_xlsx:
                    # Path C: User uploaded edited DVS → translate to XLSForms
                    print("Path C: translating DVS changes to XLSForms...", flush=True)
                    await append_log(item_id, "Translating DVS input to XLSForm updates.")
                    dvs_text = _dvs_xlsx_to_text(edited_dvs_xlsx)

                    # Claude generates XLSForm JSON directly (no file generation)
                    build_prompt_json = (
                        "You need to generate the XLSForm structure as JSON.\n"
                        "Return a single valid JSON object with this structure:\n"
                        '{"forms": {"<filename>.xlsx": {"survey": [...], "choices": [...], "settings": {...}}}}\n'
                        "No text before or after. No markdown code fences.\n\n"
                    )
                    struct_slim = {
                        "study_meta": struct_json.get("study_meta", {}) if struct_json else {},
                        "forms":      struct_json.get("forms", []) if struct_json else [],
                    }
                    base_build_text = await call_claude(
                        build_prompt_json,
                        extra_text="Study Specification JSON:\n" + json.dumps(struct_slim),
                    )
                    try:
                        base_build = extract_json(base_build_text)
                        if isinstance(base_build, list):
                            base_build = {"forms": {}}
                    except ValueError:
                        base_build = {"forms": {}}

                    updated_text = await call_claude(
                        DVS_TRANSLATE_PROMPT,
                        extra_text=("Current XLSForm JSON:\n" + json.dumps(base_build) +
                                    "\n\nDVS Changes:\n" + dvs_text),
                    )
                    try:
                        build_json_holder[0] = extract_json(updated_text)
                    except ValueError:
                        build_json_holder[0] = base_build
                    build_zip_holder[0] = _xlsform_zip(build_json_holder[0])

                else:
                    # Fresh run: EDC builder (local scripts)
                    print("Chain C: Running edc-builder (local)...", flush=True)
                    try:
                        loop = asyncio.get_event_loop()
                        zip_bytes, edc_log, forms_json = await loop.run_in_executor(
                            None, lambda: run_edc_build(struct_json)
                        )
                        build_zip_holder[0]  = zip_bytes
                        build_json_holder[0] = forms_json
                        # Summarise validation results if present
                        v_results = edc_log.get('validation_results', [])
                        if v_results:
                            n_total  = len(v_results)
                            n_err    = sum(1 for r in v_results if r.get('errors'))
                            n_warn   = sum(1 for r in v_results if r.get('warnings'))
                            n_skip   = sum(1 for r in v_results if r.get('skipped'))
                            v_msg    = (f"validation: {n_total} forms, "
                                        f"{n_err} with errors, {n_warn} with warnings"
                                        + (f", {n_skip} skipped" if n_skip else ""))
                        else:
                            v_msg = "validation: not run"
                        print(f"EDC Build complete — "
                              f"built={len(edc_log.get('forms_built', []))} "
                              f"zip={len(zip_bytes) if zip_bytes else 0} bytes "
                              f"| {v_msg}",
                              flush=True)
                    except Exception as e:
                        print(f"EDC Build error: {e}", flush=True)
                        traceback.print_exc()
                        await append_log(item_id, f"EDC Build error: {e}")
                        raise   # B6: propagate to chain_c

                if build_zip_holder[0] and not suppress_uploads:
                    await upload_file(item_id, COL["edc_build"],
                                      f"{protocol_num}_EDC_Build_{version}.zip",
                                      build_zip_holder[0])
                    await append_log(item_id, "EDC Build complete — ZIP uploaded.")
                elif build_zip_holder[0] and suppress_uploads:
                    print("EDC Build complete (preview-only mode — zip not uploaded)",
                          flush=True)

                # DVS — skip when called as a silent prerequisite for Build Preview
                if suppress_uploads:
                    print("DVS skipped (preview-only mode)", flush=True)
                    return

                print("Chain C: Running DVS (local)...", flush=True)
                try:
                    loop = asyncio.get_event_loop()
                    dvs_xlsx = await loop.run_in_executor(
                        None,
                        lambda: run_dvs_xlsx(
                            struct_json if struct_json
                                else {"study_meta": {"protocol_number": protocol_num}},
                            build_json_holder[0] or {"forms": {}},
                        ),
                    )
                    if dvs_xlsx:
                        await upload_file(item_id, COL["dvs_output"],
                                          f"{protocol_num}_DVS_{version}.xlsx",
                                          dvs_xlsx)
                        await append_log(item_id, "DVS complete.")
                    else:
                        await append_log(item_id, "DVS skipped — builder unavailable.")
                except Exception as e:
                    print(f"DVS error: {e}", flush=True)
                    traceback.print_exc()
                    await append_log(item_id, f"DVS error: {e}")
                    raise   # B6: propagate to chain_c

                # Calendaring Rules (local) — Tier 1 mechanical rules. Gated on
                # the "calendaring rules" output-requested label (_want matches
                # the dropdown TEXT, not the numeric label ID).
                if _want("calendaring rules"):
                    print("Chain C: Running Calendaring Rules (local)...", flush=True)
                    try:
                        loop = asyncio.get_event_loop()
                        cal_zip = await loop.run_in_executor(
                            None,
                            lambda: run_calendaring_rules(
                                struct_json if struct_json
                                    else {"study_meta": {"protocol_number": protocol_num}},
                                build_json_holder[0] or {"forms": {}},
                            ),
                        )
                        if cal_zip:
                            await upload_file(item_id, COL["calendaring_output"],
                                              f"{protocol_num}_Calendaring_{version}.zip",
                                              cal_zip)
                            await append_log(item_id, "Calendaring Rules complete.")
                        else:
                            await append_log(item_id,
                                             "Calendaring Rules skipped — skill unavailable.")
                    except Exception as e:
                        print(f"Calendaring Rules error: {e}", flush=True)
                        traceback.print_exc()
                        await append_log(item_id, f"Calendaring Rules error: {e}")
                        # Calendaring is additive — log and continue, do not abort the chain

            # ── Chain D: Create OC Study (parallel, only needs struct_json) ───
            async def chain_d():
                if not (create_study and oc_subdomain and struct_json):
                    if create_study and not oc_subdomain:
                        await append_log(item_id, "Create Study requested but no OC Subdomain — skipped.")
                    return
                env_label = "production" if oc_production else "test"

                # ── Spec-change detection ─────────────────────────────────
                # If a study is already published (Study UUID column populated)
                # and the spec JSON hasn't changed since the last upload, skip
                # form re-upload and re-publish entirely — the study is current.
                # Only re-upload if the spec actually changed.
                _existing_uuid_d = (cols.get(COL["study_uuid"], {})
                                    .get("text") or "").strip()
                if _existing_uuid_d and struct_json:
                    import hashlib as _hashlib
                    _spec_bytes = json.dumps(struct_json, sort_keys=True,
                                            ensure_ascii=False).encode()
                    _spec_hash  = _hashlib.sha256(_spec_bytes).hexdigest()[:16]
                    _saved_rec  = _read_upload_record(str(item_id))
                    _saved_hash = _saved_rec.get("spec_hash", "")
                    if _saved_hash and _saved_hash == _spec_hash:
                        print(f"[chain-d] Spec hash unchanged ({_spec_hash}) "
                              f"— skipping form upload and publish, study "
                              f"is already current.", flush=True)
                        await append_log(item_id,
                            "Study Spec unchanged since last upload — "
                            "form re-upload and publish skipped.")
                        # Populate forms_publish_holder with a no-op result
                        # so downstream status logic doesn't false-alarm.
                        forms_publish_holder[0] = {
                            "forms_uploaded": 0,
                            "forms_total": 0,
                            "errors": [],
                            "uploaded_oids": [],
                            "skipped_reason": "spec_unchanged",
                        }
                        return

                await append_log(item_id, f"Creating study in OpenClinica {env_label} ({oc_subdomain})...")
                try:
                    # Wait for Chain C (EDC Build) to complete before fetching
                    # the EDC Build ZIP URL — prevents race condition where we
                    # look for the asset before Chain C uploads it to Monday.
                    print("Chain D: waiting for Chain C EDC build…", flush=True)
                    await edc_build_event.wait()
                    
                    # Fetch the EDC build ZIP URL from monday so
                    # create_oc_study can upload XLSForm files after the
                    # board import (required to clear "No form version
                    # defined" before publish_to_test). Best-effort: if
                    # we can't find the asset, just skip form upload and
                    # log — study creation still proceeds.
                    edc_zip_url = None
                    try:
                        from monday_client import get_asset_url
                        assets = await get_asset_url(item_id)
                        for asset in assets or []:
                            name = (asset.get("name") or "")
                            if "EDC_Build" in name and name.endswith(".zip"):
                                edc_zip_url = (asset.get("public_url")
                                               or asset.get("url"))
                                break
                        if not edc_zip_url:
                            print("Chain D: no EDC build ZIP asset found on "
                                  "item — form publish will be skipped",
                                  flush=True)
                    except Exception as _e:
                        print(f"Chain D: edc_zip_url lookup failed ({_e}); "
                              f"form publish will be skipped", flush=True)

                    result = await create_oc_study(oc_subdomain, struct_json,
                                                    is_production=oc_production,
                                                    edc_zip_url=edc_zip_url,
                                                    oc_email=oc_email,
                                                    item_id=item_id,
                                                    fast_rerun=fast_rerun)
                    study_url      = result["study_url"]
                    board_imported = result["board_imported"]
                    board_error    = result.get("board_error", "")
                    forms_publish  = result.get("forms_publish")
                    forms_publish_holder[0] = forms_publish
                    await set_text(item_id, COL["oc_study_url"], study_url)
                    # Persist the raw UUID separately so publish_to_test can
                    # read it without parsing the URL (which historically
                    # carried board_id, not study_uuid).
                    await set_text(item_id, COL["study_uuid"], result["study_uuid"])
                    if board_imported:
                        await append_log(item_id,
                            f"Study + design board created: {study_url}")
                    else:
                        await append_log(item_id,
                            f"Study shell created: {study_url}  |  "
                            f"Design board import skipped — {board_error}")
                    if forms_publish:
                        fp = forms_publish
                        await append_log(item_id,
                            f"Form-version upload: {fp['forms_uploaded']}/"
                            f"{fp['forms_total']} succeeded"
                            + (f" — errors: {fp['errors'][:2]}"
                               if fp['errors'] else ""))
                    print(f"Chain D complete: {study_url} "
                          f"(board_imported={board_imported})", flush=True)
                except Exception as e:
                    print(f"OC Study error: {e}", flush=True)
                    await append_log(item_id, f"OC Study creation failed: {e}")
                    raise   # B6: propagate to the chain-outcome tracker

            # ── Chain E: Build Preview PDF (local renderer, no Claude API) ────
            # Consumes the in-memory `struct_json` and the EDC build zip held by
            # `build_zip_holder[0]` (populated by chain_c). If the user only
            # selected "build preview" in the dropdown (chain_c skipped), we
            # trigger _run_edc_and_dvs() inline so the build zip exists before
            # we render. Renderer is fully local — no Claude API calls.
            async def chain_e():
                if not _want("build preview"):
                    return
                if not struct_json:
                    print("Chain E: skipped — no struct_json available", flush=True)
                    return
                # Build Preview needs the EDC zip from chain_c. If chain_c isn't
                # producing one (chain_c gated off), trigger the build inline.
                if not build_zip_holder[0]:
                    if _want("study build zip"):
                        # Chain C is running — wait for it rather than triggering
                        # a duplicate build (which would double-upload the zip/DVS)
                        print("Chain E: waiting for Chain C EDC build…", flush=True)
                        await edc_build_event.wait()
                    else:
                        # Build Preview selected without Study Build ZIP —
                        # run a silent build (no uploads to monday) just to get
                        # the zip bytes needed for rendering.
                        print("Chain E: building EDC zip (preview-only, no upload)…",
                              flush=True)
                        try:
                            await _run_edc_and_dvs(suppress_uploads=True)
                        except Exception as e:
                            print(f"Chain E: EDC build failed: {e}", flush=True)
                            await append_log(item_id,
                                f"Build Preview skipped — EDC build failed: {e}")
                            return
                if not build_zip_holder[0]:
                    print("Chain E: skipped — EDC zip unavailable", flush=True)
                    await append_log(item_id,
                        "Build Preview skipped — no EDC build zip produced.")
                    return

                print("Chain E: rendering Build Preview PDF (local)...",
                      flush=True)
                await append_log(item_id, "Build Preview started.")
                try:
                    from build_preview import render_build_preview_from_spec
                    loop = asyncio.get_event_loop()

                    # Defensive unpack — render returns (pdf_bytes, html_zip_bytes)
                    # Use indexed access to avoid ValueError if shape is unexpected.
                    _render_out = await loop.run_in_executor(
                        None,
                        lambda: render_build_preview_from_spec(
                            struct_json, build_zip_holder[0], protocol_num),
                    )
                    print(f"Chain E: render returned type={type(_render_out).__name__} "
                          f"len={len(_render_out) if hasattr(_render_out,'__len__') else 'N/A'}",
                          flush=True)

                    # Accept tuple, list, or plain bytes (legacy fallback)
                    if isinstance(_render_out, (tuple, list)) and len(_render_out) >= 2:
                        pdf_bytes      = _render_out[0]
                        html_zip_bytes = _render_out[1]
                    elif isinstance(_render_out, (tuple, list)) and len(_render_out) == 1:
                        pdf_bytes      = _render_out[0]
                        html_zip_bytes = b""
                    else:
                        pdf_bytes      = _render_out if isinstance(_render_out, bytes) else b""
                        html_zip_bytes = b""

                    # Upload PDF
                    if pdf_bytes:
                        await upload_file(item_id, COL["build_preview"],
                            f"{protocol_num}_Build_Preview_{version}.pdf",
                            pdf_bytes)

                    # Upload interactive ZIP to the same column if produced
                    if html_zip_bytes:
                        await upload_file(item_id, COL["build_preview"],
                            f"{protocol_num}_Form_Simulator_{version}.zip",
                            html_zip_bytes)

                    await append_log(item_id,
                        f"Build Preview complete — PDF {len(pdf_bytes):,} bytes "
                        + (f"+ Simulator ZIP {len(html_zip_bytes):,} bytes" if html_zip_bytes else ""))
                    print(f"Chain E complete — PDF {len(pdf_bytes):,}b "
                          f"+ ZIP {len(html_zip_bytes):,}b",
                          flush=True)
                except Exception as e:
                    import io as _io
                    _tb_buf = _io.StringIO()
                    traceback.print_exc(file=_tb_buf)
                    _tb_str = _tb_buf.getvalue()
                    # Print each line separately so Railway doesn't truncate
                    for _line in _tb_str.splitlines():
                        print(f"[chain_e_tb] {_line}", flush=True)
                    await append_log(item_id, f"Build Preview error: {e}")

            # ── Launch all four chains in parallel ─────────────────────────────
            # return_exceptions=True prevents one chain's failure from cancelling
            # others. B7: the Study Spec JSON upload runs concurrently here too
            # (saves 1-3s vs blocking before chain launch).
            tasks       = [chain_a(), chain_b(), chain_c(), chain_d(), chain_e()]
            task_names  = ["A", "B", "C", "D", "E"]
            if spec_json_upload_task is not None:
                tasks.append(spec_json_upload_task)
                task_names.append("spec_json_upload")
            results = await asyncio.gather(*tasks, return_exceptions=True)

            # B2: track outcomes so the final status reflects reality.
            failed_chains = []
            for name, result in zip(task_names, results):
                if isinstance(result, Exception):
                    failed_chains.append(name)
                    print(f"Task {name} exception escaped: {result}", flush=True)
                    await append_log(item_id, f"Task {name} error: {result}")

            # Read the post-completion checkboxes UP FRONT so we can
            # decide whether to set pipeline_status="All Complete" now
            # or defer it until publish-to-test succeeds. Previously
            # this read happened later in the auto-trigger block, so
            # the row briefly showed "All Complete" while publish was
            # still in progress — confusing for operators and outright
            # wrong if publish then failed.
            publish_checked  = False
            load_uat_checked = False
            try:
                _item_pre = await get_item(item_id)
                _cols_pre = {c["id"]: c for c in
                             _item_pre.get("column_values", [])}
                publish_checked  = (_cols_pre.get("boolean_mm3g2vzf", {})
                                    .get("text") == "v")
                load_uat_checked = (_cols_pre.get("boolean_mm3gxe49", {})
                                    .get("text") == "v")
            except Exception as _ce:
                print(f"[pipeline] checkbox read pre-status failed: "
                      f"{_ce}", flush=True)

            if failed_chains:
                final_status = STATUS["failed"]
                final_log    = (f"Pipeline finished with errors in chains: "
                                f"{', '.join(failed_chains)}. Check uploaded files "
                                f"and logs above for details.")
                await asyncio.gather(
                    set_status(item_id, COL["pipeline_status"], final_status),
                    append_log(item_id, final_log),
                )
            elif (forms_publish_holder[0]
                  and isinstance(forms_publish_holder[0], dict)
                  and forms_publish_holder[0].get("errors")
                  and forms_publish_holder[0].get("forms_uploaded", 0) == 0
                  and forms_publish_holder[0].get("forms_total", 0) > 0):
                # Form upload was attempted but EVERY form failed (e.g.
                # the SSO session died mid-run). The build chains
                # "succeeded" but the study is not usable — do NOT report
                # "Build Complete", which falsely implies success. Mark
                # Failed so the row reflects the real outcome.
                _fp = forms_publish_holder[0]
                final_status = STATUS["failed"]
                print(f"[pipeline] Form upload failed "
                      f"({_fp.get('forms_uploaded')}/"
                      f"{_fp.get('forms_total')} uploaded, "
                      f"{len(_fp.get('errors') or [])} error(s)) — "
                      f"setting Failed, not Build Complete", flush=True)
                await asyncio.gather(
                    set_status(item_id, COL["pipeline_status"],
                               final_status),
                    append_log(item_id,
                        "Pipeline FAILED: form upload to OpenClinica did "
                        "not succeed (0 forms uploaded). See logs above."),
                )
            elif publish_checked:
                # Chains succeeded but publish-to-test is queued.
                # Defer "All Complete" until publish reports
                # success — set it down in the auto-trigger block
                # below after publish_to_test returns and
                # published_status reads "Published".
                # Without this set_status the column would be left
                # at whatever in-progress label the last chain wrote
                # (typically "Build + Pricing Running"), which then
                # sticks if publish later fails. "Build Complete"
                # is the correct interstitial state: the build IS
                # complete, only publish-to-test remains.
                print("[pipeline] Build chains complete — deferring "
                      "'All Complete' until publish-to-test succeeds",
                      flush=True)
                await asyncio.gather(
                    set_status(item_id, COL["pipeline_status"],
                               STATUS["build_complete"]),
                    append_log(item_id,
                               "Pipeline build complete. "
                               "Awaiting publish to test."),
                )
            else:
                await asyncio.gather(
                    set_status(item_id, COL["pipeline_status"],
                               STATUS["all_complete"]),
                    append_log(item_id,
                               "Pipeline complete. All outputs uploaded."),
                )

            # ── Migration post-build hook: gap analysis + Migrations Hub upsert ──
            # Runs only when this row was routed through Path M. Wrapped
            # in try/except so any failure here cannot retroactively fail
            # a build that has already posted its all_complete / failed
            # status above. The Migrations Hub row is the long-lived
            # per-study record; we upsert by study OID, write the gap
            # report file, and post the Syndeo URL for human review.
            if _migration_post_build_state and struct_json:
                try:
                    from migration_pipeline import run_gap_analysis_and_hub_upsert
                    await run_gap_analysis_and_hub_upsert(
                        item_id,
                        odm_metadata=_migration_post_build_state[
                            "odm_metadata"],
                        spec_json=struct_json,
                        source_system=_migration_post_build_state[
                            "source_system"] or "UNKNOWN",
                        source_odm_bytes=_migration_post_build_state.get(
                            "source_odm_bytes"),
                        source_odm_filename=_migration_post_build_state.get(
                            "source_odm_filename"),
                    )
                except Exception as _ga_exc:
                    # Non-fatal — the per-step error handling inside the
                    # helper already logs to monday. Belt-and-braces.
                    print(f"[gap-analysis] post-build hook crashed "
                          f"(non-fatal): {type(_ga_exc).__name__}: "
                          f"{_ga_exc}", flush=True)

            # ── Auto-trigger post-completion actions if checkboxes checked ──
            # The /webhook/publish_test handler only fires when the checkbox
            # CHANGES — if it was already checked when the pipeline started,
            # it never re-fires after run_pipeline ends. publish_checked /
            # load_uat_checked were read up above (before the chain-result
            # status decision) so we'd know whether to defer 'All
            # Complete'; reuse those values here.
            try:
                _cal_publish_already_complete = False
                if publish_checked:
                    # Forward the just-uploaded OIDs to publish_to_test so
                    # its pre-flight doesn't false-positive on the OC REST
                    # API's propagation delay for newly-uploaded versions.
                    # forms_publish_holder is populated by chain_d; None if
                    # Chain D was skipped or failed before uploads ran.
                    _fp = forms_publish_holder[0]
                    _uploaded_oids = (
                        set(_fp.get("uploaded_oids") or [])
                        if isinstance(_fp, dict) else None
                    )
                    print(f"[auto-publish] Publish to Test checkbox is "
                          f"checked — starting publish "
                          f"(trusting {len(_uploaded_oids or [])} "
                          f"just-uploaded OIDs)", flush=True)
                    await publish_to_test(
                        item_id, uploaded_oids=_uploaded_oids)

                    # publish_to_test never raises — it writes its
                    # outcome to COL["published_status"]. Read that
                    # column back to decide pipeline_status. Only
                    # promote pipeline_status to 'All Complete' on a
                    # "Published" result. On "Failed" we leave
                    # pipeline_status alone — published_status="Failed"
                    # already tells operators publish broke, and we
                    # don't want to overwrite the "Awaiting publish
                    # to test" state with a misleading 'All Complete'.
                    try:
                        _post_item = await get_item(item_id)
                        _post_cols = {c["id"]: c for c in
                                      _post_item.get("column_values", [])}
                        _published_status = (
                            _post_cols.get(COL["published_status"], {})
                            .get("text") or ""
                        ).strip()
                    except Exception as _pe:
                        _published_status = ""
                        print(f"[auto-publish] post-publish status "
                              f"read failed: {_pe}", flush=True)
                    if _published_status == "Published":
                        if load_uat_checked:
                            # UAT load is also queued — don't jump to
                            # "All Complete" yet or the UAT loader will
                            # immediately overwrite it with "Loading UAT
                            # Data", creating a confusing flash. Let the
                            # UAT loader set the final pipeline status.
                            print(f"[auto-publish] publish-to-test "
                                  f"succeeded — deferring 'All Complete' "
                                  f"until UAT load finishes", flush=True)
                            await append_log(item_id,
                                             "Publish to test succeeded. "
                                             "Starting UAT data load...")
                        else:
                            _cal_publish_already_complete = True
                            await asyncio.gather(
                                set_status(item_id,
                                           COL["pipeline_status"],
                                           STATUS["all_complete"]),
                                append_log(item_id,
                                           "Pipeline complete. Publish to "
                                           "test succeeded."),
                            )
                            print(f"[auto-publish] publish-to-test "
                                  f"succeeded → pipeline_status="
                                  f"'All Complete'", flush=True)
                    else:
                        print(f"[auto-publish] publish-to-test "
                              f"published_status={_published_status!r}"
                              f" — leaving pipeline_status as-is",
                              flush=True)

                # ── Publish Calendaring Rules (optional, gated on checkbox) ─────────
                _publish_cal = (cols.get(COL["publish_cal_rules"], {})
                                .get("text") or "").strip() == "v"
                # study_uuid is written to Monday by Chain D — read it fresh here
                _post_cal_item = await get_item(item_id)
                _post_cal_cols = {c["id"]: c for c in
                                  _post_cal_item.get("column_values", [])}
                _study_uuid_for_cal = (
                    _post_cal_cols.get(COL["study_uuid"], {}).get("text") or ""
                ).strip()
                if _publish_cal and oc_subdomain and _study_uuid_for_cal:
                    print("[cal-publish] Publish Calendaring Rules checkbox is checked — starting upload", flush=True)
                    await set_status(item_id, COL["pipeline_status"], "Publishing Calendaring Rules")
                    try:
                        # Download cal zip from Monday board output column
                        _cal_zip = await download_column_file(item_id, COL["calendaring_output"])
                        if not _cal_zip:
                            await append_log(item_id, "Calendaring publish skipped — no calendaring output found on board.")
                        else:
                            _cal_summary = await publish_calendaring_rules(
                                oc_subdomain, _study_uuid_for_cal, _cal_zip
                            )
                            _msg = (
                                f"Calendaring rules published: "
                                f"{_cal_summary['uploaded']} uploaded, "
                                f"{_cal_summary['skipped']} skipped, "
                                f"{_cal_summary['failed']} failed."
                            )
                            await append_log(item_id, _msg)
                            if _cal_summary["errors"]:
                                for _e in _cal_summary["errors"][:3]:
                                    await append_log(item_id, f"  CAL ERROR: {_e}")
                    except Exception as _cal_exc:
                        print(f"[cal-publish] error: {_cal_exc}", flush=True)
                        await append_log(item_id, f"Calendaring publish error: {_cal_exc}")
                    finally:
                        # Only reset if publish_to_test didn't already reach All Complete
                        if not _cal_publish_already_complete:
                            await set_status(item_id, COL["pipeline_status"], "Build Complete")

                if load_uat_checked:
                    print(f"[auto-uat] Load UAT checkbox is checked — "
                          f"launching uat_loader...", flush=True)
                    try:
                        await set_status(
                            item_id, COL["pipeline_status"],
                            UAT_STATUS["loading"]
                        )
                        await append_log(
                            item_id,
                            "UAT Loader: starting UAT data load..."
                        )

                        _uat_result = await run_uat_loader(item_id)

                        if _uat_result["success"]:
                            # UAT load is the last step — promote to
                            # "All Complete" so the row shows a clean
                            # terminal success state.
                            await asyncio.gather(
                                set_status(
                                    item_id, COL["pipeline_status"],
                                    STATUS["all_complete"]
                                ),
                                append_log(
                                    item_id,
                                    f"Pipeline complete. UAT load "
                                    f"succeeded. "
                                    f"Site: {_uat_result['site_oid']}. "
                                    f"Participants: "
                                    f"{len(_uat_result['participants_created'])}."
                                ),
                            )
                            print(
                                f"[auto-uat] UAT load succeeded → "
                                f"pipeline_status='All Complete' "
                                f"site={_uat_result['site_oid']}",
                                flush=True
                            )
                        else:
                            _uat_errs = "; ".join(_uat_result["errors"])
                            await set_status(
                                item_id, COL["pipeline_status"],
                                UAT_STATUS["failed"]
                            )
                            await append_log(
                                item_id,
                                f"UAT Load FAILED: {_uat_errs}"
                            )
                            print(
                                f"[auto-uat] UAT load failed: {_uat_errs}",
                                flush=True
                            )
                    except Exception as _uat_exc:
                        print(
                            f"[auto-uat] UAT loader crashed: {_uat_exc}",
                            flush=True
                        )
                        print(traceback.format_exc(), flush=True)
                        await append_log(
                            item_id,
                            f"UAT Loader ERROR: {_uat_exc}"
                        )
                        await set_status(
                            item_id, COL["pipeline_status"],
                            UAT_STATUS["failed"]
                        )
            except Exception as e:
                print(f"[auto-publish] Error checking post-completion "
                      f"flags: {e}", flush=True)

    except Exception as e:
        print(f"PIPELINE CRASHED: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        await append_log(item_id, f"PIPELINE ERROR: {e}")
        await set_status(item_id, COL["pipeline_status"], STATUS["failed"])
        raise


async def run_design_change_intake(item_id, source_type, source_text,
                                    protocol_hint=None):
    """
    Design Change Intake handler. Triggered automatically from main.py
    when a [DESIGN_CHANGE] update is posted on an AI Hub board row.
    Calls the design-change-intake skill to apply changes to the spec XLSX,
    save the transcript, notify the assignee, and route convention proposals.
    """
    try:
        await set_status(item_id, COL["pipeline_status"],
                         STATUS["change_intake_running"])
        await append_log(item_id, "Design change intake started.")

        from prompts import DESIGN_CHANGE_PROMPT
        import os as _os
        skill_md_path = _os.path.join(
            _os.path.dirname(_os.path.abspath(__file__)),
            "skills", "design-change-intake", "SKILL.md"
        )
        ref_md_path = _os.path.join(
            _os.path.dirname(_os.path.abspath(__file__)),
            "skills", "design-change-intake", "references",
            "spec-xls-format.md"
        )
        skill_instructions = ""
        try:
            skill_instructions = open(skill_md_path).read()
            skill_instructions += "\n\n---\n\n" + open(ref_md_path).read()
        except Exception as _e:
            print(f"DESIGN_CHANGE_INTAKE: could not read skill files: {_e}",
                  flush=True)

        full_prompt = DESIGN_CHANGE_PROMPT.format(
            source_type=source_type,
            protocol_hint=protocol_hint or "",
            source_text=source_text,
        )
        response = await run_skill(
            full_prompt,
            [],
            extra_text=skill_instructions,
        )
        summary = extract_json(response)
        if summary:
            await append_log(item_id,
                f"Design change intake complete. "
                f"{summary.get('changes_applied', 0)} change(s) applied, "
                f"{summary.get('changes_unresolved', 0)} unresolved, "
                f"{summary.get('conventions_proposed', 0)} convention(s) proposed.")
            await set_status(item_id, COL["pipeline_status"],
                             STATUS["change_intake_complete"])
        else:
            await append_log(item_id,
                "Design change intake: no summary returned from skill.")
            await set_status(item_id, COL["pipeline_status"],
                             STATUS["change_intake_failed"])

    except Exception as e:
        import traceback
        print(f"DESIGN_CHANGE_INTAKE CRASHED: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        await append_log(item_id, f"DESIGN CHANGE INTAKE ERROR: {e}")
        await set_status(item_id, COL["pipeline_status"],
                         STATUS["change_intake_failed"])
        raise


async def run_email_change_intake(member_id=None):
    """
    Email Change Intake handler.
    Polls Gmail inboxes for active PS team members hourly (triggered via
    Monday.com automation → POST /admin/run-email-intake).
    member_id: optional Monday user ID to run for one member only.
    """
    import sys as _sys
    _sys.path.insert(0, os.path.join(os.path.dirname(
        os.path.abspath(__file__)),
        "skills", "email-change-intake", "scripts"))
    from email_change_intake import (
        run_email_change_intake as _run,
        handle_review_decision  as _handle,
    )
    return await _run(member_id)

async def handle_email_review_decision(item_id, decision_label):
    """
    Called from /webhook/email-change-decision when Review Decision
    column changes on Change Requests board.
    """
    import sys as _sys
    _sys.path.insert(0, os.path.join(os.path.dirname(
        os.path.abspath(__file__)),
        "skills", "email-change-intake", "scripts"))
    from email_change_intake import handle_review_decision as _handle
    return await _handle(item_id, decision_label)


async def generate_gmail_auth_link(monday_user_id: str,
                                    member_name: str,
                                    staff_item_id: str):
    """
    Called by email_change_intake when it detects a missing Gmail token.
    Generates the Gmail OAuth initiation URL and posts a bell notification
    to the team member with a link to connect their Gmail.
    Uses monday_client.get_headers() / MONDAY_API_URL directly —
    monday_client has no generic gql_request helper.
    """
    try:
        import httpx as _httpx
        from gmail_oauth import build_initiation_url
        from monday_client import get_headers, MONDAY_API_URL

        auth_url = build_initiation_url(monday_user_id)
        notif_text = (
            f"Action needed: connect your Gmail to enable email monitoring. "
            f"Click here to connect: {auth_url}"
        )
        update_body = (
            f"Gmail connection required for email monitoring\n\n"
            f"Hi {member_name} — to activate automatic email change request "
            f"detection for your inbox, please connect your Gmail:\n\n"
            f"{auth_url}\n\n"
            f"This grants read-only access to your Gmail inbox. "
            f"You only need to do this once."
        )

        async with _httpx.AsyncClient(timeout=30) as c:
            # Bell notification
            await c.post(MONDAY_API_URL, headers=get_headers(), json={
                "query": """mutation($u: ID!, $t: ID!, $tx: String!) {
                    create_notification(user_id: $u, target_id: $t,
                                        text: $tx, target_type: Project) { text }
                }""",
                "variables": {
                    "u": str(monday_user_id),
                    "t": str(staff_item_id),
                    "tx": notif_text,
                },
            })
            # Item update (triggers email notification)
            await c.post(MONDAY_API_URL, headers=get_headers(), json={
                "query": """mutation($id: ID!, $b: String!) {
                    create_update(item_id: $id, body: $b) { id }
                }""",
                "variables": {
                    "id": str(staff_item_id),
                    "b": update_body,
                },
            })

        print(f"GMAIL_AUTH_LINK sent to {member_name} ({monday_user_id})",
              flush=True)

    except Exception as e:
        print(f"generate_gmail_auth_link failed for {monday_user_id}: {e}",
              flush=True)

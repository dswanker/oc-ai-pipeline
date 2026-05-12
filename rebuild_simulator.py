#!/usr/bin/env python3
"""
rebuild_simulator.py
====================
Inject the latest build_preview/interactive.py JS + CSS into an existing
simulator ZIP without running the full pipeline.  Optionally also expands
select dropdown choices by reading the original XLSForm files from the
EDC build ZIP.

Workflow:
  1. Edit build_preview/interactive.py
  2. python rebuild_simulator.py SIMULATOR.zip [EDC_BUILD.zip] [--open]
  3. Open the generated _rebuilt/ folder in Chrome
  4. Repeat from step 1

Usage:
    # JS/CSS fixes only (fast, ~2 seconds):
    python rebuild_simulator.py ~/Downloads/PrTK05_Form_Simulator_V0501.zip

    # JS/CSS fixes AND dropdown choices from EDC build ZIP:
    python rebuild_simulator.py ~/Downloads/PrTK05_Form_Simulator_V0501.zip \
                                ~/Downloads/PrTK05_EDC_Build_V0501.zip --open
"""
import sys, re, zipfile, io, os, shutil, subprocess
from pathlib import Path

# ── Load latest INTERACTIVE_JS and INTERACTIVE_CSS from the project ────────
_repo_root = Path(__file__).resolve().parent
sys.path.insert(0, str(_repo_root / 'build_preview'))

try:
    from interactive import INTERACTIVE_JS, INTERACTIVE_CSS
    print(f"✓ Loaded interactive.py from build_preview/interactive.py")
except ImportError as e:
    print(f"✗ Could not import interactive.py: {e}")
    print("  Run this script from the repo root.")
    sys.exit(1)

CSS_MARKER = '/* ── OC Simulator chrome'
JS_MARKER  = '/* OC Form Simulator'


# ── Read choices from an XLSForm .xlsx ────────────────────────────────────

def read_choices_from_xlsform(xlsx_bytes: bytes) -> dict:
    """
    Read the choices sheet from an XLSForm .xlsx.
    Returns {list_name: [{name: str, label: str}, ...]}
    """
    try:
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), data_only=True)
        if 'choices' not in wb.sheetnames:
            return {}
        ws = wb['choices']
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        # Find column indices
        header = [str(c).strip().lower() if c else '' for c in rows[0]]
        try:
            li = header.index('list_name')
            ni = header.index('name')
            # label column: 'label' or 'label::english (en)' etc.
            lbl_candidates = [i for i, h in enumerate(header)
                              if h.startswith('label')]
            if not lbl_candidates:
                return {}
            lbli = lbl_candidates[0]
        except ValueError:
            return {}
        choices = {}
        for row in rows[1:]:
            ln = str(row[li] or '').strip()
            nm = str(row[ni] or '').strip()
            lb = str(row[lbli] or '').strip()
            if ln and nm:
                choices.setdefault(ln, []).append({'name': nm, 'label': lb or nm})
        return choices
    except Exception as e:
        print(f"  Warning: could not read choices: {e}")
        return {}


def load_all_choices_from_edc_zip(edc_zip_path: str) -> dict:
    """
    Read all XLSForm files from the EDC build ZIP and merge their choices.
    Returns {list_name: [{name, label}, ...]}
    """
    all_choices = {}
    with zipfile.ZipFile(edc_zip_path, 'r') as zf:
        xlsx_files = [n for n in zf.namelist()
                      if n.endswith('.xlsx') and '/forms/' in n]
        print(f"  Reading choices from {len(xlsx_files)} XLSForm(s)...")
        for name in xlsx_files:
            choices = read_choices_from_xlsform(zf.read(name))
            for k, v in choices.items():
                if k not in all_choices:
                    all_choices[k] = v
    total = sum(len(v) for v in all_choices.values())
    print(f"  ✓ {len(all_choices)} choice lists, {total} options total")
    return all_choices


# ── Expand <option class="itemset-template"> in HTML ──────────────────────

def expand_select_options(html: str, choices: dict) -> str:
    """
    Replace <option class="itemset-template" data-items-path="instance('LIST')/root/item">
    with concrete <option> elements for each choice in LIST.
    This mirrors the Python fix in render.py's expand_itemsets().
    """
    pattern = re.compile(
        r'<option\s+class="itemset-template"\s+[^>]*'
        r'data-items-path="instance\(\'([^\']+)\'\)/root/item"[^>]*>.*?</option>',
        re.DOTALL,
    )
    expanded = 0

    def replacer(m):
        nonlocal expanded
        list_id = m.group(1)
        opts = choices.get(list_id, [])
        if not opts:
            return m.group(0)
        expanded += 1
        parts = ['<option value="">...</option>']   # keep placeholder
        for opt in opts:
            parts.append('<option value="' + opt['name'] + '">'
                         + opt['label'] + '</option>')
        return ''.join(parts)

    result = pattern.sub(replacer, html)
    if expanded:
        print(f"    expanded {expanded} select list(s)")
    return result


# ── JS / CSS patching ──────────────────────────────────────────────────────

def patch_html(html: str, filename: str,
               choices: dict = None) -> tuple:
    issues = []

    # CSS
    css_idx = html.find(CSS_MARKER)
    if css_idx == -1:
        issues.append(f"CSS marker not found in {filename}")
    else:
        style_end = html.find('</style>', css_idx)
        if style_end != -1:
            html = html[:css_idx] + INTERACTIVE_CSS + '\n' + html[style_end:]

    # JS
    js_idx = html.find(JS_MARKER)
    if js_idx == -1:
        issues.append(f"JS marker not found in {filename}")
    else:
        script_end = html.find('</script>', js_idx)
        if script_end != -1:
            html = html[:js_idx] + INTERACTIVE_JS + '\n' + html[script_end:]

    # Expand select options if choices provided
    if choices:
        html = expand_select_options(html, choices)

    return html, issues


# ── Main ───────────────────────────────────────────────────────────────────

def rebuild(sim_zip: str, edc_zip: str = None, auto_open: bool = False):
    src = Path(sim_zip).expanduser().resolve()
    if not src.exists():
        print(f"✗ Simulator ZIP not found: {src}")
        sys.exit(1)

    dst_zip = src.parent / (src.stem + '_rebuilt.zip')
    dst_dir = src.parent / (src.stem + '_rebuilt')

    print(f"\nSimulator: {src.name}")

    # Load choices from EDC build ZIP if provided
    choices = {}
    if edc_zip:
        edc_path = Path(edc_zip).expanduser().resolve()
        if not edc_path.exists():
            print(f"✗ EDC build ZIP not found: {edc_path}")
            sys.exit(1)
        print(f"EDC build: {edc_path.name}")
        choices = load_all_choices_from_edc_zip(str(edc_path))
    else:
        print("No EDC ZIP provided — skipping dropdown expansion")
        print("  Tip: add the EDC build ZIP as a second argument to fix dropdowns")

    print()

    buf = io.BytesIO()
    patched = 0
    all_issues = []

    with zipfile.ZipFile(src, 'r') as zin, \
         zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:

        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith('.html'):
                html = data.decode('utf-8', errors='replace')
                fname = Path(item.filename).name
                if fname != 'index.html':
                    print(f"  Patching {fname}...")
                patched_html, issues = patch_html(
                    html, item.filename,
                    choices if fname != 'index.html' else None,
                )
                all_issues.extend(issues)
                if not issues or fname == 'index.html':
                    patched += 1
                data = patched_html.encode('utf-8')
            zout.writestr(item, data)

    buf.seek(0)
    dst_zip.write_bytes(buf.getvalue())

    if dst_dir.exists():
        shutil.rmtree(dst_dir)
    with zipfile.ZipFile(dst_zip, 'r') as zf:
        zf.extractall(dst_dir)

    index_files = list(dst_dir.rglob('index.html'))
    print(f"\n✓ Patched {patched} HTML files → {dst_zip.name}")

    if index_files:
        index = index_files[0]
        print(f"\nOpen in Chrome:")
        print(f"  open '{index}'")
        if auto_open:
            subprocess.run(['open', str(index)])


if __name__ == '__main__':
    positional = [a for a in sys.argv[1:] if not a.startswith('--')]
    flags      = [a for a in sys.argv[1:] if a.startswith('--')]

    if not positional:
        print(__doc__)
        sys.exit(0)

    sim_zip = positional[0]
    edc_zip = positional[1] if len(positional) > 1 else None
    rebuild(sim_zip, edc_zip, auto_open='--open' in flags)

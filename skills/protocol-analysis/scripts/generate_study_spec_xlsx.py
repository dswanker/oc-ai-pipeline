"""
generate_xlsx.py — EDC Structure Specification XLSX Generator
Produces a human-editable specification workbook from the EDC structure JSON.

Workbook structure per form:
  - INDEX sheet (workbook-level summary)
  - For each form:
      [FORMID]_survey   — editable XLSForm survey rows
      [FORMID]_choices  — editable choice lists
      [FORMID]_settings — editable settings + metadata tab

When a human edits this file and feeds it back to the EDC structure skill,
the skill reads the changes and regenerates all three outputs (PDF, JSON, XLSX).

Usage:
    from generate_xlsx import build_edc_xlsx
    build_edc_xlsx(data_dict, output_path)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from dep_utils import (
    extract_all_form_dependencies, extract_row_dependencies,
    annotate_survey_with_dependencies, format_deps_short
)

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE_HEX   = "1B3A6B"
MID_BLUE_HEX    = "2E6DA4"
LIGHT_BLUE_HEX  = "D6E4F0"
WHITE_HEX       = "FFFFFF"
GREY_LIGHT_HEX  = "F5F5F5"
GREY_MID_HEX    = "CCCCCC"
AMBER_HEX       = "FFF3CD"
RED_LIGHT_HEX   = "FADBD8"
GREEN_LIGHT_HEX = "D5F5E3"
YELLOW_HEX      = "FFFF99"
ORANGE_HEX      = "FFD580"
TEAL_HEX        = "D0F0F0"

# ── Style helpers ─────────────────────────────────────────────────────────────
def hdr_font(bold=True, color=WHITE_HEX, size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)

def body_font(bold=False, color="1A1A1A", size=8):
    return Font(name="Arial", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=GREY_MID_HEX)
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(horizontal="left"):
    return Alignment(wrap_text=True, vertical="top", horizontal=horizontal)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row_num, col_count, bg=DARK_BLUE_HEX):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = hdr_font(color=WHITE_HEX)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = wrap_align("center")

def style_data_row(ws, row_num, col_count, bg=WHITE_HEX):
    alt = GREY_LIGHT_HEX if row_num % 2 == 0 else WHITE_HEX
    row_bg = bg if bg != WHITE_HEX else alt
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE_HEX):
            cell.fill = fill(row_bg)
        cell.font = body_font()
        cell.border = thin_border()
        cell.alignment = wrap_align()

def status_fill(status):
    s = str(status).upper()
    if "COMPLETE"    in s: return fill(GREEN_LIGHT_HEX)
    if "FLAGGED"     in s: return fill(AMBER_HEX)
    if "PLACEHOLDER" in s: return fill(RED_LIGHT_HEX)
    return fill(WHITE_HEX)

def safe(val):
    return str(val) if val is not None else ""


# ── INDEX sheet ───────────────────────────────────────────────────────────────
def build_index_sheet(wb, data):
    ws = wb.active
    ws.title = "INDEX"

    meta = data.get("study_meta", {})
    forms = data.get("forms", [])

    # Title banner
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = f"EDC STRUCTURE SPECIFICATION — {meta.get('protocol_number','')}  |  Generated: {datetime.date.today().strftime('%d %b %Y')}  |  PENDING HUMAN REVIEW"
    c.font = hdr_font(size=11, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Study meta block
    meta_rows = [
        ("Protocol Number",  meta.get("protocol_number", "")),
        ("Study ID",         meta.get("study_id", "")),
        ("Input Mode",       meta.get("input_mode", "PROTOCOL_ONLY")),
        ("Generated Date",   meta.get("generated_date", "")),
        ("Review Status",    "PENDING HUMAN REVIEW — DO NOT BUILD UNTIL APPROVED"),
        ("Total Forms",      str(len(forms))),
        ("Total Unique CRFs",str(len(forms))),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k).font = hdr_font(color="1A1A1A", bold=True)
        ws.cell(row=i, column=1).fill = fill(LIGHT_BLUE_HEX)
        ws.cell(row=i, column=2, value=v).font = body_font()
        ws.cell(row=i, column=2).fill = fill(WHITE_HEX)
        for col in [1, 2]:
            ws.cell(row=i, column=col).border = thin_border()
            ws.cell(row=i, column=col).alignment = wrap_align()

    # Instructions box
    instr_row = len(meta_rows) + 3
    ws.merge_cells(f"A{instr_row}:L{instr_row}")
    c = ws.cell(row=instr_row, column=1)
    c.value = "HOW TO USE THIS DOCUMENT"
    c.font = hdr_font(color=WHITE_HEX, size=10)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")

    instructions = [
        "1. REVIEW each form tab (named [FORMID]_survey, [FORMID]_choices, [FORMID]_settings)",
        "2. In the survey tab: FLAGGED rows are amber, PLACEHOLDER rows are red — these need your attention",
        "3. To ADD a field: insert a new row and fill in type, name, label, and any other columns needed",
        "4. To DELETE a field: clear the entire row or add 'DELETE' in the ACTION column",
        "5. To CHANGE a field property: edit the relevant cell directly (constraint, relevant, label, etc.)",
        "6. To ADD a choice: go to the [FORMID]_choices tab and add a new row",
        "7. ACTION column: leave blank (no change), write 'DELETE' (remove this row), or 'ADD' (new row)",
        "8. When done: save this file and upload it back to Claude to regenerate all 3 outputs (PDF, JSON, XLSX)",
        "9. The [FORMID]_meta tab is READ-ONLY — it shows technical metadata for reference only",
    ]
    for i, instr in enumerate(instructions, start=instr_row + 1):
        ws.merge_cells(f"A{i}:L{i}")
        c = ws.cell(row=i, column=1)
        c.value = instr
        c.font = body_font(size=8)
        c.fill = fill(GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX)
        c.alignment = wrap_align()
        c.border = thin_border()

    # Form inventory table
    inv_row = instr_row + len(instructions) + 2
    ws.merge_cells(f"A{inv_row}:L{inv_row}")
    c = ws.cell(row=inv_row, column=1)
    c.value = "FORM INVENTORY"
    c.font = hdr_font(color=WHITE_HEX, size=10)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")

    inv_headers = ["#", "Form ID", "Form Title", "Category", "CDASH Domain",
                   "Arm", "Complexity", "Repeating", "ePRO",
                   "Re-uses", "Survey Tab", "Status"]
    inv_cols = [3, 10, 22, 12, 12, 12, 10, 9, 6, 8, 18, 16]

    hdr_row = inv_row + 1
    for col, (h, w) in enumerate(zip(inv_headers, inv_cols), start=1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = hdr_font(color=WHITE_HEX)
        c.fill = fill(MID_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        set_col_width(ws, get_column_letter(col), w)

    for i, form in enumerate(forms, start=1):
        r = hdr_row + i
        lm = form.get("library_match", {})
        n_flagged = sum(1 for s in form.get("survey", [])
                        if s.get("completion_status") in ("FLAGGED", "PLACEHOLDER"))
        status_str = "✓ Ready" if n_flagged == 0 else f"⚠ {n_flagged} items need review"
        row_vals = [
            str(i),
            form.get("form_id", ""),
            form.get("form_title", ""),
            form.get("form_category", "").replace("CDASH_CLINICAL", "CDASH").replace("INFRASTRUCTURE", "INFRA"),
            form.get("cdash_domain", "") or "—",
            form.get("arm_applicability", ""),
            form.get("complexity", ""),
            "Yes" if form.get("has_repeating_group") else "No",
            "Yes" if form.get("is_epro") else "No",
            str(form.get("reuse_count", 0) or 0),
            f"{form.get('form_id','')}_survey",
            status_str,
        ]
        for col, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = body_font()
            c.border = thin_border()
            c.alignment = wrap_align()
            bg = GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX
            c.fill = fill(bg)
        # colour status cell
        status_cell = ws.cell(row=r, column=12)
        if "⚠" in status_str:
            status_cell.fill = fill(AMBER_HEX)
            status_cell.font = Font(name="Arial", bold=True, color="7D5A00", size=8)
        else:
            status_cell.fill = fill(GREEN_LIGHT_HEX)
            status_cell.font = Font(name="Arial", bold=True, color="1A6B3A", size=8)

    ws.freeze_panes = "A2"


# ── Survey sheet ──────────────────────────────────────────────────────────────
SURVEY_EDITABLE_COLS = [
    ("ACTION",           "Leave blank = no change | DELETE = remove row | ADD = new row", 12),
    ("type",             "XLSForm field type (e.g. text, integer, select_one NY)", 18),
    ("name",             "Machine-readable field ID (no spaces)", 18),
    ("label",            "User-visible question text", 32),
    ("bind::oc:itemgroup","CDASH domain group (e.g. AE, VS, LB)", 14),
    ("appearance",       "Layout hint (w1-w6, horizontal, minimal, multiline)", 16),
    ("relevant",         "Show/hide condition (XPath expression)", 30),
    ("required",         "yes / true() / expression", 12),
    ("constraint",       "Validation rule (XPath expression)", 30),
    ("constraint_message","Error message shown when constraint fails", 28),
    ("calculation",      "Auto-calculated value (XPath expression)", 30),
    ("readonly",         "yes or blank", 8),
    ("hint",             "Helper text shown below label", 24),
    ("repeat_count",     "Integer or expression (repeating groups only)", 12),
    ("bind::oc:external","External data source: clinicaldata / labranges / tpt", 20),
    ("choice_filter",    "Filter choices based on expression", 24),
    ("DEPENDENCIES",     "Auto-derived: [FormOID].[ItemOID] cross-form references (read-only)", 32),
    ("REVIEW_NOTES",     "Your review notes or change instructions", 30),
]

def build_survey_sheet(wb, form):
    form_id = form.get("form_id", "FORM")
    sheet_name = f"{form_id}_survey"[:31]
    ws = wb.create_sheet(title=sheet_name)

    survey = form.get("survey", [])
    n_editable = len(SURVEY_EDITABLE_COLS)

    # Tab colour by status
    n_flagged = sum(1 for r in survey if r.get("completion_status") in ("FLAGGED", "PLACEHOLDER"))
    ws.sheet_properties.tabColor = "E67E22" if n_flagged > 0 else "27AE60"

    # Title banner
    ws.merge_cells(f"A1:{get_column_letter(n_editable)}1")
    c = ws["A1"]
    c.value = f"{form.get('form_title','')}  ({form_id})  —  SURVEY SHEET  |  Arm: {form.get('arm_applicability','')}  |  Visits: {len(form.get('visits_assigned',[]))}  |  Repeating: {'Yes' if form.get('has_repeating_group') else 'No'}  |  ePRO: {'Yes' if form.get('is_epro') else 'No'}"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    # Colour legend
    legend_items = [
        ("GREEN = COMPLETE", GREEN_LIGHT_HEX, "1A6B3A"),
        ("AMBER = FLAGGED — needs review", AMBER_HEX, "7D5A00"),
        ("RED = PLACEHOLDER — must be completed", RED_LIGHT_HEX, "8B1A1A"),
        ("YELLOW = ACTION required (edited by reviewer)", YELLOW_HEX, "5A4000"),
    ]
    for i, (txt, bg, fg) in enumerate(legend_items):
        col = i * 3 + 1
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
        c = ws.cell(row=2, column=col, value=txt)
        c.font = Font(name="Arial", bold=True, size=7, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 14

    # Header row
    for col, (hdr, hint, width) in enumerate(SURVEY_EDITABLE_COLS, start=1):
        c = ws.cell(row=3, column=col, value=hdr)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(DARK_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        c.comment = None
        set_col_width(ws, get_column_letter(col), width)
    ws.row_dimensions[3].height = 16

    # Data rows
    col_keys = [c[0] for c in SURVEY_EDITABLE_COLS]

    # Pre-annotate survey rows with extracted dependencies
    annotate_survey_with_dependencies(survey, form)

    for row_i, row in enumerate(survey, start=4):
        status = row.get("completion_status", "")
        base_fill = status_fill(status)

        for col_i, key in enumerate(col_keys, start=1):
            c = ws.cell(row=row_i, column=col_i)

            if key == "ACTION":
                c.value = ""
                c.fill = fill(WHITE_HEX)
            elif key == "REVIEW_NOTES":
                c.value = row.get("flag_reason", "") or ""
                c.fill = fill(AMBER_HEX) if row.get("flag_reason") else fill(WHITE_HEX)
            elif key == "DEPENDENCIES":
                # Auto-derived — show extracted deps, grey background (read-only)
                row_deps = row.get("dependencies", [])
                c.value = ", ".join(row_deps) if row_deps else ""
                c.fill = fill("E8F4FD") if row_deps else fill(GREY_LIGHT_HEX)
                c.font = body_font(bold=False, color="1B3A6B" if row_deps else "888888")
                c.border = thin_border()
                c.alignment = wrap_align()
                ws.row_dimensions[row_i].height = 14
                continue
            else:
                # Map JSON key to XLSForm column name
                json_key = key.replace("::", "__").replace(":", "_")
                # Try direct key match, then underscored version
                val = row.get(key) or row.get(json_key) or row.get(key.replace("::", "__")) or ""
                c.value = safe(val)
                c.fill = base_fill

            c.font = body_font(bold=(key == "name"))
            c.border = thin_border()
            c.alignment = wrap_align()

        # Row height
        ws.row_dimensions[row_i].height = 14

    # Freeze header
    ws.freeze_panes = "B4"

    # Add validation for ACTION column
    dv = DataValidation(type="list", formula1='"ADD,DELETE,"', allow_blank=True)
    dv.sqref = f"A4:A{len(survey)+10}"
    ws.add_data_validation(dv)

    # Add validation for type column
    dv_type = DataValidation(
        type="list",
        formula1='"text,integer,decimal,date,select_one,select_multiple,note,calculate,begin group,end group,begin repeat,end repeat"',
        allow_blank=True
    )
    dv_type.sqref = f"B4:B{len(survey)+10}"
    ws.add_data_validation(dv_type)

    return sheet_name


# ── Choices sheet ─────────────────────────────────────────────────────────────
CHOICES_EDITABLE_COLS = [
    ("ACTION",      "Leave blank | DELETE | ADD", 10),
    ("list_name",   "Choice list identifier (no spaces)", 16),
    ("label",       "User-visible option text", 28),
    ("name",        "Machine-readable option value (no spaces)", 20),
    ("source",      "STANDARD or PROTOCOL_SPECIFIC", 18),
    ("filter_column","Column name used for choice_filter", 14),
    ("filter_value", "Value for choice_filter matching", 20),
    ("REVIEW_NOTES", "Your notes", 28),
]

def build_choices_sheet(wb, form):
    form_id = form.get("form_id", "FORM")
    sheet_name = f"{form_id}_choices"[:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = "2E6DA4"

    choices = form.get("choices", [])
    n_cols = len(CHOICES_EDITABLE_COLS)

    # Title
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = f"{form.get('form_title','')}  ({form_id})  —  CHOICES SHEET  |  {len(choices)} choices in {len(set(c.get('list_name','') for c in choices))} lists"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    # Headers
    for col, (hdr, hint, width) in enumerate(CHOICES_EDITABLE_COLS, start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(DARK_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        set_col_width(ws, get_column_letter(col), width)
    ws.row_dimensions[2].height = 16

    # Group choices by list_name with alternating colours per list
    list_colours = [LIGHT_BLUE_HEX, TEAL_HEX, GREY_LIGHT_HEX, WHITE_HEX]
    list_order = []
    for ch in choices:
        ln = ch.get("list_name", "")
        if ln not in list_order:
            list_order.append(ln)

    for row_i, ch in enumerate(choices, start=3):
        ln = ch.get("list_name", "")
        list_idx = list_order.index(ln) % len(list_colours)
        row_bg = list_colours[list_idx]

        # Highlight protocol-specific choices
        if ch.get("source", "") == "PROTOCOL_SPECIFIC":
            row_bg = AMBER_HEX

        vals = [
            "",
            ch.get("list_name", ""),
            ch.get("label", ""),
            ch.get("name", ""),
            ch.get("source", ""),
            ch.get("filter_column", ""),
            ch.get("filter_value", ""),
            "",
        ]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=row_i, column=col_i, value=safe(val))
            c.font = body_font()
            c.fill = fill(row_bg)
            c.border = thin_border()
            c.alignment = wrap_align()
        ws.row_dimensions[row_i].height = 13

    ws.freeze_panes = "B3"

    # Add validation for ACTION
    dv = DataValidation(type="list", formula1='"ADD,DELETE,"', allow_blank=True)
    dv.sqref = f"A3:A{len(choices)+20}"
    ws.add_data_validation(dv)


# ── Settings + Meta sheet ─────────────────────────────────────────────────────
def build_settings_sheet(wb, form):
    form_id = form.get("form_id", "FORM")
    sheet_name = f"{form_id}_settings"[:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = "1B3A6B"

    settings = form.get("settings", {})
    lm = form.get("library_match", {})
    xdeps = form.get("cross_form_dependencies", [])

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"{form.get('form_title','')}  ({form_id})  —  SETTINGS & METADATA"
    c.font = hdr_font(size=10, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── EDITABLE SETTINGS ──
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "EDITABLE — XLSForm Settings"
    c.font = hdr_font(color=WHITE_HEX, size=9)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")

    editable_settings = [
        ("form_title",          settings.get("form_title", ""),   "Human-readable form name"),
        ("form_id",             settings.get("form_id", ""),      "CDASH code or custom ID (no spaces)"),
        ("version",             settings.get("version", "1"),     "Increment when form is updated"),
        ("style",               settings.get("style", "theme-grid"), "Always theme-grid for OpenClinica"),
        ("namespaces",          settings.get("namespaces", ""),   "OpenClinica namespace declaration"),
        ("crossform_references",settings.get("crossform_references", ""), "Leave blank or 'current_event'"),
    ]
    for i, (key, val, hint) in enumerate(editable_settings, start=3):
        ws.cell(row=i, column=1, value=key).font = hdr_font(color="1A1A1A", bold=True, size=8)
        ws.cell(row=i, column=1).fill = fill(LIGHT_BLUE_HEX)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = wrap_align()

        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        c = ws.cell(row=i, column=2, value=val)
        c.font = body_font()
        c.fill = fill(YELLOW_HEX)   # yellow = editable
        c.border = thin_border()
        c.alignment = wrap_align()

        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=6)
        c = ws.cell(row=i, column=5, value=hint)
        c.font = Font(name="Arial", italic=True, size=7, color="666666")
        c.fill = fill(GREY_LIGHT_HEX)
        c.border = thin_border()
        c.alignment = wrap_align()

    # ── CHOICE LISTS ──
    choices = form.get("choices", [])
    if choices:
        ch_start = len(editable_settings) + 4
        ws.merge_cells(f"A{ch_start}:F{ch_start}")
        c = ws.cell(row=ch_start, column=1)
        list_names = list(dict.fromkeys(ch.get("list_name","") for ch in choices))
        c.value = f"CHOICE LISTS  ({len(list_names)} lists, {len(choices)} options)"
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E6DA4")
        c.alignment = Alignment(horizontal="left", vertical="center")

        # Choice list headers
        ch_hdrs = ["list_name", "label", "name", "source", "filter_column", "filter_value"]
        ch_widths = [16, 28, 22, 18, 14, 20]
        for col_i, (h, w) in enumerate(zip(ch_hdrs, ch_widths), start=1):
            c = ws.cell(row=ch_start+1, column=col_i, value=h)
            c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1B3A6B")
            c.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )
            c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            ws.column_dimensions[get_column_letter(col_i)].width = w

        list_colours = ["D6E4F0", "D0F0F0", "F5F5F5", "FFFFFF"]
        list_idx_map = {ln: i % len(list_colours) for i, ln in enumerate(list_names)}
        for r_i, ch in enumerate(choices):
            r = ch_start + 2 + r_i
            ln = ch.get("list_name","")
            src = ch.get("source","")
            row_bg = "FFF3CD" if src == "PROTOCOL_SPECIFIC" else list_colours[list_idx_map.get(ln, 0)]
            for col_i, key in enumerate(ch_hdrs, start=1):
                val = ch.get(key, "")
                c = ws.cell(row=r, column=col_i, value=val)
                c.font = Font(name="Arial", size=8)
                c.fill = PatternFill("solid", fgColor=row_bg)
                c.border = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC")
                )
                c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 13

        # Offset meta_start past choices block
        meta_start_offset = ch_start + 2 + len(choices) + 2
    else:
        meta_start_offset = len(editable_settings) + 4

    # ── DEPENDENCIES SUMMARY ──
    form_all_deps = extract_all_form_dependencies(form)
    dep_start = meta_start_offset
    ws.merge_cells(f"A{dep_start}:F{dep_start}")
    c = ws.cell(row=dep_start, column=1)
    c.value = f"DEPENDENCIES  ({len(form_all_deps)} cross-form references)"
    c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1B3A6B")
    c.alignment = Alignment(horizontal="left", vertical="center")
    if form_all_deps:
        for d_i, dep in enumerate(form_all_deps, start=dep_start+1):
            ws.merge_cells(f"A{d_i}:F{d_i}")
            c = ws.cell(row=d_i, column=1, value=dep)
            c.font = Font(name="Arial", size=8, color="1B3A6B")
            c.fill = PatternFill("solid", fgColor="EBF5FB" if d_i % 2 == 0 else "FFFFFF")
            c.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )
            c.alignment = Alignment(wrap_text=True, vertical="top")
        meta_start_offset = dep_start + 1 + len(form_all_deps) + 2
    else:
        c_none = ws.cell(row=dep_start+1, column=1, value="No cross-form dependencies identified")
        c_none.font = Font(name="Arial", italic=True, size=8, color="888888")
        ws.merge_cells(f"A{dep_start+1}:F{dep_start+1}")
        meta_start_offset = dep_start + 3

    # ── READ-ONLY METADATA ──
    meta_start = meta_start_offset
    ws.merge_cells(f"A{meta_start}:F{meta_start}")
    c = ws.cell(row=meta_start, column=1)
    c.value = "READ-ONLY — Technical Metadata (for reference)"
    c.font = hdr_font(color=WHITE_HEX, size=9)
    c.fill = fill(GREY_MID_HEX.replace("CC", "88") + "FF"[:0] or "888888")
    c.fill = fill("888888")
    c.alignment = Alignment(horizontal="left", vertical="center")

    metadata = [
        ("form_category",    form.get("form_category", "")),
        ("cdash_domain",     form.get("cdash_domain", "") or "—"),
        ("arm_applicability",form.get("arm_applicability", "")),
        ("complexity",       form.get("complexity", "")),
        ("has_repeating_group", str(form.get("has_repeating_group", False))),
        ("is_epro",          str(form.get("is_epro", False))),
        ("reuse_count",      str(form.get("reuse_count", 0))),
        ("pricing_summary_source", str(form.get("pricing_summary_source", False))),
        ("library_match_status", lm.get("status", "PROTOCOL_ONLY")),
        ("library_source_type",  lm.get("source_type", "NONE")),
        ("visits_assigned",  ", ".join(form.get("visits_assigned", []))),
    ]
    for i, (key, val) in enumerate(metadata, start=meta_start + 1):
        ws.cell(row=i, column=1, value=key).font = body_font(bold=True, size=8)
        ws.cell(row=i, column=1).fill = fill(GREY_LIGHT_HEX)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = wrap_align()

        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        c = ws.cell(row=i, column=2, value=val)
        c.font = body_font(color="555555", size=8)
        c.fill = fill(GREY_LIGHT_HEX)
        c.border = thin_border()
        c.alignment = wrap_align()

    # ── CROSS-FORM DEPENDENCIES ──
    if xdeps:
        dep_start = meta_start + len(metadata) + 2
        ws.merge_cells(f"A{dep_start}:F{dep_start}")
        c = ws.cell(row=dep_start, column=1)
        c.value = "CROSS-FORM DEPENDENCIES — All require OID confirmation after study configuration"
        c.font = hdr_font(color=WHITE_HEX, size=9)
        c.fill = fill(MID_BLUE_HEX)
        c.alignment = Alignment(horizontal="left", vertical="center")

        dep_hdrs = ["Source Form", "Source Field", "Purpose", "Visit Context", "XPath Pattern", "Status"]
        dep_widths = [12, 14, 26, 14, 30, 28]
        for col, (h, w) in enumerate(zip(dep_hdrs, dep_widths), start=1):
            c = ws.cell(row=dep_start+1, column=col, value=h)
            c.font = hdr_font(color=WHITE_HEX, size=8)
            c.fill = fill(DARK_BLUE_HEX)
            c.border = thin_border()
            c.alignment = wrap_align("center")
            set_col_width(ws, get_column_letter(col), w)

        for i, dep in enumerate(xdeps, start=dep_start+2):
            row_bg = AMBER_HEX  # all deps are flagged
            vals = [
                dep.get("source_form",""),
                dep.get("source_field",""),
                dep.get("purpose",""),
                dep.get("visit_context",""),
                dep.get("xpath_pattern",""),
                dep.get("status","FLAGGED — OID CONFIRMATION REQUIRED"),
            ]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=i, column=col, value=safe(val))
                c.font = body_font(size=7)
                c.fill = fill(row_bg)
                c.border = thin_border()
                c.alignment = wrap_align()

    # Column widths
    set_col_width(ws, "A", 22)
    set_col_width(ws, "B", 28)
    set_col_width(ws, "C", 28)
    set_col_width(ws, "D", 14)
    set_col_width(ws, "E", 18)
    set_col_width(ws, "F", 22)
    ws.freeze_panes = "A3"


# ── Supporting sheets ─────────────────────────────────────────────────────────
def build_timepoint_sheet(wb, tpt_csv):
    ws = wb.create_sheet(title="TIMEPOINTS")
    ws.sheet_properties.tabColor = "1B3A6B"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"TIMEPOINT CSV — {tpt_csv.get('filename','')}  |  Feed this file to OpenClinica alongside the XLSForms"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    for col, h in enumerate(["event", "timepoint", "REVIEW_NOTES"], start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(MID_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")

    set_col_width(ws, "A", 24)
    set_col_width(ws, "B", 34)
    set_col_width(ws, "C", 28)

    for i, row in enumerate(tpt_csv.get("rows", []), start=3):
        bg = GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        for col, val in enumerate([row.get("event",""), row.get("timepoint",""), ""], start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = wrap_align()

    ws.freeze_panes = "A3"


def build_labranges_sheet(wb, labranges):
    ws = wb.create_sheet(title="LAB_RANGES")
    ws.sheet_properties.tabColor = "E67E22"

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "LAB RANGES CSV — PLACEHOLDER — Complete all highlighted cells with site-specific values before building"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    cols = labranges.get("columns", ["lab_name","test_code","test_name","lower","upper","unit","sex_filter","age_lower","age_upper"])
    widths = [20, 10, 24, 12, 12, 16, 10, 10, 10]

    for col_i, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=2, column=col_i, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(DARK_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        set_col_width(ws, get_column_letter(col_i), w)

    for i, row in enumerate(labranges.get("rows", []), start=3):
        bg = GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        for col_i, key in enumerate(cols, start=1):
            val = row.get(key, "")
            c = ws.cell(row=i, column=col_i, value=val)
            c.font = body_font()
            c.border = thin_border()
            c.alignment = wrap_align()
            # Highlight placeholders
            if "PLACEHOLDER" in str(val).upper():
                c.fill = fill(RED_LIGHT_HEX)
                c.font = Font(name="Arial", bold=True, size=8, color="8B1A1A")
            else:
                c.fill = fill(bg)

    ws.freeze_panes = "A3"


def build_review_flags_sheet(wb, flags):
    ws = wb.create_sheet(title="REVIEW_FLAGS")
    ws.sheet_properties.tabColor = "C0392B"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "REVIEW FLAGS — All items must be resolved before passing to edc-builder"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    cat_colours = {
        "site_specific":        ("SITE SPECIFIC",         RED_LIGHT_HEX,   "8B1A1A"),
        "oid_confirmation":     ("OID CONFIRMATION",      AMBER_HEX,       "7D5A00"),
        "protocol_ambiguous":   ("PROTOCOL AMBIGUOUS",    AMBER_HEX,       "7D5A00"),
        "constraint_review":    ("CONSTRAINT REVIEW",     AMBER_HEX,       "7D5A00"),
        "choice_list_review":   ("CHOICE LIST REVIEW",    LIGHT_BLUE_HEX,  "1B3A6B"),
        "custom_domain":        ("CUSTOM DOMAIN",         LIGHT_BLUE_HEX,  "1B3A6B"),
        "pdf_mapping_uncertain":("PDF MAPPING UNCERTAIN", AMBER_HEX,       "7D5A00"),
        "name_deviation":       ("NAME DEVIATION",        LIGHT_BLUE_HEX,  "1B3A6B"),
    }

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 22)
    set_col_width(ws, "C", 80)
    set_col_width(ws, "D", 30)

    row = 2
    for key, (label, bg, fg) in cat_colours.items():
        items = flags.get(key, [])
        if not items:
            continue

        # Category header
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1)
        c.value = f"  {label}  ({len(items)} items)"
        c.font = Font(name="Arial", bold=True, size=9, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()
        ws.row_dimensions[row].height = 16
        row += 1

        for item in items:
            ws.cell(row=row, column=1, value="").fill = fill(bg)
            ws.cell(row=row, column=2, value="→").font = body_font(color=fg, bold=True)
            ws.cell(row=row, column=2).fill = fill(bg)
            ws.cell(row=row, column=2).border = thin_border()
            ws.cell(row=row, column=2).alignment = wrap_align("center")

            ws.merge_cells(f"C{row}:D{row}")
            c = ws.cell(row=row, column=3, value=str(item))
            c.font = body_font(size=8)
            c.fill = fill(GREY_LIGHT_HEX if row % 2 == 0 else WHITE_HEX)
            c.border = thin_border()
            c.alignment = wrap_align()
            ws.row_dimensions[row].height = 13
            row += 1

        row += 1  # gap between categories

    ws.freeze_panes = "A2"


# ── Main builder ──────────────────────────────────────────────────────────────
def build_edc_xlsx(data: dict, output_path: str):
    wb = Workbook()

    # INDEX sheet (uses default active sheet)
    build_index_sheet(wb, data)

    # Supporting sheets
    build_timepoint_sheet(wb, data.get("timepoint_csv", {}))
    build_labranges_sheet(wb, data.get("labranges_csv", {}))
    build_review_flags_sheet(wb, data.get("review_flags", {}))

    # One set of tabs per form
    for form in data.get("forms", []):
        build_survey_sheet(wb, form)
        build_choices_sheet(wb, form)
        build_settings_sheet(wb, form)

    wb.save(output_path)
    print(f"EDC Structure XLSX written to: {output_path}")


# ── Test run ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import json, sys
    sys.path.insert(0, '/home/claude')
    # Load the full PrTK05 data from the existing test script
    exec(open('/home/claude/prtk05_edc_data.py').read().split('build_edc_pdf')[0])
    build_edc_xlsx(data, "/mnt/user-data/outputs/PrTK05_EDC_Structure.xlsx")

# ── Alias so the function can also be imported by its skill-level name ────
build_study_spec_xlsx = build_edc_xlsx


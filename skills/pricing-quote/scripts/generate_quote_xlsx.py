"""
generate_quote_xlsx.py — OpenClinica Quote XLSX Generator
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, os

OC_DARK  = "1B3A6B"
OC_MID   = "2E6DA4"
OC_LIGHT = "D6E4F0"
OC_TEAL  = "00A99D"
WHITE    = "FFFFFF"
GREY_L   = "F5F5F5"
GREY_M   = "CCCCCC"
AMBER    = "FFF3CD"
AMBER_T  = "CC6600"
RED_C    = "CC0000"

def _fl(h):  return PatternFill("solid", fgColor=h)
def _fn(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, italic=italic, color=color, size=size)
def _bd():
    s = Side(style="thin", color=GREY_M)
    return Border(left=s, right=s, top=s, bottom=s)
def _al(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _ccy(): return '"$"#,##0.00'
def _fh(h): return f"{h} hr{'s' if h != 1 else ''}"
def _fmt(v, tbd=False): return v if not tbd else "TBD"

def _cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _cell(ws, r, c, val, bold=False, color="000000", bg=WHITE, size=10,
          italic=False, h="left", fmt=None, bdr=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _fn(bold=bold, color=color, size=size, italic=italic)
    cell.fill = _fl(bg)
    cell.alignment = _al(h=h)
    if bdr: cell.border = _bd()
    if fmt: cell.number_format = fmt
    return cell

def _sec(ws, r, text, nc, bg=OC_DARK):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(row=r, column=1, value=text)
    c.font = _fn(bold=True, color=WHITE, size=10)
    c.fill = _fl(bg); c.alignment = _al(h="left", v="center")
    ws.row_dimensions[r].height = 18
    return r + 1

def _build_ws(wb, quote, title, is_internal):
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = OC_DARK if is_internal else OC_TEAL
    meta  = quote.get('study_meta', {})
    bf    = quote['build_fee']
    fa    = quote['flag_analysis']
    tots  = quote['totals']
    mods  = quote.get('modules', [])
    sym   = quote.get('currency_symbol', '$')
    dur   = quote['study_duration']
    today = datetime.date.today().strftime('%B %d, %Y')
    NC    = 5
    _cw(ws, [30, 14, 16, 14, 20])
    row = 1

    # Title banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=1,
        value=f"OpenClinica EDC Services — {'INTERNAL CONFIDENTIAL' if is_internal else 'Commercial Proposal'}")
    c.font = _fn(bold=True, color=WHITE, size=13)
    c.fill = _fl(OC_DARK); c.alignment = _al(h="left", v="center")
    ws.row_dimensions[row].height = 28; row += 1

    if is_internal:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        c = ws.cell(row=row, column=1,
            value="⚠  INTERNAL USE ONLY — CONFIDENTIAL PRICING — DO NOT DISTRIBUTE")
        c.font = _fn(bold=True, color=RED_C, size=9)
        c.fill = _fl(AMBER); c.alignment = _al(h="center", v="center")
        ws.row_dimensions[row].height = 16; row += 1
    row += 1

    # Study info
    row = _sec(ws, row, "STUDY INFORMATION", NC, bg=OC_MID)
    info = [("Protocol",    meta.get('protocol_number','')),
            ("Study Title", meta.get('study_title','—')),
            ("Sponsor",     meta.get('sponsor','—')),
            ("Phase",       meta.get('study_phase','—')),
            ("Indication",  meta.get('indication','—')),
            ("Study Duration", f"{dur['months']} months"),
            ("Quote Date",  today),
            ("Valid For",   "30 days from quote date")]
    for k, v in info:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NC)
        _cell(ws, row, 1, k, bold=True, bg=OC_LIGHT, size=9)
        c = ws.cell(row=row, column=2, value=str(v) if v else "—")
        c.font = _fn(size=9); c.fill = _fl(WHITE); c.alignment = _al(); c.border = _bd()
        ws.row_dimensions[row].height = 15; row += 1
    row += 1

    # Flag analysis (internal only)
    if is_internal:
        row = _sec(ws, row, "SCOPE ANALYSIS — FLAGGED ITEMS", NC)
        for col, h in enumerate(["Flag Category","Items","% of Counted","",""], start=1):
            _cell(ws, row, col, h, bold=True, color=WHITE, bg=OC_MID, size=9, h="center")
        ws.row_dimensions[row].height = 16; row += 1
        total = fa['total_flagged_counted']
        for cat, count in fa['category_counts'].items():
            pct = (count / max(total,1)) * 100
            bg = GREY_L if row % 2 == 0 else WHITE
            _cell(ws, row, 1, cat.replace('_',' ').title(), bg=bg, size=9)
            _cell(ws, row, 2, count, bold=(count>0), bg=bg, size=9, h="center",
                  color=OC_DARK if count else "888888")
            _cell(ws, row, 3, f"{pct:.0f}%" if count else "—", bg=bg, size=9, h="right")
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=NC)
            ws.row_dimensions[row].height = 14; row += 1
        # Excluded
        excl = "Excluded: " + ", ".join(c.replace('_',' ').title()
                                         for c in fa['excluded_categories'])
        _cell(ws, row, 1, excl, italic=True, color="888888", bg=GREY_L, size=8)
        _cell(ws, row, 2, fa['total_flagged_excluded'], italic=True,
              color="888888", bg=GREY_L, size=8, h="center")
        _cell(ws, row, 3, "—", italic=True, color="888888", bg=GREY_L, size=8, h="right")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=NC)
        ws.row_dimensions[row].height = 14; row += 1; row += 1
        # Basis note
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        c = ws.cell(row=row, column=1,
            value=f"Basis: {fa['total_flagged_counted']} items × {bf['minutes_per_item']:.0f} min "
                  f"= {bf['raw_hours']:.2f} raw hrs → {bf['base_hours']} base + "
                  f"{bf['pm_hours']} PM = {bf['pre_cont_hours']} pre-contingency + "
                  f"{bf['contingency_hours']} contingency ({bf['contingency_pct_display']}) = "
                  f"{bf['total_hours']} total hrs @ ${bf['hourly_rate']:.2f}/hr")
        c.font = _fn(italic=True, size=8, color="555555")
        c.fill = _fl(GREY_L); c.alignment = _al()
        ws.row_dimensions[row].height = 14; row += 1; row += 1

    # One-time fees
    row = _sec(ws, row, "ONE-TIME FEES", NC)
    for col, h in enumerate(["Service","Hours","Rate","Amount",""], start=1):
        _cell(ws, row, col, h, bold=True, color=WHITE, bg=OC_MID, size=9, h="center")
    ws.row_dimensions[row].height = 16; row += 1

    if is_internal:
        fee_rows = [
            ("CRF Configuration — Specialist Effort",
             _fh(bf['base_hours']), bf['hourly_rate'], bf['base_fee']),
            ("Project Management, Review & Support",
             _fh(bf['pm_hours']), bf['hourly_rate'], bf['pm_fee']),
            (f"Contingency ({bf['contingency_pct_display']}) on {_fh(bf['pre_cont_hours'])}",
             _fh(bf['contingency_hours']), bf['hourly_rate'], bf['contingency_fee']),
        ]
    else:
        fee_rows = [
            ("Study Configuration & Build *",
             _fh(bf['base_hours']), bf['hourly_rate'], bf['base_fee']),
            ("Project Management, Review & Support",
             _fh(bf['pm_hours']), bf['hourly_rate'], bf['pm_fee']),
            (f"Contingency ({bf['contingency_pct_display']}) *",
             _fh(bf['contingency_hours']), bf['hourly_rate'], bf['contingency_fee']),
        ]

    for i, (svc, hrs, rate, amt) in enumerate(fee_rows):
        bg = GREY_L if i % 2 == 0 else WHITE
        _cell(ws, row, 1, svc, bg=bg, size=9)
        _cell(ws, row, 2, hrs, bg=bg, size=9, h="center")
        _cell(ws, row, 3, rate, bg=bg, size=9, h="right", fmt=_ccy())
        _cell(ws, row, 4, amt, bold=True, color=OC_DARK, bg=bg, size=9, h="right", fmt=_ccy())
        ws.cell(row=row, column=5).fill = _fl(bg)
        ws.row_dimensions[row].height = 15; row += 1

    if not is_internal:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        c = ws.cell(row=row, column=1,
            value="* Contingency (20%) applied to all specialist and project management hours combined")
        c.font = _fn(italic=True, size=8, color="555555")
        c.fill = _fl(WHITE); c.alignment = _al()
        ws.row_dimensions[row].height = 13; row += 1

    # Total one-time
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value="TOTAL ONE-TIME FEE")
    c.font = _fn(bold=True, color=WHITE, size=10)
    c.fill = _fl(OC_DARK); c.alignment = _al(h="left", v="center")
    _cell(ws, row, 4, bf['total_fee'], bold=True, color=OC_TEAL, bg=OC_DARK,
          size=11, h="right", fmt=_ccy())
    ws.cell(row=row, column=5).fill = _fl(OC_DARK)
    ws.row_dimensions[row].height = 20; row += 1; row += 1

    # Subscriptions
    if mods:
        pc     = quote.get('pricing_context', {})
        seg    = pc.get('segment','COMMERCIAL').replace('_',' ').title()
        vol_d  = pc.get('volume_discount_display','0%')
        plat_d = pc.get('platform_discount_display','0%')
        bundle = pc.get('use_bundle', False)
        rates_date = pc.get('rates_effective_date', 'unknown')
        src    = f"(rates effective {rates_date})"

        hdr_text = (f"SUBSCRIPTION FEES  —  {seg} | {dur['months']} mo | "
                    f"Vol disc: {vol_d}" +
                    (f" | Plat disc: {plat_d}" if pc.get('use_platform_discount') else "") +
                    (f" | Bundle" if bundle else "") +
                    f"  {src}")
        row = _sec(ws, row, hdr_text, NC)

        if is_internal:
            hdrs = ["Module","List/mo","Vol Disc","Plat Disc","Net/mo","Total"]
        else:
            hdrs = ["Module","Term","Monthly Fee","Months","Total",""]
        for col, h in enumerate(hdrs, start=1):
            _cell(ws, row, col, h, bold=True, color=WHITE, bg=OC_MID, size=9, h="center")
        ws.row_dimensions[row].height = 16; row += 1

        for i, m in enumerate(mods):
            bg = GREY_L if i % 2 == 0 else WHITE
            if is_internal:
                lp = m.get('list_price', 0)
                vd = m.get('vol_discount', 0)
                pd = m.get('plat_discount', 0)
                _cell(ws, row, 1, m['name'], bold=True, color=OC_DARK, bg=bg, size=9)
                _cell(ws, row, 2, lp, bg=bg, size=9, h="right", fmt=_ccy())
                _cell(ws, row, 3, f"{int(vd*100)}%", bg=bg, size=9, h="center")
                _cell(ws, row, 4, f"{int(pd*100)}%", bg=bg, size=9, h="center")
                _cell(ws, row, 5, m['monthly_fee'], bold=True, color=OC_DARK,
                      bg=bg, size=9, h="right", fmt=_ccy())
                _cell(ws, row, 6, m['total_fee'], bold=True, color=OC_DARK,
                      bg=bg, size=9, h="right", fmt=_ccy())
            else:
                _cell(ws, row, 1, m['name'], bold=True, color=OC_DARK, bg=bg, size=9)
                _cell(ws, row, 2, "Monthly", bg=bg, size=9, h="center")
                _cell(ws, row, 3, m['monthly_fee'], bg=bg, size=9, h="right", fmt=_ccy())
                _cell(ws, row, 4, dur['months'], bg=bg, size=9, h="center")
                _cell(ws, row, 5, m['total_fee'], bold=True, color=OC_DARK,
                      bg=bg, size=9, h="right", fmt=_ccy())
                ws.cell(row=row, column=6).fill = _fl(bg)
            ws.row_dimensions[row].height = 15; row += 1
        row += 1

    # Grand total
    row = _sec(ws, row, "QUOTE SUMMARY", NC)
    summary = [("One-Time Fees (Build + Project Management)", tots['build_fee']),
               ("Subscription Fees (all modules × duration)", tots['module_total'])]
    for k, v in summary:
        bg = GREY_L if row % 2 == 0 else WHITE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        _cell(ws, row, 1, k, size=9, bg=bg)
        tbd = v == 0 and k.startswith("Sub")
        c4 = ws.cell(row=row, column=4, value=v if not tbd else "TBD")
        c4.font = _fn(size=9, color=AMBER_T if tbd else "000000")
        c4.fill = _fl(bg); c4.border = _bd(); c4.alignment = _al(h="right")
        if not tbd: c4.number_format = _ccy()
        ws.cell(row=row, column=5).fill = _fl(bg)
        ws.row_dimensions[row].height = 15; row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value="ESTIMATED TOTAL")
    c.font = _fn(bold=True, color=WHITE, size=11)
    c.fill = _fl(OC_DARK); c.alignment = _al(h="left", v="center")
    grand = tots['grand_total']
    tbd_g = grand == 0
    c4 = ws.cell(row=row, column=4, value=grand if not tbd_g else "TBD")
    c4.font = _fn(bold=True, color=OC_TEAL, size=12)
    c4.fill = _fl(OC_DARK); c4.border = _bd(); c4.alignment = _al(h="right")
    if not tbd_g: c4.number_format = _ccy()
    ws.cell(row=row, column=5).fill = _fl(OC_DARK)
    ws.row_dimensions[row].height = 22; row += 1

    ws.freeze_panes = "A2"



def _build_appendix_ws(wb, quote, is_internal):
    """
    Appendix sheet — one row per flagged item.
    Columns: # | Category | Item | Comment (when EDC structure supplied) | Est. Effort
    """
    fa  = quote['flag_analysis']
    raw = quote.get('_raw_flags', {})

    cat_labels = {
        'site_specific':         'Site Specific',
        'oid_confirmation':      'OID Confirmation',
        'protocol_ambiguous':    'Protocol Ambiguous',
        'constraint_review':     'Constraint Review',
        'custom_domain':         'Custom Domain',
        'pdf_mapping_uncertain': 'PDF Mapping Uncertain',
        'name_deviation':        'Name Deviation',
    }
    cat_bgs = {
        'site_specific':         'FADBD8',
        'oid_confirmation':      'FDEBD0',
        'protocol_ambiguous':    'FDEBD0',
        'constraint_review':     'FEF9E7',
        'custom_domain':         'EBF5FB',
        'pdf_mapping_uncertain': 'FDEBD0',
        'name_deviation':        'EBF5FB',
    }

    # Collect all items — handle plain strings and {item, comment} dicts
    all_items = []   # (cat_label, item_str, comment_str, bg)
    for cat in fa['counted_categories']:
        entries = raw.get(cat, [])
        if not isinstance(entries, list) or not entries:
            continue
        label = cat_labels.get(cat, cat.replace('_',' ').title())
        bg    = cat_bgs.get(cat, GREY_L)
        for entry in entries:
            if isinstance(entry, dict):
                item    = entry.get('item', '')
                comment = entry.get('comment', '')
            else:
                s = str(entry).strip()
                if ' — ' in s:
                    parts = s.split(' — ', 1)
                    item, comment = parts[0].strip(), parts[1].strip()
                elif ' - ' in s:
                    parts = s.split(' - ', 1)
                    item, comment = parts[0].strip(), parts[1].strip()
                else:
                    item, comment = s, ''
            all_items.append((label, item, comment, bg))

    any_comments = any(c for _, _, c, _ in all_items)

    ws = wb.create_sheet(title="Scope Item Detail")
    ws.sheet_properties.tabColor = "2D3561"

    # Column widths — 5-col with comment, 4-col without
    if any_comments:
        _cw(ws, [4, 22, 28, 42, 10])
        NC = 5
    else:
        _cw(ws, [4, 24, 58, 10])
        NC = 4

    row = 1

    # Title banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=1, value="APPENDIX — SCOPE ITEM DETAIL")
    c.font = _fn(bold=True, color=WHITE, size=11)
    c.fill = _fl(OC_DARK); c.alignment = _al(h="left", v="center")
    ws.row_dimensions[row].height = 22; row += 1

    # Intro note
    src = " (with comments from EDC structure analysis)" if any_comments else ""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=1,
        value=(f"{fa['total_flagged_counted']} items identified during protocol "
               f"analysis{src} — each requires 1 hr specialist effort."))
    c.font = _fn(size=9); c.fill = _fl(GREY_L); c.alignment = _al()
    ws.row_dimensions[row].height = 15; row += 1; row += 1

    # Color legend — one mini-row per category that appears in this quote
    legend_cats = []
    for cat in fa.get('counted_categories', []):
        # Only include categories that actually have items in all_items
        if any(lbl == cat_labels.get(cat, cat.replace('_',' ').title())
               for lbl, _, _, _ in all_items):
            legend_cats.append(cat)
    if legend_cats:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        c = ws.cell(row=row, column=1, value="COLOR KEY")
        c.font = _fn(bold=True, size=8, color="555555")
        c.fill = _fl(WHITE); c.alignment = _al()
        ws.row_dimensions[row].height = 12; row += 1

        # Show legend as one row with all categories side by side
        # Each legend cell: left half = colored swatch with label
        legend_col = 1
        for cat in legend_cats:
            label = cat_labels.get(cat, cat.replace('_',' ').title())
            bg    = cat_bgs.get(cat, GREY_L)
            if legend_col > NC:
                row += 1
                legend_col = 1
            c = ws.cell(row=row, column=legend_col, value=f"  {label}")
            c.font = _fn(bold=True, size=8, color=OC_DARK)
            c.fill = _fl(bg); c.border = _bd(); c.alignment = _al(h="left", v="center")
            legend_col += 1
        ws.row_dimensions[row].height = 16; row += 1; row += 1

    # Column headers
    if any_comments:
        hdrs = ["#", "Category", "Item", "Comment / Action Required", "Effort"]
    else:
        hdrs = ["#", "Category", "Item Description", "Effort"]

    for col, h in enumerate(hdrs, start=1):
        _cell(ws, row, col, h, bold=True, color=WHITE, bg=OC_MID, size=9, h="center")
    ws.row_dimensions[row].height = 16; row += 1

    # Data rows
    for i, (label, item, comment, bg) in enumerate(all_items, start=1):
        _cell(ws, row, 1, i, bg=bg, size=8, h="center")
        _cell(ws, row, 2, label, bold=True, color=OC_DARK, bg=bg, size=8)
        _cell(ws, row, 3, item, bg=bg, size=8)
        if any_comments:
            c = ws.cell(row=row, column=4, value=comment)
            c.font = _fn(italic=bool(comment), size=8,
                         color="444444" if comment else "888888")
            c.fill = _fl(bg); c.border = _bd(); c.alignment = _al()
            _cell(ws, row, 5, "1 hr", bg=bg, size=8, h="center")
        else:
            _cell(ws, row, NC, "1 hr", bg=bg, size=8, h="center")
        ws.row_dimensions[row].height = 18; row += 1

    row += 1
    # Total row
    n = len(all_items)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC-1)
    c = ws.cell(row=row, column=1, value=f"TOTAL BILLABLE ITEMS: {n}")
    c.font = _fn(bold=True, color=WHITE, size=10)
    c.fill = _fl(OC_DARK); c.alignment = _al(h="left", v="center")
    _cell(ws, row, NC, f"{n} hrs", bold=True, color=OC_DARK,
          bg=OC_DARK, size=10, h="center")
    ws.cell(row=row, column=NC).font = _fn(bold=True, color="00A99D", size=10)
    ws.row_dimensions[row].height = 20; row += 1; row += 1

    excl = fa['total_flagged_excluded']
    if excl:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        c = ws.cell(row=row, column=1,
            value=(f"Note: {excl} Choice List Review item(s) excluded — "
                   f"resolved during client review, not billable build effort."))
        c.font = _fn(italic=True, size=8, color="555555")
        c.fill = _fl(GREY_L); c.alignment = _al()
        ws.row_dimensions[row].height = 14

    ws.freeze_panes = "A4"

def build_quote_xlsx(quote, internal_path, client_path):
    wb_i = Workbook()
    # Remove the default blank sheet — _build_ws creates its own
    del wb_i[wb_i.sheetnames[0]]
    _build_ws(wb_i, quote, "INTERNAL QUOTE", is_internal=True)
    _build_appendix_ws(wb_i, quote, is_internal=True)
    wb_i.save(internal_path)
    print(f"Internal XLSX: {internal_path}")

    wb_c = Workbook()
    # Remove the default blank sheet — _build_ws creates its own
    del wb_c[wb_c.sheetnames[0]]
    _build_ws(wb_c, quote, "PROPOSAL", is_internal=False)
    _build_appendix_ws(wb_c, quote, is_internal=False)
    wb_c.save(client_path)
    print(f"Client XLSX: {client_path}")


if __name__ == '__main__':
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    from pricing_engine import calculate_quote
    sample = {
        'study_meta': {'protocol_number':'PrTK05','study_title':'CAN-2409 Phase 2a',
                       'sponsor':'Candel Therapeutics','study_phase':'Phase 2a',
                       'indication':'Prostate Cancer','total_study_duration_months':24},
        'review_flags': {
            'site_specific':      ['Lab ranges','LBNAM','Site count'],
            'oid_confirmation':   [],
            'protocol_ambiguous': ['BE qPCR','Biomarker list','BES type','SE_UNSCH','DC'],
            'constraint_review':  ['VS window','LB window','EBRT date','EC dates','EXDOSE'],
            'choice_list_review': ['IE003CD','DSDECOD'],
            'custom_domain':      ['BE Lab Manual','EC_DIARY','DC sponsor'],
            'pdf_mapping_uncertain': [], 'name_deviation': [],
        },
        'is_epro_required': True,
    }
    quote = calculate_quote(sample)
    build_quote_xlsx(quote,
        '/mnt/user-data/outputs/PrTK05_Quote_Internal.xlsx',
        '/mnt/user-data/outputs/PrTK05_Quote_Client.xlsx')

"""
generate_dvs.py — DVS Specification Generator
Builds a completed OpenClinica 4 DVS xlsx from a dvs_data dict.
Opens the template, preserves reference sheets, writes all editable sheets.

Usage:
    from generate_dvs import build_dvs
    build_dvs(dvs_data, output_path)
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, datetime

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "references", "DVS_Template.xlsx"
)

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1B3A6B"
MID_BLUE    = "2E6DA4"
LIGHT_BLUE  = "D6E4F0"
WHITE       = "FFFFFF"
GREY_LIGHT  = "F5F5F5"
GREY_MID    = "CCCCCC"
AMBER       = "FFF3CD"
GREEN_LIGHT = "D5F5E3"
RED_LIGHT   = "FADBD8"

# Reference sheets that must be preserved exactly from the template
REFERENCE_SHEETS = {"README", "Lookups", "OC4_Syntax_Guide", "Examples"}

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=9, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def _border(color=GREY_MID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def _hdr(cell, value, width=20):
    cell.value = value
    cell.font = _font(bold=True, color=WHITE, size=9)
    cell.fill = _fill(DARK_BLUE)
    cell.border = _border()
    cell.alignment = _align(h="center", v="center")

def _data(cell, value, row_i=0, amber=False, green=False, red=False):
    cell.value = value if value is not None else ""
    cell.font = _font(size=8)
    if amber:
        cell.fill = _fill(AMBER)
    elif green:
        cell.fill = _fill(GREEN_LIGHT)
    elif red:
        cell.fill = _fill(RED_LIGHT)
    else:
        cell.fill = _fill(GREY_LIGHT if row_i % 2 == 0 else WHITE)
    cell.border = _border()
    cell.alignment = _align()

def _title(ws, text, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = _font(bold=True, color=WHITE, size=11)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align(h="left", v="center")
    ws.row_dimensions[1].height = 22

def _sub_title(ws, text, n_cols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, color=WHITE, size=8)
    c.fill = _fill(MID_BLUE)
    c.alignment = _align(h="left", v="center")
    ws.row_dimensions[row].height = 14

def _write_sheet(ws, headers, rows, col_widths, title_text=None, start_data_row=3):
    """Generic sheet writer: title (row 1), optional sub-row (row 2), headers, data."""
    n = len(headers)
    if title_text:
        _title(ws, title_text, n)
        ws.row_dimensions[1].height = 20

    # Header row
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_data_row - 1, column=col_i)
        _hdr(c, h)
        ws.column_dimensions[get_column_letter(col_i)].width = col_widths.get(h, 18)
    ws.row_dimensions[start_data_row - 1].height = 16

    # Data rows
    for row_i, row in enumerate(rows):
        for col_i, h in enumerate(headers, start=1):
            val = row.get(h, "")
            is_amber  = val and "[PLACEHOLDER" in str(val).upper()
            is_draft  = h == "Status" and val == "Draft"
            c = ws.cell(row=start_data_row + row_i, column=col_i)
            _data(c, val, row_i, amber=is_amber)
        ws.row_dimensions[start_data_row + row_i].height = 13

    ws.freeze_panes = ws.cell(row=start_data_row, column=1)


# ── Sheet column definitions ──────────────────────────────────────────────────
PROTOCOL_EXTRACTION_COLS = [
    "Source Section", "Protocol Reference", "Category",
    "Structured Requirement / Fact", "Raw Protocol Text Summary",
    "Downstream Build Object", "Potential Check Needed?", "Candidate Check ID",
    "Related Event OID", "Related Form OID", "Related Item Name / OID",
    "Priority", "Owner", "Status", "Notes"
]
PROTOCOL_EXTRACTION_WIDTHS = {
    "Source Section": 20, "Protocol Reference": 16, "Category": 18,
    "Structured Requirement / Fact": 40, "Raw Protocol Text Summary": 30,
    "Downstream Build Object": 18, "Potential Check Needed?": 18,
    "Candidate Check ID": 14, "Related Event OID": 18, "Related Form OID": 16,
    "Related Item Name / OID": 28, "Priority": 12, "Owner": 14,
    "Status": 12, "Notes": 35,
}

DVS_OC4_COLS = [
    "Check ID", "Status", "Check Name", "Business Purpose",
    "Protocol Reference", "Source Section", "Check Type", "Severity",
    "Trigger Point", "Event Scope", "Source Event OID(s)", "Current Event Needed?",
    "crossform_references", "Target Form OID", "Target Item Name", "Target Item OID",
    "Source Form OID(s)", "Source Item Name(s)", "Source Item OID(s)",
    "Helper Calculate Item Needed?", "Helper Item OID", "OC4 Logic Pattern",
    "Expression / Calculation", "Constraint / Required / Relevant Message",
    "Query Text ID", "Expected Site Action", "Build Owner", "Priority",
    "UAT Case ID(s)", "Notes"
]
DVS_OC4_WIDTHS = {
    "Check ID": 12, "Status": 12, "Check Name": 32, "Business Purpose": 36,
    "Protocol Reference": 16, "Source Section": 20, "Check Type": 22,
    "Severity": 14, "Trigger Point": 26, "Event Scope": 16,
    "Source Event OID(s)": 20, "Current Event Needed?": 18,
    "crossform_references": 20, "Target Form OID": 16, "Target Item Name": 28,
    "Target Item OID": 22, "Source Form OID(s)": 18, "Source Item Name(s)": 24,
    "Source Item OID(s)": 22, "Helper Calculate Item Needed?": 22,
    "Helper Item OID": 20, "OC4 Logic Pattern": 28,
    "Expression / Calculation": 40, "Constraint / Required / Relevant Message": 40,
    "Query Text ID": 12, "Expected Site Action": 36, "Build Owner": 14,
    "Priority": 12, "UAT Case ID(s)": 16, "Notes": 35,
}

QT_COLS = [
    "Query Text ID", "Status", "Standard Message", "Audience",
    "When to Use", "Avoid / Notes", "Related Check ID(s)",
    "Priority", "Owner", "Version Notes"
]
QT_WIDTHS = {
    "Query Text ID": 14, "Status": 12, "Standard Message": 50, "Audience": 12,
    "When to Use": 30, "Avoid / Notes": 28, "Related Check ID(s)": 18,
    "Priority": 12, "Owner": 14, "Version Notes": 28,
}

UAT_COLS = [
    # ── Human-authored test spec (cols 1–16) ───────────────────────────
    "UAT Case ID", "Status", "Related Check ID", "Scenario",
    "Preconditions", "Test Steps", "Input Data", "Expected Result",
    "Actual Result", "Test Result", "Tester", "Execution Date",
    "Defect / Ticket", "Retest Needed?", "Priority", "Notes",
    # ── ODM load coordinates (cols 17–25) ─────────────────────────────
    # Cols 17-18 are RUNTIME POPULATED (blank in DVS, stamped by uat_loader)
    "Site_OID", "Participant_Key",
    # Cols 19-25 are populated by DVS skill from XLSForm metadata
    "Study_Event_OID", "Event_Repeat_Key", "Form_OID",
    "Item_Group_OID", "Item_OID", "Participant_ID", "Load_Order", "Load_Value",
]
UAT_WIDTHS = {
    "UAT Case ID": 12, "Status": 12, "Related Check ID": 14, "Scenario": 32,
    "Preconditions": 28, "Test Steps": 36, "Input Data": 22,
    "Expected Result": 36, "Actual Result": 28, "Test Result": 14,
    "Tester": 14, "Execution Date": 16, "Defect / Ticket": 16,
    "Retest Needed?": 14, "Priority": 12, "Notes": 28,
    # ODM columns
    "Site_OID": 28, "Participant_Key": 30,
    "Study_Event_OID": 22, "Event_Repeat_Key": 16, "Form_OID": 16,
    "Item_Group_OID": 22, "Item_OID": 28, "Participant_ID": 14, "Load_Order": 12,
    "Load_Value": 22,
}

# ── Calendaring_Rules columns ─────────────────────────────────────────
CAL_RULES_COLS = [
    "Rule ID", "Status", "Rule Name", "Business Purpose",
    "Protocol Reference", "Trigger Type", "Trigger OID",
    "Schedule", "Schedule Time",
    "Condition (XPath)", "Condition (Plain English)",
    "Action Type", "Target Event OID", "Target Form OID",
    "Action Parameters", "Rule Result To Trigger On",
    "JSON Output", "Build Owner", "Priority", "UAT Case ID(s)",
]
CAL_RULES_WIDTHS = {
    "Rule ID": 12, "Status": 12, "Rule Name": 36, "Business Purpose": 40,
    "Protocol Reference": 16, "Trigger Type": 26, "Trigger OID": 20,
    "Schedule": 12, "Schedule Time": 14,
    "Condition (XPath)": 50, "Condition (Plain English)": 40,
    "Action Type": 22, "Target Event OID": 22, "Target Form OID": 18,
    "Action Parameters": 40, "Rule Result To Trigger On": 22,
    "JSON Output": 60, "Build Owner": 14, "Priority": 12,
    "UAT Case ID(s)": 16,
}

# ── Calendaring_UAT columns ───────────────────────────────────────────
CAL_UAT_COLS = [
    "UAT Case ID", "Status", "Related Rule ID", "Scenario",
    "Preconditions", "Setup Steps", "Trigger Action",
    "Expected Outcome", "Verification Steps",
    "Actual Result", "Test Result", "Tester", "Execution Date",
    "Defect / Ticket", "Notes",
]
CAL_UAT_WIDTHS = {
    "UAT Case ID": 12, "Status": 12, "Related Rule ID": 14, "Scenario": 36,
    "Preconditions": 36, "Setup Steps": 40, "Trigger Action": 36,
    "Expected Outcome": 40, "Verification Steps": 40,
    "Actual Result": 28, "Test Result": 14, "Tester": 14,
    "Execution Date": 16, "Defect / Ticket": 16, "Notes": 28,
}


# ── Info-tab / special sheet writers ─────────────────────────────────────────

# Fixed UAT_Setup topic/description rows (written identically on every run)
_UAT_SETUP_ROWS = [
    ("Purpose",
     "This tab documents how the automated UAT data loader works. "
     "It is a reference for anyone running or troubleshooting UAT."),
    ("Site Naming",
     "A new site is created in the OC4 Test environment on every UAT run. "
     "Site Name: UAT Automation Site - YYYY-MM-DD HH:MM  |  "
     "Site ID (OID): UAT-YYYYMMDD-HHMMSS. "
     "This timestamp is derived from the moment the pipeline triggers. "
     "Creating a new site per run prevents data contamination between test runs "
     "and provides a visible audit trail in OC4's site list."),
    ("Finding a Test Run",
     "In OC4, navigate to the study in the Test environment and open the Sites list. "
     "Each UAT run appears as a dated site. "
     "The most recent run is the site with the latest timestamp in its name."),
    ("Participant Naming",
     "Logical participant IDs in this DVS are UAT-P001, UAT-P002, etc. "
     "At runtime the loader maps these to run-scoped keys: "
     "UAT-YYYYMMDD-HHMMSS-P001. "
     "This ensures participants from different runs never collide."),
    ("Cross-Form Test Cases",
     "All data for a single test case — including baseline data from prerequisite "
     "forms — must load into the same participant. For example, if an AE form checks "
     "consent date from DM, the DM data and AE data for that test both load for "
     "UAT-P001. A second participant (UAT-P002) is only used when the test scenario "
     "requires a genuinely different baseline state (e.g. different sex)."),
    ("Load Order",
     "Data loads in the order defined by the Load_Order column in UAT_Cases. "
     "DM and ICF always load first to establish the baseline participant record. "
     "Clinical forms follow in visit order. The loader respects cross-form "
     "dependencies: source form data always loads before forms that reference it."),
    ("Runtime Columns",
     "Two columns in UAT_Cases are blank in this DVS and are stamped by the loader "
     "at runtime: Site_OID (the OID of the dated site created for this run) and "
     "Participant_Key (the run-scoped full participant key). These are written back "
     "to the DVS Results file uploaded to monday.com after the run completes."),
    ("If a Run Fails",
     "If the loader fails partway through, a partial site and participants may exist "
     "in OC4 Test. These can be left in place — they are harmless, clearly dated, "
     "and do not affect subsequent runs. The next run creates a new dated site. "
     "Do not attempt to delete partial sites manually unless storage is a concern."),
    ("Calendaring Tests",
     "Calendaring rule tests (on the Calendaring_UAT tab) are MANUAL STEPS ONLY. "
     "The loader does not execute calendaring UAT. "
     "See the Calendaring_UAT tab for step-by-step instructions to verify each rule."),
]


def _write_uat_setup_sheet(ws):
    """Write the fixed UAT_Setup information tab (2-column topic/description table)."""
    n_cols = 2
    # Title row
    _title(ws, "UAT Data Loading — Setup Reference", n_cols)
    ws.row_dimensions[1].height = 22
    # Sub-title
    _sub_title(ws, "Read-only reference. Describes how the automated UAT loader works.", n_cols, row=2)
    ws.row_dimensions[2].height = 14
    # Column headers (row 3)
    _hdr(ws.cell(row=3, column=1), "Topic")
    _hdr(ws.cell(row=3, column=2), "Description")
    ws.row_dimensions[3].height = 16
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    # Data rows
    for i, (topic, desc) in enumerate(_UAT_SETUP_ROWS):
        row_num = 4 + i
        tc = ws.cell(row=row_num, column=1, value=topic)
        tc.font = _font(bold=True, size=9)
        tc.fill = _fill(LIGHT_BLUE)
        tc.border = _border()
        tc.alignment = _align(h="left", v="top")
        dc = ws.cell(row=row_num, column=2, value=desc)
        dc.font = _font(size=8)
        dc.fill = _fill(GREY_LIGHT if i % 2 == 0 else WHITE)
        dc.border = _border()
        dc.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        ws.row_dimensions[row_num].height = 48
    ws.freeze_panes = ws.cell(row=4, column=1)


def _write_cal_rules_sheet(ws, rows, meta, today):
    """Write the Calendaring_Rules sheet."""
    _write_sheet(
        ws,
        CAL_RULES_COLS,
        rows,
        CAL_RULES_WIDTHS,
        title_text=f"Calendaring Rules  |  {meta.get('protocol_number', '')}  |  {today}",
        start_data_row=4,
    )


def _write_cal_uat_sheet(ws, rows, meta, today):
    """Write the Calendaring_UAT sheet."""
    _write_sheet(
        ws,
        CAL_UAT_COLS,
        rows,
        CAL_UAT_WIDTHS,
        title_text=f"Calendaring UAT Cases (Manual)  |  {meta.get('protocol_number', '')}  |  {today}",
        start_data_row=4,
    )


# ── Main build function ───────────────────────────────────────────────────────
def build_dvs(dvs_data, output_path):
    """
    Build DVS xlsx from dvs_data dict.
    dvs_data structure:
    {
        'study_meta': { 'protocol_number': ..., 'study_id': ..., 'build_date': ... },
        'protocol_extraction': [ {col: val, ...}, ... ],
        'dvs_oc4': [ {col: val, ...}, ... ],
        'query_text_library': [ {col: val, ...}, ... ],
        'uat_cases': [ {col: val, ...}, ... ],  # 25 cols; ODM cols 17-25 populated by DVS skill
        'calendaring_rules': [ {col: val, ...}, ... ],  # optional; from Study Spec JSON
        'calendaring_uat': [ {col: val, ...}, ... ],    # optional; one per CAL rule
    }
    """
    # Load template to preserve reference sheets
    wb = load_workbook(TEMPLATE_PATH)
    today = datetime.date.today().isoformat()
    meta  = dvs_data.get("study_meta", {})

    # Update README prepared date
    if "README" in wb.sheetnames:
        ws_rm = wb["README"]
        for row in ws_rm.iter_rows(min_row=1, max_row=15):
            for cell in row:
                if cell.value and "Prepared date" in str(cell.value):
                    # Update the cell to the right
                    pass  # keep as generated date from template

    # ── Protocol_Extraction ───────────────────────────────────────────────
    ws_pe = wb["Protocol_Extraction"] if "Protocol_Extraction" in wb.sheetnames \
            else wb.create_sheet("Protocol_Extraction")
    # Clear existing data rows (keep row 1 title and row 3 headers)
    for row in ws_pe.iter_rows(min_row=4):
        for cell in row:
            cell.value = None
    # Rewrite
    _write_sheet(
        ws_pe,
        PROTOCOL_EXTRACTION_COLS,
        dvs_data.get("protocol_extraction", []),
        PROTOCOL_EXTRACTION_WIDTHS,
        title_text=f"Protocol-to-DVS Extraction Log  |  {meta.get('protocol_number','')}  |  {today}",
        start_data_row=4,
    )

    # ── DVS_OC4 ───────────────────────────────────────────────────────────
    ws_dvs = wb["DVS_OC4"] if "DVS_OC4" in wb.sheetnames \
             else wb.create_sheet("DVS_OC4")
    for row in ws_dvs.iter_rows(min_row=4):
        for cell in row:
            cell.value = None
    _write_sheet(
        ws_dvs,
        DVS_OC4_COLS,
        dvs_data.get("dvs_oc4", []),
        DVS_OC4_WIDTHS,
        title_text=f"OpenClinica 4 Data Validation Specification  |  {meta.get('protocol_number','')}  |  {today}",
        start_data_row=4,
    )

    # ── Query_Text_Library ────────────────────────────────────────────────
    ws_qt = wb["Query_Text_Library"] if "Query_Text_Library" in wb.sheetnames \
            else wb.create_sheet("Query_Text_Library")
    for row in ws_qt.iter_rows(min_row=4):
        for cell in row:
            cell.value = None
    _write_sheet(
        ws_qt,
        QT_COLS,
        dvs_data.get("query_text_library", []),
        QT_WIDTHS,
        title_text=f"Query Text Library  |  {meta.get('protocol_number','')}  |  {today}",
        start_data_row=4,
    )

    # ── UAT_Cases ─────────────────────────────────────────────────────────
    ws_uat = wb["UAT_Cases"] if "UAT_Cases" in wb.sheetnames \
             else wb.create_sheet("UAT_Cases")
    for row in ws_uat.iter_rows(min_row=4):
        for cell in row:
            cell.value = None
    _write_sheet(
        ws_uat,
        UAT_COLS,
        dvs_data.get("uat_cases", []),
        UAT_WIDTHS,
        title_text=f"User Acceptance Test Cases  |  {meta.get('protocol_number','')}  |  {today}",
        start_data_row=4,
    )

    # ── UAT_Setup (info tab — fixed content, no dvs_data input) ──────────
    ws_setup = wb["UAT_Setup"] if "UAT_Setup" in wb.sheetnames \
               else wb.create_sheet("UAT_Setup")
    # Clear and rewrite on every run
    ws_setup.delete_rows(1, ws_setup.max_row + 1)
    _write_uat_setup_sheet(ws_setup)

    # ── Calendaring_Rules ─────────────────────────────────────────────────
    ws_cal = wb["Calendaring_Rules"] if "Calendaring_Rules" in wb.sheetnames \
             else wb.create_sheet("Calendaring_Rules")
    for row in ws_cal.iter_rows(min_row=4):
        for cell in row:
            cell.value = None
    _write_cal_rules_sheet(ws_cal, dvs_data.get("calendaring_rules", []), meta, today)

    # ── Calendaring_UAT ───────────────────────────────────────────────────
    ws_cuat = wb["Calendaring_UAT"] if "Calendaring_UAT" in wb.sheetnames \
              else wb.create_sheet("Calendaring_UAT")
    for row in ws_cuat.iter_rows(min_row=4):
        for cell in row:
            cell.value = None
    _write_cal_uat_sheet(ws_cuat, dvs_data.get("calendaring_uat", []), meta, today)

    # Ensure sheet order:
    # README, Lookups, Protocol_Extraction, DVS_OC4, Query_Text_Library,
    # UAT_Cases, UAT_Setup, Calendaring_Rules, Calendaring_UAT,
    # OC4_Syntax_Guide, Examples
    desired_order = [
        "README", "Lookups", "Protocol_Extraction", "DVS_OC4",
        "Query_Text_Library", "UAT_Cases", "UAT_Setup",
        "Calendaring_Rules", "Calendaring_UAT",
        "OC4_Syntax_Guide", "Examples",
    ]
    # Move sheets to correct positions
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i
                          if wb.sheetnames.index(name) != i else 0)

    wb.save(output_path)
    print(f"DVS written to: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 3:
        print("Usage: python generate_dvs.py <dvs_data.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        data = json.load(f)
    build_dvs(data, sys.argv[2])

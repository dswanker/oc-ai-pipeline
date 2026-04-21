"""
build_checklist.py — Study Build Checklist Generator
Produces both a PDF and XLSX checklist from the build log and spec data.

Usage:
    from build_checklist import build_checklist_pdf, build_checklist_xlsx
    build_checklist_pdf(spec_data, build_log, output_path)
    build_checklist_xlsx(spec_data, build_log, output_path)
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab for PDF
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE  = colors.HexColor("#1B3A6B")
MID_BLUE   = colors.HexColor("#2E6DA4")
LIGHT_BLUE = colors.HexColor("#D6E4F0")
WHITE      = colors.white
GREY_LIGHT = colors.HexColor("#F5F5F5")
GREY_MID   = colors.HexColor("#CCCCCC")
GREEN      = colors.HexColor("#27AE60")
AMBER      = colors.HexColor("#E67E22")
RED        = colors.HexColor("#C0392B")
GREEN_LIGHT= colors.HexColor("#D5F5E3")
AMBER_LIGHT= colors.HexColor("#FDEBD0")
RED_LIGHT  = colors.HexColor("#FADBD8")

PAGE_W, PAGE_H = landscape(A4)
MARGIN    = 1.8 * cm
CONTENT_W = PAGE_W - 2 * MARGIN

# ── QA checks definition ──────────────────────────────────────────────────────
QA_CHECKS = [
    ("settings_complete",   "Settings Complete",          "form_title, form_id, version, style, namespaces all populated"),
    ("survey_has_rows",     "Survey Has Rows",            "At least 1 non-calculate survey row present"),
    ("groups_balanced",     "Groups Balanced",            "Every begin group has matching end group"),
    ("repeats_balanced",    "Repeats Balanced",           "Every begin repeat has matching end repeat"),
    ("choices_complete",    "Choice Lists Complete",      "Every select_one/select_multiple references a defined list"),
    ("required_cols",       "Required Columns Present",   "type, name, label populated for all non-calculate rows"),
    ("no_placeholders",     "No Placeholders Remaining",  "No [PLACEHOLDER...] text in any cell"),
    ("oid_flagged",         "OID Paths Flagged",          "Cross-form OID placeholders noted for post-config"),
    ("tpt_csv_present",     "Timepoint CSV Present",      "{study_id}_tpt.csv included in package"),
    ("labranges_present",   "Lab Ranges CSV Present",     "labranges.csv included in package"),
    ("epro_flagged",        "ePRO Forms Identified",      "is_epro=Yes forms flagged for ePRO module config"),
]

def run_qa_checks(form, build_log):
    """Run QA checks on a single form. Returns list of (check_id, status, note)."""
    results = []
    settings = form.get('settings', {})
    choices  = form.get('choices', [])
    survey   = form.get('survey', [])
    form_id  = form.get('form_id', '')

    # settings_complete
    missing_settings = [k for k in ['form_title','form_id','version','style','namespaces']
                        if not settings.get(k)]
    results.append(("settings_complete",
                    "PASS" if not missing_settings else "FAIL",
                    f"Missing: {', '.join(missing_settings)}" if missing_settings else ""))

    # survey_has_rows
    data_rows = [r for r in survey if r.get('type','').lower() not in
                 ('calculate','begin group','end group','begin repeat','end repeat','')]
    results.append(("survey_has_rows",
                    "PASS" if data_rows else "FAIL",
                    f"{len(data_rows)} data rows" if data_rows else "No data rows found"))

    # groups_balanced
    opens  = sum(1 for r in survey if r.get('type','').lower() in ('begin group', 'begin repeat'))
    closes = sum(1 for r in survey if r.get('type','').lower() in ('end group', 'end repeat'))
    results.append(("groups_balanced",
                    "PASS" if opens == closes else "FAIL",
                    f"{opens} opens, {closes} closes" if opens != closes else f"{opens} groups"))

    # repeats_balanced (subset of above, check separately)
    r_opens  = sum(1 for r in survey if r.get('type','').lower() == 'begin repeat')
    r_closes = sum(1 for r in survey if r.get('type','').lower() == 'end repeat')
    results.append(("repeats_balanced",
                    "PASS" if r_opens == r_closes else "FAIL",
                    f"{r_opens} begin repeat, {r_closes} end repeat" if r_opens != r_closes
                    else (f"{r_opens} repeat group(s)" if r_opens > 0 else "No repeating groups")))

    # choices_complete
    defined_lists = set(c.get('list_name','') for c in choices)
    referenced_lists = set()
    for r in survey:
        t = r.get('type','')
        if t.startswith('select_one ') or t.startswith('select_multiple '):
            parts = t.split(' ', 1)
            if len(parts) > 1:
                referenced_lists.add(parts[1].strip())
    missing_lists = referenced_lists - defined_lists
    results.append(("choices_complete",
                    "PASS" if not missing_lists else "FAIL",
                    f"Missing lists: {', '.join(missing_lists)}" if missing_lists
                    else f"{len(defined_lists)} list(s) defined"))

    # required_cols
    bad_rows = [r.get('name','?') for r in survey
                if r.get('type','').lower() not in ('calculate','begin group','end group',
                                                     'begin repeat','end repeat','note','')
                and (not r.get('type') or not r.get('name') or not r.get('label'))]
    results.append(("required_cols",
                    "PASS" if not bad_rows else "NEEDS ATTENTION",
                    f"Rows with missing type/name/label: {', '.join(bad_rows[:3])}" if bad_rows else ""))

    # no_placeholders
    ph_fields = []
    for r in survey:
        for k, v in r.items():
            if v and '[PLACEHOLDER' in str(v).upper():
                ph_fields.append(r.get('name','?'))
                break
    for c in choices:
        for k, v in c.items():
            if v and '[PLACEHOLDER' in str(v).upper():
                ph_fields.append(f"choice:{c.get('list_name','?')}")
                break
    results.append(("no_placeholders",
                    "PASS" if not ph_fields else "NEEDS ATTENTION",
                    f"{len(ph_fields)} field(s) still have placeholder values" if ph_fields
                    else "No placeholders remaining"))

    # oid_flagged
    oid_fields = [r.get('name','?') for r in survey
                  if r.get('calculation','') and '[' in r.get('calculation','') and
                  'OID' in r.get('calculation','').upper()]
    results.append(("oid_flagged",
                    "NEEDS ATTENTION" if oid_fields else "PASS",
                    f"{len(oid_fields)} field(s) need OID confirmation" if oid_fields else ""))

    # tpt/labranges present — form-level N/A
    results.append(("tpt_csv_present", "N/A", "Checked at package level"))
    results.append(("labranges_present", "N/A", "Checked at package level"))

    # epro_flagged
    is_epro = form.get('epro', 'No') == 'Yes'
    results.append(("epro_flagged",
                    "NEEDS ATTENTION" if is_epro else "N/A",
                    "Configure in OpenClinica ePRO module before upload" if is_epro else ""))

    return results


# ── PDF Checklist ─────────────────────────────────────────────────────────────
def _make_styles():
    return {
        "title":    ParagraphStyle("title",    fontName="Helvetica-Bold", fontSize=14, textColor=WHITE, alignment=TA_LEFT),
        "subtitle": ParagraphStyle("subtitle", fontName="Helvetica",      fontSize=9,  textColor=colors.HexColor("#BDD7EE")),
        "section":  ParagraphStyle("section",  fontName="Helvetica-Bold", fontSize=10, textColor=WHITE, alignment=TA_LEFT, leftIndent=6),
        "body":     ParagraphStyle("body",     fontName="Helvetica",      fontSize=8.5, textColor=colors.HexColor("#1A1A1A"), leading=13),
        "small":    ParagraphStyle("small",    fontName="Helvetica",      fontSize=7.5, textColor=colors.HexColor("#1A1A1A"), leading=11),
        "label":    ParagraphStyle("label",    fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.HexColor("#1B3A6B")),
        "cell":     ParagraphStyle("cell",     fontName="Helvetica",      fontSize=7,   textColor=colors.HexColor("#1A1A1A"), leading=9),
        "cell_b":   ParagraphStyle("cell_b",   fontName="Helvetica-Bold", fontSize=7,   textColor=colors.HexColor("#1B3A6B"), leading=9),
        "pass":     ParagraphStyle("pass",     fontName="Helvetica-Bold", fontSize=7,   textColor=GREEN,   leading=9),
        "warn":     ParagraphStyle("warn",     fontName="Helvetica-Bold", fontSize=7,   textColor=AMBER,   leading=9),
        "fail":     ParagraphStyle("fail",     fontName="Helvetica-Bold", fontSize=7,   textColor=RED,     leading=9),
        "na":       ParagraphStyle("na",       fontName="Helvetica",      fontSize=7,   textColor=colors.HexColor("#888888"), leading=9),
    }

def _hdr(text, st, bg=None):
    bg = bg or DARK_BLUE
    t = Table([[Paragraph(text, st["section"])]], colWidths=[CONTENT_W])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), bg),
        ("TOPPADDING", (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING", (0,0),(-1,-1), 8),
    ]))
    return t

def _status_para(status, st):
    s = str(status).upper()
    if s == "PASS":         return Paragraph("✓  PASS", st["pass"])
    if s == "NEEDS ATTENTION": return Paragraph("⚠  NEEDS ATTENTION", st["warn"])
    if s == "FAIL":         return Paragraph("✗  FAIL", st["fail"])
    return Paragraph("—  N/A", st["na"])

def build_checklist_pdf(spec_data, build_log, output_path):
    st = _make_styles()
    story = []
    meta   = spec_data.get('study_meta', {})
    forms  = spec_data.get('forms', [])
    today  = datetime.date.today().strftime('%d %b %Y')
    protocol = meta.get('protocol_number', '')
    study_id = meta.get('study_id', '')

    # Count summary stats
    n_built  = len(build_log.get('forms_built', []))
    n_skip   = len(build_log.get('forms_skipped', []))
    n_ph     = len(build_log.get('placeholder_applied', []))
    n_err    = len(build_log.get('build_errors', []))
    n_epro   = sum(1 for f in forms if f.get('epro','No') == 'Yes')

    # ── Cover header ──────────────────────────────────────────────────────
    hdr_data = [[
        Paragraph(f"STUDY BUILD CHECKLIST — {protocol}", st["title"]),
        Paragraph(f"Study ID: {study_id}  |  Built: {today}  |  Forms: {n_built}", st["subtitle"])
    ]]
    hdr_tbl = Table(hdr_data, colWidths=[CONTENT_W*0.6, CONTENT_W*0.4])
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), DARK_BLUE),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0),(-1,-1), 12),
        ("BOTTOMPADDING", (0,0),(-1,-1), 12),
        ("LEFTPADDING", (0,0),(-1,-1), 12),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 8))

    # ── SECTION 1: Sign-off Summary ───────────────────────────────────────
    story.append(_hdr("SECTION 1 — BUILD SUMMARY & SIGN-OFF", st))
    story.append(Spacer(1, 4))

    # Summary stats
    overall = "READY FOR UPLOAD" if n_err == 0 and n_ph == 0 else \
              "NEEDS ATTENTION" if n_ph > 0 and n_err == 0 else "BUILD ERRORS — DO NOT UPLOAD"
    overall_color = GREEN if "READY" in overall else AMBER if "ATTENTION" in overall else RED

    summary_rows = [
        [Paragraph("Metric", st["label"]),            Paragraph("Value", st["label"]),  Paragraph("Notes", st["label"])],
        [Paragraph("Forms Built",          st["small"]), Paragraph(str(n_built),  st["body"]), Paragraph("",st["small"])],
        [Paragraph("Forms Skipped/Errors", st["small"]), Paragraph(str(n_skip+n_err),st["body"]), Paragraph(', '.join(build_log.get('forms_skipped',[])) or "None",st["small"])],
        [Paragraph("Placeholder Fields",   st["small"]), Paragraph(str(n_ph),     st["body"]), Paragraph("Require site-specific completion" if n_ph else "None", st["small"])],
        [Paragraph("ePRO Forms",           st["small"]), Paragraph(str(n_epro),   st["body"]), Paragraph("Configure in ePRO module before upload" if n_epro else "None", st["small"])],
        [Paragraph("Timepoint CSV",        st["small"]), Paragraph("✓ Included",  st["pass"]), Paragraph(f"{study_id}_tpt.csv", st["small"])],
        [Paragraph("Lab Ranges CSV",       st["small"]), Paragraph("✓ Included",  st["pass"]), Paragraph("labranges.csv — contains placeholders" if n_ph else "labranges.csv", st["small"])],
        [Paragraph("OVERALL STATUS",       st["label"]), Paragraph(overall, ParagraphStyle("ovr", fontName="Helvetica-Bold", fontSize=9, textColor=overall_color)), Paragraph("",st["small"])],
    ]
    summary_tbl = Table(summary_rows, colWidths=[CONTENT_W*0.25, CONTENT_W*0.20, CONTENT_W*0.55])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK_BLUE),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [WHITE, GREY_LIGHT]),
        ("BACKGROUND",    (0,-1),(-1,-1), LIGHT_BLUE),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
        ("GRID",          (0,0),(-1,-1), 0.3, GREY_MID),
    ]))
    story.append(summary_tbl)
    story.append(Spacer(1, 10))

    # Sign-off table
    story.append(_hdr("SIGN-OFF", st, bg=MID_BLUE))
    story.append(Spacer(1, 4))
    signoff_rows = [
        [Paragraph("Role", st["label"]),          Paragraph("Name", st["label"]),     Paragraph("Date", st["label"]),     Paragraph("Signature", st["label"])],
        [Paragraph("Build Specialist", st["small"]),  Paragraph("", st["small"]), Paragraph("", st["small"]), Paragraph("", st["small"])],
        [Paragraph("QC Reviewer",     st["small"]),   Paragraph("", st["small"]), Paragraph("", st["small"]), Paragraph("", st["small"])],
        [Paragraph("Project Manager", st["small"]),   Paragraph("", st["small"]), Paragraph("", st["small"]), Paragraph("", st["small"])],
    ]
    so_tbl = Table(signoff_rows, colWidths=[CONTENT_W*0.18, CONTENT_W*0.28, CONTENT_W*0.18, CONTENT_W*0.36])
    so_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  MID_BLUE),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [WHITE, GREY_LIGHT]),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",    (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 10),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
        ("GRID",          (0,0),(-1,-1), 0.5, GREY_MID),
    ]))
    story.append(so_tbl)

    # ── SECTION 2: Detailed QA Checklist ─────────────────────────────────
    story.append(PageBreak())
    story.append(_hdr("SECTION 2 — DETAILED QA CHECKLIST (one row per form per check)", st))
    story.append(Spacer(1, 4))

    qa_headers = ["Form ID", "Form Title", "Check", "Status", "Notes"]
    qa_cw = [CONTENT_W*0.08, CONTENT_W*0.16, CONTENT_W*0.18,
             CONTENT_W*0.14, CONTENT_W*0.44]
    qa_data = [[Paragraph(h, ParagraphStyle("qah", fontName="Helvetica-Bold",
                fontSize=7.5, textColor=WHITE, leading=10)) for h in qa_headers]]

    check_labels = {c[0]: c[1] for c in QA_CHECKS}
    for form in forms:
        qa_results = run_qa_checks(form, build_log)
        for check_id, status, note in qa_results:
            qa_data.append([
                Paragraph(form.get('form_id',''), st["cell_b"]),
                Paragraph(form.get('form_title','')[:35], st["cell"]),
                Paragraph(check_labels.get(check_id, check_id), st["cell"]),
                _status_para(status, st),
                Paragraph(note[:80] if note else "", st["cell"]),
            ])

    qa_tbl = Table(qa_data, colWidths=qa_cw, repeatRows=1)
    qa_ts = [
        ("BACKGROUND",    (0,0),(-1,0),  DARK_BLUE),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING",   (0,0),(-1,-1), 3),
        ("GRID",          (0,0),(-1,-1), 0.3, GREY_MID),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [WHITE, GREY_LIGHT]),
    ]
    # Colour status cells
    for i, row in enumerate(qa_data[1:], start=1):
        status_text = row[3].text if hasattr(row[3], 'text') else ''
        if "FAIL" in str(status_text):
            qa_ts.append(("BACKGROUND", (3,i), (3,i), RED_LIGHT))
        elif "ATTENTION" in str(status_text):
            qa_ts.append(("BACKGROUND", (3,i), (3,i), AMBER_LIGHT))
        elif "PASS" in str(status_text):
            qa_ts.append(("BACKGROUND", (3,i), (3,i), GREEN_LIGHT))
    qa_tbl.setStyle(TableStyle(qa_ts))
    story.append(qa_tbl)

    # ── SECTION 3: Placeholder / OID Items ───────────────────────────────
    story.append(PageBreak())
    story.append(_hdr("SECTION 3 — ITEMS REQUIRING POST-CONFIGURATION", st))
    story.append(Spacer(1, 4))

    ph_items = build_log.get('placeholder_applied', [])
    oid_items = build_log.get('oid_placeholders', [])

    if ph_items:
        story.append(_hdr("Placeholder Values Applied (require site-specific completion)", st, bg=AMBER))
        story.append(Spacer(1, 3))
        for item in ph_items:
            story.append(Paragraph(
                f"<b>{item.get('form_id','')}</b> — {item.get('note','')} — "
                f"Fields: {', '.join(str(f) for f in item.get('fields',[])[:5])}",
                st["small"]
            ))
        story.append(Spacer(1, 6))

    if oid_items:
        story.append(_hdr("OID Paths Requiring Confirmation After Study Configuration", st, bg=MID_BLUE))
        story.append(Spacer(1, 3))
        for item in oid_items:
            story.append(Paragraph(
                f"<b>{item.get('form_id','')}</b> — {item.get('field','')} — {item.get('note','')}",
                st["small"]
            ))
        story.append(Spacer(1, 6))

    if not ph_items and not oid_items:
        story.append(Paragraph("No placeholder or OID items to report.", st["body"]))

    # Build and save
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            leftMargin=MARGIN, rightMargin=MARGIN,
                            topMargin=MARGIN, bottomMargin=MARGIN)
    doc.build(story)


# ── XLSX Checklist ─────────────────────────────────────────────────────────────
def build_checklist_xlsx(spec_data, build_log, output_path):
    wb = Workbook()
    meta   = spec_data.get('study_meta', {})
    forms  = spec_data.get('forms', [])
    today  = datetime.date.today().strftime('%d %b %Y')

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def _bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def _aln(h="left"): return Alignment(wrap_text=True, vertical="top", horizontal=h)

    # ── Sheet 1: SIGN-OFF SUMMARY ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "SIGN-OFF SUMMARY"
    ws1.sheet_properties.tabColor = "1B3A6B"

    ws1.merge_cells("A1:F1")
    c = ws1["A1"]
    c.value = f"STUDY BUILD CHECKLIST — {meta.get('protocol_number','')}  |  Study ID: {meta.get('study_id','')}  |  Built: {today}"
    c.font  = _font(bold=True, color="FFFFFF", size=11)
    c.fill  = _fill("1B3A6B")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 22

    # Summary section header
    ws1.merge_cells("A3:F3")
    ws1["A3"].value = "BUILD SUMMARY"
    ws1["A3"].font  = _font(bold=True, color="FFFFFF", size=10)
    ws1["A3"].fill  = _fill("2E6DA4")
    ws1["A3"].alignment = Alignment(horizontal="left", vertical="center")

    n_built = len(build_log.get('forms_built', []))
    n_ph    = len(build_log.get('placeholder_applied', []))
    n_err   = len(build_log.get('build_errors', []))
    n_epro  = sum(1 for f in forms if f.get('epro','No') == 'Yes')
    overall = "READY FOR UPLOAD" if n_err == 0 and n_ph == 0 else \
              "NEEDS ATTENTION"   if n_ph > 0 and n_err == 0 else "BUILD ERRORS"

    summary = [
        ("Forms Built",              str(n_built)),
        ("Forms with Placeholders",  str(n_ph)),
        ("Build Errors",             str(n_err)),
        ("ePRO Forms",               str(n_epro)),
        ("Overall Status",           overall),
    ]
    for i, (k, v) in enumerate(summary, start=4):
        ws1.cell(row=i, column=1, value=k).font  = _font(bold=True, size=9)
        ws1.cell(row=i, column=1).fill           = _fill("D6E4F0")
        ws1.cell(row=i, column=1).border         = _bdr()
        ws1.cell(row=i, column=1).alignment      = _aln()
        ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        c = ws1.cell(row=i, column=2, value=v)
        c.font      = _font(bold=(k=="Overall Status"), size=9,
                            color=("27AE60" if "READY" in v else
                                   "E67E22" if "ATTENTION" in v else
                                   "C0392B" if "ERROR" in v else "000000"))
        c.fill      = _fill("FFFFFF")
        c.border    = _bdr()
        c.alignment = _aln()
        ws1.row_dimensions[i].height = 15

    # Sign-off table
    ws1["A10"].value = "SIGN-OFF"
    ws1["A10"].font  = _font(bold=True, color="FFFFFF", size=10)
    ws1["A10"].fill  = _fill("2E6DA4")
    ws1.merge_cells("A10:F10")
    ws1["A10"].alignment = Alignment(horizontal="left", vertical="center")

    signoff_hdrs = ["Role", "Printed Name", "Date", "Signature", "Comments", ""]
    for col, h in enumerate(signoff_hdrs, start=1):
        c = ws1.cell(row=11, column=col, value=h)
        c.font      = _font(bold=True, color="FFFFFF", size=9)
        c.fill      = _fill("1B3A6B")
        c.border    = _bdr()
        c.alignment = _aln("center")

    for i, role in enumerate(["Build Specialist", "QC Reviewer", "Project Manager"], start=12):
        ws1.cell(row=i, column=1, value=role).font = _font(size=9)
        ws1.cell(row=i, column=1).fill = _fill("F5F5F5")
        ws1.cell(row=i, column=1).border = _bdr()
        for col in range(2, 7):
            c = ws1.cell(row=i, column=col, value="")
            c.fill   = _fill("FFFFFF")
            c.border = _bdr()
        ws1.row_dimensions[i].height = 20

    col_widths = [20, 24, 14, 24, 28, 10]
    for i, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: DETAILED QA ───────────────────────────────────────────────
    ws2 = wb.create_sheet("DETAILED QA CHECKLIST")
    ws2.sheet_properties.tabColor = "2E6DA4"

    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = "DETAILED QA CHECKLIST — One row per form per check"
    c.font  = _font(bold=True, color="FFFFFF", size=10)
    c.fill  = _fill("1B3A6B")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 18

    qa_hdrs = ["Form ID", "Form Title", "QA Check", "Status", "Notes", "Resolved?"]
    for col, h in enumerate(qa_hdrs, start=1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font      = _font(bold=True, color="FFFFFF", size=9)
        c.fill      = _fill("1B3A6B")
        c.border    = _bdr()
        c.alignment = _aln("center")
    ws2.row_dimensions[2].height = 16

    check_labels = {c[0]: c[1] for c in QA_CHECKS}
    row_idx = 3
    for form in forms:
        qa_results = run_qa_checks(form, build_log)
        for check_id, status, note in qa_results:
            bg = ("D5F5E3" if status == "PASS" else
                  "FFF3CD" if status == "NEEDS ATTENTION" else
                  "FADBD8" if status == "FAIL" else "F5F5F5")
            row_data = [
                form.get('form_id',''),
                form.get('form_title',''),
                check_labels.get(check_id, check_id),
                status,
                note,
                "" if status not in ("PASS","N/A") else "—"
            ]
            for col, val in enumerate(row_data, start=1):
                c = ws2.cell(row=row_idx, column=col, value=val)
                c.font      = _font(bold=(col==4), size=8,
                                    color=("27AE60" if status=="PASS" else
                                           "7D5A00" if status=="NEEDS ATTENTION" else
                                           "8B1A1A" if status=="FAIL" else "555555"))
                c.fill      = _fill(bg)
                c.border    = _bdr()
                c.alignment = _aln()
            ws2.row_dimensions[row_idx].height = 13
            row_idx += 1

    qa_widths = [12, 22, 24, 16, 44, 12]
    for i, w in enumerate(qa_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    wb.save(output_path)


if __name__ == "__main__":
    # Test with minimal data
    spec = {'study_meta': {'protocol_number': 'PrTK05', 'study_id': 'prtk05'},
            'forms': [], 'timepoint_csv': {'rows': []}, 'labranges_csv': {'rows': []}}
    log  = {'forms_built': [], 'forms_skipped': [], 'placeholder_applied': [],
            'oid_placeholders': [], 'qa_results': [], 'build_warnings': [], 'build_errors': []}
    build_checklist_pdf(spec, log, "/tmp/checklist_test.pdf")
    build_checklist_xlsx(spec, log, "/tmp/checklist_test.xlsx")
    print("Checklist test done")

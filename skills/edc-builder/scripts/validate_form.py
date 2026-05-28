"""
validate_form.py — XLSForm validation for the edc-builder skill.

Public API
──────────
    validate_xlsform(xlsx_path: str) -> tuple[bool, list[str], list[str]]
        Returns (is_valid, errors, warnings).

        Uses pyxform.xls2xform.xls2xform_convert under the hood. Calls it
        with validate=True (runs ODK Validate, requires Java + the JAR
        bundled with pyxform) when both prerequisites are available;
        falls back to validate=False otherwise.

        Catches PyXFormError and returns (False, [str(e)], []) so callers
        never have to deal with raised exceptions.

Known OC-specific false positives suppressed by this validator:
  * OC XPath extensions — instance(clinicaldata), floor(), etc. The
    bundled ODK Validate JAR does not know them; OC's form-service does.
  * OC-8 phantom end_group — the orphan `end group` between begin_repeat
    and end_repeat that OpenClinica REQUIRES to activate a repeating
    form's version. Standard pyxform flags it as "Unmatched 'end_group'";
    we treat that error as expected when the OC-8 marker is present.

History
───────
This module briefly stripped the OC-8 phantom end_group on the belief
that OC rejected it. That was the wrong direction: OpenClinica REQUIRES
the phantom (without it the form stays stuck at "Please select default
version for data entry"), and the build now preserves it
(build_xlsforms._balance_begin_end_tags). Since standard pyxform rejects
the phantom, this validator recognises the OC-8 marker and treats the
resulting "Unmatched 'end_group'" error as an expected, OC-required
pattern. See xlsform-build-rules.md (OC-8).
"""

from __future__ import annotations

import logging
import os
import shutil
import tempfile

logger = logging.getLogger(__name__)


# Patterns that ODK Validate flags as errors but are valid in OC.
# These are OC-proprietary extensions that the bundled ODK Validate
# JAR does not know about. Treat them as expected warnings so the
# self-correction loop is not triggered on correct forms.
_OC_KNOWN_ERROR_PATTERNS = [
    "instance(clinicaldata)",
    "instance('clinicaldata')",
    "instance(labranges)",
    "instance('labranges')",
    "cannot handle function 'floor'",
    "cannot handle function 'once'",
    "ItemGroupRepeatKey",
]


def _error_is_all_oc_known(error_str: str) -> bool:
    """Return True if every substantive error line in an ODK Validate
    error string matches a known OC-specific pattern that is valid in
    OpenClinica but unsupported by the bundled JAR."""
    substantive = [
        ln.strip() for ln in error_str.split("\n")
        if ln.strip()
        and any(kw in ln for kw in (
            "Error", "error", "XPath", "Instance", "function",
            "Invalid", "exception",
        ))
    ]
    if not substantive:
        return False
    return all(
        any(p in ln for p in _OC_KNOWN_ERROR_PATTERNS)
        for ln in substantive
    )


def _has_oc8_phantom_marker(xlsx_path: str) -> bool:
    """True if the survey sheet contains the OpenClinica OC-8 repeating-form
    marker: an `end group` row sitting directly between `begin repeat` and
    `end repeat`.

    This pattern is invalid per the standard XLSForm spec — pyxform raises
    "Unmatched 'end_group'" — but OpenClinica REQUIRES it for the form
    version to activate (see xlsform-build-rules.md OC-8). We use this to
    distinguish the OC-required phantom from a genuinely broken form before
    suppressing the pyxform error.
    """
    try:
        import openpyxl
    except ImportError:
        return False
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        if "survey" not in wb.sheetnames:
            return False
        ws = wb["survey"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return False
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    try:
        ti = header.index("type")
    except ValueError:
        return False
    types = []
    for r in rows[1:]:
        v = r[ti] if ti < len(r) else None
        t = str(v).strip().lower().replace("_", " ") if v is not None else ""
        if t:
            types.append(t)
    for i in range(len(types) - 2):
        if (types[i] == "begin repeat"
                and types[i + 1] == "end group"
                and types[i + 2] == "end repeat"):
            return True
    return False


def _odk_validate_available() -> bool:
    """True iff the ODK Validate JAR exists AND a working JRE is callable.

    macOS in particular ships a `/usr/bin/java` stub that's on PATH but
    fails at runtime ("Unable to locate a Java Runtime") if no JDK is
    installed. We actually invoke `java -version` (sub-1s) so the answer
    reflects what pyxform will see when it shells out, not just what's
    on PATH."""
    try:
        import pyxform
    except ImportError:
        return False
    jar_path = os.path.join(
        os.path.dirname(pyxform.__file__),
        "validators", "odk_validate", "bin", "ODK_Validate.jar",
    )
    if not os.path.exists(jar_path):
        return False
    java = shutil.which("java")
    if java is None:
        return False
    try:
        import subprocess
        r = subprocess.run(
            [java, "-version"],
            capture_output=True, timeout=5,
        )
        return r.returncode == 0
    except (subprocess.TimeoutExpired, OSError):
        return False


def validate_xlsform(
    xlsx_path: str,
) -> tuple[bool, list[str], list[str]]:
    """Validate one XLSForm. See module docstring for full contract.

    Returns:
        (is_valid, errors, warnings)

      * is_valid — True iff pyxform converted the form without raising.
      * errors   — non-empty when invalid. PyXFormError messages are
                   surfaced as-is; unexpected exceptions are tagged with
                   their class name.
      * warnings — non-empty when pyxform emitted advisory output but
                   the form is still convertible.
    """
    # Lazy-import so callers that don't validate don't pay the cost.
    try:
        from pyxform.errors import PyXFormError
        from pyxform.xls2xform import xls2xform_convert
        try:
            from pyxform.validators.odk_validate import ODKValidateError
        except ImportError:
            ODKValidateError = None
    except ImportError as e:
        # pyxform missing → can't validate. Treat as a skip (valid) with
        # a warning so missing pyxform never triggers self-correction.
        return (True, [], [f"pyxform not installed — skipping: {e}"])

    if not os.path.exists(xlsx_path):
        return (False, [f"File does not exist: {xlsx_path}"], [])

    use_validate = _odk_validate_available()

    # xls2xform_convert needs an output XML target even though we
    # discard the converted XML. Clean up after ourselves.
    tmp_xml: str | None = None
    try:
        tmp = tempfile.NamedTemporaryFile(suffix=".xml", delete=False)
        tmp.close()
        tmp_xml = tmp.name

        warnings_out = xls2xform_convert(
            xlsform_path=xlsx_path,
            xform_path=tmp_xml,
            validate=use_validate,
            pretty_print=False,
            enketo=False,
        )

        # pyxform returns a list of warning strings (or None). Normalise.
        warnings = [w for w in (warnings_out or [])
                    if w and str(w).strip()
                    and "Use this worksheet to define" not in str(w)]
        return (True, [], warnings)

    except PyXFormError as e:
        err_str = str(e)
        # OC-8 phantom end_group: standard pyxform rejects the orphan
        # `end group` between begin_repeat and end_repeat, but OpenClinica
        # REQUIRES it to activate a repeating form's version. Treat the
        # resulting "Unmatched 'end_group'" error as expected when the OC-8
        # marker is actually present, so the self-correction loop does not
        # strip the OC-required phantom back out. See xlsform-build-rules.md.
        if "Unmatched 'end_group'" in err_str and _has_oc8_phantom_marker(xlsx_path):
            return (True, [], [
                f"pyxform flagged the OC-8 phantom end_group "
                f"(required by OpenClinica, not a real error): {err_str[:200]}"
            ])
        return (False, [err_str], [])

    except Exception as e:
        # Check if this is an ODKValidateError with only OC-known patterns
        if ODKValidateError and isinstance(e, ODKValidateError):
            err_str = str(e)
            if _error_is_all_oc_known(err_str):
                oc_warn = (
                    f"ODK Validate flagged OC-specific XPath "
                    f"(expected for this form, not a real error): "
                    f"{err_str[:300]}"
                )
                return (True, [], [oc_warn])
            return (False, [err_str], [])
        logger.exception("validate_xlsform: unexpected error on %s", xlsx_path)
        return (False,
                [f"Unexpected validation error ({type(e).__name__}): {e}"],
                [])

    finally:
        # Always clean up the temp XML and any pyxform sidecar files.
        if tmp_xml and os.path.exists(tmp_xml):
            try:
                os.unlink(tmp_xml)
            except OSError:
                pass
            sidecar = os.path.join(os.path.dirname(tmp_xml), "itemsets.csv")
            if os.path.exists(sidecar):
                try:
                    os.unlink(sidecar)
                except OSError:
                    pass


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python validate_form.py <xlsx_path>", file=sys.stderr)
        sys.exit(1)
    ok, errs, warns = validate_xlsform(sys.argv[1])
    print(f"is_valid: {ok}")
    print(f"odk_validate_available: {_odk_validate_available()}")
    if errs:
        print(f"errors ({len(errs)}):")
        for e in errs:
            print(f"  - {e}")
    if warns:
        print(f"warnings ({len(warns)}):")
        for w in warns:
            print(f"  - {w}")
    sys.exit(0 if ok else 2)

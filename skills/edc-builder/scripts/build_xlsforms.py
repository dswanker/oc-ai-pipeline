"""
build_xlsforms.py — OpenClinica XLSForm Generator
Generates one production-ready .xlsx file per CRF form from the EDC
structure specification.

Each output file is built from form_template.xlsx (must live alongside
this script in the same directory).  The template supplies:
  • Correctly structured settings / choices / survey sheets with OC headers
  • Six reference/instruction tabs (Cross-Form Examples, Custom Annotation
    Examples, Contact Info Examples, eConsent Examples, Hard Checks
    Examples, Offline Forms Examples)
  • bind::oc:external dropdown in the survey sheet

If form_template.xlsx is not found, the script falls back to building
the three functional sheets from scratch (no reference tabs, no dropdowns).

Usage:
    from build_xlsforms import build_all_xlsforms, build_single_xlsform
    results = build_all_xlsforms(spec_data, output_dir)
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os, re, datetime

# Path to the OC form template (must be in the same folder as this script)
_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_SCRIPT_DIR, 'form_template.xlsx')

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1B3A6B"
MID_BLUE   = "2E6DA4"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
GREY_LIGHT = "F5F5F5"
GREY_MID   = "CCCCCC"
AMBER      = "FFF3CD"

# ── Exact column order from OpenClinica template ──────────────────────────────
SURVEY_COLS = [
    "type", "name", "label", "bind::oc:itemgroup", "hint", "appearance",
    "bind::oc:briefdescription", "bind::oc:description", "relevant",
    "required", "required_message", "constraint", "constraint_message",
    "default", "calculation", "trigger", "readonly", "image",
    "repeat_count", "bind::oc:external"
]

CHOICES_COLS = ["list_name", "label", "name", "image"]

SETTINGS_COLS = [
    "form_title", "form_id", "version", "style",
    "crossform_references", "namespaces"
]

# Column widths for survey sheet
SURVEY_WIDTHS = {
    "type": 18, "name": 20, "label": 36, "bind::oc:itemgroup": 14,
    "hint": 24, "appearance": 18, "bind::oc:briefdescription": 20,
    "bind::oc:description": 22, "relevant": 32, "required": 10,
    "required_message": 26, "constraint": 32, "constraint_message": 28,
    "default": 18, "calculation": 32, "trigger": 14, "readonly": 9,
    "image": 12, "repeat_count": 12, "bind::oc:external": 16,
    "choice_filter": 26, "bind::oc:constraint-type": 20,
    "bind::oc:required-type": 18,
}

CHOICES_WIDTHS = {"list_name": 18, "label": 32, "name": 24, "image": 10}
SETTINGS_WIDTHS = {
    "form_title": 28, "form_id": 16, "version": 8, "style": 14,
    "crossform_references": 24, "namespaces": 55
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=9, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color=GREY_MID)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def _header_cell(cell, value):
    cell.value = value
    cell.font = _font(bold=True, color=WHITE, size=9)
    cell.fill = _fill(DARK_BLUE)
    cell.border = _border()
    cell.alignment = _align(h="center", v="center")

def _coerce_cell_value(value):
    """Coerce a value to something openpyxl can write to a cell.
    Lists/tuples become pipe-delimited strings; dicts become JSON;
    anything else is left as-is."""
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return " | ".join(str(v) for v in value)
    if isinstance(value, dict):
        try:
            import json as _json
            return _json.dumps(value)
        except Exception:
            return str(value)
    return value


def _data_cell(cell, value, row_idx=0, flagged=False):
    cell.value = _coerce_cell_value(value)
    cell.font = _font(size=8)
    bg = AMBER if flagged else (GREY_LIGHT if row_idx % 2 == 0 else WHITE)
    cell.fill = _fill(bg)
    cell.border = _border()
    cell.alignment = _align()


# ── XLSX column name → XLSForm column name mapping ────────────────────────────
XLSX_TO_XLSFORM = {
    "bind__oc_itemgroup":        "bind::oc:itemgroup",
    "bind__oc_external":         "bind::oc:external",
    "bind__oc_briefdescription": "bind::oc:briefdescription",
    "bind__oc_description":      "bind::oc:description",
    "bind__oc_constraint_type":  "bind::oc:constraint-type",
    "bind__oc_required_type":    "bind::oc:required-type",
}
# Inverse: given the XLSForm column name, find the JSON key that carries
# its value. Used by _resolve_cell_value when the XLSForm name isn't
# present as a direct key in the row dict.
XLSFORM_TO_JSON = {v: k for k, v in XLSX_TO_XLSFORM.items()}


def _resolve_cell_value(row, col):
    """
    Given a row dict and an XLSForm column name, return the value.
    Tries both the XLSForm column name (e.g. 'bind::oc:itemgroup') AND its
    JSON underscore equivalent (e.g. 'bind__oc_itemgroup'), because
    Claude's extraction emits underscore keys (colons aren't friendly in
    JSON) while the XLSForm spec mandates colon keys.
    """
    # Try exact XLSForm name first
    if col in row and row[col] not in (None, ''):
        return row[col]
    # Fall back to JSON underscore variant
    json_key = XLSFORM_TO_JSON.get(col)
    if json_key and json_key in row:
        return row[json_key]
    return ''

# Columns from spec XLSX to strip (never include in output XLSForms)
STRIP_COLS = {
    "ACTION", "REVIEW_NOTES", "completion_status", "library_source",
    "pdf_original_label", "cdash_standard_name", "cdash_name_deviation",
    "cdash_name_confidence", "flag_reason", "choice_filter_meta",
    "source", "filter_column", "filter_value"
}

# ── Read spec XLSX ─────────────────────────────────────────────────────────────
def read_spec_xlsx(spec_path):
    """
    Read the EDC structure specification XLSX and return a dict of form data.
    Returns: {
        'study_meta': {...},
        'timepoint_csv': {'filename': ..., 'rows': [...]},
        'labranges_csv': {'filename': ..., 'columns': [...], 'rows': [...]},
        'forms': [{settings, choices, survey, form_id, form_title, ...}, ...]
    }
    """
    import openpyxl
    wb = openpyxl.load_workbook(spec_path, data_only=True)
    result = {'study_meta': {}, 'timepoint_csv': {}, 'labranges_csv': {}, 'forms': []}

    # ── INDEX sheet ────────────────────────────────────────────────────────
    if 'INDEX' in wb.sheetnames:
        ws = wb['INDEX']
        meta_map = {}
        for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
            if row[0] and row[1]:
                key = str(row[0]).strip()
                val = str(row[1]).strip() if row[1] else ''
                meta_map[key] = val
        result['study_meta'] = {
            'protocol_number': meta_map.get('Protocol Number', ''),
            'study_id':        meta_map.get('Study ID', ''),
            'input_mode':      meta_map.get('Input Mode', ''),
        }
        # Read form inventory from INDEX
        form_list = []
        reading_forms = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0] == '#':
                reading_forms = True
                continue
            if reading_forms and row and row[0] and str(row[0]).strip().isdigit():
                form_list.append({
                    'num':        str(row[0]).strip(),
                    'form_id':    str(row[1]).strip() if row[1] else '',
                    'form_title': str(row[2]).strip() if row[2] else '',
                    'category':   str(row[3]).strip() if row[3] else '',
                    'cdash':      str(row[4]).strip() if row[4] else '',
                    'arm':        str(row[5]).strip() if row[5] else '',
                    'complexity': str(row[6]).strip() if row[6] else '',
                    'repeating':  str(row[7]).strip() if row[7] else 'No',
                    'epro':       str(row[8]).strip() if row[8] else 'No',
                    'reuse':      str(row[9]).strip() if row[9] else '0',
                    'survey_tab': str(row[10]).strip() if row[10] else '',
                })
        result['_form_list'] = form_list

    # ── TIMEPOINTS sheet ───────────────────────────────────────────────────
    if 'TIMEPOINTS' in wb.sheetnames:
        ws = wb['TIMEPOINTS']
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
            if row[0] and row[1]:
                action = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ''
                if action != 'DELETE':
                    rows.append({'event': str(row[0]).strip(), 'timepoint': str(row[1]).strip()})
        study_id = result['study_meta'].get('study_id', 'study')
        result['timepoint_csv'] = {'filename': f"{study_id}_tpt.csv", 'rows': rows}

    # ── LAB_RANGES sheet ───────────────────────────────────────────────────
    if 'LAB_RANGES' in wb.sheetnames:
        ws = wb['LAB_RANGES']
        headers = []
        data_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i == 0:
                headers = [str(c).strip() for c in row if c is not None]
            else:
                if any(c is not None for c in row[:len(headers)]):
                    row_dict = {}
                    for j, h in enumerate(headers):
                        val = row[j] if j < len(row) else None
                        row_dict[h] = str(val).strip() if val is not None else ''
                    data_rows.append(row_dict)
        result['labranges_csv'] = {
            'filename': 'labranges.csv',
            'columns': headers,
            'rows': data_rows
        }

    # ── Per-form tabs ──────────────────────────────────────────────────────
    form_list = result.get('_form_list', [])
    sheet_names = wb.sheetnames

    for form_meta in form_list:
        form_id = form_meta['form_id']
        survey_tab = form_meta.get('survey_tab', f"{form_id}_survey")
        choices_tab = f"{form_id}_choices"
        settings_tab = f"{form_id}_settings"

        # Try to find tabs (tab names are truncated to 31 chars)
        s_tab = next((s for s in sheet_names if s == survey_tab or
                      s == survey_tab[:31]), None)
        c_tab = next((s for s in sheet_names if s == choices_tab or
                      s == choices_tab[:31]), None)
        st_tab = next((s for s in sheet_names if s == settings_tab or
                       s == settings_tab[:31]), None)

        if not s_tab:
            continue  # form tabs not found, skip

        # Read survey
        survey_rows = []
        extra_cols = []
        ws = wb[s_tab]
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=3, values_only=True)):
            header_row = [str(c).strip() if c else '' for c in row]
        if header_row:
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not any(c is not None for c in row):
                    continue
                action = str(row[0]).strip().upper() if row[0] else ''
                if action == 'DELETE':
                    continue
                row_dict = {}
                for j, h in enumerate(header_row):
                    if h in STRIP_COLS or not h:
                        continue
                    val = row[j] if j < len(row) else None
                    # Remap XLSX col names to XLSForm col names
                    xlsform_key = XLSX_TO_XLSFORM.get(h, h)
                    row_dict[xlsform_key] = str(val).strip() if val is not None else ''
                survey_rows.append(row_dict)
                # Track any extra columns (choice_filter, hard checks etc.)
                for k in row_dict:
                    if k not in SURVEY_COLS and k not in STRIP_COLS and k not in extra_cols:
                        extra_cols.append(k)

        # Read choices
        choices_rows = []
        if c_tab:
            ws = wb[c_tab]
            c_header = None
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                c_header = [str(c).strip() if c else '' for c in row]
            if c_header:
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not any(c is not None for c in row):
                        continue
                    action = str(row[0]).strip().upper() if row[0] else ''
                    if action == 'DELETE':
                        continue
                    row_dict = {}
                    for j, h in enumerate(c_header):
                        if h in STRIP_COLS or not h:
                            continue
                        val = row[j] if j < len(row) else None
                        row_dict[h] = str(val).strip() if val is not None else ''
                    if row_dict.get('list_name') or row_dict.get('name'):
                        choices_rows.append(row_dict)

        # Read settings
        settings_dict = {
            'form_title': form_meta.get('form_title', ''),
            'form_id': form_id,
            'version': '1',
            'style': 'theme-grid',
            'crossform_references': '',
            'namespaces': 'oc="http://openclinica.org/xforms" , OpenClinica="http://openclinica.com/odm"'
        }
        if st_tab:
            ws = wb[st_tab]
            setting_keys = ['form_title', 'form_id', 'version', 'style',
                            'crossform_references', 'namespaces']
            for i, row in enumerate(ws.iter_rows(min_row=3, max_row=8, values_only=True)):
                if i < len(setting_keys) and row[1] is not None:
                    val = str(row[1]).strip()
                    if val:
                        settings_dict[setting_keys[i]] = val

        result['forms'].append({
            'form_id':    form_id,
            'form_title': form_meta.get('form_title', ''),
            'category':   form_meta.get('category', ''),
            'arm':        form_meta.get('arm', 'BOTH'),
            'complexity': form_meta.get('complexity', ''),
            'repeating':  form_meta.get('repeating', 'No'),
            'epro':       form_meta.get('epro', 'No'),
            'reuse':      form_meta.get('reuse', '0'),
            'settings':   settings_dict,
            'choices':    choices_rows,
            'survey':     survey_rows,
            'extra_cols': extra_cols,
            'source_tab': survey_tab,
        })

    return result


# ── Survey row normalization ──────────────────────────────────────────────────
def _normalize_survey_rows(rows):
    """
    Normalize survey rows before writing to XLSForm.

    XLSForm spec (and pyxform validation) requires that end group and
    end repeat rows have a BLANK name. Claude sometimes generates these
    rows with names (e.g. 'AE_REPEAT_END') which causes pyxform to look
    for a matching begin_group by name and fail.

    NOTE: begin/end TYPE pairing (begin_repeat closes with end_repeat,
    begin_group closes with end_group) is enforced separately by
    _balance_begin_end_tags, which runs before this normalizer. Earlier
    comments here treated a phantom `end_group` inside a repeat as
    intentional ("OC-8") — that was wrong; OC's form-service silently
    rejects forms with that pattern. See xlsform-build-rules.md for the
    full diagnosis.
    """
    END_TYPES = {'end group', 'end repeat', 'end_group', 'end_repeat'}
    normalized = []
    for row in rows:
        r = dict(row)
        t = str(r.get('type', '') or '').strip().lower()
        if t in END_TYPES:
            r['name'] = ''
        normalized.append(r)

    _BEGIN_TYPES = {
        'begin group', 'begin repeat',
        'begin_group', 'begin_repeat',
    }
    for r in normalized:
        t_norm = str(r.get('type', '') or '').strip().lower()
        if t_norm in _BEGIN_TYPES:
            if not str(r.get('label', '') or '').strip():
                r['label'] = (
                    str(r.get('name', '') or '').strip() or 'Repeat'
                )

    return normalized


# ── Begin/end tag pairing balancer ────────────────────────────────────────────
def _balance_begin_end_tags(rows, form_id, build_log=None):
    """Walk the survey rows tracking a begin/end stack. Correct any end
    tag whose type doesn't match its opener (e.g. begin_repeat closed by
    end_group), drop orphan end rows whose stack is empty, and append
    closer rows for any unclosed begins at the form tail.

    Why this is a hard correctness requirement (not a defensive nicety):
    OC's form-service silently rejects forms with mismatched begin/end
    tags — uploads return HTTP 200 but no version object ever appears
    in minimongo. The symptom looks like propagation lag but the form
    was actually rejected at parse time. See xlsform-build-rules.md.

    Correction policy:
      * begin_repeat / end_group adjacency → rewrite end row's type to
        'end repeat' (opener wins).
      * begin_group / end_repeat adjacency → rewrite to 'end group'.
      * end row with empty stack → drop (orphan, can't pair).
      * begin row with no matching end before form end → auto-append
        the matching end row (minimal — only `type`, blank `name`).

    Returns the corrected row list. Logs every correction made into
    build_log['tag_balance_corrections'] when build_log is provided.

    Raises ValueError only if the input is genuinely unsalvageable
    (currently no such case is reachable — the open/close model is
    single-tier within one form's survey sheet).
    """
    def _classify(t: str):
        u = (t or '').strip().lower().replace('_', ' ')
        if u == 'begin group':  return ('begin', 'group')
        if u == 'begin repeat': return ('begin', 'repeat')
        if u == 'end group':    return ('end',   'group')
        if u == 'end repeat':   return ('end',   'repeat')
        return (None, None)

    stack: list[str] = []
    corrections: list[str] = []
    out: list[dict] = []

    for idx, row in enumerate(rows):
        t = str(row.get('type', '') or '')
        action, kind = _classify(t)

        if action == 'begin':
            stack.append(kind)
            out.append(dict(row))

        elif action == 'end':
            if not stack:
                corrections.append(
                    f"row {idx + 2}: dropped orphan end ({t!r}) — "
                    f"no matching begin on the stack"
                )
                continue
            opener_kind = stack.pop()
            corrected = 'end repeat' if opener_kind == 'repeat' else 'end group'
            new_row = dict(row)
            if kind != opener_kind:
                corrections.append(
                    f"row {idx + 2}: rewrote {t!r} → {corrected!r} "
                    f"(opener was begin_{opener_kind})"
                )
                new_row['type'] = corrected
            out.append(new_row)

        else:
            out.append(dict(row))

    # Auto-close any unclosed begins at the tail.
    while stack:
        opener_kind = stack.pop()
        closer = 'end repeat' if opener_kind == 'repeat' else 'end group'
        out.append({'type': closer, 'name': ''})
        corrections.append(
            f"tail: appended {closer!r} to close unclosed begin_{opener_kind}"
        )

    if corrections:
        print(f"[edc-builder] balanced begin/end tags for {form_id}:",
              flush=True)
        for c in corrections:
            print(f"  {c}", flush=True)
        if build_log is not None:
            build_log.setdefault('tag_balance_corrections', []).append({
                'form_id': form_id,
                'corrections': corrections,
            })

    return out


# ── Choice-list back-fill (CRS-136 fix) ───────────────────────────────────────
def _ensure_referenced_choice_lists(survey, choices):
    """
    Guarantee every list_name referenced by a survey row's `type`
    column also appears in the choices sheet for this form.

    XLSForm requires each form to be self-contained — pyxform validates
    each .xlsx independently against its own choices sheet. Upstream
    Claude sometimes drops a boilerplate list (notably `yn`) when a
    form has exactly one yn reference plus its own study-specific
    list (CRS-136 root cause: DV and PE shipped without yn defined).

    Behaviour:
      - For each survey row whose type is "select_one X" or
        "select_multiple X", record X as referenced.
      - Compare against list_name values present in `choices`.
      - For any referenced list missing from choices:
          * If the list is `yn` → inject canonical Y/N rows.
          * Otherwise → return it as a hard error to the caller.

    Returns a tuple: (augmented_choices, missing_non_yn_lists).
    The caller raises on a non-empty missing_non_yn_lists.

    Doctest (acts as inline unit test, runnable with python -m doctest):
    >>> survey = [{'type': 'select_one yn', 'name': 'DVYN'}]
    >>> choices = [{'list_name': 'dvcat', 'name': 'C1', 'label': 'Cat 1'}]
    >>> aug, missing = _ensure_referenced_choice_lists(survey, choices)
    >>> sorted({c['list_name'] for c in aug})
    ['dvcat', 'yn']
    >>> missing
    []
    >>> survey = [{'type': 'select_one rel', 'name': 'AEREL'}]
    >>> aug, missing = _ensure_referenced_choice_lists(survey, [])
    >>> missing
    ['rel']
    """
    SELECT_PREFIXES = ('select_one ', 'select_multiple ')
    YN_ROWS = [
        {'list_name': 'yn', 'name': 'Y', 'label': 'Yes', 'image': ''},
        {'list_name': 'yn', 'name': 'N', 'label': 'No',  'image': ''},
    ]

    referenced = []
    seen_ref = set()
    for row in survey:
        t = str(row.get('type', '') or '').strip()
        for prefix in SELECT_PREFIXES:
            if t.startswith(prefix):
                list_name = t[len(prefix):].strip().split()[0] if t[len(prefix):].strip() else ''
                if list_name and list_name not in seen_ref:
                    seen_ref.add(list_name)
                    referenced.append(list_name)
                break

    defined = {str(c.get('list_name', '') or '').strip() for c in choices}
    defined.discard('')

    augmented = list(choices)
    missing_non_yn = []
    for list_name in referenced:
        if list_name in defined:
            continue
        if list_name == 'yn':
            augmented.extend(YN_ROWS)
        else:
            missing_non_yn.append(list_name)

    return augmented, missing_non_yn


# ── Build a single XLSForm .xlsx ───────────────────────────────────────────────
def build_single_xlsform(form_data, output_path, build_log):
    """Build one XLSForm .xlsx file from form_data dict.

    When form_template.xlsx is present alongside this script, the output is
    based on that template so the human receives:
      - The six OC reference/instruction tabs (Cross-Form Examples, etc.)
      - The bind::oc:external dropdown in the survey sheet
      - The official OC header row styling
    If the template is missing, falls back to building the three functional
    sheets from scratch (original behaviour).
    """
    settings = form_data.get('settings', {}) or {}
    choices  = form_data.get('choices', [])
    survey   = form_data.get('survey', [])
    extra_c  = form_data.get('extra_cols', [])

    # CRS-136 fix: back-fill `yn` and detect any other missing list_names
    # BEFORE the choices/survey loops run, so the rest of the function
    # operates on consistent data.
    choices, missing_lists = _ensure_referenced_choice_lists(survey, choices)
    if missing_lists:
        form_id_for_err = form_data.get('settings', {}).get('form_id') or output_path
        raise ValueError(
            f"Form {form_id_for_err}: survey references choice list(s) "
            f"{missing_lists!r} but they are not defined in the choices "
            f"sheet. yn is auto-templated; all other lists must be supplied."
        )

    # The rest of the function reads `choices` and `survey` only via the
    # local variables above, so the back-fill applies to both template
    # and scratch paths.
    form_id  = form_data.get('form_id', 'FORM')

    # Apply OpenClinica XLSForm defaults for missing settings fields.
    # OpenClinica's XLSForm validator requires these — without them the
    # form won't load. Claude may omit them in the JSON extraction.
    OC_SETTINGS_DEFAULTS = {
        'form_title':           form_data.get('form_title', form_id),
        'form_id':               form_id,
        'version':               '1',
        'style':                 'theme-grid',
        'crossform_references': '',
        'namespaces':
            'oc="http://openclinica.org/xforms" , '
            'OpenClinica="http://openclinica.com/odm"',
    }
    for k, default in OC_SETTINGS_DEFAULTS.items():
        if not settings.get(k):
            settings[k] = default

    def _ascii_safe_title(title: str) -> str:
        """Replace known non-ASCII characters in XLSForm settings
        fields with safe ASCII equivalents.

        Scope: applied only to form_title in the settings sheet.
        Survey labels are left as-is — they are display text and
        OC renders them correctly. The issue is specifically with
        non-ASCII in the settings form_title field.

        Confirmed by engineering feedback for OpenClinica 4.
        """
        _REPLACEMENTS = {
            '°': 'deg',   # ° DEGREE SIGN → deg
            '®': '(R)',   # ® REGISTERED SIGN
            '©': '(C)',   # © COPYRIGHT SIGN
            '™': '(TM)',  # ™ TRADE MARK SIGN
            '’': "'",     # ' RIGHT SINGLE QUOTATION MARK
            '‘': "'",     # ' LEFT SINGLE QUOTATION MARK
            '“': '"',     # " LEFT DOUBLE QUOTATION MARK
            '”': '"',     # " RIGHT DOUBLE QUOTATION MARK
            '–': '-',     # – EN DASH
            '—': '-',     # — EM DASH
            '…': '...',   # … HORIZONTAL ELLIPSIS
        }
        result = str(title or '')
        for char, repl in _REPLACEMENTS.items():
            result = result.replace(char, repl)
        return result

    use_template = os.path.exists(TEMPLATE_PATH)

    # ── Load workbook ──────────────────────────────────────────────────────
    if use_template:
        wb = load_workbook(TEMPLATE_PATH)
        ws_set = wb['settings']
        ws_ch  = wb['choices']
        ws_sv  = wb['survey']

        # Clean up LibreOffice conversion artefacts that cause Excel to show
        # a "repaired content" warning on every open:
        #
        # 1. Invalid sheetView selection elements — LO writes 4 <selection>
        #    pane entries (topRight, bottomLeft×2, bottomRight) for a simple
        #    horizontal freeze. Excel only accepts 2 for a row-only freeze.
        #    Wipe all selections/panes now; freeze_panes='A2' below rewrites
        #    them cleanly.
        #
        # 2. type=None DataValidations — internal LO markers, not real rules.
        #
        # 3. formula2='0' and operator='between' on list validations — invalid
        #    for list type; Excel flags and repairs them.
        # Trim LibreOffice's extra <selection> pane elements from sheetViews.
        # LO writes 4 selections for a simple horizontal freeze; Excel only
        # accepts 1. openpyxl's freeze_panes setter only modifies selection[0]
        # and leaves the extras intact, so we strip them here explicitly.
        # The actual freeze ('A2') is applied later per-sheet.
        for ws in (ws_set, ws_ch, ws_sv):
            ws.sheet_view.selection = ws.sheet_view.selection[:1]

        for ws in (ws_set, ws_ch, ws_sv):
            bad = [dv for dv in ws.data_validations.dataValidation
                   if dv.type is None]
            for dv in bad:
                ws.data_validations.dataValidation.remove(dv)
            for dv in ws.data_validations.dataValidation:
                if dv.type == 'list':
                    dv.formula2 = None
                    dv.operator = None

        # Clear any data rows left in the template (keep row 1 = headers)
        for ws in (ws_set, ws_ch, ws_sv):
            for row_cells in ws.iter_rows(min_row=2):
                for cell in row_cells:
                    cell.value = None
    else:
        wb     = Workbook()
        ws_set = wb.active
        ws_set.title = 'settings'
        ws_ch  = wb.create_sheet('choices')
        ws_sv  = wb.create_sheet('survey')

    # ── Sheet: settings ────────────────────────────────────────────────────
    if not use_template:
        for col_i, col in enumerate(SETTINGS_COLS, start=1):
            _header_cell(ws_set.cell(row=1, column=col_i), col)

    # Write settings values; sanitize form_title for OC compatibility.
    for col_i, col in enumerate(SETTINGS_COLS, start=1):
        val = settings.get(col, '') or ''
        if col == 'form_title':
            val = _ascii_safe_title(val)
        ws_set.cell(row=2, column=col_i).value = val
        ws_set.column_dimensions[get_column_letter(col_i)].width = \
            SETTINGS_WIDTHS.get(col, 20)

    ws_set.row_dimensions[1].height = 16
    ws_set.row_dimensions[2].height = 14
    ws_set.freeze_panes = 'A2'

    # ── Sheet: choices ─────────────────────────────────────────────────────
    choice_extra_cols = []
    _seen_choice_extra = set()
    for ch in choices:
        for k in ch:
            if (k not in CHOICES_COLS
                    and k not in STRIP_COLS
                    and k not in _seen_choice_extra
                    and ' ' not in str(k)):   # strips phantom instruction col
                choice_extra_cols.append(k)
                _seen_choice_extra.add(k)
    all_choice_cols = CHOICES_COLS + choice_extra_cols

    if not use_template:
        for col_i, col in enumerate(all_choice_cols, start=1):
            _header_cell(ws_ch.cell(row=1, column=col_i), col)
    else:
        # Always rewrite every choices header so the template's
        # phantom instruction column (col 5) is replaced cleanly.
        for col_i, col in enumerate(all_choice_cols, start=1):
            _header_cell(ws_ch.cell(row=1, column=col_i), col)
        # Null out any residual template columns beyond our valid set.
        for col_i in range(len(all_choice_cols) + 1,
                           ws_ch.max_column + 1):
            cell = ws_ch.cell(row=1, column=col_i)
            if cell.value is not None:
                cell.value = None

    for col_i in range(1, len(all_choice_cols) + 1):
        col = all_choice_cols[col_i - 1]
        ws_ch.column_dimensions[get_column_letter(col_i)].width = \
            CHOICES_WIDTHS.get(col, 18)

    for row_i, ch in enumerate(choices, start=2):
        for col_i, col in enumerate(all_choice_cols, start=1):
            val = ch.get(col, '')
            _data_cell(ws_ch.cell(row=row_i, column=col_i), val, row_i - 2)

    ws_ch.row_dimensions[1].height = 16
    ws_ch.freeze_panes = 'A2'

    # ── Sheet: survey ──────────────────────────────────────────────────────
    additional_survey_cols = [c for c in extra_c
                               if c not in SURVEY_COLS and c not in STRIP_COLS]
    all_survey_cols = SURVEY_COLS + additional_survey_cols

    if not use_template:
        for col_i, col in enumerate(all_survey_cols, start=1):
            _header_cell(ws_sv.cell(row=1, column=col_i), col)
    else:
        # Extra columns beyond the template's 20 still need headers written
        for col_i, col in enumerate(all_survey_cols, start=1):
            if col_i > len(SURVEY_COLS):
                _header_cell(ws_sv.cell(row=1, column=col_i), col)

    for col_i, col in enumerate(all_survey_cols, start=1):
        ws_sv.column_dimensions[get_column_letter(col_i)].width = \
            SURVEY_WIDTHS.get(col, 16)

    # Balance begin/end tag pairs BEFORE the name-blanking normalizer:
    # the balancer can drop or rewrite type, and the normalizer's name-
    # blanking pass only operates correctly on a balanced row sequence.
    survey = _balance_begin_end_tags(survey, form_id, build_log)
    survey = _normalize_survey_rows(survey)

    placeholders_in_form = []
    for row_i, row in enumerate(survey, start=2):
        has_placeholder = any(
            '[PLACEHOLDER' in str(v).upper() or '[UNIT]' in str(v).upper()
            for v in row.values() if v
        )
        if has_placeholder:
            placeholders_in_form.append(row.get('name', f'row_{row_i}'))

        for col_i, col in enumerate(all_survey_cols, start=1):
            val = _resolve_cell_value(row, col)
            _data_cell(ws_sv.cell(row=row_i, column=col_i), val,
                       row_i - 2, flagged=has_placeholder)
        ws_sv.row_dimensions[row_i].height = 13

    ws_sv.row_dimensions[1].height = 16
    ws_sv.freeze_panes = 'A2'

    # bind::oc:external dropdown — only needed when building from scratch;
    # the template already carries this validation for rows 2:10000.
    if not use_template:
        external_col = len(SURVEY_COLS)   # column 20 = bind::oc:external
        dv = DataValidation(
            type="list",
            formula1='"clinicaldata,contactdata,signature,identifier"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"{get_column_letter(external_col)}2:{get_column_letter(external_col)}10000"
        ws_sv.add_data_validation(dv)

    # Log any placeholders found
    if placeholders_in_form:
        build_log['placeholder_applied'].append({
            'form_id': form_id,
            'fields': placeholders_in_form,
            'note': 'Best-guess values applied — require site-specific completion'
        })

    wb.save(output_path)
    return True


# ── Build all XLSForms ─────────────────────────────────────────────────────────

# Anthropic model used for the validation self-correction loop. Pinned per
# the operator spec so behavior is reproducible across runs. If/when this
# rolls forward, update both here and the prompt template.
_SELF_CORRECTION_MODEL = "claude-sonnet-4-20250514"
_SELF_CORRECTION_MAX_ATTEMPTS = 3


def _regenerate_survey_via_ai(form, error_msg, attempt_num):
    """Ask the Anthropic API to repair this form's survey rows given the
    exact pyxform / ODK Validate error message.

    Returns the new list of survey row dicts, or raises if the API call
    fails or the response can't be parsed as a JSON array. Callers treat
    a raise here as one failed attempt — the surrounding loop decides
    whether to retry.

    The prompt is intentionally tight: model gets enough context to
    repair (form metadata + current rows + error) and is told to return
    JSON only with no commentary. Begin/end tag pairing rules are stated
    explicitly so the model doesn't reintroduce the very bug this loop
    is fixing.
    """
    import json
    try:
        import anthropic
    except ImportError as e:
        raise RuntimeError(
            f"anthropic SDK not installed — cannot self-correct "
            f"validation failures: {e}"
        ) from e

    form_id    = form.get('form_id', '?')
    form_title = form.get('form_title', form_id)
    survey     = form.get('survey', [])

    prompt = (
        "You are correcting an XLSForm survey table that failed "
        "validation for OpenClinica 4.\n\n"
        f"Form ID: {form_id}\n"
        f"Form title: {form_title}\n"
        f"Attempt: {attempt_num} of {_SELF_CORRECTION_MAX_ATTEMPTS}\n\n"
        f"Validation error:\n{error_msg}\n\n"
        f"Current (invalid) survey rows:\n"
        f"{json.dumps(survey, indent=2, default=str)}\n\n"
        "Return ONLY the corrected survey rows as a JSON array. Apply "
        "these rules:\n"
        "  - begin_repeat MUST be closed by end_repeat (never end_group).\n"
        "  - begin_group MUST be closed by end_group (never end_repeat).\n"
        "  - end_group / end_repeat rows must have an empty name field.\n"
        "  - Do not introduce new fields that were not in the input.\n"
        "  - Preserve every original field's keys; only correct values "
        "needed to pass validation.\n\n"
        "Output: a single JSON array. No markdown fence, no commentary, "
        "no preamble."
    )

    client = anthropic.Anthropic()
    resp = client.messages.create(
        model=_SELF_CORRECTION_MODEL,
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}],
    )
    # Concatenate all text blocks. Anthropic responses are a list of
    # content blocks; with a plain text-only response there's normally one.
    raw = "".join(
        b.text for b in resp.content if getattr(b, "type", "") == "text"
    ).strip()
    # Strip an accidental markdown fence if the model added one despite
    # being told not to.
    if raw.startswith("```"):
        raw = re.sub(r"^```[a-zA-Z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw.rstrip())
    parsed = json.loads(raw)
    if not isinstance(parsed, list):
        raise ValueError(
            f"AI returned a {type(parsed).__name__}, expected JSON array"
        )
    return parsed


def build_all_xlsforms(spec_data, output_dir, build_log):
    """
    Build all XLSForm files from spec_data.
    Returns list of (form_id, output_path) tuples.

    After each form is built we validate it (pyxform → ODK Validate when
    Java is present). On validation failure we attempt up to
    _SELF_CORRECTION_MAX_ATTEMPTS rounds of AI-assisted re-generation of
    the form's survey rows; each round logs:
        [edc-builder] VALIDATION FAILED: {form_id} — {error} (attempt N/3)
    and on eventual success:
        [edc-builder] VALIDATION PASSED: {form_id} after N correction(s)

    Forms that fail every correction attempt are removed from the output
    directory and recorded in build_log['build_errors'] with a
    `hard_error: True` flag — they're excluded from the final ZIP.

    Validation results (tuple form: (is_valid, errors, warnings)) are
    accumulated in build_log['validation_results'] for downstream
    surfacing in BUILD_README and the Build Checklist PDF.
    """
    # Lazy import to keep build_xlsforms.py lean.
    try:
        from validate_form import validate_xlsform
    except ImportError:
        validate_xlsform = None

    os.makedirs(output_dir, exist_ok=True)
    built = []
    build_log.setdefault('validation_results', [])
    build_log.setdefault('forms_excluded', [])

    # Track form_ids to handle variants (same form_id, different designs)
    seen_form_ids = {}

    for form in spec_data.get('forms', []):
        form_id = form.get('form_id', 'FORM')
        source_tab = form.get('source_tab', form_id)

        # Determine output filename
        # If source tab has a suffix (e.g., IE_TRT_survey), use that as filename
        tab_prefix = source_tab.replace('_survey', '').replace('_choices', '')

        if form_id in seen_form_ids:
            # Variant — use tab prefix as filename
            filename = f"{tab_prefix}.xlsx"
        else:
            seen_form_ids[form_id] = True
            # Check if there will be a variant (another form with same form_id)
            same_id_count = sum(1 for f in spec_data['forms'] if f.get('form_id') == form_id)
            if same_id_count > 1:
                filename = f"{tab_prefix}.xlsx"
            else:
                filename = f"{form_id}.xlsx"

        output_path = os.path.join(output_dir, filename)

        try:
            build_single_xlsform(form, output_path, build_log)
            built.append((tab_prefix, output_path))
            build_log['forms_built'].append(form_id)

            # Validate, then self-correct if the form is invalid.
            if validate_xlsform is None:
                continue

            is_valid, errors, warnings = validate_xlsform(output_path)
            build_log['validation_results'].append({
                'form_id':  tab_prefix,
                'is_valid': is_valid,
                'errors':   errors,
                'warnings': warnings,
            })

            if is_valid:
                continue

            # ── Self-correction loop ──────────────────────────────
            corrected = False
            working_form = dict(form)
            for attempt in range(1, _SELF_CORRECTION_MAX_ATTEMPTS + 1):
                err_preview = (errors[0] if errors else "<no detail>")[:200]
                print(
                    f"[edc-builder] VALIDATION FAILED: {tab_prefix} — "
                    f"{err_preview} (attempt {attempt}/"
                    f"{_SELF_CORRECTION_MAX_ATTEMPTS})",
                    flush=True,
                )
                try:
                    new_survey = _regenerate_survey_via_ai(
                        working_form, errors[0] if errors else "", attempt,
                    )
                except Exception as ai_err:
                    print(
                        f"[edc-builder] correction attempt {attempt} "
                        f"crashed: {type(ai_err).__name__}: {ai_err}",
                        flush=True,
                    )
                    continue

                working_form = dict(working_form)
                working_form['survey'] = new_survey
                try:
                    build_single_xlsform(working_form, output_path, build_log)
                except Exception as build_err:
                    print(
                        f"[edc-builder] correction attempt {attempt} "
                        f"rebuild failed: {build_err}", flush=True,
                    )
                    continue

                is_valid, errors, warnings = validate_xlsform(output_path)
                build_log['validation_results'].append({
                    'form_id':  tab_prefix,
                    'is_valid': is_valid,
                    'errors':   errors,
                    'warnings': warnings,
                    'attempt':  attempt,
                })
                if is_valid:
                    print(
                        f"[edc-builder] VALIDATION PASSED: {tab_prefix} "
                        f"after {attempt} correction(s)", flush=True,
                    )
                    corrected = True
                    break

            if not corrected:
                # Exhausted retries — exclude this form from the ZIP and
                # record a hard error so the build report flags it.
                build_log['build_errors'].append({
                    'form_id':    form_id,
                    'hard_error': True,
                    'error':      (
                        f"Validation failed after "
                        f"{_SELF_CORRECTION_MAX_ATTEMPTS} self-correction "
                        f"attempts. Last error: "
                        f"{errors[0] if errors else '<none>'}"
                    ),
                })
                build_log['forms_excluded'].append(tab_prefix)
                # Remove the bad file from the output dir and the built
                # list so downstream packaging doesn't ship it.
                try:
                    os.remove(output_path)
                except OSError:
                    pass
                built = [b for b in built if b[1] != output_path]
                if form_id in build_log['forms_built']:
                    build_log['forms_built'].remove(form_id)
                print(
                    f"[edc-builder] EXCLUDED from ZIP: {tab_prefix} "
                    f"(unrecoverable validation failure)", flush=True,
                )

        except Exception as e:
            build_log['build_errors'].append({
                'form_id': form_id,
                'error': str(e)
            })
            build_log['forms_skipped'].append(form_id)

    return built


# ── Write CSV files ────────────────────────────────────────────────────────────
def write_timepoint_csv(tpt_data, output_path, build_log):
    """Write the timepoint CSV file."""
    rows = tpt_data.get('rows', [])
    with open(output_path, 'w', newline='') as f:
        f.write('event,timepoint\r\n')
        for row in rows:
            event = row.get('event', '').replace(',', '')
            tpt = row.get('timepoint', '').replace(',', '')
            f.write(f'{event},{tpt}\r\n')
    build_log['build_warnings'] = build_log.get('build_warnings', [])
    if not rows:
        build_log['build_warnings'].append('Timepoint CSV has no rows — check TIMEPOINTS sheet')


def write_labranges_csv(lr_data, output_path, build_log):
    """Write the lab ranges CSV file."""
    cols = lr_data.get('columns', [])
    rows = lr_data.get('rows', [])
    placeholders = []

    with open(output_path, 'w', newline='') as f:
        f.write(','.join(cols) + '\r\n')
        for row in rows:
            values = []
            for c in cols:
                val = row.get(c, '')
                if '[PLACEHOLDER' in str(val).upper():
                    placeholders.append(f"{c}: {val}")
                    # Apply best-guess
                    if 'lower' in c.lower():
                        val = '0'
                    elif 'upper' in c.lower():
                        val = '999'
                    elif 'unit' in c.lower():
                        val = '[UNIT]'
                    elif 'lab_name' in c.lower():
                        val = '[LAB_NAME]'
                values.append(str(val).replace(',', ';'))
            f.write(','.join(values) + '\r\n')

    if placeholders:
        build_log['placeholder_applied'].append({
            'form_id': 'labranges.csv',
            'fields': placeholders[:5],
            'note': 'Lab ranges require site-specific values from each participating lab'
        })


if __name__ == '__main__':
    import sys, json
    if len(sys.argv) < 3:
        print("Usage: python build_xlsforms.py <spec_xlsx_path> <output_dir>")
        sys.exit(1)

    spec_path  = sys.argv[1]
    output_dir = sys.argv[2]

    build_log = {
        'forms_built': [], 'forms_skipped': [],
        'placeholder_applied': [], 'oid_placeholders': [],
        'qa_results': [], 'build_warnings': [], 'build_errors': []
    }

    print(f"Reading spec: {spec_path}")
    spec_data = read_spec_xlsx(spec_path)
    print(f"Found {len(spec_data['forms'])} forms")

    forms_dir = os.path.join(output_dir, 'forms')
    csv_dir   = os.path.join(output_dir, 'csv')
    os.makedirs(forms_dir, exist_ok=True)
    os.makedirs(csv_dir,   exist_ok=True)

    built = build_all_xlsforms(spec_data, forms_dir, build_log)
    print(f"Built {len(built)} XLSForms")

    study_id = spec_data['study_meta'].get('study_id', 'study')
    tpt_path = os.path.join(csv_dir, f"{study_id}_tpt.csv")
    write_timepoint_csv(spec_data['timepoint_csv'], tpt_path, build_log)
    print(f"Written: {tpt_path}")

    lr_path = os.path.join(csv_dir, 'labranges.csv')
    write_labranges_csv(spec_data['labranges_csv'], lr_path, build_log)
    print(f"Written: {lr_path}")

    print(f"\nBuild log: {json.dumps(build_log, indent=2)}")

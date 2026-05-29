"""
generate_rule_artifacts.py — Artifact writer for OC4 calendaring rules

Takes the rule_data dict (from extract_calendar_rules + validate_rules) and
writes a directory tree of deliverables, then zips it:

    rules/      one validated rule per .json file, stripped of _meta
    reports/    validation_report.md + simple_rule_recommendations.md
    review/     calendaring_spec.xlsx (tabular review)
    rationale/  calendaring_rationale.pdf (per-rule rationale)

reportlab (PDF) and openpyxl (XLSX) are optional — if either is missing the
corresponding artifact is skipped with a build_log warning rather than raising.
"""

import json
import os
import io
import zipfile
import datetime


def _warn(build_log, msg):
    build_log.setdefault("build_warnings", []).append(msg)


def _protocol_number(rule_data):
    return (rule_data.get("study_meta", {}) or {}).get("protocol_number") or "STUDY"


def _generated_stamp():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M")


def _first_action(rule):
    actions = rule.get("actions") or []
    return actions[0] if actions else {}


# ── rules/*.json ──────────────────────────────────────────────────────────────

def _write_rules_json(rule_data, rules_dir, build_log):
    for rule in rule_data.get("rules", []):
        meta = rule.get("_meta", {}) or {}
        errors = meta.get("validation_errors") or []
        name = rule.get("name") or "UNNAMED_RULE"
        if errors:
            _warn(build_log,
                  f"Rule '{name}' has {len(errors)} validation error(s) — "
                  f"excluded from rules/*.json output.")
            continue
        clean_rule = {k: v for k, v in rule.items() if k != "_meta"}
        path = os.path.join(rules_dir, f"{name}.json")
        with open(path, "w") as f:
            f.write(json.dumps(clean_rule, indent=2))


# ── reports/validation_report.md ──────────────────────────────────────────────

def _write_validation_report(rule_data, reports_dir):
    protocol = _protocol_number(rule_data)
    summary = rule_data.get("validation_summary", {}) or {}
    lines = []
    lines.append(f"# Calendaring Validation Report — {protocol}")
    lines.append("")
    lines.append(f"Generated: {_generated_stamp()}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Total rules: {summary.get('total', 0)}")
    lines.append(f"- Passed: {summary.get('passed', 0)}")
    lines.append(f"- Failed: {summary.get('failed', 0)}")
    lines.append("")
    lines.append("## Failed Rules")
    lines.append("")
    failed = summary.get("errors", []) or []
    if failed:
        for entry in failed:
            lines.append(f"### {entry.get('rule_name')}")
            for err in entry.get("errors", []):
                lines.append(f"- {err}")
            lines.append("")
    else:
        lines.append("All rules passed static validation.")
        lines.append("")
    lines.append("## Notes")
    lines.append("")
    lines.append(
        "Errors 5-10 (epoch/calendar existence, permissions, study UUID, "
        "update-mode lookups) are runtime-only and not checked statically. "
        "See validation_summary.runtime_only_checks."
    )
    lines.append("")
    lines.append("## Warnings")
    lines.append("")
    warnings = rule_data.get("warnings", []) or []
    if warnings:
        for w in warnings:
            lines.append(f"- {w}")
    else:
        lines.append("None.")
    lines.append("")
    with open(os.path.join(reports_dir, "validation_report.md"), "w") as f:
        f.write("\n".join(lines))


# ── reports/simple_rule_recommendations.md ────────────────────────────────────

def _write_recommendations(rule_data, reports_dir):
    protocol = _protocol_number(rule_data)
    recs = rule_data.get("simple_rule_recommendations", []) or []
    reminders = [r for r in recs if r.get("type") == "REMINDER"]
    conditionals = [r for r in recs if r.get("type") == "CONDITIONAL"]

    lines = []
    lines.append(f"# Simple Rule Recommendations — {protocol}")
    lines.append("")
    lines.append(
        "These items are NOT auto-built by Tier 1. Set them up manually in the "
        "OpenClinica Study Designer."
    )
    lines.append("")

    if not reminders and not conditionals:
        lines.append("No simple rule recommendations for this study.")
        lines.append("")
    else:
        if reminders:
            lines.append("## Reminders (Visit Windows)")
            lines.append("")
            for r in reminders:
                lines.append(f"- **{r.get('event_oid')}**: {r.get('description')}")
            lines.append("")
        if conditionals:
            lines.append("## Conditional / Dynamic Events")
            lines.append("")
            for r in conditionals:
                lines.append(f"- **{r.get('event_oid')}**: {r.get('description')}")
            lines.append("")

    with open(os.path.join(reports_dir, "simple_rule_recommendations.md"), "w") as f:
        f.write("\n".join(lines))


# ── review/calendaring_spec.xlsx ──────────────────────────────────────────────

def _write_xlsx(rule_data, review_dir, build_log):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        _warn(build_log, f"openpyxl unavailable — skipped calendaring_spec.xlsx ({e})")
        return

    red_fill   = PatternFill("solid", fgColor="FFC7CE")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"

    headers = [
        "Rule Name", "Event OID", "Trigger", "Condition", "Action Type",
        "Relative Event OID", "Offset Days", "Window Lower", "Window Upper",
        "Arm", "Confidence", "Validation",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    for rule in rule_data.get("rules", []):
        meta = rule.get("_meta", {}) or {}
        action = _first_action(rule)
        errors = meta.get("validation_errors") or []
        trigger = ", ".join(rule.get("triggerType") or [])
        row = [
            rule.get("name"),
            action.get("targetEventOid"),
            trigger,
            rule.get("condition"),
            action.get("type"),
            action.get("relativeEventOid"),
            action.get("startDateRelativeDays"),
            meta.get("window_lower_days"),
            meta.get("window_upper_days"),
            meta.get("arm"),
            meta.get("confidence"),
            "; ".join(errors) if errors else "OK",
        ]
        ws.append(row)
        confidence = meta.get("confidence")
        conf_cell = ws.cell(row=ws.max_row, column=11)  # Confidence column
        if confidence == "NEEDS_REVIEW":
            conf_cell.fill = red_fill
        elif confidence == "HIGH":
            conf_cell.fill = green_fill

    wb.save(os.path.join(review_dir, "calendaring_spec.xlsx"))


# ── rationale/calendaring_rationale.pdf ───────────────────────────────────────

def _write_pdf(rule_data, rationale_dir, build_log):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
    except ImportError as e:
        _warn(build_log, f"reportlab unavailable — skipped calendaring_rationale.pdf ({e})")
        return

    protocol = _protocol_number(rule_data)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(
        f"{protocol} Calendaring Rules — Tier 1 Rationale", styles["Title"]))
    story.append(Paragraph(f"Generated: {_generated_stamp()}", styles["Normal"]))
    story.append(Spacer(1, 12))

    for rule in rule_data.get("rules", []):
        meta = rule.get("_meta", {}) or {}
        action = _first_action(rule)
        src = meta.get("source_event", {}) or {}

        story.append(Paragraph(rule.get("name", "UNNAMED_RULE"), styles["Heading2"]))

        window = f"{meta.get('window_lower_days')} to {meta.get('window_upper_days')}"
        table_data = [
            ["Event OID", str(action.get("targetEventOid"))],
            ["Anchor", str(src.get("anchor_event_oid"))],
            ["Offset", str(action.get("startDateRelativeDays"))],
            ["Window", window],
            ["Arm", str(meta.get("arm"))],
            ["Confidence", str(meta.get("confidence"))],
            ["Source entry", json.dumps(src)],
        ]
        tbl = Table(table_data, colWidths=[110, 360])
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(tbl)

        if meta.get("confidence") == "NEEDS_REVIEW":
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                "<b>⚠ NEEDS REVIEW — verify timing fields manually</b>",
                styles["Normal"]))
        story.append(Spacer(1, 12))

    warnings = rule_data.get("warnings", []) or []
    if warnings:
        story.append(Paragraph("Warnings", styles["Heading2"]))
        for w in warnings:
            story.append(Paragraph(w, styles["Normal"]))

    doc = SimpleDocTemplate(
        os.path.join(rationale_dir, "calendaring_rationale.pdf"), pagesize=letter)
    doc.build(story)


# ── Main artifact generator ───────────────────────────────────────────────────

def generate_rule_artifacts(rule_data, output_dir, build_log):
    """Write all calendaring artifacts under output_dir and return a zip (bytes)."""
    rules_dir     = os.path.join(output_dir, "rules")
    reports_dir   = os.path.join(output_dir, "reports")
    review_dir    = os.path.join(output_dir, "review")
    rationale_dir = os.path.join(output_dir, "rationale")
    for d in (rules_dir, reports_dir, review_dir, rationale_dir):
        os.makedirs(d, exist_ok=True)

    _write_rules_json(rule_data, rules_dir, build_log)
    _write_validation_report(rule_data, reports_dir)
    _write_recommendations(rule_data, reports_dir)
    _write_xlsx(rule_data, review_dir, build_log)
    _write_pdf(rule_data, rationale_dir, build_log)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(output_dir):
            for fname in files:
                fpath = os.path.join(root, fname)
                arcname = os.path.relpath(fpath, output_dir)
                zf.write(fpath, arcname)
    return buf.getvalue()


if __name__ == "__main__":
    import sys, tempfile
    if len(sys.argv) < 2:
        print("Usage: python generate_rule_artifacts.py <rule_data.json> [out.zip]")
        sys.exit(1)
    rule_data = json.load(open(sys.argv[1]))
    build_log = {"build_warnings": []}
    with tempfile.TemporaryDirectory() as tmp:
        data = generate_rule_artifacts(rule_data, tmp, build_log)
    out_zip = sys.argv[2] if len(sys.argv) > 2 else "calendaring_artifacts.zip"
    with open(out_zip, "wb") as f:
        f.write(data)
    for w in build_log["build_warnings"]:
        print("WARNING:", w)
    print(f"Wrote {out_zip} ({len(data)} bytes)")

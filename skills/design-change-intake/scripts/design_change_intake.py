"""
design_change_intake.py
Design Change Intake Skill — implementation script

Accepts unstructured text describing study design changes, applies them to
the Study Specification XLSX on the AI Hub Monday.com board, saves the
source transcript, notifies the assigned PS team member, and routes any
convention proposals to the Convention Rulebook board for OC team review.

Entry point: run_design_change_intake(payload_dict) -> summary_dict
"""

import asyncio
import base64
import io
import json
import os
import uuid
from datetime import datetime, timezone

import httpx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ─────────────────────────────────────────────────────────────────

BOARD_ID              = "18409146946"
CONV_BOARD_ID         = "18411236453"
CONV_GROUP_ID         = "group_mm3xt80m"
COL_SPEC_XLSX         = "file_mm2gjqgx"   # Protocol Specification (xlsx) — always write here
COL_TRANSCRIPTS       = "file_mm3tntz9"
COL_ASSIGNEE          = "dup__of_requester__1"
COL_PROTOCOL_NUM      = "text_mm2hcfre"
COL_CLIENT            = "text7__1"
CONV_COL_PROTOCOL     = "text_mm2yzxg3"
CONV_COL_STATUS       = "color_mm2yp992"
CONV_COL_SOURCE_TYPE  = "color_mm3xbjjv"
CONV_COL_SCOPE        = "color_mm3xx99e"
CONV_COL_TEXT         = "long_text_mm3x37ty"
CONV_COL_LINK         = "link_mm3xyxk3"

MONDAY_API_URL  = "https://api.monday.com/v2"
MONDAY_FILE_URL = "https://api.monday.com/v2/file"

DOMAIN_ABBREVS = {
    "adverse events": "AE", "ae": "AE",
    "concomitant medications": "CM", "cm": "CM",
    "demographics": "DM", "dm": "DM",
    "date of visit": "DOV", "dov": "DOV",
    "disposition": "DS", "ds": "DS",
    "protocol deviations": "DV", "dv": "DV",
    "exposure dosing": "EC", "ec": "EC",
    "exposure": "EX", "ex": "EX",
    "inclusion exclusion": "IE", "ie": "IE",
    "laboratory": "LB", "lb": "LB",
    "medical history": "MH", "mh": "MH",
    "physical examination": "PE", "pe": "PE",
    "procedures": "PR", "pr": "PR",
    "psa": "PSA",
    "vital signs": "VS", "vs": "VS",
    "sponsor eligibility": "SPELIG", "spelig": "SPELIG",
    "sleep": "SLEEP", "sleep diary": "SLEEP",
    "biospecimen": "BE", "be": "BE",
    "pregnancy": "PREGPART", "pregpart": "PREGPART",
}

SOURCE_TYPE_LABELS = {
    "meeting_notes": "meeting notes",
    "email": "email",
    "transcript": "voice transcript",
}


# ── Monday helpers ────────────────────────────────────────────────────────────

def _get_token():
    return os.environ.get("MONDAY_API_TOKEN", "").strip()

def _headers():
    return {
        "Authorization": _get_token(),
        "Content-Type": "application/json",
        "API-Version": "2024-01",
    }

async def _gql(query: str, variables: dict = None) -> dict:
    payload = {"query": query}
    if variables:
        payload["variables"] = variables
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post(MONDAY_API_URL, headers=_headers(), json=payload)
    r.raise_for_status()
    return r.json()

async def _upload_file(item_id: str, col_id: str, filename: str,
                       content: bytes) -> bool:
    mutation = f"""
    mutation ($file: File!) {{
        add_file_to_column(item_id: {item_id}, column_id: "{col_id}",
                           file: $file) {{ id }}
    }}
    """
    async with httpx.AsyncClient(timeout=120) as c:
        r = await c.post(
            MONDAY_FILE_URL,
            headers={"Authorization": _get_token(), "API-Version": "2023-10"},
            files={
                "query":     (None, mutation),
                "variables": (None, '{"file": null}'),
                "map":       (None, '{"file": ["variables.file"]}'),
                "file":      (filename, content, "application/octet-stream"),
            },
        )
    return r.status_code == 200

async def _download(url: str) -> bytes:
    is_s3 = "s3.amazonaws.com" in url or "files-monday-com" in url
    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as c:
        if is_s3:
            r = await c.get(url)
        else:
            r = await c.get(url, headers={"Authorization": _get_token()})
    if r.status_code == 200 and len(r.content) > 100:
        return r.content
    return b""


# ── Step 1: Parse source text via Claude ─────────────────────────────────────

async def _parse_source_text(source_type: str, source_text: str,
                              protocol_hint: str) -> dict:
    system = """You are an expert OpenClinica EDC study build analyst.
Extract all design change requests from the provided text.

For each change identify:
- form: CRF form name or CDASH domain
- field: specific field or element name
- change_type: add_field | remove_field | rename | change_validation |
  change_visit | change_choices | change_logic | other
- description: clear unambiguous description of what changes
- rationale: why this change is needed (null if not stated)
- is_convention: true if explicitly flagged as a convention with phrases like
  "this is a convention", "always do this", "add as convention",
  "for all our studies", "for this study always"
- convention_scope: "study" | "customer" | null

Also extract:
- protocol_id: protocol or study identifier (e.g. CRS-136, PrTK05)
- study_name: human-readable study name if mentioned
- summary: one-sentence summary of all changes

Return ONLY valid JSON, no markdown, no preamble."""

    user_msg = (f"Source type: {source_type}\n"
                + (f"Protocol hint: {protocol_hint}\n" if protocol_hint else "")
                + f"Text:\n{source_text}")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    async with httpx.AsyncClient(timeout=60) as c:
        r = await c.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 2000,
                "system": system,
                "messages": [{"role": "user", "content": user_msg}],
            },
        )
    data = r.json()
    raw = "".join(b.get("text", "") for b in data.get("content", [])
                  if b.get("type") == "text")
    clean = raw.replace("```json", "").replace("```", "").strip()
    return json.loads(clean)


# ── Step 2: Find board row ────────────────────────────────────────────────────

async def _find_board_row(protocol_id: str) -> dict:
    q = """
    query {
      boards(ids: [18409146946]) {
        items_page(limit: 200) {
          items {
            id name updated_at
            column_values(ids: ["text_mm2hcfre", "dup__of_requester__1",
                                 "text7__1"]) {
              id text value
            }
          }
        }
      }
    }
    """
    resp = await _gql(q)
    items = (resp.get("data", {})
                 .get("boards", [{}])[0]
                 .get("items_page", {})
                 .get("items", []))

    pid_lower = protocol_id.lower().replace("-", "").replace(" ", "")
    matches = []
    for item in items:
        for cv in item.get("column_values", []):
            if cv["id"] == "text_mm2hcfre":
                val = (cv.get("text") or "").lower().replace("-", "").replace(" ", "")
                if pid_lower in val or val in pid_lower:
                    matches.append(item)
                    break

    if not matches:
        return {}

    matches.sort(key=lambda x: x.get("updated_at", ""), reverse=True)
    best = matches[0]

    col_map = {cv["id"]: cv for cv in best.get("column_values", [])}
    return {
        "item_id": best["id"],
        "assignee_value": col_map.get("dup__of_requester__1", {}).get("value"),
        "customer_name": col_map.get("text7__1", {}).get("text", ""),
    }


# ── Step 3: Download spec XLSX ────────────────────────────────────────────────

# Columns to search for the spec XLSX — read from whichever has the
# most recently uploaded file; always write back to COL_SPEC_XLSX_OUT.
COL_SPEC_XLSX_IN  = "file_mm2gjqgx"   # Protocol Specification (xlsx) — pipeline output / skill write-back
COL_SPEC_XLSX_ALT = "file_mm2n3x71"   # Study Specification Update Input — human upload slot


async def _download_spec_xlsx(item_id: str) -> tuple[bytes, str]:
    """
    Download the most recently uploaded spec XLSX from the board row.

    Checks both COL_SPEC_XLSX_IN (file_mm2gjqgx) and COL_SPEC_XLSX_ALT
    (file_mm2n3x71) and returns the file with the latest created_at
    timestamp. This handles the case where a PS team member manually
    uploads a revised spec to the Update Input column after the skill
    has already written a version to the output column.

    Returns (xlsx_bytes, source_col_id). Returns (b"", "") if no xlsx found.

    Example timeline this handles correctly:
      08:00 — pipeline writes spec to file_mm2gjqgx
      09:00 — email change request → skill reads file_mm2gjqgx, applies
               changes, writes back to file_mm2gjqgx
      10:30 — second email change request → skill reads file_mm2gjqgx
               (still newest), applies changes, writes back
      11:00 — PS team member uploads manual revision to file_mm2n3x71
      11:30 — third email change request → skill finds file_mm2n3x71 is
               newer, reads from there, writes result to file_mm2gjqgx
    """
    q = """
    query($i: [ID!]) {
      items(ids: $i) {
        assets { id name url public_url created_at column_id }
      }
    }
    """
    resp = await _gql(q, {"i": [item_id]})
    assets = (resp.get("data", {})
                  .get("items", [{}])[0]
                  .get("assets", []))

    # Filter to xlsx files from either of the two spec columns
    spec_cols = {COL_SPEC_XLSX_IN, COL_SPEC_XLSX_ALT}
    xlsx_assets = [
        a for a in assets
        if (a.get("name") or "").lower().endswith(".xlsx")
        and a.get("column_id") in spec_cols
    ]

    # If no column_id in assets (older Monday API responses), fall back
    # to all xlsx assets on the row
    if not xlsx_assets:
        xlsx_assets = [
            a for a in assets
            if (a.get("name") or "").lower().endswith(".xlsx")
        ]

    if not xlsx_assets:
        return b"", ""

    # Pick the most recently uploaded
    xlsx_assets.sort(key=lambda a: a.get("created_at", ""), reverse=True)
    best = xlsx_assets[0]
    source_col = best.get("column_id", COL_SPEC_XLSX_IN)

    url = best.get("public_url") or best.get("url", "")
    if not url:
        return b"", ""

    data = await _download(url)
    print(f"_download_spec_xlsx: read from col={source_col} "
          f"file={best.get('name','')} "
          f"created_at={best.get('created_at','')}", flush=True)
    return data, source_col


# ── Notification helpers ─────────────────────────────────────────────────────

async def _post_item_update(item_id: str, body: str):
    q = """mutation($id: ID!, $b: String!) {
        create_update(item_id: $id, body: $b) { id }
    }"""
    try:
        await _gql(q, {"id": str(item_id), "b": body})
    except Exception as e:
        print(f"_post_item_update failed: {e}", flush=True)


async def _bell_assignee(assignee_value: str, item_id: str, text: str):
    """Send bell notification to the assignee if one is set on the row."""
    if not assignee_value:
        return
    try:
        parsed = json.loads(assignee_value)
        persons = parsed.get("personsAndTeams", [])
        if not persons:
            return
        uid = str(persons[0]["id"])
        q = """mutation($u: ID!, $t: ID!, $tx: String!) {
            create_notification(user_id: $u, target_id: $t,
                                text: $tx, target_type: Project) { text }
        }"""
        await _gql(q, {"u": uid, "t": str(item_id), "tx": text})
    except Exception as e:
        print(f"_bell_assignee failed: {e}", flush=True)


async def _create_unmatched_review_item(source_text: str, source_type: str,
                                         summary: str, reason: str,
                                         protocol_id: str = "") -> str | None:
    """
    Creates a review item on the Change Requests New board when the skill
    cannot match the change request to a board row. Routes to the
    'Email Change Requests (AI)' group so a PS team member can identify
    the correct study and re-submit.

    reason: 'no_protocol_id' | 'no_board_row'
    """
    from datetime import datetime, timezone as _tz
    ts = datetime.now(_tz.utc).strftime("%Y-%m-%d %H:%M UTC")
    src_label = {
        "meeting_notes": "meeting notes",
        "email": "email",
        "transcript": "voice transcript",
    }.get(source_type, source_type)

    if reason == "no_protocol_id":
        item_name = f"[UNMATCHED] No protocol ID — {src_label} {ts}"
        ai_summary = (f"Change request received via {src_label} but no "
                      f"protocol ID could be identified. Human must match "
                      f"to the correct study row and re-submit.")
    else:
        item_name = f"[UNMATCHED] {protocol_id} — row not found — {ts}"
        ai_summary = (f"Change request for protocol '{protocol_id}' received "
                      f"via {src_label} but no matching AI Hub board row found. "
                      f"Check protocol number spelling or create the study row "
                      f"first, then re-submit.")

    proposed_update = (
        f"[DESIGN_CHANGE] [SOURCE_TYPE:{source_type}]"
        + (f" [PROTOCOL:{protocol_id}]" if protocol_id else "")
        + f"\n\n{source_text}"
    )

    col_values = json.dumps({
        "project_status":         {"label": "Ready To Start"},
        "project_priority":       {"label": "High"},
        "long_text_mm3zvw2q":     {"text": source_text},
        "long_text_mm3z80v1":     {"text": ai_summary},
        "long_text_mm3z9m21":     {"text": proposed_update},
        "color_mm3zkh2y":         {"label": "Awaiting Review"},
        "text_mm3zkmkw":          protocol_id or "Not identified",
    })

    q = """
    mutation($name: String!, $col: JSON!) {
        create_item(
            board_id: 18395557554,
            group_id: "group_mm3zj7yj",
            item_name: $name,
            column_values: $col
        ) { id }
    }
    """
    try:
        resp = await _gql(q, {"name": item_name, "col": col_values})
        new_id = (resp.get("data", {})
                      .get("create_item", {})
                      .get("id"))
        if new_id:
            await _post_item_update(new_id, (
                f"⚠️ Unmatched design change request\n\n"
                f"Reason: {ai_summary}\n\n"
                f"To action this:\n"
                f"1. Identify the correct AI Hub board row for this study\n"
                f"2. Copy the Proposed Update text from this item\n"
                f"3. Post it as an update on the correct AI Hub row\n"
                f"(The [PROTOCOL:] tag will ensure the pipeline routes it correctly)"
            ))
        return new_id
    except Exception as e:
        print(f"_create_unmatched_review_item failed: {e}", flush=True)
        return None


# ── Step 4: Apply changes to spec XLSX ───────────────────────────────────────

def _get_header_map(ws, header_row: int) -> dict:
    """Return {header_text: col_index} for a given row."""
    hmap = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            hmap[str(val).strip()] = col
    return hmap

def _find_form_tab(wb, form_name: str):
    """Return (survey_tab_name, form_id) or (None, None)."""
    form_lower = form_name.lower().strip()

    # 1. Direct tab prefix match
    for name in wb.sheetnames:
        if name.lower().endswith("_survey"):
            prefix = name[:-7]
            if prefix.lower() == form_lower:
                return name, prefix

    # 2. Settings form_title match
    for name in wb.sheetnames:
        if name.lower().endswith("_settings"):
            prefix = name[:-9]
            ws = wb[name]
            # form_title is in row 3, col B (row index 3, col 2)
            try:
                title = str(ws.cell(row=3, column=2).value or "").lower()
                if form_lower in title or title in form_lower:
                    survey_name = f"{prefix}_survey"
                    if survey_name in wb.sheetnames:
                        return survey_name, prefix
            except Exception:
                pass

    # 3. Domain abbreviation lookup
    abbrev = DOMAIN_ABBREVS.get(form_lower)
    if abbrev:
        tab = f"{abbrev}_survey"
        if tab in wb.sheetnames:
            return tab, abbrev

    # 4. Partial substring
    for name in wb.sheetnames:
        if name.lower().endswith("_survey"):
            prefix = name[:-7]
            if form_lower in prefix.lower() or prefix.lower() in form_lower:
                return name, prefix

    return None, None

def _find_field_row(ws, hmap: dict, field_name: str, data_start: int) -> int:
    """Find the row where name or label matches field_name. Returns row num or 0."""
    name_col = hmap.get("name")
    label_col = hmap.get("label")
    field_lower = field_name.lower().strip()

    for row in range(data_start, ws.max_row + 1):
        if name_col:
            val = str(ws.cell(row=row, column=name_col).value or "").lower()
            if val == field_lower or field_lower in val:
                return row
        if label_col:
            val = str(ws.cell(row=row, column=label_col).value or "").lower()
            if val == field_lower or field_lower in val:
                return row
    return 0

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_new_row(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        c = ws.cell(row=row_num, column=col)
        c.font = Font(name="Arial", size=8)
        c.fill = PatternFill("solid", fgColor="FFF3CD")  # amber = FLAGGED
        c.border = _thin_border()
        c.alignment = Alignment(wrap_text=True, vertical="top")

def apply_changes(wb, changes: list, run_id: str,
                  source_type: str, source_filename: str,
                  timestamp: str) -> list:
    """
    Apply all changes to the workbook in-place.
    Returns list of change_log dicts.
    """
    change_log = []

    for change in changes:
        form_name    = change.get("form", "")
        field_name   = change.get("field", "")
        change_type  = change.get("change_type", "other")
        description  = change.get("description", "")
        rationale    = change.get("rationale") or ""

        survey_tab, form_id = _find_form_tab(wb, form_name)
        if not survey_tab:
            change_log.append({
                "timestamp": timestamp, "run_id": run_id,
                "change_type": change_type, "form_id": form_name,
                "field_name": field_name, "description": description,
                "action_taken": "UNRESOLVED — form not found in spec",
                "resolved": False,
                "log_note": f"No tab matching '{form_name}' in workbook",
                "source_type": source_type, "source_filename": source_filename,
            })
            continue

        ws_survey   = wb[survey_tab]
        hmap        = _get_header_map(ws_survey, row_num=3)
        data_start  = 4  # rows 1-3 are banner/legend/headers

        action_col  = hmap.get("ACTION", 1)
        notes_col   = hmap.get("NOTES_FOR_AI", 2)
        name_col    = hmap.get("name")
        label_col   = hmap.get("label")
        type_col    = hmap.get("type")
        constr_col  = hmap.get("constraint")
        constr_msg_col = hmap.get("constraint_message")
        calc_col    = hmap.get("calculation")
        rel_col     = hmap.get("relevant")
        ncols       = ws_survey.max_column

        action_taken = ""
        resolved = True
        log_note = ""

        # ── add_field ──────────────────────────────────────────────────────
        if change_type == "add_field":
            new_row = ws_survey.max_row + 1
            _style_new_row(ws_survey, new_row, ncols)
            ws_survey.cell(row=new_row, column=action_col, value="ADD")
            # Infer type from description
            desc_lower = description.lower()
            inferred_type = "text"
            if any(w in desc_lower for w in ["date", "day", "month", "year"]):
                inferred_type = "date"
            elif any(w in desc_lower for w in ["integer", "number", "count", "age"]):
                inferred_type = "integer"
            elif any(w in desc_lower for w in ["select", "dropdown", "choice", "yes/no",
                                                 "yes or no", "radio"]):
                inferred_type = "select_one"
            if type_col:
                ws_survey.cell(row=new_row, column=type_col, value=inferred_type)
            if name_col:
                snake = field_name.lower().replace(" ", "_").replace("-", "_")
                ws_survey.cell(row=new_row, column=name_col, value=snake)
            if label_col:
                ws_survey.cell(row=new_row, column=label_col, value=field_name)
            if notes_col:
                ws_survey.cell(row=new_row, column=notes_col,
                               value=f"Added by design-change-intake: {description}")
            action_taken = f"Added new row {new_row} in {survey_tab} (ACTION=ADD, type={inferred_type})"
            log_note = "Verify type, validation, and itemgroup before next build"

        # ── remove_field ───────────────────────────────────────────────────
        elif change_type == "remove_field":
            row_num = _find_field_row(ws_survey, hmap, field_name, data_start)
            if row_num:
                ws_survey.cell(row=row_num, column=action_col, value="DELETE")
                if notes_col:
                    ws_survey.cell(row=row_num, column=notes_col,
                                   value=f"Marked for deletion by design-change-intake: {description}")
                action_taken = f"Set ACTION=DELETE on row {row_num} in {survey_tab}"
            else:
                resolved = False
                action_taken = "UNRESOLVED — field not found"
                log_note = f"No row matching '{field_name}' in {survey_tab}"

        # ── rename ─────────────────────────────────────────────────────────
        elif change_type == "rename":
            row_num = _find_field_row(ws_survey, hmap, field_name, data_start)
            if row_num and label_col:
                # New label is likely in description — extract after "to" or "as"
                new_label = field_name  # fallback
                for marker in [" to ", " as ", " rename to ", " renamed to "]:
                    if marker in description.lower():
                        new_label = description.lower().split(marker)[-1].strip().capitalize()
                        break
                ws_survey.cell(row=row_num, column=label_col, value=new_label)
                if notes_col:
                    ws_survey.cell(row=row_num, column=notes_col,
                                   value=f"Label renamed by design-change-intake: {description}")
                action_taken = f"Updated label on row {row_num} in {survey_tab} to '{new_label}'"
            else:
                resolved = False
                action_taken = "UNRESOLVED — field not found"
                log_note = f"No row matching '{field_name}' in {survey_tab}"

        # ── change_validation ──────────────────────────────────────────────
        elif change_type == "change_validation":
            row_num = _find_field_row(ws_survey, hmap, field_name, data_start)
            if row_num:
                # If description contains XPath-like syntax, apply directly
                has_xpath = any(tok in description for tok in
                                ["instance(", "concat(", "format-date(", ".", "${"])
                if has_xpath and constr_col:
                    ws_survey.cell(row=row_num, column=constr_col, value=description)
                    action_taken = f"Updated constraint on row {row_num} in {survey_tab}"
                else:
                    # Natural language — note for human authoring
                    if notes_col:
                        existing = str(ws_survey.cell(row=row_num, column=notes_col).value or "")
                        ws_survey.cell(row=row_num, column=notes_col,
                                       value=f"{existing}\nVALIDATION_CHANGE_REQUEST (needs XPath): {description}".strip())
                    action_taken = f"Noted validation change request on row {row_num} — needs human XPath authoring"
                    log_note = "Natural language validation request — human must write XPath"
                    resolved = False
            else:
                resolved = False
                action_taken = "UNRESOLVED — field not found"
                log_note = f"No row matching '{field_name}' in {survey_tab}"

        # ── change_choices ─────────────────────────────────────────────────
        elif change_type == "change_choices":
            choices_tab = f"{form_id}_choices"
            if choices_tab in wb.sheetnames:
                ws_ch = wb[choices_tab]
                ch_hmap = _get_header_map(ws_ch, row_num=2)
                list_col = ch_hmap.get("list_name")
                act_col_ch = ch_hmap.get("ACTION", 1)
                notes_col_ch = ch_hmap.get("NOTES_FOR_AI", 2)
                ncols_ch = ws_ch.max_column

                new_row = ws_ch.max_row + 1
                _style_new_row(ws_ch, new_row, ncols_ch)
                if act_col_ch:
                    ws_ch.cell(row=new_row, column=act_col_ch, value="ADD")
                if list_col:
                    ws_ch.cell(row=new_row, column=list_col, value=field_name.upper())
                if notes_col_ch:
                    ws_ch.cell(row=new_row, column=notes_col_ch,
                               value=f"Choice change by design-change-intake: {description}")
                action_taken = f"Added placeholder row in {choices_tab} for choice change — complete list_name, label, name"
                log_note = "Partial: human must fill in exact choice values"
                resolved = False
            else:
                resolved = False
                action_taken = "UNRESOLVED — choices tab not found"
                log_note = f"{choices_tab} not in workbook"

        # ── change_visit ───────────────────────────────────────────────────
        elif change_type == "change_visit":
            if "TIMEPOINTS" in wb.sheetnames:
                ws_tp = wb["TIMEPOINTS"]
                new_row = ws_tp.max_row + 1
                ws_tp.cell(row=new_row, column=1,
                           value=f"DESIGN_CHANGE_REQUEST: {description}")
                ws_tp.cell(row=new_row, column=1).font = Font(
                    name="Arial", size=8, bold=True, color="7D5A00")
                ws_tp.cell(row=new_row, column=1).fill = PatternFill(
                    "solid", fgColor="FFF3CD")
                action_taken = f"Visit change noted in TIMEPOINTS row {new_row}"
                log_note = "Requires manual schedule update"
                resolved = False
            else:
                action_taken = "UNRESOLVED — TIMEPOINTS sheet not found"
                resolved = False

        # ── change_logic ───────────────────────────────────────────────────
        elif change_type == "change_logic":
            row_num = _find_field_row(ws_survey, hmap, field_name, data_start)
            if row_num:
                has_xpath = any(tok in description for tok in
                                ["instance(", "${", "= '", "!= '", "and ", "or "])
                if has_xpath and rel_col:
                    ws_survey.cell(row=row_num, column=rel_col, value=description)
                    action_taken = f"Updated relevant on row {row_num} in {survey_tab}"
                else:
                    if notes_col:
                        existing = str(ws_survey.cell(row=row_num, column=notes_col).value or "")
                        ws_survey.cell(row=row_num, column=notes_col,
                                       value=f"{existing}\nLOGIC_CHANGE_REQUEST (needs XPath): {description}".strip())
                    action_taken = f"Noted logic change on row {row_num} — needs human XPath authoring"
                    log_note = "Natural language logic request — human must write XPath"
                    resolved = False
            else:
                resolved = False
                action_taken = "UNRESOLVED — field not found"
                log_note = f"No row matching '{field_name}' in {survey_tab}"

        # ── other ──────────────────────────────────────────────────────────
        else:
            if notes_col and ws_survey.max_row >= data_start:
                last_row = ws_survey.max_row
                existing = str(ws_survey.cell(row=last_row, column=notes_col).value or "")
                ws_survey.cell(row=last_row, column=notes_col,
                               value=f"{existing}\nDESIGN_CHANGE_REQUEST: {description}".strip())
            action_taken = f"Unstructured change noted in {survey_tab}"
            log_note = "Human review required"
            resolved = False

        change_log.append({
            "timestamp": timestamp, "run_id": run_id,
            "change_type": change_type, "form_id": form_id or form_name,
            "field_name": field_name, "description": description,
            "action_taken": action_taken, "resolved": resolved,
            "log_note": log_note,
            "source_type": source_type, "source_filename": source_filename,
        })

    return change_log


def write_change_log(wb, change_log: list):
    if "CHANGE_LOG" not in wb.sheetnames:
        ws = wb.create_sheet("CHANGE_LOG")
        headers = ["timestamp", "run_id", "change_type", "form_id", "field_name",
                   "description", "action_taken", "resolved", "log_note",
                   "source_type", "source_filename"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1B3A6B")
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center")
    else:
        ws = wb["CHANGE_LOG"]

    next_row = ws.max_row + 1
    keys = ["timestamp", "run_id", "change_type", "form_id", "field_name",
            "description", "action_taken", "resolved", "log_note",
            "source_type", "source_filename"]
    for entry in change_log:
        for col, key in enumerate(keys, start=1):
            c = ws.cell(row=next_row, column=col, value=str(entry.get(key, "")))
            c.font = Font(name="Arial", size=8)
            c.fill = PatternFill("solid",
                                 fgColor="D5F5E3" if entry.get("resolved") else "FFF3CD")
            c.border = _thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
        next_row += 1


# ── Step 7: Notify assignee ───────────────────────────────────────────────────

async def _get_assignee(assignee_value: str) -> dict:
    if not assignee_value:
        return {}
    try:
        parsed = json.loads(assignee_value)
        persons = parsed.get("personsAndTeams", [])
        if not persons:
            return {}
        uid = persons[0]["id"]
        q = "query($ids: [ID!]) { users(ids: $ids) { id name email } }"
        resp = await _gql(q, {"ids": [str(uid)]})
        users = resp.get("data", {}).get("users", [])
        return users[0] if users else {}
    except Exception:
        return {}

async def _send_bell_notification(assignee_id: str, item_id: str,
                                   text: str):
    q = """
    mutation($uid: ID!, $tid: ID!, $txt: String!) {
      create_notification(user_id: $uid, target_id: $tid,
                          text: $txt, target_type: Project) { text }
    }
    """
    await _gql(q, {"uid": str(assignee_id), "tid": str(item_id), "txt": text})

async def _post_item_update(item_id: str, body: str):
    q = "mutation($id: ID!, $body: String!) { create_update(item_id: $id, body: $body) { id } }"
    await _gql(q, {"id": str(item_id), "body": body})


# ── Step 8: Route convention proposals ───────────────────────────────────────

async def _create_convention_row(change: dict, protocol_id: str,
                                  item_id: str, source_type_label: str):
    scope = change.get("convention_scope") or "study"
    scope_label = "Customer" if scope == "customer" else "Study"
    conv_name = f"[PROPOSED] {change['form']} / {change['field']} — {protocol_id}"
    source_url = (f"https://openclinica-customerfirst.monday.com"
                  f"/boards/18409146946/pulses/{item_id}")
    conv_text = (f"{change['description']}\n\n"
                 f"Rationale: {change.get('rationale') or 'Not stated'}\n\n"
                 f"Source: {source_type_label} for {protocol_id}")

    col_values = json.dumps({
        "text_mm2yzxg3":    protocol_id,
        "color_mm2yp992":   {"label": "Submitted"},
        "color_mm3xbjjv":   {"label": "Customer Proposed"},
        "color_mm3xx99e":   {"label": scope_label},
        "long_text_mm3x37ty": {"text": conv_text},
        "link_mm3xyxk3":    {"url": source_url, "text": f"AI Hub — {protocol_id}"},
    })
    q = """
    mutation($name: String!, $col: JSON!) {
      create_item(board_id: 18411236453, group_id: "group_mm3xt80m",
                  item_name: $name, column_values: $col) { id }
    }
    """
    resp = await _gql(q, {"name": conv_name, "col": col_values})
    new_id = (resp.get("data", {})
                  .get("create_item", {})
                  .get("id"))
    if new_id:
        await _post_item_update(
            new_id,
            f"New customer-proposed convention for review.\n\n"
            f"Protocol: {protocol_id}\nScope: {scope_label}\n"
            f"Source: {source_type_label}\n\n"
            f"Review Proposed Convention text and either:\n"
            f"- Set Submit Trigger → Submit for Review to approve\n"
            f"- Add a rejection comment\n\n"
            f"This convention will NOT affect any pipeline output until approved."
        )
    return bool(new_id)


# ── Main entry point ──────────────────────────────────────────────────────────

async def run_design_change_intake(payload: dict) -> dict:
    source_type   = payload.get("source_type", "meeting_notes")
    source_text   = payload.get("source_text", "")
    protocol_hint = payload.get("protocol_hint", "")

    run_id    = str(uuid.uuid4())[:8]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    ts_short  = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    src_label = SOURCE_TYPE_LABELS.get(source_type, source_type)

    # Step 1 — Parse
    try:
        parsed = await _parse_source_text(source_type, source_text, protocol_hint)
    except Exception as e:
        return {"status": "error", "message": f"Parse failed: {e}"}

    protocol_id = parsed.get("protocol_id") or protocol_hint
    if not protocol_id:
        # No protocol ID — route to Change Requests board for human matching
        msg = ("No protocol ID could be identified in the source text. "
               "A PS team member must match this to the correct study row "
               "and re-submit with a [PROTOCOL:XXX] tag.")
        await _create_unmatched_review_item(
            source_text=source_text,
            source_type=source_type,
            summary=parsed.get("summary", ""),
            reason="no_protocol_id",
        )
        return {"status": "error",
                "message": "No protocol ID found — review item created on Change Requests board."}

    changes = parsed.get("changes", [])
    summary = parsed.get("summary", "")

    # Step 2 — Find board row
    row_info = await _find_board_row(protocol_id)
    if not row_info:
        # Protocol identified but no matching board row — route to review
        msg = (f"No AI Hub board row found matching protocol: {protocol_id}. "
               f"The protocol number may be misspelled or the study may not "
               f"yet have a row on the AI Hub board.")
        await _create_unmatched_review_item(
            source_text=source_text,
            source_type=source_type,
            summary=parsed.get("summary", ""),
            reason="no_board_row",
            protocol_id=protocol_id,
        )
        return {"status": "error",
                "message": f"No board row for {protocol_id} — review item created."}
    item_id        = row_info["item_id"]
    assignee_value = row_info.get("assignee_value")

    # Step 3 — Download spec XLSX
    # Reads from whichever of file_mm2gjqgx or file_mm2n3x71 has the
    # most recently uploaded xlsx. Always writes result to file_mm2gjqgx.
    spec_bytes, source_col = await _download_spec_xlsx(item_id)
    if not spec_bytes:
        # No spec found in either column — notify and abort
        msg = (f"No Protocol Specification (xlsx) found on board row for "
               f"{protocol_id}. Neither file_mm2gjqgx nor file_mm2n3x71 "
               f"contains an xlsx file. Run the full pipeline first to "
               f"generate the spec, then re-submit the change request.")
        await _post_item_update(item_id, f"⚠️ Design change intake failed\n\n{msg}")
        await _bell_assignee(assignee_value, item_id,
            f"Design change intake failed for {protocol_id}: no spec XLSX found.")
        return {"status": "error", "message": msg}

    # Step 4 — Apply changes
    wb = load_workbook(io.BytesIO(spec_bytes))
    source_filename = f"{protocol_id}_ChangeRequest_{ts_short}.txt"
    change_log = apply_changes(wb, changes, run_id, source_type,
                               source_filename, timestamp)
    write_change_log(wb, change_log)

    # Step 5 — Upload updated spec
    out = io.BytesIO()
    wb.save(out)
    updated_bytes = out.getvalue()

    spec_filename = f"{protocol_id}_Study_Spec_Updated_{ts_short}.xlsx"
    spec_uploaded = False
    try:
        spec_uploaded = await _upload_file(item_id, COL_SPEC_XLSX,
                                            spec_filename, updated_bytes)
    except Exception as e:
        await asyncio.sleep(5)
        try:
            spec_uploaded = await _upload_file(item_id, COL_SPEC_XLSX,
                                                spec_filename, updated_bytes)
        except Exception:
            pass

    # Step 6 — Save transcript
    transcript_content = (
        f"Design Change Request Transcript\n"
        f"=================================\n"
        f"Source Type : {source_type}\n"
        f"Protocol    : {protocol_id}\n"
        f"Received    : {timestamp}\n"
        f"Summary     : {summary}\n"
        f"Changes     : {len(changes)}\n\n"
        f"--- SOURCE TEXT ---\n{source_text}\n\n"
        f"--- CHANGES EXTRACTED ---\n{json.dumps(changes, indent=2)}\n"
    )
    transcript_saved = False
    try:
        transcript_saved = await _upload_file(
            item_id, COL_TRANSCRIPTS, source_filename,
            transcript_content.encode("utf-8")
        )
    except Exception:
        pass

    # Step 7 — Notify assignee
    assignee = await _get_assignee(assignee_value or "")
    assignee_notified = False
    n_applied  = sum(1 for cl in change_log if cl["resolved"])
    n_unresolved = sum(1 for cl in change_log if not cl["resolved"])

    bullet_list = "\n".join(
        f"• [{cl['change_type']}] {cl['form_id']} / {cl['field_name']}: {cl['description']}"
        for cl in change_log
    )
    notif_text = (
        f"Updated spec posted for {protocol_id} — {len(changes)} change(s) "
        f"applied from {src_label}. Source transcript saved. "
        f"Review CHANGE_LOG sheet in updated spec."
    )
    update_body = (
        f"Design Change Update — {protocol_id}\n\n"
        f"{len(changes)} change(s) processed from {src_label}.\n"
        f"{n_applied} applied, {n_unresolved} unresolved (need human review).\n\n"
        f"Changes:\n{bullet_list}\n\n"
        f"Source transcript saved to Change Request Transcripts column.\n"
        f"Review CHANGE_LOG sheet in updated spec and correct any issues."
    )

    if assignee.get("id"):
        try:
            await _send_bell_notification(
                assignee["id"], item_id, notif_text)
            assignee_notified = True
        except Exception:
            pass

    try:
        await _post_item_update(item_id, update_body)
    except Exception:
        pass

    # Step 8 — Route convention proposals
    conventions_proposed = 0
    for change in changes:
        if change.get("is_convention"):
            try:
                ok = await _create_convention_row(
                    change, protocol_id, item_id, src_label)
                if ok:
                    conventions_proposed += 1
            except Exception:
                pass

    # Step 9 — Return summary
    return {
        "status": "success",
        "protocol_id": protocol_id,
        "item_id": item_id,
        "run_id": run_id,
        "changes_applied": n_applied,
        "changes_unresolved": n_unresolved,
        "spec_uploaded": spec_uploaded,
        "spec_filename": spec_filename,
        "transcript_saved": transcript_saved,
        "assignee_notified": assignee_notified,
        "conventions_proposed": conventions_proposed,
        "change_log": change_log,
    }


# ── CLI entry point for pipeline.py ──────────────────────────────────────────

if __name__ == "__main__":
    import sys
    payload = json.loads(sys.stdin.read())
    result = asyncio.run(run_design_change_intake(payload))
    out_bytes = json.dumps(result).encode()
    b64 = base64.standard_b64encode(out_bytes).decode()
    print(f"===JSON_START===\n{b64}\n===JSON_END===")

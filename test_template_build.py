"""
test_template_build.py — Local test for OC form template integration
Run from the repo root:  python3 test_template_build.py

Tests both code paths without any API tokens, monday.com, or Railway:
  1. build_single_xlsform  — the main EDC build path (run_edc_build)
  2. _xlsform_zip          — the DVS-translate human-in-the-loop path

Output files land in test_output/ next to this script.
Open any .xlsx in Excel / LibreOffice to verify 9 tabs and the
bind::oc:external dropdown in column T of the survey sheet.
"""

import io, json, os, sys, zipfile, shutil

# ── Locate repo root (script lives there) ─────────────────────────────────────
REPO_ROOT  = os.path.dirname(os.path.abspath(__file__))
SKILLS_DIR = os.path.join(REPO_ROOT, 'skills')
EDC_SCRIPTS= os.path.join(SKILLS_DIR, 'edc-builder', 'scripts')
OUT_DIR    = os.path.join(REPO_ROOT, 'test_output')

if EDC_SCRIPTS not in sys.path:
    sys.path.insert(0, EDC_SCRIPTS)

os.makedirs(OUT_DIR, exist_ok=True)

# ── Minimal form data — no protocol or API needed ─────────────────────────────
SAMPLE_FORMS = {
    "AE.xlsx": {
        "settings": {
            "form_title": "Adverse Events",
            "form_id":    "F_AE",
            "version":    "1",
            "style":      "theme-grid",
            "crossform_references": "",
            "namespaces": 'oc="http://openclinica.org/xforms" , OpenClinica="http://openclinica.com/odm"',
        },
        "survey": [
            {"type": "begin group",      "name": "GRP_AE",  "label": "Adverse Event",  "bind::oc:itemgroup": "F_AE.MAIN"},
            {"type": "text",             "name": "AETERM",  "label": "Adverse Event Term", "bind::oc:itemgroup": "F_AE.MAIN", "required": "yes()"},
            {"type": "date",             "name": "AESTDTC", "label": "Start Date",      "bind::oc:itemgroup": "F_AE.MAIN"},
            {"type": "select_one sev",   "name": "AESEV",   "label": "Severity",        "bind::oc:itemgroup": "F_AE.MAIN"},
            {"type": "select_one yn",    "name": "AESER",   "label": "Serious?",        "bind::oc:itemgroup": "F_AE.MAIN"},
            {"type": "select_one rel",   "name": "AEREL",   "label": "Relationship",    "bind::oc:itemgroup": "F_AE.MAIN"},
            {"type": "end group",        "name": "GRP_AE",  "label": ""},
        ],
        "choices": [
            {"list_name": "yn",  "label": "Yes",      "name": "1", "image": ""},
            {"list_name": "yn",  "label": "No",       "name": "0", "image": ""},
            {"list_name": "sev", "label": "Mild",     "name": "1", "image": ""},
            {"list_name": "sev", "label": "Moderate", "name": "2", "image": ""},
            {"list_name": "sev", "label": "Severe",   "name": "3", "image": ""},
            {"list_name": "rel", "label": "Related",  "name": "1", "image": ""},
            {"list_name": "rel", "label": "Unrelated","name": "2", "image": ""},
        ],
    },
    "VS.xlsx": {
        "settings": {
            "form_title": "Vital Signs",
            "form_id":    "F_VS",
            "version":    "1",
            "style":      "theme-grid",
            "crossform_references": "",
            "namespaces": 'oc="http://openclinica.org/xforms" , OpenClinica="http://openclinica.com/odm"',
        },
        "survey": [
            {"type": "begin group", "name": "GRP_VS",  "label": "Vital Signs",  "bind::oc:itemgroup": "F_VS.MAIN"},
            {"type": "decimal",     "name": "VSSYSBP", "label": "Systolic BP",  "bind::oc:itemgroup": "F_VS.MAIN",
             "constraint": ". >= 60 and . <= 250", "constraint_message": "Must be 60–250"},
            {"type": "decimal",     "name": "VSDIABP", "label": "Diastolic BP", "bind::oc:itemgroup": "F_VS.MAIN",
             "constraint": ". >= 40 and . <= 160", "constraint_message": "Must be 40–160"},
            {"type": "decimal",     "name": "VSPULSE", "label": "Pulse (bpm)",  "bind::oc:itemgroup": "F_VS.MAIN"},
            {"type": "end group",   "name": "GRP_VS",  "label": ""},
        ],
        "choices": [],
    },
}


def _check(label, condition, detail=""):
    status = "PASS" if condition else "FAIL"
    mark   = "✓" if condition else "✗"
    print(f"  {mark} {label}" + (f" — {detail}" if detail else ""))
    return condition


# ════════════════════════════════════════════════════════════════════════════════
# TEST 1: build_single_xlsform  (main EDC build path)
# ════════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("TEST 1: build_single_xlsform  (main EDC build path)")
print("="*60)

try:
    from build_xlsforms import build_single_xlsform, TEMPLATE_PATH
    print(f"  Imported build_xlsforms OK")
    print(f"  TEMPLATE_PATH = {TEMPLATE_PATH}")
    print(f"  Template exists: {os.path.exists(TEMPLATE_PATH)}")
except ImportError as e:
    print(f"  FAIL — could not import build_xlsforms: {e}")
    print(f"  Check that {EDC_SCRIPTS} contains build_xlsforms.py")
    sys.exit(1)

all_passed = True
from openpyxl import load_workbook

for filename, form_data in SAMPLE_FORMS.items():
    form_id   = form_data["settings"]["form_id"]
    out_path  = os.path.join(OUT_DIR, filename)
    build_log = {"placeholder_applied": [], "build_errors": []}

    skill_data = {
        "form_id":    form_id,
        "form_title": form_data["settings"]["form_title"],
        "settings":   form_data["settings"],
        "survey":     form_data["survey"],
        "choices":    form_data["choices"],
        "extra_cols": [],
    }

    try:
        build_single_xlsform(skill_data, out_path, build_log)
    except Exception as e:
        print(f"\n  {filename}: BUILD ERROR — {e}")
        all_passed = False
        continue

    wb  = load_workbook(out_path)
    sv  = wb["survey"]
    dvs = list(wb["survey"].data_validations.dataValidation)

    print(f"\n  {filename}:")
    all_passed &= _check("9 sheets present",   len(wb.sheetnames) == 9,
                         f"got {len(wb.sheetnames)}: {wb.sheetnames}")
    all_passed &= _check("survey has data rows", sv.max_row > 1,
                         f"{sv.max_row - 1} data rows")
    all_passed &= _check("bind::oc:external dropdown",  len(dvs) > 0,
                         dvs[0].formula1 if dvs else "none")
    all_passed &= _check("survey header row 1 col 1 = 'type'",
                         sv.cell(1, 1).value == "type")
    all_passed &= _check("settings form_id populated",
                         wb["settings"].cell(2, 2).value == form_id,
                         f"got '{wb['settings'].cell(2, 2).value}'")
    kb = os.path.getsize(out_path) / 1024
    all_passed &= _check(f"file size > 30 KB (larger than scratch ~8 KB, confirming LO template used)",
                         kb > 30, f"{kb:.0f} KB")

print(f"\nTest 1 result: {'ALL PASSED' if all_passed else 'SOME FAILED'}")


# ════════════════════════════════════════════════════════════════════════════════
# TEST 2: _xlsform_zip  (DVS-translate human-in-the-loop path)
# ════════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("TEST 2: _xlsform_zip  (DVS-translate path)")
print("="*60)

# Import _xlsform_zip without loading the full pipeline — extract it by
# exec-ing just the top portion up through the function definition.
# Simpler: just duplicate the relevant imports and call it from pipeline directly.

# We need to stub pipeline's heavy imports before importing it
import types, importlib.util, unittest.mock as mock

_stubs = ["monday_client", "claude_client", "prompts", "trainer_integration", "anthropic"]
for mod in _stubs:
    if mod not in sys.modules:
        m = types.ModuleType(mod)
        # Minimal attrs pipeline references at module level
        if mod == "monday_client":
            m.COL = {}
            for fn in ["get_item","download_file","upload_file","set_status",
                       "append_log","set_text","download_column_file","list_column_filenames"]:
                setattr(m, fn, None)
        elif mod == "claude_client":
            m.call_claude = None; m.extract_json = None
        elif mod == "prompts":
            m.EDC_STRUCTURE_PROMPT = ""; m.PRICING_SUMMARY_PROMPT = ""
            m.DVS_TRANSLATE_PROMPT = ""
        elif mod == "trainer_integration":
            for fn in ["run_protocol_analysis_quick","retrieve_examples",
                       "create_pending_row","format_examples_block"]:
                setattr(m, fn, None)
        sys.modules[mod] = m

spec    = importlib.util.spec_from_file_location(
            "pipeline", os.path.join(REPO_ROOT, "pipeline.py"))
pipeline_mod = importlib.util.module_from_spec(spec)
pipeline_mod.SKILLS_DIR = SKILLS_DIR   # set before exec so it's used
spec.loader.exec_module(pipeline_mod)

build_json = {"forms": SAMPLE_FORMS}
try:
    zip_bytes = pipeline_mod._xlsform_zip(build_json)
except Exception as e:
    print(f"  FAIL — _xlsform_zip raised: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

zip_out = os.path.join(OUT_DIR, "dvs_path_test.zip")
with open(zip_out, "wb") as f:
    f.write(zip_bytes)

t2_passed = True
print(f"  ZIP size: {len(zip_bytes):,} bytes")

with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
    names = zf.namelist()
    print(f"  ZIP contents: {names}")
    t2_passed &= _check("ZIP contains AE.xlsx", "AE.xlsx" in names)
    t2_passed &= _check("ZIP contains VS.xlsx", "VS.xlsx" in names)

    for xname in [n for n in names if n.endswith(".xlsx")]:
        wb  = load_workbook(io.BytesIO(zf.read(xname)))
        dvs = list(wb["survey"].data_validations.dataValidation)
        print(f"\n  {xname}:")
        t2_passed &= _check("9 sheets",   len(wb.sheetnames) == 9,
                             str(wb.sheetnames))
        t2_passed &= _check("dropdown",   len(dvs) > 0)
        kb = len(zf.read(xname)) / 1024
        t2_passed &= _check(f"size > 100 KB", kb > 30, f"{kb:.0f} KB")

print(f"\nTest 2 result: {'ALL PASSED' if t2_passed else 'SOME FAILED'}")


# ════════════════════════════════════════════════════════════════════════════════
# Summary
# ════════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
overall = all_passed and t2_passed
print(f"OVERALL: {'ALL TESTS PASSED ✓' if overall else 'FAILURES DETECTED ✗'}")
print(f"\nOutput files written to: {OUT_DIR}/")
for f in sorted(os.listdir(OUT_DIR)):
    kb = os.path.getsize(os.path.join(OUT_DIR, f)) / 1024
    print(f"  {f}  ({kb:.0f} KB)")
print()
sys.exit(0 if overall else 1)

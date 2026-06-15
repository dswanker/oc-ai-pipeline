from fastapi import FastAPI, Request, BackgroundTasks, HTTPException, Header
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
import hmac, hashlib, json, os, time, traceback, asyncio

# Auth manager — Chrome-extension session-capture flow
from auth_manager import (
    AuthManager,
    handle_session_upload,
    render_instructions_page,
)
from gmail_oauth import (
    build_auth_url,
    exchange_code_for_token,
    token_exists,
    delete_token,
    render_success_page,
    render_error_page,
    GmailAuthRequired,
)
from monday_client import COL, PIPELINE_CONFIG_ITEM_ID, download_column_file
from migration_pipeline import MIGRATIONS_HUB_COLUMNS

app = FastAPI()

# CORS — the Syndeo UI (mapping-ui) is a separate Railway service that
# fetches gap reports from this API via cross-origin XHR. Restricted to
# the known UI origins; webhook endpoints are server-to-server (Monday)
# and don't go through CORS preflight.
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://mapping-ui-production.up.railway.app",
        "http://localhost:3000",  # CRA dev
        "http://localhost:5173",  # Vite dev
    ],
    allow_credentials=False,
    allow_methods=["GET"],
    allow_headers=["*"],
)

# Add session middleware (required for OAuth state management)
app.add_middleware(
    SessionMiddleware,
    secret_key=os.environ["AUTH_SECRET_KEY"],
    max_age=3600  # 1 hour session lifetime
)

MONDAY_SIGNING_SECRET = os.environ.get("MONDAY_SIGNING_SECRET", "")

TRIGGER_COLUMN_ID    = "single_select5ogcb0g"
TRIGGER_LABEL_INDEX  = 0      # 0 = "Send to AI"
TRIGGER_LABEL_TEXT   = "Send to AI"

# New columns for OC study creation
CREATE_STUDY_CHECKBOX = "boolean_mm2nbn5c"   # "Would you like AI to create Study, SOE, and Form cards in OC4?"
PUBLISH_TEST_CHECKBOX = "boolean_mm3g2vzf"   # "Publish to Test" checkbox
LOAD_UAT_CHECKBOX     = "boolean_mm3gxe49"   # "Load UAT Test Data"

@app.get("/health")
async def health():
    return {"status": "ok"}


# ─────────────────────────────────────────────────────────────────────────────
# Syndeo UI — gap report fetch
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/gap-report/{item_id}")
async def get_gap_report(item_id: str):
    """Return the gap-analysis report JSON for a Migrations Hub row.

    Reads the file currently in the gap_report file column
    (file_mm3qcpnr) on the Migrations AI Hub board for the given item_id,
    downloads its bytes, parses as JSON, and returns to the browser.

    The Migrations Hub row is the long-lived per-study record produced
    by migration_pipeline.run_gap_analysis_and_hub_upsert — the file
    on this column is overwritten on every pipeline run for that study,
    so this endpoint always serves the latest version. Each report has
    a `generated_at` ISO timestamp inside it for client-side staleness
    checks.

    Errors:
      404 — column empty (pipeline hasn't run gap analysis yet, or
            item_id points at a row on a different board).
      500 — Monday API failure, or file present but not valid JSON.
    """
    try:
        blob = await download_column_file(
            item_id, MIGRATIONS_HUB_COLUMNS["gap_report"],
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Monday fetch failed for item {item_id}: {e}",
        )
    if not blob:
        raise HTTPException(
            status_code=404,
            detail=(f"No gap report uploaded on Migrations Hub item "
                    f"{item_id} — either the pipeline hasn't produced "
                    f"one yet, or this item is on a different board."),
        )
    try:
        report = json.loads(blob.decode("utf-8"))
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=(f"Gap report file on item {item_id} is not valid "
                    f"JSON: {e}"),
        )
    return JSONResponse(report)


# ─────────────────────────────────────────────────────────────────────────────
# Temporary admin endpoint — clear false-positive conflict OIDs from
# the per-item upload record (CRS-135 one-time fix, May 2026).
#
# DELETE THIS ROUTE once the operational backlog of stale records is
# cleared. Keeping a write-anywhere endpoint in production is a liability
# even when gated by a shared secret.
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/admin/clear-upload-record-oids")
async def clear_upload_record_oids(request: Request):
    """Remove specific OIDs from an item's upload record so the
    conflict detector stops flagging them on the next publish-to-test.

    Body  : {"item_id": "<numeric>", "oids": ["AE", "CM", ...]}
    Header: X-Admin-Secret must match the ADMIN_SECRET env var.

    Errors:
      503 — ADMIN_SECRET env var is not set. The endpoint refuses to
            run with an empty default so an unset env doesn't quietly
            authorise everyone.
      403 — secret header missing or mismatched.
      400 — item_id missing/non-numeric, or oids list empty.
      404 — no upload record file on disk for that item.
    """
    admin_secret = os.environ.get("ADMIN_SECRET", "")
    if not admin_secret:
        raise HTTPException(
            status_code=503,
            detail=(
                "ADMIN_SECRET env var not set — endpoint disabled. "
                "Set it on Railway before calling."
            ),
        )
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")

    body = await request.json()
    item_id = str(body.get("item_id", "")).strip()
    oids_to_remove = set(body.get("oids", []) or [])
    # item_id is interpolated into a filesystem path — keep it strictly
    # numeric to block path-traversal even though the secret check
    # already gates access.
    if not item_id or not item_id.isdigit():
        raise HTTPException(
            status_code=400,
            detail="item_id is required and must be numeric",
        )
    if not oids_to_remove:
        raise HTTPException(
            status_code=400,
            detail="oids must be a non-empty list",
        )

    path = f"/data/pipeline_upload_records/{item_id}.json"
    if not os.path.exists(path):
        raise HTTPException(
            status_code=404,
            detail=f"upload record not found: {path}",
        )

    with open(path) as f:
        rec = json.load(f)
    before = set(rec.get("uploaded_oids", []) or [])
    rec["uploaded_oids"] = sorted(before - oids_to_remove)
    # Conflict-detector store: pipeline.py reads
    # rec["forms"][oid]["pipeline_version_ids"] to decide whether an
    # OC version is pipeline-managed or human-edited. Clearing the
    # entry here lets the OID fall through the conflict detector's
    # "OID not in stored forms → unmanaged → upload fresh" branch on
    # the next run.
    if "forms" in rec and isinstance(rec["forms"], dict):
        for oid in oids_to_remove:
            rec["forms"].pop(oid, None)
    # Legacy key — older records used a flat top-level oc_version_ids
    # dict. Pop it too if present so old records don't carry stale
    # state forward after a clear.
    if "oc_version_ids" in rec and isinstance(rec["oc_version_ids"], dict):
        for oid in oids_to_remove:
            rec["oc_version_ids"].pop(oid, None)
    with open(path, "w") as f:
        json.dump(rec, f, indent=2)

    return {
        "item_id":   item_id,
        "removed":   sorted(oids_to_remove),
        "remaining": rec["uploaded_oids"],
        "forms_after": sorted((rec.get("forms") or {}).keys()),
    }


@app.post("/admin/full-reset")
async def full_reset(request: Request, body: dict):
    """Clear ALL output file columns + reset all status/text columns for a full re-run."""
    admin_secret = os.environ.get("ADMIN_SECRET", "oc-admin-2026")
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")
    item_id = str(body.get("item_id", ""))
    if not item_id:
        raise HTTPException(status_code=400, detail="item_id required")

    from monday_client import BOARD_ID, MONDAY_API_URL, get_headers, make_mutation, COL
    import httpx as _httpx

    results = {}
    async with _httpx.AsyncClient(timeout=30) as client:

        async def _set(col_id, val):
            r = await client.post(MONDAY_API_URL, headers=get_headers(), json={
                "query": make_mutation(),
                "variables": {"i": item_id, "b": BOARD_ID, "c": col_id, "v": val},
            })
            return r.status_code

        # ── Clear output file columns ────────────────────────────────────
        file_col_keys = [
            "spec_pdf", "spec_xlsx", "spec_json",
            "pricing_summary", "pricing_quote",
            "edc_build", "dvs_output", "calendaring_output", "build_preview",
        ]
        for key in file_col_keys:
            col_id = COL.get(key)
            if col_id:
                results[key] = await _set(col_id, "{\"files\": []}")

        # UAT DVS Results — not in COL dict
        results["dvs_uat_results"] = await _set("file_mm3h5s3h", "{\"files\": []}")

        # ── Reset pipeline status columns ────────────────────────────────
        results["pipeline_status"] = await _set(
            COL["pipeline_status"], '{"label": "Not Started"}')
        results["ai_trigger"] = await _set(
            COL["ai_trigger"], '{"label": "Do not Send To AI Yet"}')
        results["published_status"] = await _set(
            COL["published_status"], '{"label": "Not Published"}')

        # ── Clear text columns ───────────────────────────────────────────
        results["study_uuid"] = await _set(COL["study_uuid"], '""')
        results["study_oid"]  = await _set(COL["study_oid"],  '""')

    # ── Clear upload record on disk ──────────────────────────────────────
    from pipeline import _upload_record_path
    import json as _json
    _upload_record_path(item_id).write_text(
        _json.dumps({"study_uuid": "", "forms": {}, "uploaded_oids": []})
    )
    results["upload_record"] = "cleared"
    return results


@app.post("/admin/reset-upload-record")
async def reset_upload_record(request: Request):
    """Overwrite an item's upload record with a fresh empty record so the
    next pipeline run is treated as a first-ever run (no conflict history).

    Body  : {"item_id": "<numeric>", "study_uuid": "<uuid, optional>"}
    Header: X-Admin-Secret must match the ADMIN_SECRET env var.

    Writes (overwriting any existing record):
        {"study_uuid": <uuid or "">, "forms": {}, "uploaded_oids": []}

    Errors:
      503 — ADMIN_SECRET env var is not set (endpoint refuses to run so an
            unset env never authorises everyone).
      403 — secret header missing or mismatched.
      400 — item_id missing or non-numeric.
    """
    admin_secret = os.environ.get("ADMIN_SECRET", "")
    if not admin_secret:
        raise HTTPException(
            status_code=503,
            detail=(
                "ADMIN_SECRET env var not set — endpoint disabled. "
                "Set it on Railway before calling."
            ),
        )
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")

    body = await request.json()
    item_id = str(body.get("item_id", "")).strip()
    study_uuid = str(body.get("study_uuid", "")).strip()
    # item_id is interpolated into a filesystem path — keep it strictly
    # numeric to block path-traversal even though the secret gates access.
    if not item_id or not item_id.isdigit():
        raise HTTPException(
            status_code=400,
            detail="item_id is required and must be numeric",
        )

    record = {"study_uuid": study_uuid, "forms": {}, "uploaded_oids": []}
    os.makedirs("/data/pipeline_upload_records", exist_ok=True)
    path = f"/data/pipeline_upload_records/{item_id}.json"
    with open(path, "w") as f:
        json.dump(record, f, indent=2)

    return {"item_id": item_id, "path": path, "written": record}


@app.post("/admin/probe-board-fields")
async def probe_board_fields(request: Request):
    """Headless Playwright probe — open an OC4 designer board and return
    the Meteor `Boards` document shape (keys + selected UUID candidates).

    Body  : {"board_url": "https://…/b/…/…", "email": "user@host"}
    Header: X-Admin-Secret must match the ADMIN_SECRET env var.
    """
    admin_secret = os.environ.get("ADMIN_SECRET", "")
    if not admin_secret:
        raise HTTPException(status_code=503,
                            detail="ADMIN_SECRET env var not set")
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")

    body = await request.json()
    board_url = (body.get("board_url") or "").strip()
    email     = (body.get("email") or "").strip()
    if not board_url or not email:
        raise HTTPException(status_code=400,
                            detail="board_url and email are required")
    session_path = f"/data/browser_sessions/{email}.json"
    if not os.path.exists(session_path):
        raise HTTPException(status_code=404,
                            detail=f"session not found: {session_path}")

    from playwright.async_api import async_playwright
    _probe_js = """async () => {
    const board = Boards.findOne();
    if (!board) return {error: 'no board found'};
    const cards = Cards.find({boardId: board._id, archived: {$ne: true}}).fetch();
    const sample = cards.slice(0,5).map(c => ({title: c.title, formOcoid: c.formOcoid}));
    const allOcoids = [...new Set(cards.map(c => c.formOcoid).filter(Boolean))].sort();
    return {bucketUuid: board.bucketUuid, cardCount: cards.length, sample, allOcoids};
}"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            ctx  = await browser.new_context(storage_state=session_path)
            page = await ctx.new_page()
            await page.goto(board_url, wait_until="networkidle",
                            timeout=30000)
            await page.wait_for_timeout(5000)
            return await page.evaluate(_probe_js)
        finally:
            await browser.close()


@app.post("/admin/probe-form-service")
async def probe_form_service(request: Request):
    """Probe the OC4 form-service for a bucket's registered forms.

    Calls two candidate list endpoints with a fresh Bearer token from
    pipeline._get_oc_token (which already wraps the OAuth call against
    OC_API_USERNAME / OC_API_PASSWORD env vars) and returns the raw HTTP
    outcome of each so we can confirm which path the form-service exposes.

    Body  : {"bucket_uuid": "...", "subdomain": "...", "email": "..."}
            (email is accepted for shape consistency with
             /admin/probe-board-fields but is not used — the OAuth flow
             takes its credentials from env vars.)
    Header: X-Admin-Secret must match the ADMIN_SECRET env var.
    """
    admin_secret = os.environ.get("ADMIN_SECRET", "")
    if not admin_secret:
        raise HTTPException(status_code=503,
                            detail="ADMIN_SECRET env var not set")
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")

    body = await request.json()
    bucket_uuid = (body.get("bucket_uuid") or "").strip()
    subdomain   = (body.get("subdomain") or "").strip()
    if not bucket_uuid or not subdomain:
        raise HTTPException(status_code=400,
                            detail="bucket_uuid and subdomain are required")

    from pipeline import _get_oc_token
    import httpx as _httpx
    try:
        token = await _get_oc_token(subdomain, is_production=False)
    except Exception as e:
        raise HTTPException(status_code=500,
                            detail=f"OAuth token fetch failed: {e}")

    base = f"https://{subdomain}.build.openclinica.io/form-service/api"
    urls = [
        f"{base}/buckets/{bucket_uuid}/forms",
        f"{base}/forms?bucketUuid={bucket_uuid}",
    ]
    headers = {"Authorization": f"Bearer {token}"}
    results = []
    async with _httpx.AsyncClient(timeout=30) as c:
        for url in urls:
            try:
                r = await c.get(url, headers=headers)
                results.append({
                    "url": url,
                    "status": r.status_code,
                    "response_text_first_500_chars": r.text[:500],
                })
            except Exception as e:
                results.append({"url": url, "error": str(e)})
    return {
        "bucket_uuid": bucket_uuid,
        "subdomain":   subdomain,
        "results":     results,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Temporary diagnostic — slow-forms upload timing test
#
# Drives the upload sequence for the 7 OIDs that historically time out
# during normal publish runs (SLEEP, SF12, EX, AE, AESAE, CM, DV) and
# returns per-form timing + OC REST verification. Runs inside the
# Railway container where the SSO session JSON and the prebuilt xlsx
# files already live.
#
# DELETE THIS ROUTE once the slow-form upload timing is resolved.
# ─────────────────────────────────────────────────────────────────────────────

@app.delete("/admin/clear-session")
async def clear_session(
    request: Request,
    email: str = "dswanker@openclinica.com",
):
    """Delete the saved Playwright session so next run forces re-auth (captures EU cookies)."""
    admin_secret = os.environ.get("ADMIN_SECRET", "oc-admin-2026")
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")
    from pathlib import Path
    path = Path(f"/data/browser_sessions/{email}.json")
    if path.exists():
        path.unlink()
        return {"deleted": True, "path": str(path)}
    return {"deleted": False, "path": str(path), "reason": "file not found"}


@app.get("/admin/check-session")
async def check_session(request: Request, email: str = "dswanker@openclinica.com"):
    """Return metadata about the stored browser session file (no cookie values)."""
    admin_secret = os.environ.get("ADMIN_SECRET", "")
    if not admin_secret:
        raise HTTPException(status_code=503, detail="ADMIN_SECRET not set")
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")
    from pathlib import Path
    path = Path(f"/data/browser_sessions/{email}.json")
    if not path.exists():
        return {"exists": False, "path": str(path)}
    stat = path.stat()
    try:
        data = json.loads(path.read_text())
        cookies = data.get("cookies", [])
        origins = data.get("origins", [])
        ls_items = origins[0].get("localStorage", []) if origins else []
        ls_keys = [i.get("name") for i in ls_items]
        return {
            "exists": True,
            "path": str(path),
            "age_seconds": round(time.time() - stat.st_mtime),
            "cookie_count": len(cookies),
            "origin": origins[0].get("origin") if origins else None,
            "ls_key_count": len(ls_items),
            "ls_keys": ls_keys,
            "has_auth_token": any(
                i.get("name") in ("jhi-authenticationtoken", "jhi-idtoken")
                for i in ls_items
            ),
        }
    except Exception as e:
        return {"exists": True, "parse_error": str(e),
                "age_seconds": round(time.time() - stat.st_mtime)}


@app.get("/admin/sample-odm")
async def sample_odm(
    request: Request,
    item_id: str = "11894915699",
):
    """Generate and return the ODM XML that would be sent for a given item.
    Useful for sharing with OC engineering for debugging.
    Gated by X-Admin-Secret header.
    """
    admin_secret = os.environ.get("ADMIN_SECRET", "oc-admin-2026")
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")
    from uat_loader import (
        get_item, download_column_file, _parse_uat_cases,
        _build_odm_xml, COL
    )
    # Get item metadata
    item = await get_item(int(item_id))
    subdomain = item.get("oc_subdomain", "cust1")
    dvs_col   = COL.get("dvs_output", "file_mm2hhwmk")
    study_oid = item.get("study_oid", "S_CRS135_4530(TEST)")

    # Download DVS
    dvs_bytes = await download_column_file(int(item_id), dvs_col)
    if not dvs_bytes:
        raise HTTPException(status_code=404, detail="No DVS file found")

    # Parse UAT cases
    rows = _parse_uat_cases(dvs_bytes)
    groups = {}
    for row in rows:
        pid = row.get("Participant_ID", "UAT-P001")
        groups.setdefault(pid, []).append(row)

    # Build ODM for first participant only
    first_pid = next(iter(groups))
    first_rows = groups[first_pid]
    sample_site_oid  = "S_UAT_SAMPLE(TEST)"
    sample_p_key     = "UAT-SAMPLE-P001"
    odm_xml = _build_odm_xml(study_oid, sample_site_oid, sample_p_key, first_rows)

    from fastapi.responses import Response
    return Response(
        content=odm_xml,
        media_type="text/xml",
        headers={"Content-Disposition": "attachment; filename=sample_odm.xml"}
    )


@app.get("/admin/probe-oc-apis")
async def probe_oc_apis(
    request: Request,
    subdomain: str = "cust1",
):
    """Fetch OpenAPI docs for participant-service and data-service using a live token."""
    admin_secret = os.environ.get("ADMIN_SECRET", "oc-admin-2026")
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")
    from pipeline import _get_oc_token
    import httpx as _httpx
    token = await _get_oc_token(subdomain)
    base = f"https://{subdomain}.build.openclinica.io"
    results = {}
    async with _httpx.AsyncClient(timeout=15) as client:
        for svc in ["participant-service", "data-service"]:
            url = f"{base}/{svc}/v3/api-docs"
            r = await client.get(url, headers={"Authorization": f"Bearer {token}"})
            results[svc] = {"status": r.status_code, "body": r.text[:8000]}
        # Test Option B: does the EU clinical host accept the same Bearer token?
        import csv as _csv
        from pathlib import Path as _Path
        _csv_path = _Path(__file__).parent / "references" / "customer_uuids.csv"
        if _csv_path.exists():
            with open(_csv_path, newline="") as _f:
                for _row in _csv.DictReader(_f):
                    if _row.get("subdomain","").lower() == subdomain.lower():
                        bridge = _row.get("bridge_url","").rstrip("/")
                        if bridge:
                            # Try GET on ImportCRFData with Bearer token
                            eu_url = f"{bridge}/ImportCRFData"
                            r2 = await client.get(
                                eu_url,
                                headers={"Authorization": f"Bearer {token}"},
                                follow_redirects=False,
                            )
                            results["eu_bearer_test"] = {
                                "url": eu_url,
                                "status": r2.status_code,
                                "location": r2.headers.get("location", "n/a"),
                                "body_preview": r2.text[:200],
                            }
                        break
    return results


@app.post("/test/slow-forms")
async def test_slow_forms_endpoint(
    x_admin_secret: str = Header(None, alias="X-Admin-Secret"),
):
    """Run the slow-forms diagnostic and return the result dict.

    Gated by X-Admin-Secret header against the ADMIN_SECRET env var
    (default fallback "oc-admin-2026" so local invocations work
    without env wiring).

    Returns the dict from test_slow_forms.run_test() — see that
    function's docstring for the response shape. Per-form prints
    still flow to server stdout so Railway logs show live progress.
    """
    expected_secret = os.environ.get("ADMIN_SECRET", "oc-admin-2026")
    if x_admin_secret != expected_secret:
        raise HTTPException(status_code=403, detail="unauthorized")
    from test_slow_forms import run_test
    return await run_test()


# ─────────────────────────────────────────────────────────────────────────────
# Auth bootstrap (Chrome extension session-capture flow)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/auth")
async def auth_page(token: str = "", context: str = "pipeline",
                    clinical_host: str = ""):
    """Render the bootstrap instructions page for a one-time auth link.

    clinical_host is passed in the URL by pipeline.py when generating the
    auth link — it contains the actual subdomain for THIS study (e.g.
    'cust1.eu.openclinica.io'). We extract the subdomain from it so the
    instructions page shows the correct dynamic URL, not a hardcoded default.
    """
    if not token:
        return HTMLResponse("<h1>Missing token</h1>", status_code=400)
    am = AuthManager()
    email, error = am.validate_token(token)
    if error:
        return HTMLResponse(
            f"<h1>Auth link problem</h1><p>{error}</p>",
            status_code=400,
        )
    return HTMLResponse(render_instructions_page(token, email, context, clinical_host))


@app.post("/api/session/upload")
async def session_upload(request: Request):
    """Endpoint the OC Session Capture Chrome extension POSTs to.

    Body: {"token": "<signed token>", "storage_state": <playwright dict>}
    Response: {"ok": true, "email": ..., "cookies": N} on success;
              {"ok": false, "error": ...} on failure (HTTP 400).
    """
    data = await request.json()
    result = await handle_session_upload(
        data.get("token", ""),
        data.get("storage_state"),
    )
    status = result.pop("status", 200 if result.get("ok") else 400)
    return JSONResponse(result, status_code=status)


# ─────────────────────────────────────────────────────────────────────────────
# Extension proxy (serves the zipped Chrome extension stored in monday)
# ─────────────────────────────────────────────────────────────────────────────

# Module-level 5-min TTL cache so repeat downloads don't re-hit monday
# on every install. Resets on each Railway deploy.
_extension_zip_cache: tuple[bytes, float] | None = None
EXTENSION_CACHE_TTL_S = 300


async def _get_extension_zip_bytes() -> bytes:
    """Fetch the latest extension-zip bytes from monday (with TTL cache)."""
    global _extension_zip_cache
    now = time.time()
    if _extension_zip_cache is not None and now < _extension_zip_cache[1]:
        return _extension_zip_cache[0]
    blob = await download_column_file(
        PIPELINE_CONFIG_ITEM_ID, COL["pipeline_extension"],
    )
    if not blob:
        raise HTTPException(
            status_code=503,
            detail="Extension not uploaded to monday yet",
        )
    _extension_zip_cache = (blob, now + EXTENSION_CACHE_TTL_S)
    return blob


@app.get("/extension.zip")
async def extension_zip():
    """Serve the OC Session Capture Chrome extension (zip) to users.

    Authoritative source is the file column on monday's Pipeline
    Configuration row — drop a new zip there and it propagates within
    5 minutes (or immediately on the next Railway deploy).
    """
    blob = await _get_extension_zip_bytes()
    return Response(
        content=blob,
        media_type="application/zip",
        headers={
            "Content-Disposition":
                'attachment; filename="oc-session-capture.zip"',
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Monday.com Webhooks
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/webhook/monday")
async def monday_webhook(request: Request, background_tasks: BackgroundTasks):
    body    = await request.body()
    payload = json.loads(body)

    # Monday one-time verification challenge
    if "challenge" in payload:
        print("CHALLENGE received - responding", flush=True)
        return {"challenge": payload["challenge"]}

    event   = payload.get("event", {})
    item_id = str(event.get("pulseId", ""))
    col_id  = event.get("columnId", "")
    new_val = event.get("value", {})

    # Only care about our specific trigger column
    if col_id != TRIGGER_COLUMN_ID:
        print(f"IGNORED: column {col_id} is not the trigger column", flush=True)
        return {"status": "ignored"}

    # Extract the label index from the value
    if isinstance(new_val, str):
        try:
            new_val = json.loads(new_val)
        except:
            pass

    label_idx  = new_val.get("label", {}).get("index") if isinstance(new_val, dict) else None
    label_text = new_val.get("label", {}).get("text", "") if isinstance(new_val, dict) else ""

    print(f"COLUMN CHANGE: item={item_id} label_index={label_idx} label_text='{label_text}'", flush=True)

    # Only trigger when set to "Send to AI"
    if label_idx != TRIGGER_LABEL_INDEX:
        print(f"IGNORED: '{label_text}' is not '{TRIGGER_LABEL_TEXT}' - no action taken", flush=True)
        return {"status": f"ignored - changed to '{label_text}', only fires on '{TRIGGER_LABEL_TEXT}'"}

    # All checks passed - start the pipeline
    print(f"TRIGGERED: item {item_id} set to '{TRIGGER_LABEL_TEXT}' - starting pipeline", flush=True)

    async def safe_run_pipeline(iid):
        try:
            from pipeline import run_pipeline
            await run_pipeline(iid)
        except Exception as e:
            print(f"PIPELINE CRASHED: {e}", flush=True)
            print(traceback.format_exc(), flush=True)

    background_tasks.add_task(safe_run_pipeline, item_id)
    return {"status": "pipeline_started", "item_id": item_id}


@app.post("/webhook/create_study")
async def create_study_webhook(request: Request, background_tasks: BackgroundTasks):
    """
    Triggered when boolean_mm2nbn5c checkbox is checked.
    Creates OC study, imports board, then uploads forms to create versions.
    """
    body = await request.body()
    payload = json.loads(body)
    
    # Monday challenge response
    if "challenge" in payload:
        return {"challenge": payload["challenge"]}
    
    event = payload.get("event", {})
    item_id = str(event.get("pulseId", ""))
    col_id = event.get("columnId", "")
    new_val = event.get("value", {})
    
    # Only respond to CREATE_STUDY_CHECKBOX column
    if col_id != CREATE_STUDY_CHECKBOX:
        return {"status": "ignored"}
    
    # Parse checkbox value
    if isinstance(new_val, str):
        try:
            new_val = json.loads(new_val)
        except:
            pass
    
    # Check if checkbox is checked ({"checked": "true"})
    is_checked = new_val.get("checked") == "true" if isinstance(new_val, dict) else False
    
    if not is_checked:
        print(f"CREATE STUDY: Checkbox unchecked on item {item_id} - ignoring", flush=True)
        return {"status": "ignored - checkbox unchecked"}
    
    print(f"CREATE STUDY: Checkbox checked on item {item_id} - starting", flush=True)
    
    async def safe_create_study(iid):
        try:
            from oc_study_creator import create_oc_study_with_forms
            await create_oc_study_with_forms(iid)
        except Exception as e:
            print(f"CREATE STUDY CRASHED: {e}", flush=True)
            print(traceback.format_exc(), flush=True)
    
    background_tasks.add_task(safe_create_study, item_id)
    return {"status": "study_creation_started", "item_id": item_id}


@app.post("/webhook/publish_test")
async def publish_test_webhook(request: Request, background_tasks: BackgroundTasks):
    """
    Triggered when Publish to Test checkbox is checked.
    Waits for forms to be loaded, then publishes study to Test environment.
    """
    body = await request.body()
    payload = json.loads(body)
    
    # Monday challenge response
    if "challenge" in payload:
        return {"challenge": payload["challenge"]}
    
    event = payload.get("event", {})
    item_id = str(event.get("pulseId", ""))
    col_id = event.get("columnId", "")
    new_val = event.get("value", {})
    
    # Only respond to PUBLISH_TEST_CHECKBOX column
    if col_id != PUBLISH_TEST_CHECKBOX:
        return {"status": "ignored"}
    
    # Parse checkbox value
    if isinstance(new_val, str):
        try:
            new_val = json.loads(new_val)
        except:
            pass
    
    # Check if checkbox is checked
    is_checked = new_val.get("checked") == "true" if isinstance(new_val, dict) else False
    
    if not is_checked:
        print(f"PUBLISH TEST: Checkbox unchecked on item {item_id} - ignoring", flush=True)
        return {"status": "ignored - checkbox unchecked"}
    
    print(f"PUBLISH TEST: Checkbox checked on item {item_id} - starting", flush=True)
    
    async def safe_publish_test(iid):
        try:
            from oc_study_creator import publish_to_test_with_wait
            await publish_to_test_with_wait(iid)
        except Exception as e:
            print(f"PUBLISH TEST CRASHED: {e}", flush=True)
            print(traceback.format_exc(), flush=True)
    
    background_tasks.add_task(safe_publish_test, item_id)
    return {"status": "publish_test_started", "item_id": item_id}


@app.post("/webhook/load_uat")
async def load_uat_webhook(request: Request, background_tasks: BackgroundTasks):
    """
    Triggered when Load UAT Test Data checkbox is checked.
    Loads DVS-derived UAT test data into the published Test environment.
    NOTE: Not yet implemented — logs a clear message and returns.
    """
    body = await request.body()
    payload = json.loads(body)
    if "challenge" in payload:
        return {"challenge": payload["challenge"]}
    event = payload.get("event", {})
    item_id = str(event.get("pulseId", ""))
    col_id = event.get("columnId", "")
    new_val = event.get("value", {})
    if col_id != LOAD_UAT_CHECKBOX:
        return {"status": "ignored"}
    if isinstance(new_val, str):
        try:
            new_val = json.loads(new_val)
        except:
            pass
    is_checked = new_val.get("checked") == "true" if isinstance(new_val, dict) else False
    if not is_checked:
        return {"status": "ignored - checkbox unchecked"}
    print(f"LOAD UAT: Checkbox checked on item {item_id} — "
          f"UAT data loading not yet implemented.", flush=True)
    # TODO: implement UAT data loading via OC participant/data API
    return {"status": "load_uat_acknowledged_not_implemented",
            "item_id": item_id}


@app.post("/webhook/design-change")
async def design_change_webhook(request: Request,
                                 background_tasks: BackgroundTasks):
    """
    Receives Monday.com item update webhooks. Fires the design-change-intake
    skill when an update body starts with [DESIGN_CHANGE].

    Expected payload: Monday.com "When an update is created" event.
      event.pulseId  — item ID
      event.body     — update text (must start with [DESIGN_CHANGE])

    Optional inline metadata tags after the prefix:
      [SOURCE_TYPE:meeting_notes|email|transcript]
      [PROTOCOL:CRS-136]
    """
    body    = await request.body()
    payload = json.loads(body)

    if "challenge" in payload:
        return {"challenge": payload["challenge"]}

    event       = payload.get("event", {})
    item_id     = str(event.get("pulseId", ""))
    update_body = event.get("body", "")

    print(f"DESIGN_CHANGE_WEBHOOK: item={item_id} "
          f"body_preview='{update_body[:80]}'", flush=True)

    if not update_body.strip().startswith("[DESIGN_CHANGE]"):
        print("IGNORED: update does not start with [DESIGN_CHANGE]", flush=True)
        return {"status": "ignored"}

    text = update_body.strip()[len("[DESIGN_CHANGE]"):].strip()

    import re
    source_type   = "meeting_notes"
    protocol_hint = ""
    st_match = re.search(r"\[SOURCE_TYPE:([^\]]+)\]", text)
    ph_match = re.search(r"\[PROTOCOL:([^\]]+)\]", text)
    if st_match:
        source_type = st_match.group(1).strip()
        text = text.replace(st_match.group(0), "").strip()
    if ph_match:
        protocol_hint = ph_match.group(1).strip()
        text = text.replace(ph_match.group(0), "").strip()

    if not text:
        return {"status": "ignored - empty source text"}

    async def safe_run(iid, stype, stext, phint):
        try:
            from pipeline import run_design_change_intake
            await run_design_change_intake(iid, stype, stext, phint)
        except Exception as e:
            print(f"DESIGN_CHANGE_INTAKE CRASHED: {e}", flush=True)
            print(traceback.format_exc(), flush=True)

    background_tasks.add_task(safe_run, item_id, source_type, text,
                               protocol_hint)
    return {"status": "design_change_intake_started", "item_id": item_id}



@app.post("/admin/regen-dvs")
async def regen_dvs_route(request: Request, background_tasks: BackgroundTasks):
    """
    Re-generate the DVS XLSX from the cached spec JSON + EDC zip already on Monday,
    upload to dvs_output column, then run the UAT loader.
    Requires X-Admin-Secret header and item_id in body.
    """
    secret = request.headers.get("X-Admin-Secret", "")
    if secret != os.environ.get("ADMIN_SECRET", "oc-admin-2026"):
        return {"status": "unauthorized"}

    body_bytes = await request.body()
    try:
        payload = json.loads(body_bytes) if body_bytes else {}
    except Exception:
        payload = {}

    item_id = str(payload.get("item_id", ""))
    if not item_id:
        return {"status": "error", "detail": "item_id required"}

    async def safe_regen(iid):
        try:
            from monday_client import (
                COL, download_column_file, upload_file, get_item,
                set_status, append_log
            )
            from pipeline import run_dvs_xlsx
            import json as _json, zipfile as _zf, io as _io, tempfile as _tmp, os as _os

            await append_log(iid, "Regen DVS: starting...")
            await set_status(iid, COL["pipeline_status"], "DVS Running")

            # 1. Download spec JSON
            spec_bytes = await download_column_file(iid, COL["spec_json"])
            if not spec_bytes:
                await append_log(iid, "Regen DVS: ERROR — spec JSON not found in monday")
                await set_status(iid, COL["pipeline_status"], "Failed")
                return
            struct_json = _json.loads(spec_bytes)
            await append_log(iid, f"Regen DVS: spec JSON loaded ({len(spec_bytes)} bytes)")

            # 2. Download EDC zip and extract forms JSON
            edc_bytes = await download_column_file(iid, COL["edc_build"])
            if not edc_bytes:
                await append_log(iid, "Regen DVS: ERROR — EDC zip not found in monday")
                await set_status(iid, COL["pipeline_status"], "Failed")
                return
            import openpyxl as _opxl
            forms_json = {"forms": {}}
            with _zf.ZipFile(_io.BytesIO(edc_bytes)) as z:
                for name in z.namelist():
                    if name.endswith(".xlsx") and "/forms/" in name:
                        fname = _os.path.basename(name)
                        wb = _opxl.load_workbook(_io.BytesIO(z.read(name)),
                                                 read_only=True, data_only=True)
                        survey_rows = []
                        if "survey" in wb.sheetnames:
                            ws = wb["survey"]
                            rows = list(ws.iter_rows(values_only=True))
                            if rows:
                                headers = [str(h or "").strip() for h in rows[0]]
                                for r in rows[1:]:
                                    row_dict = {headers[i]: r[i] for i in range(len(headers))
                                                if i < len(r) and r[i] is not None}
                                    if row_dict:
                                        survey_rows.append(row_dict)
                        choice_rows = []
                        if "choices" in wb.sheetnames:
                            ws_c = wb["choices"]
                            c_rows = list(ws_c.iter_rows(values_only=True))
                            if c_rows:
                                c_hdrs = [str(h or "").strip() for h in c_rows[0]]
                                for r in c_rows[1:]:
                                    rd = {c_hdrs[i]: r[i] for i in range(len(c_hdrs))
                                          if i < len(r) and r[i] is not None}
                                    if rd:
                                        choice_rows.append(rd)
                        forms_json["forms"][fname] = {"survey": survey_rows, "choices": choice_rows}
            await append_log(iid, f"Regen DVS: EDC zip loaded, {len(forms_json['forms'])} forms")

            # 3. Regenerate DVS
            dvs_bytes = run_dvs_xlsx(struct_json, forms_json)
            if not dvs_bytes:
                await append_log(iid, "Regen DVS: ERROR — DVS generation failed")
                await set_status(iid, COL["pipeline_status"], "Failed")
                return

            # 4. Upload DVS with proper filename
            import datetime as _dt2
            _proto = (struct_json.get("study_meta") or {}).get("protocol_number", "Study")
            _proto = _proto.replace("/", "-").replace(" ", "_")
            _ts = _dt2.datetime.utcnow().strftime("%m%d.%H%M")
            fname = f"{_proto}_DVS_V{_ts}.xlsx"
            await upload_file(iid, COL["dvs_output"], fname, dvs_bytes)
            await append_log(iid, f"Regen DVS: uploaded {fname} ({len(dvs_bytes)} bytes)")

            # 5. Run UAT loader
            await append_log(iid, "Regen DVS: launching UAT loader...")
            await set_status(iid, COL["pipeline_status"], "Loading UAT Data")
            from uat_loader import run_uat_loader
            await run_uat_loader(iid)
            await set_status(iid, COL["pipeline_status"], "All Complete")

        except Exception as exc:
            import traceback
            print(f"REGEN_DVS_ERROR: {exc}\n{traceback.format_exc()}", flush=True)
            try:
                from pipeline import append_log, set_status
                await append_log(iid, f"Regen DVS: EXCEPTION — {exc}")
                await set_status(iid, COL["pipeline_status"], "Failed")
            except Exception:
                pass

    background_tasks.add_task(safe_regen, item_id)
    return {"status": "regen_dvs_started", "item_id": item_id}

@app.post("/admin/run-email-intake")
async def run_email_intake_route(request: Request,
                                  background_tasks: BackgroundTasks):
    """
    Hourly email polling trigger.
    Called by Monday.com automation every hour.
    Optional body: {"member_id": "12345678"} to run for one member only.
    Requires X-Admin-Secret header.
    """
    secret = request.headers.get("X-Admin-Secret", "")
    if secret != os.environ.get("ADMIN_SECRET", ""):
        return {"status": "unauthorized"}

    body_bytes = await request.body()
    try:
        payload = json.loads(body_bytes) if body_bytes else {}
    except Exception:
        payload = {}

    member_id = payload.get("member_id")

    async def safe_run(mid):
        try:
            from pipeline import run_email_change_intake
            result = await run_email_change_intake(mid)
            print(f"EMAIL_INTAKE_COMPLETE: {result}", flush=True)
        except Exception as e:
            print(f"EMAIL_INTAKE_CRASHED: {e}", flush=True)
            print(traceback.format_exc(), flush=True)

    background_tasks.add_task(safe_run, member_id)
    return {"status": "email_intake_started",
            "member_id": member_id or "all"}


@app.post("/webhook/email-change-decision")
async def email_change_decision_webhook(request: Request,
                                         background_tasks: BackgroundTasks):
    """
    Fires when Review Decision column changes on Change Requests board.
    Monday.com webhook payload: event.pulseId, event.columnId,
    event.value.label.text
    Routes Approve → post [DESIGN_CHANGE] to AI Hub.
    Routes Dismiss → close item with no action.
    """
    body_bytes = await request.body()
    payload    = json.loads(body_bytes)

    if "challenge" in payload:
        return {"challenge": payload["challenge"]}

    event    = payload.get("event", {})
    item_id  = str(event.get("pulseId", ""))
    col_id   = event.get("columnId", "")
    label    = (event.get("value", {})
                     .get("label", {})
                     .get("text", ""))

    print(f"EMAIL_DECISION_WEBHOOK: item={item_id} col={col_id} "
          f"label={label}", flush=True)

    if col_id != "color_mm3zkh2y":
        return {"status": "ignored - wrong column"}

    if label not in ("Approve", "Dismiss"):
        return {"status": "ignored - not an actionable label"}

    async def safe_handle(iid, lbl):
        try:
            from pipeline import handle_email_review_decision
            result = await handle_email_review_decision(iid, lbl)
            print(f"EMAIL_DECISION_RESULT: {result}", flush=True)
        except Exception as e:
            print(f"EMAIL_DECISION_CRASHED: {e}", flush=True)
            print(traceback.format_exc(), flush=True)

    background_tasks.add_task(safe_handle, item_id, label)
    return {"status": "email_decision_processing",
            "item_id": item_id, "decision": label}


@app.get("/auth/gmail/{monday_user_id}")
async def gmail_auth_start(monday_user_id: str):
    """
    Step 1 of Gmail OAuth flow.
    Team member clicks the link from their bell notification.
    Redirects to Google consent screen requesting gmail.readonly scope.
    """
    from starlette.responses import RedirectResponse
    auth_url = build_auth_url(monday_user_id)
    print(f"GMAIL_AUTH: redirecting {monday_user_id} to Google consent",
          flush=True)
    return RedirectResponse(url=auth_url)


@app.get("/auth/gmail/callback")
async def gmail_auth_callback(code: str = "", state: str = "",
                               error: str = ""):
    """
    Step 2 of Gmail OAuth flow — Google redirects here after consent.
    Exchanges the authorisation code for access + refresh tokens,
    saves to /data/gmail_sessions/{monday_user_id}.json, shows
    success/error page.
    """
    from gmail_oauth import _verify_state

    if error:
        print(f"GMAIL_AUTH_CALLBACK: Google returned error: {error}",
              flush=True)
        return HTMLResponse(
            render_error_page(
                f"Google returned an error: {error}. "
                "Please try again or contact Dan."
            )
        )

    if not code or not state:
        return HTMLResponse(
            render_error_page("Missing code or state parameter."))

    monday_user_id, state_error = _verify_state(state)
    if state_error:
        print(f"GMAIL_AUTH_CALLBACK: state error: {state_error}",
              flush=True)
        return HTMLResponse(render_error_page(state_error))

    token_data = await exchange_code_for_token(code, monday_user_id)
    if not token_data:
        return HTMLResponse(
            render_error_page(
                "Failed to exchange authorisation code for token. "
                "Please try again or contact Dan."
            )
        )

    # Look up the team member's name for the success page
    member_name = monday_user_id
    try:
        import httpx as _httpx
        from monday_client import get_headers, MONDAY_API_URL
        q = f"query {{ users(ids: [{monday_user_id}]) {{ name }} }}"
        async with _httpx.AsyncClient(timeout=10) as _c:
            _r = await _c.post(MONDAY_API_URL, headers=get_headers(),
                               json={"query": q})
        users = _r.json().get("data", {}).get("users", [])
        if users:
            member_name = users[0].get("name", monday_user_id)
    except Exception:
        pass

    gmail_address = token_data.get("gmail_address", "")
    print(f"GMAIL_AUTH_CALLBACK: success for {monday_user_id} "
          f"({gmail_address})", flush=True)
    return HTMLResponse(render_success_page(gmail_address, member_name))


@app.get("/auth/gmail/status/{monday_user_id}")
async def gmail_auth_status(monday_user_id: str, request: Request):
    """
    Check Gmail connection status for a team member.
    Requires X-Admin-Secret header.
    Returns: {connected: bool, gmail_address: str, expires_at: int}
    """
    secret = request.headers.get("X-Admin-Secret", "")
    if secret != os.environ.get("ADMIN_SECRET", ""):
        raise HTTPException(status_code=403, detail="unauthorized")

    from gmail_oauth import load_token
    token = load_token(monday_user_id)
    if not token:
        return {"connected": False, "monday_user_id": monday_user_id}

    import time
    obtained_at = token.get("obtained_at", 0)
    expires_in  = token.get("expires_in", 3600)
    return {
        "connected":       True,
        "monday_user_id":  monday_user_id,
        "gmail_address":   token.get("gmail_address", ""),
        "expires_at":      obtained_at + expires_in,
        "has_refresh":     bool(token.get("refresh_token")),
    }


@app.post("/admin/regenerate-dvs")
async def regenerate_dvs(request: Request):
    """Regenerate the DVS XLSX for an item without rerunning the full pipeline.

    Reads the existing spec_json + edc_build zip from Monday, runs the DVS
    extract + build (same path as Chain C's run_dvs_xlsx in pipeline.py),
    uploads the result to dvs_output, and appends a Monday log line.

    Body  : {"item_id": "<numeric>"}
    Header: X-Admin-Secret must match the ADMIN_SECRET env var.

    Errors:
      503 — ADMIN_SECRET env var is not set.
      403 — secret header missing or mismatched.
      400 — item_id missing/non-numeric or spec_json not valid JSON.
      404 — spec_json or edc_build column is empty on the item.
      500 — DVS scripts not loadable.
    """
    admin_secret = os.environ.get("ADMIN_SECRET", "")
    if not admin_secret:
        raise HTTPException(
            status_code=503,
            detail=(
                "ADMIN_SECRET env var not set — endpoint disabled. "
                "Set it on Railway before calling."
            ),
        )
    if request.headers.get("X-Admin-Secret", "") != admin_secret:
        raise HTTPException(status_code=403, detail="unauthorized")

    body = await request.json()
    item_id = str(body.get("item_id", "")).strip()
    if not item_id or not item_id.isdigit():
        raise HTTPException(
            status_code=400,
            detail="item_id is required and must be numeric",
        )

    from monday_client import upload_file, append_log

    # ── Download spec JSON ────────────────────────────────────────────────
    spec_bytes = await download_column_file(item_id, COL["spec_json"])
    if not spec_bytes:
        raise HTTPException(
            status_code=404,
            detail="spec_json (file_mm2gefht) is empty on this item — "
                   "run the pipeline first to populate it",
        )
    try:
        struct_json = json.loads(spec_bytes.decode("utf-8"))
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"spec_json is not valid JSON: {e}",
        )

    # ── Download EDC build ZIP ────────────────────────────────────────────
    edc_zip_bytes = await download_column_file(item_id, COL["edc_build"])
    if not edc_zip_bytes:
        raise HTTPException(
            status_code=404,
            detail="edc_build (file_mm2h51qw) is empty on this item — "
                   "run the pipeline first to produce it",
        )

    # ── Build forms_json by reading 'survey' sheet of each xlsx in zip ───
    # Mirrors run_edc_build's forms_json builder (pipeline.py:646-664).
    import io as _io
    import zipfile as _zipfile
    import openpyxl as _openpyxl

    try:
        zf = _zipfile.ZipFile(_io.BytesIO(edc_zip_bytes))
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"edc_build is not a valid zip file: {e}",
        )

    forms_json = {"forms": {}}
    for name in sorted(zf.namelist()):
        if not name.lower().endswith(".xlsx"):
            continue
        try:
            with zf.open(name) as f:
                wb_bytes = f.read()
            wb = _openpyxl.load_workbook(_io.BytesIO(wb_bytes),
                                          read_only=True, data_only=True)
            survey_rows = []
            if "survey" in wb.sheetnames:
                ws = wb["survey"]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [str(h or "").strip() for h in rows[0]]
                    for r in rows[1:]:
                        row_dict = {headers[i]: r[i]
                                    for i in range(len(headers))
                                    if i < len(r) and r[i] is not None}
                        if row_dict:
                            survey_rows.append(row_dict)
            choice_rows = []
            if "choices" in wb.sheetnames:
                ws = wb["choices"]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [str(h or "").strip() for h in rows[0]]
                    for r in rows[1:]:
                        row_dict = {headers[i]: r[i]
                                    for i in range(len(headers))
                                    if i < len(r) and r[i] is not None}
                        if row_dict:
                            choice_rows.append(row_dict)
            forms_json["forms"][os.path.basename(name)] = {
                "survey":  survey_rows,
                "choices": choice_rows,
            }
        except Exception as e:
            print(f"[regenerate-dvs] skipping {name}: {e}", flush=True)

    # ── Run DVS extract + build (mirrors run_dvs_xlsx in pipeline.py) ────
    import sys as _sys
    _sys.path.insert(0, os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "skills", "dvs-specification", "scripts"))
    import importlib
    for _mod in list(_sys.modules.keys()):
        if 'extract_dvs_from_forms' in _mod or 'generate_dvs' in _mod:
            del _sys.modules[_mod]
    try:
        from extract_dvs_from_forms import extract_dvs_data
        from generate_dvs import build_dvs
    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"DVS scripts unavailable: {e}",
        )

    print(f"[regenerate-dvs] forms_json keys sample: {list(forms_json['forms'].keys())[:5]}", flush=True)
    print(f"[regenerate-dvs] struct_json forms sample: {[(f.get('form_id'), f.get('visits_assigned')) for f in struct_json.get('forms', [])[:3]]}", flush=True)
    from extract_dvs_from_forms import _build_form_event_map
    _test_map = _build_form_event_map(struct_json)
    print(f'[regenerate-dvs] form_event_map size={len(_test_map)} sample={list(_test_map.items())[:3]}', flush=True)
    dvs_data = extract_dvs_data(struct_json, forms_json)
    sample_uat = dvs_data.get('uat_cases', [])[:3]
    print(f"[regenerate-dvs] UAT sample Study_Event_OID: {[r.get('Study_Event_OID') for r in sample_uat]}", flush=True)
    n_uat = len(dvs_data.get("uat_cases", []))
    print(f"[regenerate-dvs] {len(dvs_data.get('dvs_oc4', []))} checks, "
          f"{len(dvs_data.get('query_text_library', []))} unique messages, "
          f"{n_uat} UAT cases", flush=True)

    protocol = (struct_json.get("study_meta", {}).get("protocol_number")
                or "STUDY")

    import tempfile as _tempfile
    with _tempfile.TemporaryDirectory() as tmp:
        xlsx_path = os.path.join(tmp, f"{protocol}_DVS.xlsx")
        build_dvs(dvs_data, xlsx_path)
        with open(xlsx_path, "rb") as f:
            dvs_bytes = f.read()

    # ── Clear existing dvs_output column so the new file replaces ────────
    # the old one rather than being appended alongside it.
    from monday_client import (MONDAY_API_URL, BOARD_ID, get_headers)
    import httpx as _httpx
    _clear_mutation = (
        "mutation($i: ID!, $b: ID!, $c: String!, $v: JSON!) {"
        "  change_column_value(item_id: $i, board_id: $b, "
        "                       column_id: $c, value: $v) { id }"
        "}"
    )
    async with _httpx.AsyncClient(timeout=30) as _c:
        await _c.post(
            MONDAY_API_URL,
            headers=get_headers(),
            json={
                "query": _clear_mutation,
                "variables": {
                    "i": item_id,
                    "b": BOARD_ID,
                    "c": COL["dvs_output"],
                    "v": '{"files": []}',
                },
            },
        )

    # ── Upload + log ──────────────────────────────────────────────────────
    await upload_file(item_id, COL["dvs_output"],
                      f"{protocol}_DVS.xlsx", dvs_bytes)
    await append_log(item_id, "DVS regenerated via /admin/regenerate-dvs")

    return {"status": "ok", "dvs_rows": n_uat}


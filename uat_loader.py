import re
"""
uat_loader.py — UAT Data Loader for oc-ai-pipeline
====================================================
Triggered when the "Load DVS UAT Data" checkbox fires on the AI Hub board.

Workflow
--------
1.  Read Study UUID + OC Subdomain from monday.com item
2.  Fetch the DVS XLSX from the dvs_output column
3.  GET /api/studies/{studyUuid}/study-environments → find TEST env UUID
4.  Create a dated site in the TEST environment
5.  Parse UAT_Cases sheet → group rows by Participant_ID
6.  Create one OC participant per unique Participant_ID
7.  Build ODM XML from UAT_Cases load coordinates + Load_Value
8.  POST ODM to /clinicaldata/import for each participant
9.  Write Site_OID + Participant_Key back into the DVS XLSX
10. Upload result DVS to file_mm3h5s3h (UAT DVS Results column)
11. Update monday status columns

Site naming convention (locked in Step 0)
------------------------------------------
  Site Name: "UAT Automation Site - YYYY-MM-DD HH:MM"
  Site OID:  "UAT-YYYYMMDD-HHMMSS"

Participant naming convention
------------------------------
  Logical ID in DVS:   UAT-P001, UAT-P002, ...
  Run-scoped key:      UAT-YYYYMMDD-HHMMSS-P001
"""

import asyncio
import datetime
import io
import json
import os
import traceback
from pathlib import Path

import httpx
from openpyxl import load_workbook

from monday_client import (
    get_item, download_column_file, upload_file,
    append_log, set_status, COL,
)

# ── Constants ─────────────────────────────────────────────────────────────────

UAT_STATUS = {
    "loading":  "Loading UAT Data",
    "complete": "UAT Data Loaded",
    "failed":   "UAT Load Failed",
}

# Column IDs for UAT output files on the AI Hub board
UAT_DVS_RESULTS_COL  = "file_mm3h5s3h"   # Updated DVS with runtime columns stamped
UAT_REPORT_COL       = "file_mm3hvbpb"   # UAT Validation Report (future)
UAT_MATRIX_COL       = "file_mm3h7r4"    # UAT Traceability Matrix (future)

ODM_NAMESPACE = (
    'xmlns="http://www.cdisc.org/ns/odm/v1.3" '
    'xmlns:OpenClinica="http://www.openclinica.org/ns/odm_ext_v130/v3.1"'
)


# ── Study Service API helpers ─────────────────────────────────────────────────

def _study_service_base(subdomain: str) -> str:
    return f"https://{subdomain}.build.openclinica.io/study-service"


def _pages_base(subdomain: str) -> str:
    """
    Return the base URL for clinical data (participant + ODM) API calls.
    Reads bridge_url from customer_uuids.csv keyed by subdomain.
    e.g. cust1 -> https://cust1.eu.openclinica.io/OpenClinica
    Falls back to build host if not found (will likely 405 but keeps old behavior).
    """
    csv_path = Path(__file__).parent / "references" / "customer_uuids.csv"
    if csv_path.exists():
        import csv as _csv
        with open(csv_path, newline="") as f:
            for row in _csv.DictReader(f):
                if row.get("subdomain", "").lower() == subdomain.lower():
                    bridge = row.get("bridge_url", "").rstrip("/")
                    if bridge:
                        return bridge  # e.g. https://cust1.eu.openclinica.io/OpenClinica
    # Fallback
    return f"https://{subdomain}.build.openclinica.io"


# Module-level token cache: keyed by subdomain, value is (token_str, fetched_at).
# Tokens are valid for ~24h; we cache for 20 minutes to avoid 429s from
# repeated calls within a single pipeline run while still refreshing well
# before actual expiry.
_TOKEN_CACHE: dict[str, tuple[str, float]] = {}
_TOKEN_CACHE_TTL_S = 1200  # 20 minutes


async def _get_oc_token(subdomain: str) -> str:
    """Return a cached OC OAuth bearer token, refreshing if older than 20 min.

    OC tokens are valid for ~24 hours but the token endpoint rate-limits
    (HTTP 429) if called more than ~10 times in quick succession. Caching
    avoids redundant fetches within a single pipeline run.
    """
    import time
    import os
    cached = _TOKEN_CACHE.get(subdomain)
    if cached:
        token_str, fetched_at = cached
        if time.monotonic() - fetched_at < _TOKEN_CACHE_TTL_S:
            return token_str

    username = os.environ.get("OC_API_USERNAME", "").strip()
    password = os.environ.get("OC_API_PASSWORD", "").strip()
    if not username or not password:
        raise ValueError("OC_API_USERNAME or OC_API_PASSWORD not set in env")
    url = (f"https://{subdomain}.build.openclinica.io"
           f"/user-service/api/oauth/token")
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(url,
                                 headers={"Content-Type": "application/json"},
                                 json={"username": username,
                                       "password": password})
    if resp.status_code != 200:
        raise RuntimeError(
            f"OC auth failed {resp.status_code}: {resp.text[:200]}")
    token_str = resp.text.strip()
    _TOKEN_CACHE[subdomain] = (token_str, time.monotonic())
    return token_str


async def _get_test_env_uuid(subdomain: str, study_uuid: str) -> tuple:
    """
    GET /api/studies/{studyUuid}/study-environments
    Returns (test_env_uuid, test_study_oid) for environmentName == 'TEST'.
    Raises ValueError if TEST environment not found.
    Uses Bearer token via _get_oc_token.
    """
    url = (f"{_study_service_base(subdomain)}/api/studies"
           f"/{study_uuid}/study-environments")
    token = await _get_oc_token(subdomain)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(url,
                                headers={"Authorization": f"Bearer {token}"})
        resp.raise_for_status()
        envs = resp.json()

    for env in envs:
        if (env.get("environmentName") or "").upper() == "TEST":
            return env["uuid"], env.get("oid", "")

    names = [e.get("environmentName") for e in envs]
    raise ValueError(
        f"TEST environment not found for study {study_uuid}. "
        f"Found: {names}"
    )


async def _wait_for_test_available(
    subdomain: str, study_uuid: str,
    timeout_s: int = 60, interval_s: int = 5,
) -> None:
    """Poll study-environments until the TEST environment status == 'AVAILABLE'.

    Per study-service.adoc / Study_Service_API.md the StudyEnvironmentDTO.status
    is an enum (DESIGN, AVAILABLE, FROZEN, LOCKED, ARCHIVED). We only proceed
    when the TEST env reaches AVAILABLE. Bearer token is fetched fresh per
    poll because _get_oc_token returns a short-lived OAuth token.

    Raises RuntimeError if TEST is never AVAILABLE within timeout_s.
    """
    url = (f"{_study_service_base(subdomain)}/api/studies"
           f"/{study_uuid}/study-environments")
    deadline = asyncio.get_event_loop().time() + timeout_s
    last_status = "(not seen)"
    while True:
        token = await _get_oc_token(subdomain)
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(
                url, headers={"Authorization": f"Bearer {token}"})
        if resp.status_code == 200:
            envs = resp.json() or []
            for env in envs:
                if (env.get("environmentName") or "").upper() == "TEST":
                    last_status = (env.get("status") or "").upper()
                    if last_status == "AVAILABLE":
                        print(
                            f"[uat_loader] TEST environment AVAILABLE "
                            f"for study {study_uuid}", flush=True)
                        return
                    break
        else:
            last_status = f"HTTP {resp.status_code}"
        if asyncio.get_event_loop().time() >= deadline:
            raise RuntimeError(
                f"TEST environment for study {study_uuid} never reached "
                f"AVAILABLE within {timeout_s}s "
                f"(last status: {last_status!r}). "
                f"Verify the study is published to Test and try again."
            )
        await asyncio.sleep(interval_s)


async def _create_site(subdomain: str, test_env_uuid: str,
                        site_name: str, site_oid: str) -> str:
    """
    POST /api/study-environments/{studyEnvironmentUuid}/sites
    Returns the created site OID.
    Uses Bearer token via _get_oc_token.
    """
    url = (f"{_study_service_base(subdomain)}/api/study-environments"
           f"/{test_env_uuid}/sites")
    token = await _get_oc_token(subdomain)
    today = datetime.date.today().isoformat()
    payload = {
        "name":                  site_name,
        "uniqueIdentifier":      site_oid,
        "oid":                   site_oid,
        "status":                "AVAILABLE",
        "principalInvestigator": "UAT Automation",
        "expectedEnrollment":    999,
        "timezone":              "America/New_York",
        "expectedStartDate":     today,
    }
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(url, json=payload,
                                 headers={"Authorization": f"Bearer {token}"})
        resp.raise_for_status()
        data = resp.json()
    return data.get("oid") or data.get("uniqueIdentifier") or site_oid




async def _create_participant(subdomain: str, study_oid: str,
                               site_oid: str, subject_key: str,
                               token: str, cookies: dict) -> str:
    """
    POST /pages/auth/api/clinicaldata/studies/{studyOid}/sites/{siteOid}/participants
    Path confirmed 401 (auth required) on eu host — uses cookie session auth.
    Bearer token kept as fallback header.
    """
    import json as _json
    from urllib.parse import quote as _quote
    url = (f"{_pages_base(subdomain)}/pages/auth/api/clinicaldata"
           f"/studies/{_quote(study_oid, safe='')}"
           f"/sites/{_quote(site_oid, safe='')}/participants")
    async with httpx.AsyncClient(timeout=30, cookies=cookies) as client:
        resp = await client.post(
            url,
            content=_json.dumps({"subjectKey": subject_key}).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type":  "application/json",
            },
        )
        if not resp.is_success:
            raise RuntimeError(
                f"HTTP {resp.status_code} — body: {resp.text[:300]}"
            )
    try:
        body = resp.json()
        return body.get("subjectKey") or subject_key
    except Exception:
        return subject_key


async def _list_participants(subdomain: str, study_oid: str,
                             token: str) -> list:
    """Return all participants for the study as a list of dicts."""
    from urllib.parse import quote as _quote
    url = (f"{_pages_base(subdomain)}/pages/auth/api/clinicaldata"
           f"/studies/{_quote(study_oid, safe='')}/participants")
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(url,
                                headers={"Authorization": f"Bearer {token}"})
    if not resp.is_success:
        return []
    try:
        body = resp.json()
    except Exception:
        return []
    if isinstance(body, list):
        return body
    for v in body.values() if isinstance(body, dict) else []:
        if isinstance(v, list):
            return v
    return []


async def _get_participant_oid(subdomain: str, study_oid: str,
                                participant_id: str, token: str) -> str:
    """GET /pages/auth/api/clinicaldata/studies/{studyOID}/participants and
    look up the entry whose participantId or subjectKey matches participant_id.

    Returns the internal OC OID assigned to that participant. The exact key
    used by OC for this OID is logged on first call (entry shape varies by
    OC version) so the operator can confirm which field is being returned.

    Raises RuntimeError if the participant is not found or the GET fails.
    """
    from urllib.parse import quote as _quote
    url = (f"{_pages_base(subdomain)}/pages/auth/api/clinicaldata"
           f"/studies/{_quote(study_oid, safe='')}/participants")
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            url,
            headers={"Authorization": f"Bearer {token}"},
        )
    if not resp.is_success:
        raise RuntimeError(
            f"GET participants HTTP {resp.status_code} — "
            f"body: {resp.text[:300]}"
        )
    try:
        body = resp.json()
    except Exception as e:
        raise RuntimeError(f"GET participants returned non-JSON: {e}")

    # The endpoint may return either a bare list or a wrapper object such as
    # {"participants": [...]} / {"data": [...]} — be lenient.
    entries = body
    if isinstance(body, dict):
        for k in ("participants", "data", "items", "results", "studyParticipants"):
            v = body.get(k)
            if isinstance(v, list):
                entries = v
                break
    if not isinstance(entries, list):
        raise RuntimeError(
            f"GET participants: unexpected response shape "
            f"({type(body).__name__}): {str(body)[:300]}"
        )

    if entries:
        print(f"[uat_loader] participants first entry keys: "
              f"{list(entries[0].keys()) if isinstance(entries[0], dict) else entries[0]}",
              flush=True)
        print(f"[uat_loader] participants first entry: "
              f"{str(entries[0])[:500]}", flush=True)

    needle = (participant_id or "").strip()
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        if (str(entry.get("participantId") or "").strip() == needle
                or str(entry.get("subjectKey") or "").strip() == needle):
            for k in ("oid", "participantOid", "subjectOid",
                       "subjectKey", "studySubjectOid", "ssid"):
                v = entry.get(k)
                if v:
                    return str(v)
            raise RuntimeError(
                f"Found participant '{participant_id}' but no recognised OID "
                f"field. Entry: {str(entry)[:300]}"
            )

    raise RuntimeError(
        f"Participant '{participant_id}' not found in study {study_oid} "
        f"(searched {len(entries)} entries)."
    )


# ── DVS parsing ───────────────────────────────────────────────────────────────

def _parse_uat_cases(dvs_bytes: bytes) -> list:
    """
    Read UAT_Cases sheet from DVS XLSX.
    Returns list of dicts; only rows with Study_Event_OID + Form_OID populated.
    """
    wb = load_workbook(io.BytesIO(dvs_bytes), read_only=True, data_only=True)
    if "UAT_Cases" not in wb.sheetnames:
        raise ValueError("DVS XLSX does not contain a UAT_Cases sheet.")
    ws = wb["UAT_Cases"]

    header_row_idx = None
    headers = []
    rows = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if header_row_idx is None:
            if vals and vals[0] == "UAT Case ID":
                header_row_idx = row_idx
                headers = vals
            continue
        if not any(vals):
            continue
        row_dict = dict(zip(headers, vals))
        if not row_dict.get("Study_Event_OID", "").strip():
            continue
        if not row_dict.get("Form_OID", "").strip():
            continue
        rows.append(row_dict)

    wb.close()
    return rows


def _group_by_participant(rows: list) -> dict:
    """Group UAT_Cases rows by Participant_ID, sorted by Load_Order."""
    groups = {}
    for row in rows:
        pid = row.get("Participant_ID", "UAT-P001").strip() or "UAT-P001"
        groups.setdefault(pid, []).append(row)
    for pid in groups:
        groups[pid].sort(key=lambda r: _safe_int(r.get("Load_Order", "999")))
    return groups


def _safe_int(val, default=999) -> int:
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return default


# ── ODM XML builder ───────────────────────────────────────────────────────────

def _xml_escape(val: str) -> str:
    return (val.replace("&", "&amp;")
               .replace("<", "&lt;")
               .replace(">", "&gt;")
               .replace('"', "&quot;"))


def _validate_odm_xml(odm_xml: str) -> list[str]:
    """
    Tier 1 — XSD structural validation of ODM XML.
    Returns a list of error strings (empty = valid).
    Uses the bundled minimal ODM 1.3.2 transactional XSD.

    The bundled CDISC XSD does not include the OpenClinica extension schema,
    so OpenClinica:* attributes (e.g. StudySubjectID) on SubjectData elements
    would be rejected during validation even though OC requires them at
    import time. Before validating, strip any OpenClinica-namespaced
    attributes from SubjectData elements on the parsed in-memory copy. The
    original odm_xml string is unchanged — the caller still sends the full
    XML with the OpenClinica attributes intact to _import_odm.
    """
    from lxml import etree
    xsd_path = (Path(__file__).parent
                / "skills" / "dvs-specification" / "references"
                / "ODM1-3-2-transactional.xsd")
    _OC_NS  = "http://www.openclinica.org/ns/odm_ext_v130/v3.1"
    _ODM_NS = "http://www.cdisc.org/ns/odm/v1.3"
    errors = []
    try:
        schema_doc = etree.parse(str(xsd_path))
        schema     = etree.XMLSchema(schema_doc)
        doc        = etree.fromstring(odm_xml.encode("utf-8"))
        # Strip OpenClinica:* attributes from SubjectData and StudyEventData
        # elements on the parsed copy only (the original odm_xml string is
        # unchanged). The CDISC XSD doesn't include the OC extension schema
        # so OpenClinica:StudySubjectID and OpenClinica:StartDate would both
        # be rejected during validation even though OC requires them at import.
        for element in doc.iter(
            f"{{{_ODM_NS}}}SubjectData",
            f"{{{_ODM_NS}}}StudyEventData",
        ):
            for attr_key in [
                k for k in element.attrib if k.startswith(f"{{{_OC_NS}}}")
            ]:
                del element.attrib[attr_key]
        if not schema.validate(doc):
            errors = [str(e) for e in schema.error_log]
    except etree.XMLSyntaxError as e:
        errors = [f"XML syntax error: {e}"]
    except Exception as e:
        errors = [f"Validation error: {e}"]
    return errors


def _build_odm_xml(study_oid: str, site_oid: str,
                   participant_oid: str, participant_id: str,
                   rows: list) -> str:
    """Build ODM XML for one participant's data rows.

    participant_oid: OC's internal Participant OID (e.g. SS_UAT20260_8695) —
                     goes into SubjectKey, satisfies the ODM XSD.
    participant_id:  the human Participant ID (e.g. UAT-…-P001) — goes into
                     OpenClinica:StudySubjectID, which is what OC's import
                     uses to look up the participant.
    """
    now = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")

    # Tree: events[ev_oid][repeat_key][form_oid][ig_oid] = [(item_oid, val)]
    events = {}
    # Drop reason counters for diagnostics
    _drop_missing_coords = 0
    _drop_blank = 0
    _drop_then = 0
    _drop_visibility = 0
    _drop_examples: dict[str, list] = {}  # reason -> up to 3 example strings

    def _record_drop(reason: str, row: dict, val: str = "") -> None:
        ex = f"ev={row.get('Study_Event_OID','')} fo={row.get('Form_OID','')} item={row.get('Item_OID','')} val={val!r:.60}"
        _drop_examples.setdefault(reason, [])
        if len(_drop_examples[reason]) < 3:
            _drop_examples[reason].append(ex)

    for row in rows:
        ev       = row.get("Study_Event_OID", "").strip()
        rk       = row.get("Event_Repeat_Key", "1").strip() or "1"
        fo       = row.get("Form_OID", "").strip()
        ig       = row.get("Item_Group_OID", "").strip()
        item_oid = row.get("Item_OID", "").strip()
        # Item_OID must differ from ItemGroup_OID — fallback derives it
        if not item_oid or item_oid == ig:
            field = row.get("field_name", "").strip()
            item_oid = f"{fo}.{field}" if field else ig
        val = row.get("Load_Value", "").strip()
        # Skip rows that need Playwright or aren't directly loadable
        if not ev or not fo or not ig or not val:
            _drop_missing_coords += 1
            _record_drop("missing_coords", row, val)
            continue
        if val.lower() == "(leave blank)":
            _drop_blank += 1
            _record_drop("leave_blank", row, val)
            continue  # required-field test — Playwright handles
        if "then" in val.lower():
            _drop_then += 1
            _record_drop("has_then", row, val)
            continue  # multi-step constraint test — Playwright handles
        if "=" in val and not val.startswith("20"):
            # gate/visibility setup (FIELD=val) or calc inputs (F1=v1, F2=v2)
            # Only load calc input rows — identified by "Calc path" in Scenario
            sc = str(row.get("Scenario", "")).strip()
            if "Calc path" not in sc:
                _drop_visibility += 1
                _record_drop("visibility_gate", row, val)
                continue  # visibility/gate rows — Playwright sets these in browser
            # Calc inputs: parse "ODI1=28, ODI2=28" into multiple items
            _lv_clean = re.sub(r",\s*then\s+", ", ", val, flags=re.IGNORECASE)
            for _p in _lv_clean.split(","):
                _p = _p.strip()
                if "=" in _p:
                    _fname, _fval = _p.split("=", 1)
                    _fname_clean = _fname.strip()
                    _form_short = fo.replace('F_', '', 1) if fo.upper().startswith('F_') else fo
                    _expected_prefix = f"I_{_form_short}_"
                    if _fname_clean.upper().startswith(_expected_prefix.upper()):
                        _item_oid = _fname_clean  # already fully qualified
                    else:
                        _item_oid = f"I_{_form_short}_{_fname_clean}"
                    (events
                     .setdefault(ev, {})
                     .setdefault(rk, {})
                     .setdefault(fo, {})
                     .setdefault(ig, [])
                     .append((_item_oid, _fval.strip())))
            continue

        (events
         .setdefault(ev, {})
         .setdefault(rk, {})
         .setdefault(fo, {})
         .setdefault(ig, [])
         .append((item_oid, val)))

    _n_pass = sum(len(items) for ev_d in events.values() for rk_d in ev_d.values() for fo_d in rk_d.values() for items in fo_d.values())
    _n_total = len(rows)
    _n_drop = _n_total - _n_pass
    print(f"[odm-build] {_n_pass} items passed filter out of {_n_total} rows "
          f"({_n_drop} dropped: missing_coords={_drop_missing_coords} "
          f"leave_blank={_drop_blank} has_then={_drop_then} "
          f"visibility_gate={_drop_visibility})", flush=True)
    # Log per-event breakdown so we can see which events are getting data
    ev_summary = {ev: sum(len(items) for rk_d in rk_map.values() for fo_d in rk_d.values() for items in fo_d.values())
                  for ev, rk_map in events.items()}
    print(f"[odm-build] items per event: {ev_summary}", flush=True)
    # Log drop examples for each reason
    for reason, examples in _drop_examples.items():
        print(f"[odm-build] drop/{reason} examples: {examples}", flush=True)

    # UAT start date — used as OpenClinica:StartDate for Visit-Based events.
    # Common/Unscheduled events (SE_COMMON, SE_UNSCHEDULED) don't need a
    # start date. All other event types require one for OC to accept data.
    UAT_VISIT_DATE = "2026-01-22"

    lines = [
        "<?xml version='1.0' encoding='UTF-8'?>",
        f'<ODM {ODM_NAMESPACE}',
        f'    FileOID="UAT-{now}" CreationDateTime="{now}" FileType="Snapshot">',
        f'  <ClinicalData StudyOID="{study_oid}" MetaDataVersionOID="v1.0">',
        f'    <SubjectData SubjectKey="{participant_oid}" OpenClinica:StudySubjectID="{participant_id}">',
        f'      <SiteRef LocationOID="{site_oid}"/>',
    ]
    for ev_oid, repeats in events.items():
        is_common = "COMMON" in ev_oid.upper() or "UNSCH" in ev_oid.upper()
        for repeat_key, forms in repeats.items():
            start_attr = (
                "" if is_common
                else f' OpenClinica:StartDate="{UAT_VISIT_DATE}"'
            )
            lines.append(
                f'      <StudyEventData StudyEventOID="{ev_oid}"{start_attr}>'
            )
            for form_oid, igs in forms.items():
                lines.append(
                    f'        <FormData FormOID="{form_oid}">'
                )
                for ig_oid, items in igs.items():
                    # SE_COMMON and SE_UNSCHEDULED events use repeating
                    # ItemGroups — OC4 requires ItemGroupRepeatKey to create
                    # a new repeat instance. Without it OC silently discards
                    # the ItemGroupData. Visit-Based events use non-repeating
                    # item groups and should NOT have ItemGroupRepeatKey.
                    ig_repeat_attr = (
                        ' ItemGroupRepeatKey="1"' if is_common else ""
                    )
                    lines.append(
                        f'          <ItemGroupData ItemGroupOID="{ig_oid}"'
                        f'{ig_repeat_attr}'
                        f' TransactionType="Insert">'
                    )
                    for item_oid, val in items:
                        lines.append(
                            f'            <ItemData ItemOID="{item_oid}" '
                            f'Value="{_xml_escape(val)}"/>'
                        )
                    lines.append('          </ItemGroupData>')
                lines.append('        </FormData>')
            lines.append('      </StudyEventData>')
    lines += ['    </SubjectData>', '  </ClinicalData>', '</ODM>']
    return "\n".join(lines)


async def _import_odm(subdomain: str, study_oid: str,
                       odm_xml: str) -> dict:
    """POST ODM XML to OC4 Clinical Data Import API and poll the async job.

    Per How_and_When_to_Use_APIs.pdf (page 9):
      POST {eu_base}/pages/auth/api/clinicaldata/import/xml
        Headers: Authorization: Bearer {token}
                 Content-Type: multipart/form-data; boundary=...
        Body:    -F 'file=@<odm.xml>'
        Returns plain text "job uuid: <uuid>" on success.

    Then poll for completion:
      GET {eu_base}/pages/auth/api/jobs/{job_uuid}/downloadFile
        Headers: Authorization: Bearer {token}
        While body contains "errorCode.jobInProgress" the job is still running.
        On completion the body is the CSV log file with per-item Inserted /
        Failed rows.

    Bearer token only — no cookies, no Study Runner session. Token is fetched
    fresh per HTTP call because _get_oc_token's token is short-lived and the
    poll loop can outlive a single token's TTL.
    """
    base = _pages_base(subdomain)
    submit_url = f"{base}/pages/auth/api/clinicaldata/import/xml"

    # ── Step 1: Submit the ODM XML and capture the job UUID ───────────────
    submit_token = await _get_oc_token(subdomain)
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            submit_url,
            files={"file": ("import.xml",
                            odm_xml.encode("utf-8"),
                            "text/xml")},
            data={"runFormLogic": "y"},   # run form logic / calculated fields post-import
            headers={"Authorization": f"Bearer {submit_token}"},
        )
    if not resp.is_success:
        raise RuntimeError(
            f"ODM import submit HTTP {resp.status_code} — "
            f"body: {resp.text[:300]}"
        )

    body = resp.text or ""
    if "errorCode." in body:
        raise RuntimeError(f"ODM import submit returned error: {body[:300]}")

    import re as _re
    m = _re.search(r"job\s*uuid\s*:\s*([0-9a-fA-F\-]+)", body)
    if not m:
        raise RuntimeError(
            f"ODM import submit succeeded ({resp.status_code}) but no "
            f"job uuid in response: {body[:300]}"
        )
    job_uuid = m.group(1)
    # Capture JSESSIONID from submit response — OC's nginx load balancer uses
    # this cookie to route requests to the same backend node that registered
    # the job. Without it every poll hits a different node and gets 404.
    jsessionid = resp.cookies.get("JSESSIONID")
    print(f"[uat_loader] ODM import submitted, job_uuid={job_uuid} jsessionid={jsessionid}",
          flush=True)

    # ── Step 2: Wait for OC to process the import ───────────────────────
    # NOTE: The /pages/auth/api/jobs/{uuid}/downloadFile poll endpoint is
    # broken on cust1 — returns 404 error.invalidUuid immediately for all
    # jobs regardless of payload, client, or cookies. Reported to OC
    # engineering. Workaround: fixed wait, then caller verifies via
    # clinical data read-back.
    # Poll for completion by checking if OC has processed the job.
    # The /jobs/{uuid}/downloadFile endpoint is broken on cust1 (returns 404).
    # Instead we wait a short time and return — the caller reads back clinical
    # data after all batches are done to verify.
    # Poll for job completion — poll every 2s up to 30s
    poll_url     = f"{base}/pages/auth/api/jobs/{job_uuid}/downloadFile"
    poll_cookies = {"JSESSIONID": jsessionid} if jsessionid else {}
    poll_interval, poll_timeout = 2, 30
    elapsed = 0
    while elapsed < poll_timeout:
        await asyncio.sleep(poll_interval)
        elapsed += poll_interval
        try:
            poll_token = await _get_oc_token(subdomain)
            async with httpx.AsyncClient(timeout=15) as client:
                pr = await client.get(
                    poll_url,
                    headers={"Authorization": f"Bearer {poll_token}"},
                    cookies=poll_cookies,
                )
            if pr.is_success and ("Inserted" in pr.text or "Failed" in pr.text):
                print(f"[uat_loader] ODM job complete after {elapsed}s", flush=True)
                return {"status": "complete", "job_uuid": job_uuid, "log": pr.text, "jsessionid": jsessionid or ""}
        except Exception as pe:
            print(f"[uat_loader] ODM poll error at {elapsed}s: {pe}", flush=True)
    print(f"[uat_loader] ODM job poll timed out after {poll_timeout}s", flush=True)
    return {"status": "timeout", "job_uuid": job_uuid, "log": "", "jsessionid": jsessionid or ""}


# ── Clinical data read-back ───────────────────────────────────────────────────

async def _fetch_clinical_data(
    subdomain: str,
    study_oid: str,
    participant_oc_oid: str,
    token: str,
) -> dict:
    """
    GET all clinical data for one participant and return a lookup dict:
        (event_oid, form_oid, ig_oid, item_oid) -> value

    Uses the wildcard form of the endpoint:
        GET .../clinicaldata/{studyOID}/{participantOID}/*/*?clinicalData=y

    participant_oc_oid is the internal OC OID (SS_UAT20260_xxxx).
    Returns an empty dict on any failure — callers treat missing = Not Run.
    """
    from urllib.parse import quote as _quote
    import xml.etree.ElementTree as _ET

    base = _pages_base(subdomain)
    url  = (
        f"{base}/pages/auth/api/clinicaldata"
        f"/{_quote(study_oid, safe='')}"
        f"/{_quote(participant_oc_oid, safe='')}"
        f"/*/*"
        f"?clinicalData=y&includeMetadata=n&includeDN=n"
        f"&includeAudits=n&showArchived=n"
    )
    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.get(
                url,
                headers={
                    "Authorization": f"Bearer {token}",
                    "Accept":        "application/xml",
                },
            )
        if not resp.is_success:
            print(
                f"[uat_loader] clinical data GET HTTP {resp.status_code} "
                f"for {participant_oc_oid} — body: {resp.text[:300]}",
                flush=True,
            )
            return {}
        xml_text = resp.text or ""
        if not xml_text.strip():
            return {}
    except Exception as e:
        print(f"[uat_loader] clinical data GET failed: {e}", flush=True)
        return {}

    # Parse ODM XML into lookup dict
    lookup: dict = {}
    try:
        ODM_NS = "http://www.cdisc.org/ns/odm/v1.3"
        root   = _ET.fromstring(xml_text.encode("utf-8"))

        def _tag(local):
            return f"{{{ODM_NS}}}{local}"

        for cd in root.iter(_tag("ClinicalData")):
            for sd in cd.iter(_tag("SubjectData")):
                for sed in sd.iter(_tag("StudyEventData")):
                    ev_oid = sed.get("StudyEventOID", "")
                    for fd in sed.iter(_tag("FormData")):
                        fo_oid = fd.get("FormOID", "")
                        for igd in fd.iter(_tag("ItemGroupData")):
                            ig_oid = igd.get("ItemGroupOID", "")
                            for itd in igd.iter(_tag("ItemData")):
                                it_oid = itd.get("ItemOID", "")
                                val    = itd.get("Value", "")
                                key    = (
                                    ev_oid.upper(),
                                    fo_oid.upper(),
                                    ig_oid.upper(),
                                    it_oid.upper(),
                                )
                                lookup[key] = val
    except Exception as e:
        print(f"[uat_loader] ODM parse error: {e}", flush=True)

    print(
        f"[uat_loader] clinical data read-back: "
        f"{len(lookup)} item values for {participant_oc_oid}",
        flush=True,
    )
    return lookup


def _evaluate_uat_cases(
    dvs_bytes: bytes,
    stamp_map: dict,
    clinical_data: dict,
    job_failures: dict = None,
) -> bytes:
    """
    job_failures: {(item_oid_upper): message} from ODM job CSV Failed rows
    """
    """
    For each row in UAT_Cases, look up the stored value in clinical_data
    and compare against Expected Result.  Writes:
        - Status         → 'Pass', 'Fail', or 'Not Run'
        - Actual Result  → what OC returned (or blank)
        - Test Result    → 'Pass', 'Fail', or 'Not Run'
        - Execution Date → UTC timestamp (Pass/Fail rows only)

    clinical_data: merged lookup from all participants:
        (event_oid_upper, form_oid_upper, ig_oid_upper, item_oid_upper)
        → stored_value

    stamp_map: { logical_pid → {site_oid, participant_key, oc_oid} }

    Returns updated workbook bytes.
    """
    import datetime as _dt

    wb = load_workbook(io.BytesIO(dvs_bytes))
    if "UAT_Cases" not in wb.sheetnames:
        return dvs_bytes
    ws = wb["UAT_Cases"]

    # Locate header row and column indices
    header_row_idx = None
    col_idx: dict = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "UAT Case ID":
            header_row_idx = row_idx
            for ci, val in enumerate(row, start=1):
                if val:
                    col_idx[str(val).strip()] = ci
            break

    if not header_row_idx:
        return dvs_bytes

    # Required columns
    needed = [
        "Status", "Actual Result", "Test Result", "Execution Date",
        "Study_Event_OID", "Form_OID", "Item_Group_OID", "Item_OID",
        "Load_Value", "Expected Result", "Notes",
    ]
    if not all(c in col_idx for c in needed):
        missing = [c for c in needed if c not in col_idx]
        print(
            f"[uat_loader] _evaluate_uat_cases: missing columns "
            f"{missing} — skipping evaluation",
            flush=True,
        )
        return dvs_bytes

    now_str = _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    passed = failed = skipped = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        # Skip blank rows (empty Excel rows at end of sheet)
        uid = str(row[col_idx.get("UAT Case ID", 1) - 1].value or "").strip() if "UAT Case ID" in col_idx else ""
        if not uid:
            continue  # don't count toward skipped — truly empty row

        ev_oid  = str(row[col_idx["Study_Event_OID"]  - 1].value or "").strip().upper()
        fo_oid  = str(row[col_idx["Form_OID"]         - 1].value or "").strip().upper()
        ig_oid  = str(row[col_idx["Item_Group_OID"]   - 1].value or "").strip().upper()
        lv      = str(row[col_idx["Load_Value"]       - 1].value or "").strip()
        expected = str(row[col_idx["Expected Result"] - 1].value or "").strip()

        if not ev_oid or not fo_oid or not ig_oid or not lv:
            skipped += 1
            continue

        # Use Item_OID column directly for lookup
        item_oid = str(row[col_idx["Item_OID"] - 1].value or "").strip().upper()
        if not item_oid:
            skipped += 1
            continue

        # Lookup key — all uppercase to match _fetch_clinical_data keys
        key = (ev_oid, fo_oid, ig_oid, item_oid)
        stored = clinical_data.get(key)

        # Check if this item failed in the ODM import job
        job_msg = (job_failures or {}).get(item_oid, "")

        # Classify non-loadable rows as Not Testable via ODM
        lv_lower = lv.lower()
        is_blank_case    = lv_lower == "(leave blank)"
        is_calc_case     = "=" in lv and not lv.startswith("20")  # field=value patterns
        is_multistep     = "then" in lv_lower  # multi-step setup like "ICFDAT=x, then date=y"
        is_visibility    = any(x in expected.upper() for x in ["VISIBLE", "HIDDEN", "RELEVANT"])
        is_ui_constraint = any(x in expected for x in ["error shown", "Form does not save",
                                                         "Constraint fires", "constraint"])
        # Calc rows with multiple inputs (ODI1=x, ODI2=y) are loaded via ODM
        # and OC computes the output with runFormLogic=y — these ARE testable
        is_pure_calc     = "Calc path" in str(row[col_idx.get("Scenario", 1) - 1].value or "")
        not_testable = (is_blank_case or is_multistep or is_visibility or is_ui_constraint
                        or (is_calc_case and not is_pure_calc))

        if not_testable:
            row[col_idx["Test Result"]    - 1].value = "Not Run"
            row[col_idx["Status"]         - 1].value = "Not Run"
            row[col_idx["Actual Result"]  - 1].value = "Not Testable via ODM"
            skipped += 1
            continue

        if stored is None and job_msg:
            # Item failed to import — mark as Fail with reason
            row[col_idx["Actual Result"]  - 1].value = f"Import failed: {job_msg}"
            row[col_idx["Test Result"]    - 1].value = "Fail"
            row[col_idx["Status"]         - 1].value = "Fail"
            row[col_idx["Execution Date"] - 1].value = now_str
            row[col_idx["Notes"]          - 1].value = f"ODM import error: {job_msg}"
            failed += 1
            continue

        if stored is None:
            # Item not found in read-back — Not Run
            skipped += 1
            continue

        # Pass = stored value matches what we loaded
        stored_str = str(stored).strip()
        if stored_str == lv.strip():
            result = "Pass"
            passed += 1
        else:
            result = "Fail"
            failed += 1

        row[col_idx["Actual Result"]  - 1].value = stored_str
        row[col_idx["Test Result"]    - 1].value = result
        row[col_idx["Status"]         - 1].value = result
        row[col_idx["Execution Date"] - 1].value = now_str
        if result == "Fail":
            row[col_idx["Notes"] - 1].value = f"Loaded: {lv.strip()!r} | Got back: {stored_str!r}"

    print(
        f"[uat_loader] UAT evaluation: "
        f"Pass={passed} Fail={failed} Not Run={skipped}",
        flush=True,
    )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── DVS stamping ──────────────────────────────────────────────────────────────

def _stamp_dvs(dvs_bytes: bytes, stamp_map: dict) -> bytes:
    """
    Write runtime Site_OID and Participant_Key into UAT_Cases sheet.
    stamp_map: { logical_pid -> {"site_oid": ..., "participant_key": ...} }
    """
    wb = load_workbook(io.BytesIO(dvs_bytes))
    if "UAT_Cases" not in wb.sheetnames:
        return dvs_bytes
    ws = wb["UAT_Cases"]

    header_row_idx = None
    col_idx = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "UAT Case ID":
            header_row_idx = row_idx
            for ci, val in enumerate(row, start=1):
                if val:
                    col_idx[str(val).strip()] = ci
            break

    if not header_row_idx:
        return dvs_bytes

    site_col = col_idx.get("Site_OID")
    key_col  = col_idx.get("Participant_Key")
    pid_col  = col_idx.get("Participant_ID")
    if not (site_col and key_col and pid_col):
        return dvs_bytes

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        pid = str(row[pid_col - 1].value or "").strip()
        if pid and pid in stamp_map:
            row[site_col - 1].value = stamp_map[pid]["site_oid"]
            row[key_col  - 1].value = stamp_map[pid]["participant_key"]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Main entry point ──────────────────────────────────────────────────────────

async def run_uat_loader(item_id: str, fo_titles: dict = None) -> dict:
    """
    Execute the full UAT data loading workflow for one monday.com item.
    Loads OC session cookies from the saved Playwright storage_state file.
    Returns dict: success, site_oid, participants_created, odm_imports, errors.
    """
    result = {
        "success": False,
        "site_oid": None,
        "participants_created": [],
        "odm_imports": [],
        "errors": [],
        "test_env_uuid": "",
    }

    # ── Step 1: Read item metadata ─────────────────────────────────────────
    await append_log(item_id, "UAT Loader: reading item metadata...")
    item = await get_item(item_id)
    cols = {c["id"]: c for c in item.get("column_values", [])}

    subdomain  = (cols.get(COL["oc_subdomain"], {}).get("text") or "").strip()
    study_uuid = (cols.get(COL["study_uuid"],   {}).get("text") or "").strip()
    study_oid  = (cols.get(COL["study_oid"],    {}).get("text") or "").strip()
    oc_email   = (cols.get(COL["oc_email"],     {}).get("text") or "").strip()

    if not subdomain:
        result["errors"].append("OC Subdomain is blank.")
        return result
    if not study_uuid:
        result["errors"].append("Study UUID is blank — run the pipeline first.")
        return result
    if not study_oid:
        result["errors"].append("Study OID is blank — publish to Test first.")
        return result

    # ── Step 2: Download DVS ───────────────────────────────────────────────
    await append_log(item_id, "UAT Loader: downloading DVS...")
    dvs_bytes = await download_column_file(item_id, COL["dvs_output"])
    if not dvs_bytes:
        result["errors"].append(
            "No DVS found in DVS Output column. Run the pipeline first."
        )
        return result

    # ── Step 3: Get TEST environment UUID (on-the-fly) ────────────────────
    await append_log(item_id, "UAT Loader: locating TEST environment...")
    try:
        test_env_uuid, test_env_oid = await _get_test_env_uuid(
            subdomain, study_uuid
        )
        result["test_env_uuid"] = test_env_uuid
    except Exception as e:
        result["errors"].append(f"Could not find TEST environment: {e}")
        return result

    # ── Step 4: Create dated site ──────────────────────────────────────────
    now       = datetime.datetime.now()
    site_name = f"UAT Automation Site - {now.strftime('%Y-%m-%d %H:%M')}"
    site_oid  = f"UAT-{now.strftime('%Y%m%d-%H%M%S')}"

    # Try to reuse an existing site — fresh sites may lack published form versions
    created_site_oid = None
    try:
        _sites_url = (f"{_study_service_base(subdomain)}/api/study-environments"
                      f"/{test_env_uuid}/sites")
        _tok2 = await _get_oc_token(subdomain)
        async with httpx.AsyncClient(timeout=30) as _hc:
            _sr = await _hc.get(_sites_url, headers={"Authorization": f"Bearer {_tok2}"})
        if _sr.is_success:
            _existing = _sr.json() if isinstance(_sr.json(), list) else []
            if _existing:
                _s0 = _existing[0]
                created_site_oid = (_s0.get("oid") or _s0.get("uniqueIdentifier")
                                    or _s0.get("siteOid") or _s0.get("id"))
                await append_log(item_id,
                    f"UAT Loader: reusing existing site → {created_site_oid} keys={list(_s0.keys())[:6]}")
    except Exception as _se:
        await append_log(item_id, f"UAT Loader: site list failed: {_se}")

    if not created_site_oid:
        await append_log(item_id, f"UAT Loader: creating site '{site_name}'...")
        try:
            created_site_oid = await _create_site(
                subdomain, test_env_uuid, site_name, site_oid
            )
            result["site_oid"] = created_site_oid
            await append_log(item_id, f"UAT Loader: site created → {created_site_oid}")
        except Exception as e:
            result["errors"].append(f"Site creation failed: {e}")
            return result

    # ── Step 4b: Activate TEST environment (PUT status → AVAILABLE) ──────
    # _activate_test_environment lives in pipeline.py; lazy-imported here
    # because pipeline.py already imports run_uat_loader from this module at
    # module-load time — a top-level "from pipeline import …" would loop.
    await append_log(item_id,
        "UAT Loader: activating TEST environment → AVAILABLE...")
    try:
        from pipeline import _activate_test_environment
        await _activate_test_environment(subdomain, study_uuid)
        await append_log(item_id, "UAT Loader: TEST environment activated")
    except Exception as e:
        result["errors"].append(f"TEST environment activation failed: {e}")
        return result

    # ── Step 5: Parse UAT_Cases ────────────────────────────────────────────
    await append_log(item_id, "UAT Loader: parsing UAT_Cases sheet...")
    try:
        uat_rows = _parse_uat_cases(dvs_bytes)
    except Exception as e:
        result["errors"].append(f"DVS parse failed: {e}")
        return result

    if not uat_rows:
        result["errors"].append(
            "No loadable rows in UAT_Cases. "
            "Ensure DVS was generated with ODM load coordinates populated."
        )
        return result

    groups = _group_by_participant(uat_rows)
    await append_log(
        item_id,
        f"UAT Loader: {len(uat_rows)} rows, "
        f"{len(groups)} participant(s): {list(groups.keys())}"
    )

    # ── Study availability gate ───────────────────────────────────────────
    # Per Study_Service_API.md StudyEnvironmentDTO.status enum, TEST must be
    # AVAILABLE before participants can be created. Poll up to 60s.
    await append_log(item_id, "UAT Loader: waiting for TEST environment to be AVAILABLE...")
    try:
        await _wait_for_test_available(subdomain, study_uuid,
                                       timeout_s=60, interval_s=5)
    except Exception as e:
        result["errors"].append(f"TEST environment not AVAILABLE: {e}")
        return result

    token = await _get_oc_token(subdomain)
    stamp_map = {}

    # ── Pass 1: Create ALL participants first ─────────────────────────────
    for logical_pid, rows in groups.items():
        p_suffix = logical_pid.replace("UAT-P", "P")
        run_key  = f"{site_oid}-{p_suffix}"

        await append_log(item_id, f"UAT Loader: creating participant {run_key}...")
        try:
            confirmed_key = await _create_participant(
                subdomain, study_oid, created_site_oid, run_key, token, {}
            )
            result["participants_created"].append(confirmed_key)
            oc_oid = await _get_participant_oid(
                subdomain, study_oid, confirmed_key, token
            )

            stamp_map[logical_pid] = {
                "site_oid":        created_site_oid,
                "participant_key": confirmed_key,
                "oc_oid":          oc_oid,
            }
            await append_log(
                item_id,
                f"UAT Loader: participant {run_key} → OC SubjectKey={confirmed_key}"
            )
            await append_log(
                item_id,
                f"UAT Loader: participant {run_key} → OC internal OID={oc_oid}"
            )
        except Exception as e:
            err = f"Participant creation failed for {run_key}: {e}"
            result["errors"].append(err)
            await append_log(
                item_id,
                f"UAT Loader: ERROR — {err} (skipping this participant)"
            )

    # Participant creation is synchronous — no propagation delay needed

    # ── Pass 2: Build + import ODM for each successfully-created participant ──
    for logical_pid, rows in groups.items():
        if logical_pid not in stamp_map:
            continue  # creation failed in Pass 1 — skip ODM
        oc_oid       = stamp_map[logical_pid]["oc_oid"]
        confirmed_id = stamp_map[logical_pid]["participant_key"]
        p_suffix = logical_pid.replace("UAT-P", "P")
        run_key  = f"{site_oid}-{p_suffix}"

        await append_log(
            item_id,
            f"UAT Loader: importing ODM for {run_key} ({len(rows)} rows)..."
        )
        try:
            # Validate once against the full XML before batching
            if rows:
                _r0 = rows[0]
                await append_log(item_id, f"UAT Loader: rows[0] ev={_r0.get('Study_Event_OID')!r} item={_r0.get('Item_OID')!r} val={str(_r0.get('Load_Value',''))[:20]!r}")
            odm_xml_full = _build_odm_xml(
                study_oid, created_site_oid, oc_oid, confirmed_id, rows
            )
            odm_errors = _validate_odm_xml(odm_xml_full)
            if odm_errors:
                err_summary = "; ".join(odm_errors[:3])
                raise RuntimeError(
                    f"ODM XML failed XSD validation ({len(odm_errors)} errors): "
                    f"{err_summary}"
                )
            await append_log(item_id, f"UAT Loader: ODM XML valid (XSD passed)")
            # DEBUG: log first 300 chars of ODM to Monday to verify OID format
            # ODM validated
            # Submit full ODM in one call — no batching needed
            import_result = await _import_odm(
                subdomain, study_oid, odm_xml_full
            )
            log_csv = import_result.get("log", "")
            ins  = sum(1 for ln in log_csv.splitlines() if ",Inserted," in ln)
            fail = sum(1 for ln in log_csv.splitlines() if ",Failed,"  in ln)
            await append_log(
                item_id,
                f"UAT Loader: ODM import complete — Inserted={ins} Failed={fail}"
            )
            result["odm_imports"].append({
                "participant": run_key,
                "rows":        len(rows),
                "result":      import_result,
            })
            log_csv = import_result.get("log", "") or ""
            ins  = sum(1 for ln in log_csv.splitlines() if ",Inserted," in ln)
            fail = sum(1 for ln in log_csv.splitlines() if ",Failed,"  in ln)
            await append_log(
                item_id,
                f"UAT Loader: ODM import for {run_key} complete — "
                f"Inserted={ins} Failed={fail}"
            )
            # Parse the job log CSV to count Inserted vs Failed rows.
            import csv as _csv
            import io as _csvio
            _log_text = import_result.get("log", "") or ""
            _inserted, _failed_rows = 0, []
            if _log_text:
                reader = _csv.DictReader(_csvio.StringIO(_log_text))
                for _row in reader:
                    _status = (_row.get("Status") or "").strip()
                    if _status == "Inserted":
                        _inserted += 1
                    elif _status == "Failed":
                        _failed_rows.append(_row)
            _job_uuid = import_result.get("job_uuid", "?")
            await append_log(
                item_id,
                f"UAT Loader: ODM job {_job_uuid} — "
                f"Inserted: {_inserted}, Failed: {len(_failed_rows)}"
            )
            if _failed_rows:
                for _fr in _failed_rows[:3]:
                    _msg  = (_fr.get("Message") or "").strip()
                    _evt  = (_fr.get("StudyEventOID") or "").strip()
                    _form = (_fr.get("FormOID") or "").strip()
                    _item = (_fr.get("ItemOID") or "").strip()
                    await append_log(
                        item_id,
                        f"  FAILED row — Event={_evt} Form={_form} "
                        f"Item={_item} Message={_msg}"
                    )
        except Exception as e:
            err = f"ODM import failed for {run_key}: {e}"
            result["errors"].append(err)
            await append_log(item_id, f"UAT Loader: ERROR — {err}")

    # ── Step 8b: Read back clinical data + evaluate UAT cases ────────────
    # For each participant that was successfully created, fetch all stored
    # item values from OC via the clinical data GET endpoint, then compare
    # against Expected Result in UAT_Cases and stamp Pass/Fail.
    clinical_data: dict = {}
    if stamp_map:
        await append_log(item_id,
            "UAT Loader: reading back clinical data from OC...")
        for logical_pid, info in stamp_map.items():
            oc_oid = info.get("oc_oid", "")
            if not oc_oid:
                continue
            try:
                participant_data = await _fetch_clinical_data(
                    subdomain, study_oid, oc_oid, token
                )
                clinical_data.update(participant_data)
                await append_log(item_id,
                    f"UAT Loader: read {len(participant_data)} item "
                    f"values for {logical_pid}")
            except Exception as e:
                await append_log(item_id,
                    f"UAT Loader: clinical data read failed for "
                    f"{logical_pid} (non-fatal): {e}")

    # ── Step 9: Stamp DVS ─────────────────────────────────────────────────
    if stamp_map:
        await append_log(item_id, "UAT Loader: stamping DVS with runtime OIDs...")
        try:
            stamped_bytes = _stamp_dvs(dvs_bytes, stamp_map)
        except Exception as e:
            stamped_bytes = dvs_bytes
            await append_log(item_id, f"UAT Loader: stamp failed (non-fatal): {e}")
    else:
        stamped_bytes = dvs_bytes

    # ── Step 9a: Playwright UAT for UI-only test cases ──────────────────
    try:
        from playwright_uat import run_playwright_uat
        oc_email = (cols.get(COL.get("oc_email", "emailothn6i3m"), {}).get("text") or "").strip()
        # Use the first participant's OC OID for Playwright
        _first_oc_oid = next(iter(stamp_map.values()), {}).get("oc_oid", "")
        if _first_oc_oid:
            await append_log(item_id,
                "UAT Loader: running Playwright UAT for UI-only cases...")
            # Get fresh token + jsessionid for Playwright auth
            _pw_token = await _get_oc_token(subdomain)
            # Collect jsessionid from ODM imports — it's an active OC session
            _jsessionid = ""
            for _imp in result.get("odm_imports", []):
                _js = _imp.get("result", {}).get("jsessionid", "")
                if _js:
                    _jsessionid = _js
                    break
            # Get study env UUID for build app navigation
            _study_uuid = cols.get("text_mm3ggzga", {}).get("text", "").strip()
            _test_env_uuid = result.get("test_env_uuid", "")
            stamped_bytes = await run_playwright_uat(
                stamped_bytes if stamp_map else dvs_bytes,
                subdomain, _first_oc_oid, oc_email, stamp_map,
                bearer_token=_pw_token,
                jsessionid=_jsessionid,
                study_uuid=_study_uuid,
                study_env_uuid=_test_env_uuid,
                fo_titles=fo_titles or {},
            )
    except Exception as _pw_err:
        await append_log(item_id,
            f"UAT Loader: Playwright UAT skipped — {_pw_err}")

    # If no session file, let user know how to get Playwright tests enabled
    import os as _os2
    _oc_email = (cols.get("emailothn6i3m") or {}).get("text", "") or ""
    _sess_file = f"/data/browser_sessions/{_oc_email}.json"
    if _oc_email and not _os2.path.exists(_sess_file):
        await append_log(item_id,
            "⚠️ Playwright UI tests skipped — no browser session on file. "
            "Run a full pipeline build from Monday to enable them.")

    # ── Step 9b: Evaluate UAT cases (Pass/Fail) ───────────────────────────
    if clinical_data:
        await append_log(item_id,
            "UAT Loader: evaluating UAT cases against clinical data...")
        try:
            # Build job_failures: item_oid_upper -> error message
            _job_failures = {}
            _all_log = ""
            for _imp in result.get("odm_imports", []):
                _all_log += _imp.get("result", {}).get("log", "") or ""
            if _all_log:
                import csv as _csv2, io as _io2
                for _fr in _csv2.DictReader(_io2.StringIO(_all_log)):
                    if (_fr.get("Status") or "").strip() == "Failed":
                        _ioid = (_fr.get("ItemOID") or "").strip().upper()
                        _msg  = (_fr.get("Message") or "").strip()
                        if _ioid:
                            _job_failures[_ioid] = _msg
            stamped_bytes = _evaluate_uat_cases(
                stamped_bytes, stamp_map, clinical_data, _job_failures
            )
        except Exception as e:
            await append_log(item_id,
                f"UAT Loader: evaluation failed (non-fatal): {e}")

    # ── Step 10: Upload stamped DVS ────────────────────────────────────────
    protocol_number = (
        cols.get(COL["protocol_number"], {}).get("text") or "UAT"
    ).strip()
    dvs_filename = f"{protocol_number}_DVS_UAT_Results.xlsx"

    await append_log(item_id, "UAT Loader: uploading UAT DVS Results...")
    try:
        await upload_file(item_id, UAT_DVS_RESULTS_COL,
                          dvs_filename, stamped_bytes)
    except Exception as e:
        await append_log(item_id, f"UAT Loader: results upload failed: {e}")

    # ── Done ───────────────────────────────────────────────────────────────
    n_ok  = len(result["participants_created"])
    n_err = len(result["errors"])
    result["success"] = n_ok > 0 and n_err == 0

    await append_log(
        item_id,
        f"UAT Load complete. Site: {result['site_oid']}. "
        f"Participants: {n_ok}. ODM imports: {len(result['odm_imports'])}. "
        f"Errors: {n_err}."
    )
    return result

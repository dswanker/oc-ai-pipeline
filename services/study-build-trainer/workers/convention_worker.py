"""
convention_worker.py — Processes accuracy review appendix submissions.

Flow:
  1. Human uploads edited Diff Appendix XLSX to Convention Rulebook board
     and sets Submit Trigger → "Submit for Review".
  2. Webhook fires → convention_worker.process() runs.
  3. Worker downloads XLSX, parses diff rows, calls Claude.
  4a. If Claude needs clarification:
       - Writes questions into column G of the XLSX for each ambiguous row
       - Re-uploads updated XLSX to the board
       - Writes summary of questions to "Claude Questions" column
       - Sets status → Needs Clarification, trigger → Awaiting Human
       - Increments Round counter
  4b. If all clear:
       - Extracts structured conventions from Claude response
       - Writes them to /data/rulebook/conventions.json (Railway volume)
       - Sets Conventions Extracted count, Date Completed
       - Sets status → Added to Rulebook, trigger → Awaiting Human
"""
from __future__ import annotations

import asyncio
import json
import os
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, UTC
from pathlib import Path
from typing import Any

from convention_client import ConventionMondayClient, RulebookItem

try:
    import structlog
    logger = structlog.get_logger(__name__)
except ImportError:
    import logging
    _log = logging.getLogger(__name__)
    class _L:
        def info(self, e, **kw):  _log.info(f"{e} {kw}")
        def warning(self, e, **kw): _log.warning(f"{e} {kw}")
        def error(self, e, **kw): _log.error(f"{e} {kw}")
        def exception(self, e, **kw): _log.exception(f"{e} {kw}")
    logger = _L()

# Path to the conventions JSON file on the Railway persistent volume.
# Override via CONVENTIONS_PATH env var for local dev.
CONVENTIONS_PATH = Path(
    os.environ.get("CONVENTIONS_PATH", "/data/rulebook/conventions.json")
)

# Claude model used for convention extraction
CLAUDE_MODEL     = "claude-opus-4-7"
MAX_TOKENS       = 8000


# ── Appendix XLSX parser ──────────────────────────────────────────────────────

@dataclass
class DiffEntry:
    """One row from the Diff Appendix sheet."""
    row_num:         int
    layer:           str
    element:         str
    ai_generated:    str
    human_approved:  str
    notes:           str
    convention:      str   # "This customer only" | "All customers" | "Skip..." | ""
    claude_question: str   # previously written question (may be empty)
    human_response:  str   # human's answer to Claude's question (may be empty)


def parse_appendix(xlsx_bytes: bytes) -> list[DiffEntry]:
    """
    Read the Diff Appendix sheet from the accuracy report XLSX.
    Returns list of DiffEntry for rows that are tagged with a Convention value.
    Skips group-header rows (layer banner rows) and empty rows.
    """
    import openpyxl

    wb  = openpyxl.load_workbook(
        __import__("io").BytesIO(xlsx_bytes), data_only=True, read_only=True
    )
    # Find the Diff Appendix sheet (Sheet 2)
    sheet_name = None
    for name in wb.sheetnames:
        if "appendix" in name.lower() or "diff" in name.lower():
            sheet_name = name
            break
    if sheet_name is None and len(wb.sheetnames) >= 2:
        sheet_name = wb.sheetnames[1]
    if sheet_name is None:
        wb.close()
        return []

    ws   = wb[sheet_name]
    rows = list(ws.values)
    wb.close()

    # Find the header row (contains "Layer", "Element", "AI Generated" etc.)
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v).strip().lower() if v else "" for v in row]
        if "layer" in vals and "element" in " ".join(vals):
            header_row = i
            break

    if header_row is None:
        return []

    entries = []
    # Skip the header + sub-header rows (typically 2 rows after header)
    data_start = header_row + 2  # skip sub-header explanatory row

    SKIP_CONVENTIONS = {"", "skip — not a convention", "skip"}

    for i, row in enumerate(rows[data_start:], start=data_start + 1):
        if not row or not any(v for v in row):
            continue
        cols = [str(v).strip() if v is not None else "" for v in row]
        # Pad to at least 8 columns
        while len(cols) < 8:
            cols.append("")

        layer      = cols[0]
        element    = cols[1]
        ai_val     = cols[2]
        human_val  = cols[3]
        notes      = cols[4]
        convention = cols[5]
        claude_q   = cols[6]
        human_resp = cols[7]

        # Skip group header rows (all-caps layer banners, merged cells)
        if not element and not ai_val and not human_val:
            continue
        # Skip rows without a convention tag
        if convention.lower() in SKIP_CONVENTIONS:
            continue

        entries.append(DiffEntry(
            row_num        = i,
            layer          = layer,
            element        = element,
            ai_generated   = ai_val,
            human_approved = human_val,
            notes          = notes,
            convention     = convention,
            claude_question= claude_q,
            human_response = human_resp,
        ))

    return entries


# ── Write Claude questions back into XLSX ─────────────────────────────────────

def write_questions_to_xlsx(
    xlsx_bytes: bytes,
    questions: list[dict],  # [{"row_num": int, "question": str}]
) -> bytes:
    """
    Open the XLSX, write Claude's questions into column G for the relevant rows,
    and return the modified bytes.
    """
    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    # Find the diff sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "appendix" in name.lower() or "diff" in name.lower():
            sheet_name = name
            break
    if sheet_name is None and len(wb.sheetnames) >= 2:
        sheet_name = wb.sheetnames[1]
    if sheet_name is None:
        return xlsx_bytes

    ws = wb[sheet_name]
    q_by_row = {q["row_num"]: q["question"] for q in questions}

    for row_num, question in q_by_row.items():
        cell = ws.cell(row=row_num, column=7)
        cell.value = question
        # Light blue background to signal this is Claude's write
        from openpyxl.styles import PatternFill, Font
        cell.fill  = PatternFill("solid", fgColor="E8F4FD")
        cell.font  = Font(name="Arial", size=8, italic=True, color="003366")

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ── Conventions JSON store ────────────────────────────────────────────────────

def load_conventions() -> dict:
    """Load conventions.json from the Railway volume. Creates if missing."""
    CONVENTIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not CONVENTIONS_PATH.exists():
        initial = {
            "version":          "1.0",
            "last_updated":     date.today().isoformat(),
            "global":           [],
            "customer_specific": {},
        }
        CONVENTIONS_PATH.write_text(json.dumps(initial, indent=2))
        return initial
    return json.loads(CONVENTIONS_PATH.read_text())


def save_conventions(data: dict) -> None:
    """Write conventions.json back to disk."""
    data["last_updated"] = date.today().isoformat()
    CONVENTIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    CONVENTIONS_PATH.write_text(json.dumps(data, indent=2))


def apply_conventions(
    conventions_data: dict,
    new_conventions: list[dict],
    customer_uuid: str | None,
) -> int:
    """
    Add new conventions to the store.
    Returns count of conventions added.
    """
    added = 0
    for conv in new_conventions:
        scope = conv.get("scope", "global")
        entry = {
            "id":           _new_id(conventions_data),
            "layer":        conv.get("layer", ""),
            "source_study": conv.get("source_study", ""),
            "rule":         conv.get("rule", ""),
            "rationale":    conv.get("rationale", ""),
            "created":      date.today().isoformat(),
        }
        if scope == "global":
            conventions_data["global"].append(entry)
            added += 1
        elif scope == "customer" and customer_uuid:
            if customer_uuid not in conventions_data["customer_specific"]:
                conventions_data["customer_specific"][customer_uuid] = []
            conventions_data["customer_specific"][customer_uuid].append(entry)
            added += 1
    return added


def _new_id(conventions_data: dict) -> str:
    """Generate next sequential ID."""
    existing = [c.get("id", "") for c in conventions_data.get("global", [])]
    for clist in conventions_data.get("customer_specific", {}).values():
        existing.extend(c.get("id", "") for c in clist)
    nums = []
    for eid in existing:
        try:
            nums.append(int(eid.split("-")[-1]))
        except (ValueError, IndexError):
            pass
    next_num = (max(nums) + 1) if nums else 1
    return f"CONV-{next_num:04d}"


# ── Claude call ───────────────────────────────────────────────────────────────

_EXTRACTION_PROMPT = """\
You are processing an accuracy review appendix from a clinical trial EDC build comparison.

The human reviewer has gone through a diff between an AI-predicted EDC build and the
human-approved actual build. For each difference, they have:
 - Noted what the AI generated vs what the human approved
 - Added notes explaining why the human version is correct
 - Tagged it as "This customer only" or "All customers"
 - Possibly provided responses to previous questions

Your task: extract structured conventions from these tagged rows.

RULES:
1. A convention is a reusable rule the AI should follow in future builds.
2. Write the rule as a clear, actionable statement (e.g. "Use select_one for
   all Yes/No consent items across all forms").
3. "All customers" → scope: "global"
4. "This customer only" → scope: "customer"
5. If a row is ambiguous (missing notes, unclear what rule to derive, or the
   human response doesn't fully answer your previous question), ask a specific
   numbered question referencing the row element. Be concise.
6. If the human has already answered a previous question in Human Response,
   use that answer to resolve the convention — do not ask again.

Return ONLY valid JSON in this exact structure:
{
  "status": "complete" | "needs_clarification",
  "questions": [
    {"row_num": <int>, "element": "<element field>", "question": "<your question>"}
  ],
  "conventions": [
    {
      "scope": "global" | "customer",
      "layer": "<Study|Events|Form Placement|Forms|Items|Choices|Logic>",
      "source_study": "<protocol number>",
      "rule": "<clear actionable rule statement>",
      "rationale": "<one sentence why this rule is correct>"
    }
  ]
}

If status is "complete", questions must be an empty array [].
If status is "needs_clarification", include both questions (for unclear rows) AND
conventions (for rows that ARE clear enough to codify now).

DIFF ROWS TO PROCESS:
"""


async def call_claude_for_conventions(
    entries: list[DiffEntry],
    protocol_number: str,
) -> dict:
    """
    Call Claude to extract conventions from the tagged diff entries.
    Returns the parsed JSON response.
    """
    import anthropic

    client = anthropic.AsyncAnthropic(
        api_key=os.environ.get("ANTHROPIC_API_KEY", "").strip()
    )

    # Build the rows text
    rows_text = []
    for e in entries:
        row_text = (
            f"Row {e.row_num} | Layer: {e.layer} | Element: {e.element}\n"
            f"  AI Generated:   {e.ai_generated}\n"
            f"  Human Approved: {e.human_approved}\n"
            f"  Notes:          {e.notes or '(none)'}\n"
            f"  Convention Tag: {e.convention}\n"
        )
        if e.claude_question:
            row_text += f"  Previous Question: {e.claude_question}\n"
        if e.human_response:
            row_text += f"  Human Response:    {e.human_response}\n"
        rows_text.append(row_text)

    prompt = (
        _EXTRACTION_PROMPT
        + f"Protocol: {protocol_number}\n\n"
        + "\n".join(rows_text)
    )

    response = await client.messages.create(
        model      = CLAUDE_MODEL,
        max_tokens = MAX_TOKENS,
        messages   = [{"role": "user", "content": prompt}],
    )
    text = response.content[0].text

    # Strip markdown fences
    import re
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```", "", text).strip()

    return json.loads(text)


# ── Main worker ───────────────────────────────────────────────────────────────

class ConventionWorker:

    def __init__(self, monday: ConventionMondayClient) -> None:
        self.monday = monday

    async def process(self, item_id: int) -> None:
        log_ctx = {"item_id": item_id}
        logger.info("convention.start", **log_ctx)

        try:
            item = await self.monday.get_item(item_id)
            logger.info("convention.row_loaded", name=item.name, **log_ctx)

            # Mark as processing immediately
            await self.monday.set_review_status(item_id, "processing")
            await self.monday.set_date(item_id, "date_submitted", date.today())

            await self._run(item, log_ctx)

        except Exception as exc:
            logger.exception("convention.failed", **log_ctx)
            try:
                await self.monday.set_review_status(item_id, "needs_clarification")
                await self.monday.set_claude_questions(
                    item_id,
                    f"Processing error — please re-submit or contact admin.\n"
                    f"Details: {type(exc).__name__}: {str(exc)[:500]}",
                )
            except Exception:
                pass

    async def _run(self, item: RulebookItem, log_ctx: dict) -> None:
        item_id = item.item_id

        # 1. Download the latest submitted appendix
        if not item.appendix_files:
            await self.monday.set_review_status(item_id, "needs_clarification")
            await self.monday.set_claude_questions(
                item_id,
                "No appendix file found on this row. Please upload the completed "
                "Accuracy Report XLSX and re-set Submit Trigger to 'Submit for Review'."
            )
            return

        # Download most recent appendix (last in list)
        latest_asset = item.appendix_files[-1]
        asset_id     = latest_asset["asset_id"]
        url          = item.asset_urls.get(asset_id)
        if not url:
            raise RuntimeError(f"No download URL for asset {asset_id}")

        with tempfile.TemporaryDirectory() as tmp:
            dest      = Path(tmp) / (latest_asset.get("name") or "appendix.xlsx")
            await self.monday.download_asset(url, dest)
            xlsx_bytes = dest.read_bytes()

        logger.info("convention.appendix_downloaded",
                    bytes=len(xlsx_bytes), **log_ctx)

        # 2. Parse the appendix
        entries = parse_appendix(xlsx_bytes)
        logger.info("convention.entries_parsed",
                    count=len(entries), **log_ctx)

        if not entries:
            await self.monday.set_review_status(item_id, "needs_clarification")
            await self.monday.set_claude_questions(
                item_id,
                "No rows with convention tags found in the appendix. "
                "Please fill in the Convention column (columns F) for rows you want "
                "codified, then re-submit."
            )
            return

        # 3. Call Claude
        protocol = item.protocol_number or item.name
        result   = await call_claude_for_conventions(entries, protocol)
        logger.info("convention.claude_response",
                    status=result.get("status"),
                    questions=len(result.get("questions", [])),
                    conventions=len(result.get("conventions", [])),
                    **log_ctx)

        # 4. Handle response
        if result.get("status") == "needs_clarification":
            await self._handle_needs_clarification(
                item_id, item, xlsx_bytes, result, log_ctx
            )
        else:
            await self._handle_complete(
                item_id, item, result, protocol, log_ctx
            )

    async def _handle_needs_clarification(
        self,
        item_id: int,
        item: RulebookItem,
        xlsx_bytes: bytes,
        result: dict,
        log_ctx: dict,
    ) -> None:
        """Write questions into the XLSX col G, re-upload, update board."""
        questions = result.get("questions", [])

        # Write questions into the XLSX
        updated_xlsx = write_questions_to_xlsx(xlsx_bytes, questions)

        # Re-upload with questions embedded
        protocol  = item.protocol_number or item.name
        round_num = item.round_number + 1
        filename  = f"{protocol}_Accuracy_Report_Round{round_num}.xlsx"
        await self.monday.upload_file_to_column(
            item_id, "submitted_appendix", filename, updated_xlsx
        )

        # Write summary of questions to the Claude Questions column
        q_summary_lines = [
            f"Round {round_num} — {len(questions)} question(s):\n"
        ]
        for q in questions:
            q_summary_lines.append(
                f"• Row {q['row_num']} ({q['element']}): {q['question']}"
            )
        # Also note any conventions that WERE extracted this round
        if result.get("conventions"):
            q_summary_lines.append(
                f"\n{len(result['conventions'])} convention(s) are ready "
                f"and will be applied once clarifications are resolved."
            )
        await self.monday.set_claude_questions(item_id, "\n".join(q_summary_lines))
        await self.monday.set_round(item_id, round_num)
        await self.monday.set_review_status(item_id, "needs_clarification")
        await self.monday.set_trigger(item_id, "awaiting_human")

        logger.info("convention.needs_clarification",
                    round=round_num, questions=len(questions), **log_ctx)

    async def _handle_complete(
        self,
        item_id: int,
        item: RulebookItem,
        result: dict,
        protocol: str,
        log_ctx: dict,
    ) -> None:
        """Write conventions to JSON store and mark row complete."""
        conventions = result.get("conventions", [])
        for conv in conventions:
            conv["source_study"] = protocol

        # Determine customer UUID from customer_uuids.csv if scope is customer
        customer_uuid = self._resolve_customer_uuid(item)

        # Load, update, save conventions
        loop = asyncio.get_event_loop()
        added = await loop.run_in_executor(
            None,
            lambda: self._apply_and_save(conventions, customer_uuid),
        )

        await self.monday.set_conventions_extracted(item_id, added)
        await self.monday.set_date(item_id, "date_completed", date.today())
        await self.monday.set_review_status(item_id, "added_to_rulebook")
        await self.monday.set_trigger(item_id, "awaiting_human")
        # Clear questions column now that we're done
        await self.monday.set_claude_questions(
            item_id,
            f"✓ Complete — {added} convention(s) written to conventions.json "
            f"on {date.today().isoformat()}."
        )

        logger.info("convention.complete",
                    added=added, protocol=protocol, **log_ctx)

    @staticmethod
    def _apply_and_save(conventions: list[dict], customer_uuid: str | None) -> int:
        data = load_conventions()
        added = apply_conventions(data, conventions, customer_uuid)
        save_conventions(data)
        return added

    @staticmethod
    def _resolve_customer_uuid(item: RulebookItem) -> str | None:
        """
        Try to find the customer UUID from customer_uuids.csv matching
        the protocol's sponsor. Best-effort — returns None if not found.
        """
        csv_path = Path(
            os.environ.get("CUSTOMER_UUIDS_PATH",
                           "/app/references/customer_uuids.csv")
        )
        if not csv_path.exists():
            return None
        import csv
        protocol = (item.protocol_number or "").upper()
        try:
            with open(csv_path) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Match by subdomain or external_customer_id containing protocol
                    if protocol and protocol in str(row.get("name", "")).upper():
                        return row.get("uuid")
        except Exception:
            pass
        return None


# ── Module entry point (called by queue/webhook) ──────────────────────────────

async def process_convention_job(item_id: int) -> None:
    """Called by the webhook handler when Submit Trigger fires on rulebook board."""
    monday = ConventionMondayClient()
    worker = ConventionWorker(monday=monday)
    try:
        await worker.process(item_id)
    finally:
        await monday.aclose()

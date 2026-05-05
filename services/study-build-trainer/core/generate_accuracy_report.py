"""
generate_accuracy_report.py — Accuracy Scorer for Study Build Trainer

Compares AI-predicted EDC build against human-approved actual build.
Produces an XLSX with two sheets:
  Sheet 1: "Accuracy Scorecard" — per-layer scores + overall %
  Sheet 2: "Diff Appendix"     — row-by-row diffs, human-editable

Layers scored (and their sources):
  1. Study          — ODM XML study OID/name  vs  Study Spec JSON study_meta
  2. Events         — ODM XML StudyEventDef   vs  Study Spec JSON visits_assigned
  3. Form Placement — ODM XML FormRef/event   vs  Study Spec JSON form visits_assigned
  4. Forms          — actual XLSForm ZIP       vs  predicted EDC ZIP (settings sheet)
  5. Items          — actual XLSForm ZIP       vs  predicted EDC ZIP (survey sheet)
  6. Choices        — actual XLSForm ZIP       vs  predicted EDC ZIP (choices sheet)
  7. Logic          — actual XLSForm ZIP       vs  predicted EDC ZIP (relevant/constraint/calc)

Overall score = average of all 7 layer scores.

Entry point:
  generate_accuracy_report(
      actual_xml_bytes,        # ODM XML — human approved final build
      actual_xls_bytes,        # XLSForm ZIP — human approved final forms
      predicted_spec_json,     # Study Spec JSON dict from protocol-analysis skill
      predicted_edc_zip_bytes, # EDC build ZIP from edc-builder skill
      output_path,             # where to write the .xlsx
      study_name="",           # display name for the header
  ) -> dict   # {"overall_pct": float, "layer_scores": {...}}
"""

from __future__ import annotations

import io
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def _safe_val(v) -> str:
    """
    Safely convert a worksheet cell value to a plain string.
    Handles openpyxl MergedCell objects (non-top-left cells of a merged range)
    which appear when iterating over worksheets with merged cells.
    """
    try:
        from openpyxl.cell.cell import MergedCell
        if isinstance(v, MergedCell):
            return ""
    except ImportError:
        pass
    return str(v).strip() if v is not None else ""


# ── Brand colours (mirror pricing XLSX palette) ──────────────────────────────

OC_DARK   = "1B3A6B"
OC_MID    = "2E6DA4"
OC_LIGHT  = "D6E4F0"
OC_TEAL   = "00A99D"
OC_ORANGE = "F47920"
WHITE     = "FFFFFF"
GREY_L    = "F5F5F5"
GREY_M    = "CCCCCC"
AMBER     = "FFF3CD"
RED_C     = "CC0000"
GREEN_C   = "1A7A1A"

LAYER_COLORS = {
    "Study":          "E8EAF0",
    "Events":         "E3F2FD",
    "Form Placement": "E8F5E9",
    "Forms":          "FFF8E1",
    "Items":          "FCE4EC",
    "Choices":        "F3E5F5",
    "Logic":          "E0F7FA",
}

CONVENTION_OPTIONS = '"This customer only,All customers,Skip — not a convention"'


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fl(h): return PatternFill("solid", fgColor=h)
def _fn(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, italic=italic, color=color, size=size)
def _bd(color=GREY_M):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def _al(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_cell(ws, r, c, val, bg=OC_DARK, fg=WHITE, size=10, bold=True, span=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _fn(bold=bold, color=fg, size=size)
    cell.fill = _fl(bg)
    cell.alignment = _al()
    cell.border = _bd()
    if span:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r, end_column=c + span - 1)
    return cell

def _data_cell(ws, r, c, val, bg=WHITE, bold=False, color="000000",
               h="left", size=9, italic=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _fn(bold=bold, color=color, size=size, italic=italic)
    cell.fill = _fl(bg)
    cell.alignment = _al(h=h)
    cell.border = _bd()
    if fmt:
        cell.number_format = fmt
    return cell

def _cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── ODM XML parser ────────────────────────────────────────────────────────────

_ODM_NS = "http://www.cdisc.org/ns/odm/v1.3"

def _tag(local): return f"{{{_ODM_NS}}}{local}"

def _parse_odm(xml_bytes: bytes) -> dict:
    """
    Parse ODM XML into a structured dict.

    Returns:
      {
        "study_oid": str,
        "study_name": str,
        "events": {oid: {"name": str, "forms": [form_oid]}},
        "forms": {oid: {"name": str}},
        "items": {oid: {"name": str, "data_type": str, "label": str,
                        "codelist_oid": str|None}},
        "codelists": {oid: {"name": str,
                            "items": [{"value": str, "label": str}]}},
      }
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as e:
        raise ValueError(f"Invalid ODM XML: {e}") from e

    # Detect namespace
    ns = _ODM_NS if root.tag.startswith("{") else ""
    def t(local): return f"{{{ns}}}{local}" if ns else local

    study_el = root.find(t("Study")) or root.find("Study")
    if study_el is None:
        # Try direct children
        for child in root:
            if child.tag.endswith("Study"):
                study_el = child
                break
    if study_el is None:
        return {"study_oid": "", "study_name": "", "events": {},
                "forms": {}, "items": {}, "codelists": {}}

    study_oid  = study_el.get("OID", "")
    gv         = study_el.find(t("GlobalVariables")) or study_el.find("GlobalVariables")
    study_name = ""
    if gv is not None:
        sn = gv.find(t("StudyName")) or gv.find("StudyName")
        if sn is not None and sn.text:
            study_name = sn.text.strip()

    mdv = study_el.find(t("MetaDataVersion")) or study_el.find("MetaDataVersion")
    if mdv is None:
        return {"study_oid": study_oid, "study_name": study_name,
                "events": {}, "forms": {}, "items": {}, "codelists": {}}

    events    = {}
    forms     = {}
    items     = {}
    codelists = {}

    # Events
    for ev in mdv.findall(t("StudyEventDef")) or mdv.findall("StudyEventDef"):
        oid  = ev.get("OID", "")
        name = ev.get("Name", "")
        form_refs = []
        for fr in ev.findall(t("FormRef")) or ev.findall("FormRef"):
            form_refs.append(fr.get("FormOID", ""))
        events[oid] = {"name": name, "forms": form_refs}

    # Forms
    for fd in mdv.findall(t("FormDef")) or mdv.findall("FormDef"):
        oid  = fd.get("OID", "")
        name = fd.get("Name", "")
        forms[oid] = {"name": name}

    # Items
    for itd in mdv.findall(t("ItemDef")) or mdv.findall("ItemDef"):
        oid       = itd.get("OID", "")
        name      = itd.get("Name", "")
        data_type = itd.get("DataType", "")
        label     = ""
        q = itd.find(t("Question")) or itd.find("Question")
        if q is not None:
            tt = q.find(t("TranslatedText")) or q.find("TranslatedText")
            if tt is not None and tt.text:
                label = tt.text.strip()
        cl_ref = itd.find(t("CodeListRef")) or itd.find("CodeListRef")
        cl_oid = cl_ref.get("CodeListOID", "") if cl_ref is not None else ""
        items[oid] = {"name": name, "data_type": data_type,
                      "label": label, "codelist_oid": cl_oid}

    # Codelists
    for cl in mdv.findall(t("CodeList")) or mdv.findall("CodeList"):
        oid  = cl.get("OID", "")
        name = cl.get("Name", "")
        cl_items = []
        for cli in cl.findall(t("CodeListItem")) or cl.findall("CodeListItem"):
            val   = cli.get("CodedValue", "")
            lbl   = ""
            dc = cli.find(t("Decode")) or cli.find("Decode")
            if dc is not None:
                tt = dc.find(t("TranslatedText")) or dc.find("TranslatedText")
                if tt is not None and tt.text:
                    lbl = tt.text.strip()
            cl_items.append({"value": val, "label": lbl})
        codelists[oid] = {"name": name, "items": cl_items}

    return {
        "study_oid":  study_oid,
        "study_name": study_name,
        "events":     events,
        "forms":      forms,
        "items":      items,
        "codelists":  codelists,
    }


# ── XLSForm ZIP parser ────────────────────────────────────────────────────────

def _parse_xlsform_zip(zip_bytes: bytes) -> dict:
    """
    Parse a ZIP of XLSForm files. Supports both modern .xlsx (openpyxl)
    and legacy .xls (xlrd) workbooks. Apple resource-fork garbage in
    __MACOSX/ is skipped.

    Returns:
      {
        form_id: {
          "form_title": str,
          "survey":  [{"name": str, "type": str, "label": str,
                       "relevant": str, "constraint": str, "calculation": str}],
          "choices": [{"list_name": str, "name": str, "label": str}],
        }
      }
    """
    import openpyxl
    try:
        import xlrd
        _have_xlrd = True
    except ImportError:
        _have_xlrd = False

    def _read_sheets_openpyxl(content: bytes) -> dict[str, list[list]]:
        """Return {sheet_name: rows} from an .xlsx workbook."""
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        out = {sn: list(wb[sn].values) for sn in wb.sheetnames}
        wb.close()
        return out

    def _read_sheets_xlrd(content: bytes) -> dict[str, list[list]]:
        """Return {sheet_name: rows} from a legacy .xls workbook."""
        wb = xlrd.open_workbook(file_contents=content)
        out = {}
        for sn in wb.sheet_names():
            sh = wb.sheet_by_name(sn)
            rows = []
            for r in range(sh.nrows):
                rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
            out[sn] = rows
        return out

    result = {}
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            for name in zf.namelist():
                # Skip macOS resource forks and other non-data zip entries
                if name.startswith("__") or "/__MACOSX/" in name or "/._" in name \
                        or Path(name).name.startswith("._"):
                    continue

                lower = name.lower()
                if lower.endswith(".xlsx"):
                    reader = _read_sheets_openpyxl
                elif lower.endswith(".xls"):
                    if not _have_xlrd:
                        # xlrd not available — skip with no error so the rest of
                        # the zip still parses. Caller sees fewer forms than expected.
                        continue
                    reader = _read_sheets_xlrd
                else:
                    continue

                try:
                    sheets = reader(zf.read(name))
                except Exception:
                    continue

                # settings
                settings   = {}
                form_id    = ""
                form_title = ""
                if "settings" in sheets:
                    rows = sheets["settings"]
                    if len(rows) >= 2:
                        hdrs = [_safe_val(h) for h in rows[0]]
                        vals = [_safe_val(v) for v in rows[1]]
                        settings   = dict(zip(hdrs, vals))
                        form_id    = settings.get("form_id", "")
                        form_title = settings.get("form_title", "")

                # Strip F_ prefix from form_id for matching
                raw_form_id = form_id
                if form_id.upper().startswith("F_"):
                    form_id = form_id[2:]

                # Use filename as fallback form_id
                if not form_id:
                    form_id = Path(name).stem

                # survey
                survey = []
                if "survey" in sheets:
                    rows = sheets["survey"]
                    if rows:
                        hdrs = [str(h).strip().lower() if h else "" for h in rows[0]]
                        for row in rows[1:]:
                            if not any(v for v in row):
                                continue
                            d = {h: (str(v).strip() if v is not None else "")
                                 for h, v in zip(hdrs, row)}
                            if d.get("name") or d.get("type"):
                                survey.append({
                                    "name":        d.get("name", ""),
                                    "type":        d.get("type", ""),
                                    "label":       d.get("label", ""),
                                    "relevant":    d.get("relevant", ""),
                                    "constraint":  d.get("constraint", ""),
                                    "calculation": d.get("calculation", ""),
                                })

                # choices
                choices = []
                if "choices" in sheets:
                    rows = sheets["choices"]
                    if rows:
                        hdrs = [str(h).strip().lower() if h else "" for h in rows[0]]
                        for row in rows[1:]:
                            if not any(v for v in row):
                                continue
                            d = {h: (str(v).strip() if v is not None else "")
                                 for h, v in zip(hdrs, row)}
                            if d.get("list_name") or d.get("name"):
                                choices.append({
                                    "list_name": d.get("list_name", ""),
                                    "name":      d.get("name", ""),
                                    "label":     d.get("label", ""),
                                })

                result[form_id.upper()] = {
                    "form_title": form_title,
                    "raw_form_id": raw_form_id,
                    "survey":  survey,
                    "choices": choices,
                }
    except Exception as e:
        raise ValueError(f"Cannot parse XLSForm ZIP: {e}") from e

    return result


# ── Predicted build extractor ─────────────────────────────────────────────────

def _extract_predicted(spec_json: dict) -> dict:
    """
    Extract predicted Study/Event/Form Placement from Study Spec JSON.

    Returns:
      {
        "study_oid":  str,
        "study_name": str,
        "events":     {event_oid: True},         # set of known events
        "placements": {(event_oid, form_id): True},
        "forms":      {form_id: {"form_title": str}},
      }
    """
    meta       = spec_json.get("study_meta", {})
    study_oid  = meta.get("study_oid", meta.get("protocol_number", ""))
    study_name = meta.get("study_title", meta.get("protocol_number", ""))

    events     = {}
    placements = {}
    forms_out  = {}

    form_list = spec_json.get("forms", [])
    if isinstance(form_list, dict):
        # Some skill outputs use a dict keyed by form_id
        form_list = [{"form_id": k, **v} for k, v in form_list.items()]

    for form in form_list:
        form_id    = str(form.get("form_id", "")).upper()
        form_title = (form.get("form_title", "")
                      or form.get("settings", {}).get("form_title", ""))
        visits     = form.get("visits_assigned", [])

        forms_out[form_id] = {"form_title": form_title}

        for ev in visits:
            ev = str(ev).strip()
            if ev:
                events[ev] = True
                placements[(ev, form_id)] = True

    return {
        "study_oid":  study_oid,
        "study_name": study_name,
        "events":     events,
        "placements": placements,
        "forms":      forms_out,
    }


# ── Scoring ───────────────────────────────────────────────────────────────────

@dataclass
class DiffRow:
    layer:    str
    element:  str
    ai_value: str
    hu_value: str          # human/actual value
    note:     str = ""     # human-fillable notes
    match_info: str = ""   # scorer-written: normalized/fuzzy match details


def _score(matched: int, total: int) -> float:
    if total == 0:
        return 100.0
    return round(100.0 * matched / total, 1)


# ── Normalized + fuzzy matching helpers ───────────────────────────────────────
# Used by Forms / Items / Choices scorers to relax matching beyond exact-string.
# Strategy per layer (incremental rollout):
#   - Forms layer: normalized + fuzzy at threshold 0.75
#   - Items / Choices: planned (separate patches)
# Match credit tiers:
#   1.0  exact          — original strings identical
#   1.0  normalized     — equal after stripping prefixes/separators/case
#   <1.0 fuzzy          — rapidfuzz ratio above threshold; partial credit
#   miss                — below threshold; reported as unmatched

FUZZY_FORMS_THRESHOLD = 0.75

# Common form-OID prefixes seen in OpenClinica builds. Stripped before comparison.
_OID_PREFIX_RX = re.compile(r"^(?:f_|form_|ig_)+", re.IGNORECASE)


def _norm_oid(s: str) -> str:
    """Normalize an identifier for comparison.
    Lowercase, strip whitespace, remove common form prefixes (F_, FORM_, IG_),
    collapse separators (_, -, space) so AETERM, AE_TERM, AE-TERM all match.
    Returns the normalized string. Empty input returns empty.
    """
    if not s:
        return ""
    out = s.strip().lower()
    out = _OID_PREFIX_RX.sub("", out)
    out = re.sub(r"[_\-\s]+", "", out)
    return out


try:
    from rapidfuzz import fuzz as _rapidfuzz_fuzz
    _HAVE_RAPIDFUZZ = True
except ImportError:
    _HAVE_RAPIDFUZZ = False


def _fuzzy_score(a: str, b: str) -> float:
    """Return similarity score in 0.0-1.0. Uses rapidfuzz when available,
    falls back to difflib so missing dependency degrades gracefully."""
    if not a or not b:
        return 0.0
    if _HAVE_RAPIDFUZZ:
        # token_set_ratio handles word-order and substring variations well
        return _rapidfuzz_fuzz.token_set_ratio(a, b) / 100.0
    from difflib import SequenceMatcher
    return SequenceMatcher(None, a, b).ratio()


def _best_fuzzy_match(target: str, candidates: list[str], threshold: float
                      ) -> tuple[str | None, float]:
    """Find the candidate with highest similarity to target.
    Returns (candidate, score) if score >= threshold, else (None, best_score).
    Comparison is done on normalized forms; original string is returned.
    Empty target/candidates -> (None, 0.0).
    """
    if not target or not candidates:
        return None, 0.0
    norm_target = _norm_oid(target)
    if not norm_target:
        return None, 0.0
    best_cand = None
    best_score = 0.0
    for cand in candidates:
        score = _fuzzy_score(norm_target, _norm_oid(cand))
        if score > best_score:
            best_score, best_cand = score, cand
    if best_score >= threshold:
        return best_cand, best_score
    return None, best_score


# ── Title fuzzy + semantic matching helpers ───────────────────────────────────
# When OID-based matching fails (Tiers 1-3), the matcher tries to bridge naming
# conventions by comparing form titles. Two strategies:
#
#   Tier 4 — Title fuzzy: rapidfuzz token_set_ratio on the form_title strings.
#       Catches "Subject demographics" ↔ "Demographics and Baseline Characteristics".
#       Threshold 0.65, credit factor 0.85 (less confident than OID match).
#
#   Tier 5 — Semantic: cosine similarity of sentence embeddings of "OID — title".
#       Catches "DNBP — Diagnostic Nerve Block Procedure" ↔ "PROC — Procedures",
#       which fuzzy strings cannot bridge.
#       Threshold 0.70, credit factor 0.75 (least confident; semantic guesses
#       benefit most from human review).
#
# The embedder is optional — if no embed_fn is passed to score_forms, Tier 5
# is silently skipped. Tier 4 always runs as long as form_titles exist.

TITLE_FUZZY_THRESHOLD     = 0.65
TITLE_FUZZY_CREDIT_FACTOR = 0.85   # confidence discount

SEMANTIC_THRESHOLD        = 0.70
SEMANTIC_CREDIT_FACTOR    = 0.75   # confidence discount

# Type for a sync embedder: takes texts, returns one float vector per text.
EmbedFn = Callable[[list[str]], list[list[float]]] | None


def _norm_title(s: str) -> str:
    """Normalize a free-text title for comparison: lowercase, collapse
    whitespace and punctuation. Less aggressive than _norm_oid because
    titles are natural language."""
    if not s:
        return ""
    out = s.strip().lower()
    out = re.sub(r"\s+", " ", out)
    return out


def _title_fuzzy_score(a: str, b: str) -> float:
    """Similarity in 0.0-1.0 of two free-text titles. Reuses _fuzzy_score
    on normalized titles."""
    return _fuzzy_score(_norm_title(a), _norm_title(b))


def _best_title_match(target_title: str,
                      candidate_titles: dict[str, str],
                      threshold: float,
                      ) -> tuple[str | None, float]:
    """Find the candidate form_id whose title best matches target_title.
    Args:
      target_title: title of the form we're looking for a match for
      candidate_titles: {form_id: title} for unclaimed predicted forms
      threshold: minimum similarity to qualify as a match
    Returns:
      (best_form_id, score) where score >= threshold, else (None, best_score).
    """
    if not target_title or not candidate_titles:
        return None, 0.0
    best_id = None
    best_score = 0.0
    for cand_id, cand_title in candidate_titles.items():
        if not cand_title:
            continue
        score = _title_fuzzy_score(target_title, cand_title)
        if score > best_score:
            best_score, best_id = score, cand_id
    if best_score >= threshold:
        return best_id, best_score
    return None, best_score


def _cosine_similarity(v1: list[float], v2: list[float]) -> float:
    """Cosine similarity of two vectors. Assumes vectors may be unnormalized;
    handles the normalization inline. Returns 0.0 for any zero-length vector."""
    if not v1 or not v2 or len(v1) != len(v2):
        return 0.0
    dot = sum(a * b for a, b in zip(v1, v2))
    n1  = sum(a * a for a in v1) ** 0.5
    n2  = sum(b * b for b in v2) ** 0.5
    if n1 == 0.0 or n2 == 0.0:
        return 0.0
    return dot / (n1 * n2)


def _form_embed_text(form_id: str, form_title: str) -> str:
    """Canonical text representation of a form for embedding.
    Combines OID and title so the embedder sees both signals."""
    fid = (form_id or "").strip()
    ttl = (form_title or "").strip()
    if fid and ttl:
        return f"{fid} — {ttl}"
    return fid or ttl


def _best_semantic_match(target_text: str,
                         candidate_texts: dict[str, str],
                         embed_fn: EmbedFn,
                         threshold: float,
                         ) -> tuple[str | None, float]:
    """Find the candidate form_id whose embedding is closest to target's.
    Returns (None, 0.0) if embed_fn is None or any other failure.
    Falls through silently on errors so the scorer never crashes when
    semantic matching is unavailable.
    """
    if embed_fn is None or not target_text or not candidate_texts:
        return None, 0.0
    cand_ids   = list(candidate_texts.keys())
    cand_texts = [candidate_texts[cid] for cid in cand_ids]
    try:
        all_vecs = embed_fn([target_text] + cand_texts)
    except Exception:
        # Embedder unavailable mid-run — degrade silently
        return None, 0.0
    if not all_vecs or len(all_vecs) != 1 + len(cand_texts):
        return None, 0.0
    target_vec = all_vecs[0]
    best_id    = None
    best_score = 0.0
    for cid, vec in zip(cand_ids, all_vecs[1:]):
        sim = _cosine_similarity(target_vec, vec)
        if sim > best_score:
            best_score, best_id = sim, cid
    if best_score >= threshold:
        return best_id, best_score
    return None, best_score


def score_study(actual_odm: dict, predicted: dict) -> tuple[float, list[DiffRow]]:
    diffs   = []
    matched = 0
    total   = 2   # OID + name

    if actual_odm["study_oid"] == predicted["study_oid"]:
        matched += 1
    else:
        diffs.append(DiffRow("Study", "Study OID",
                             predicted["study_oid"], actual_odm["study_oid"]))

    # Normalise names for comparison
    a_name = actual_odm["study_name"].strip().lower()
    p_name = predicted["study_name"].strip().lower()
    if a_name == p_name or not a_name or not p_name:
        matched += 1
    else:
        diffs.append(DiffRow("Study", "Study Name",
                             predicted["study_name"], actual_odm["study_name"]))

    return _score(matched, total), matched, total, diffs


def score_events(actual_odm: dict, predicted: dict) -> tuple[float, list[DiffRow]]:
    """Score event-level match. Matches by OID."""
    actual_ev    = set(actual_odm["events"].keys())
    predicted_ev = set(predicted["events"].keys())
    diffs        = []

    total   = len(actual_ev)
    matched = 0

    for ev in sorted(actual_ev):
        name = actual_odm["events"][ev]["name"]
        if ev in predicted_ev:
            matched += 1
        else:
            diffs.append(DiffRow("Events", f"Event: {name} ({ev})",
                                 "— missing —", ev))

    for ev in sorted(predicted_ev - actual_ev):
        diffs.append(DiffRow("Events", f"Extra event: {ev}",
                             ev, "— not in actual —"))

    return _score(matched, total), matched, total, diffs


def score_form_placement(actual_odm: dict, predicted: dict) -> tuple[float, list[DiffRow]]:
    """Score (event, form) assignment pairs."""
    # Build actual placements from ODM
    actual_pairs = set()
    for ev_oid, ev in actual_odm["events"].items():
        for form_oid in ev["forms"]:
            # Normalise form OID: strip F_ prefix
            fid = form_oid[2:].upper() if form_oid.upper().startswith("F_") else form_oid.upper()
            actual_pairs.add((ev_oid, fid))

    predicted_pairs = set(predicted["placements"].keys())
    diffs           = []
    total           = len(actual_pairs)
    matched         = 0

    for (ev, fid) in sorted(actual_pairs):
        if (ev, fid) in predicted_pairs:
            matched += 1
        else:
            ev_name   = actual_odm["events"].get(ev, {}).get("name", ev)
            diffs.append(DiffRow(
                "Form Placement",
                f"{fid} → {ev_name}",
                "— missing —",
                f"{fid} assigned to {ev}",
            ))

    for (ev, fid) in sorted(predicted_pairs - actual_pairs):
        diffs.append(DiffRow(
            "Form Placement",
            f"Extra: {fid} → {ev}",
            f"{fid} assigned to {ev}",
            "— not in actual —",
        ))

    return _score(matched, total), matched, total, diffs


def score_forms(actual_xls: dict, predicted_xls: dict,
                embed_fn: EmbedFn = None,
                ) -> tuple[float, list[DiffRow]]:
    """Score form-level match by form_id.

    Match tiers (from strongest to weakest):
      1. EXACT       — actual form_id == predicted form_id verbatim
      2. NORMALIZED  — equal after _norm_oid (strips F_/FORM_/IG_ prefixes,
                       lowercase, collapses separators). Full credit (1.0).
      3. FUZZY       — rapidfuzz token_set_ratio on OIDs >= FUZZY_FORMS_THRESHOLD.
                       Partial credit equal to similarity score.
      4. TITLE       — fuzzy match on form_title text >= TITLE_FUZZY_THRESHOLD.
                       Partial credit = score * TITLE_FUZZY_CREDIT_FACTOR.
      5. SEMANTIC    — cosine similarity of "OID — title" embeddings >=
                       SEMANTIC_THRESHOLD. Requires embed_fn; silently skipped
                       when embed_fn is None.
                       Partial credit = score * SEMANTIC_CREDIT_FACTOR.
      6. MISS        — no candidate qualifies. Reports best near-misses in
                       match_info for reviewer context.

    A predicted form may only be claimed by one actual form. The first actual
    form (sorted) gets the strongest available match.

    Args:
      actual_xls:    parsed XLSForm zip from human-approved build
      predicted_xls: parsed XLSForm zip from Claude-generated predicted build
      embed_fn:      optional sync embedding function. Takes a list of strings,
                     returns one float-vector per string. When None, Tier 5
                     is skipped silently. Failures inside embed_fn are caught
                     and treated as "no semantic match" so the scorer never
                     crashes on embedder unavailability.
    """
    actual_ids    = sorted(actual_xls.keys())
    predicted_ids = list(predicted_xls.keys())
    diffs         = []
    total         = len(actual_ids)
    matched_score = 0.0   # fractional sum (allows partial credit for fuzzy)
    claimed       = set()  # predicted_ids already matched to an actual

    # Pre-build normalized index of predicted ids for fast tier-2 lookups
    pred_by_norm: dict[str, list[str]] = {}
    for pid in predicted_ids:
        pred_by_norm.setdefault(_norm_oid(pid), []).append(pid)

    for fid in actual_ids:
        a_title = actual_xls[fid]["form_title"]

        # Tier 1: exact
        if fid in predicted_xls and fid not in claimed:
            claimed.add(fid)
            matched_score += 1.0
            p_title = predicted_xls[fid]["form_title"]
            if (a_title.strip().lower() != p_title.strip().lower()
                    and a_title and p_title):
                diffs.append(DiffRow(
                    "Forms", f"Title mismatch: {fid}",
                    p_title, a_title,
                    match_info="exact",
                ))
            continue

        # Tier 2: normalized equality
        norm = _norm_oid(fid)
        norm_candidates = [c for c in pred_by_norm.get(norm, [])
                           if c not in claimed]
        if norm_candidates:
            matched_pid = norm_candidates[0]
            claimed.add(matched_pid)
            matched_score += 1.0
            p_title = predicted_xls[matched_pid]["form_title"]
            diffs.append(DiffRow(
                "Forms", f"Form: {fid}",
                matched_pid, fid,
                match_info=f"normalized → {matched_pid}",
            ))
            continue

        # Tier 3: fuzzy on OID
        unclaimed = [pid for pid in predicted_ids if pid not in claimed]
        best_pid, best_score = _best_fuzzy_match(
            fid, unclaimed, FUZZY_FORMS_THRESHOLD,
        )
        if best_pid is not None:
            claimed.add(best_pid)
            matched_score += best_score   # partial credit
            p_title = predicted_xls[best_pid]["form_title"]
            diffs.append(DiffRow(
                "Forms", f"Form: {fid} ({a_title})",
                best_pid, fid,
                match_info=f"fuzzy {best_score:.2f} → {best_pid}",
            ))
            continue

        # Tier 4: title fuzzy match
        unclaimed_titles = {pid: predicted_xls[pid]["form_title"]
                            for pid in predicted_ids if pid not in claimed}
        title_pid, title_score = _best_title_match(
            a_title, unclaimed_titles, TITLE_FUZZY_THRESHOLD,
        )
        if title_pid is not None:
            claimed.add(title_pid)
            credit = title_score * TITLE_FUZZY_CREDIT_FACTOR
            matched_score += credit
            p_title = predicted_xls[title_pid]["form_title"]
            # Truncate titles for the match_info display so the cell stays readable
            a_show = (a_title[:35] + "…") if len(a_title) > 36 else a_title
            p_show = (p_title[:35] + "…") if len(p_title) > 36 else p_title
            diffs.append(DiffRow(
                "Forms", f"Form: {fid} ({a_title})",
                title_pid, fid,
                match_info=(f'title-match {title_score:.2f} → {title_pid} '
                            f'("{p_show}" ≈ "{a_show}")'),
            ))
            continue

        # Tier 5: semantic embedding match (optional)
        unclaimed_for_semantic = {
            pid: _form_embed_text(pid, predicted_xls[pid]["form_title"])
            for pid in predicted_ids if pid not in claimed
        }
        sem_pid, sem_score = _best_semantic_match(
            _form_embed_text(fid, a_title),
            unclaimed_for_semantic,
            embed_fn,
            SEMANTIC_THRESHOLD,
        )
        if sem_pid is not None:
            claimed.add(sem_pid)
            credit = sem_score * SEMANTIC_CREDIT_FACTOR
            matched_score += credit
            p_title = predicted_xls[sem_pid]["form_title"]
            diffs.append(DiffRow(
                "Forms", f"Form: {fid} ({a_title})",
                sem_pid, fid,
                match_info=(f'semantic {sem_score:.2f} → {sem_pid} '
                            f'({p_title})'),
            ))
            continue

        # Tier 6: miss — report best near-misses for reviewer context
        miss_parts = []
        unclaimed_now = [pid for pid in predicted_ids if pid not in claimed]
        if unclaimed_now:
            _, near_oid = _best_fuzzy_match(fid, unclaimed_now, 0.0)
            if near_oid > 0:
                near_cand = max(unclaimed_now,
                                key=lambda c: _fuzzy_score(_norm_oid(fid),
                                                            _norm_oid(c)))
                miss_parts.append(f"oid {near_cand} ({near_oid:.2f})")
            # Best title near-miss
            t_titles = {pid: predicted_xls[pid]["form_title"]
                        for pid in unclaimed_now}
            _, near_title = _best_title_match(a_title, t_titles, 0.0)
            if near_title > 0:
                miss_parts.append(f"title {near_title:.2f}")
            # Best semantic near-miss (only if embedder available)
            if embed_fn is not None:
                t_semantic = {pid: _form_embed_text(pid, predicted_xls[pid]["form_title"])
                              for pid in unclaimed_now}
                _, near_sem = _best_semantic_match(
                    _form_embed_text(fid, a_title),
                    t_semantic, embed_fn, 0.0,
                )
                if near_sem > 0:
                    miss_parts.append(f"semantic {near_sem:.2f}")
        miss_info = ("no match — best: " + ", ".join(miss_parts)) if miss_parts else ""
        diffs.append(DiffRow(
            "Forms", f"Form: {fid} ({a_title})",
            "— missing —", fid,
            match_info=miss_info,
        ))

    # Forms in predicted that no actual claimed
    for pid in sorted(set(predicted_ids) - claimed):
        p_title = predicted_xls[pid]["form_title"]
        diffs.append(DiffRow(
            "Forms", f"Extra form: {pid} ({p_title})",
            pid, "— not in actual —",
        ))

    # Round score so it stays an int-like 0-100 percentage
    pct = 100.0 if total == 0 else round(100.0 * matched_score / total, 1)
    # Note: matched is reported as int for the scorecard; we round the
    # weighted matched_score for display purposes.
    matched_int = int(round(matched_score))
    return pct, matched_int, total, diffs


def score_items(actual_xls: dict, predicted_xls: dict) -> tuple[float, list[DiffRow]]:
    """Score item-level match across all forms.
    Caps diff rows at 30 per form so a single complex form (e.g. Biospecimen
    with 500+ custom items) doesn't swamp the diff sheet and hide other layers.
    Scoring still uses ALL items — only the diff row display is capped.
    """
    diffs   = []
    total   = 0
    matched = 0
    MAX_DIFFS_PER_FORM = 30

    DATA_TYPES = {"text", "integer", "decimal", "date", "datetime",
                  "time", "select_one", "select_multiple", "note", "calculate"}

    for fid in sorted(actual_xls.keys()):
        form_diff_count = 0
        a_items = {
            r["name"]: r for r in actual_xls[fid]["survey"]
            if r["name"] and r["type"].split(" ")[0].lower() in DATA_TYPES
        }
        p_items = {}
        if fid in predicted_xls:
            p_items = {
                r["name"]: r for r in predicted_xls[fid]["survey"]
                if r["name"] and r["type"].split(" ")[0].lower() in DATA_TYPES
            }

        for name in sorted(a_items.keys()):
            total += 1
            a = a_items[name]
            if name in p_items:
                p = p_items[name]
                a_type = a["type"].split(" ")[0].lower()
                p_type = p["type"].split(" ")[0].lower()
                if a_type == p_type:
                    matched += 1
                elif form_diff_count < MAX_DIFFS_PER_FORM:
                    diffs.append(DiffRow(
                        "Items", f"{fid}.{name} — data type",
                        f"{p_type}",
                        f"{a_type}",
                        "Data type differs",
                    ))
                    form_diff_count += 1
            elif form_diff_count < MAX_DIFFS_PER_FORM:
                diffs.append(DiffRow(
                    "Items", f"{fid}.{name}",
                    "— missing —",
                    f"{name} ({a['type']})",
                ))
                form_diff_count += 1

        if form_diff_count >= MAX_DIFFS_PER_FORM:
            diffs.append(DiffRow(
                "Items", f"{fid} — (capped)",
                f"… {len(a_items) - MAX_DIFFS_PER_FORM}+ more diffs",
                "See full form comparison",
                f"Diff capped at {MAX_DIFFS_PER_FORM} rows per form",
            ))

        for name in sorted(set(p_items) - set(a_items)):
            if form_diff_count >= MAX_DIFFS_PER_FORM:
                break
            diffs.append(DiffRow(
                "Items", f"Extra: {fid}.{name}",
                f"{name} ({p_items[name]['type']})",
                "— not in actual —",
            ))
            form_diff_count += 1

    return _score(matched, total), matched, total, diffs


# ── Choice list pairing helpers ───────────────────────────────────────────────
# The Choices layer aggregates {list_name: {value: label}} across all forms,
# then needs to pair each actual list to a predicted list before comparing
# values inside. Naming-convention drift (RACE_OPT vs RACE) used to collapse
# entire lists to "missing." This pairing layer relaxes that.
#
# Tier ladder for list-name pairing (mirrors Forms but adds set-overlap):
#   1. EXACT          list_name verbatim
#   2. NORMALIZED     equal under _norm_oid
#   3. FUZZY          rapidfuzz token_set_ratio on list_name >= 0.75
#   4. VALUE-OVERLAP  Jaccard on the SET of value codes >= 0.50.
#                     Two lists named differently but containing the same
#                     codes (e.g. {Y,N} vs {Y,N}) almost certainly match.
#   5. SEMANTIC       cosine sim of "list_name + value labels" embeddings
#                     >= 0.70. Optional; needs embed_fn.
#
# Once a list pair is chosen, value-level comparison runs as before:
# matched += (val in paired_list); the "credit" comes from how many values
# inside the paired list actually align. This means a wrong pairing hurts
# the score automatically — you don't need partial-credit math at the list
# level.

CHOICE_LIST_FUZZY_THRESHOLD     = 0.75
CHOICE_LIST_JACCARD_THRESHOLD   = 0.50
CHOICE_LIST_SEMANTIC_THRESHOLD  = 0.70


def _jaccard_overlap(a: set, b: set) -> float:
    """Jaccard coefficient |a∩b|/|a∪b|. Returns 0 for empty sets."""
    union = a | b
    if not union:
        return 0.0
    return len(a & b) / len(union)


def _list_embed_text(list_name: str, values: dict) -> str:
    """Canonical text rep of a choice list for embedding.
    Combines list_name with up to 20 sorted value/label pairs so the
    embedder sees both the structural tag and the semantic content."""
    parts = [list_name]
    if values:
        for v, lbl in sorted(values.items())[:20]:
            if lbl and lbl != v:
                parts.append(f"{v}: {lbl}")
            else:
                parts.append(v)
    return " | ".join(parts)


def _pair_choice_lists(actual_lists: dict, predicted_lists: dict,
                       ) -> dict:
    """For each actual list_name, find the best matching predicted list_name.

    Returns:
      {actual_ln: (predicted_ln_or_None, tier_label, score)}

    A predicted list may only be claimed once. First actual (sorted) wins.
    Tier_label is the human-readable string for the Match Info column;
    'no match — best: ...' messages communicate misses.

    Pairing tiers (semantic Tier 5 was removed in Patch 3.2 because batch
    embedding of 100+ choice list strings caused OOM on the trainer
    container — see Patch 3 commit history. If you want semantic matching
    here in the future, the right approach is to chunk the embedding batch,
    not retry the all-at-once approach):
      1. EXACT          list_name verbatim
      2. NORMALIZED     equal under _norm_oid
      3. FUZZY          rapidfuzz token_set_ratio on list_name >= 0.75
      4. VALUE-OVERLAP  Jaccard on the SET of value codes >= 0.50
    """
    pairings: dict = {}
    claimed: set  = set()

    actual_keys    = sorted(actual_lists.keys())
    predicted_keys = list(predicted_lists.keys())

    # Pre-build normalized index of predicted list names for Tier 2
    pred_by_norm: dict[str, list[str]] = {}
    for pln in predicted_keys:
        pred_by_norm.setdefault(_norm_oid(pln), []).append(pln)

    # ── Per-list pairing loop ───────────────────────────────────────────────
    for ln in actual_keys:
        a_vals = actual_lists[ln]
        a_value_set = set(a_vals.keys())

        # Tier 1: exact list_name
        if ln in predicted_lists and ln not in claimed:
            claimed.add(ln)
            pairings[ln] = (ln, "exact", 1.0)
            continue

        # Tier 2: normalized list_name
        norm = _norm_oid(ln)
        norm_candidates = [c for c in pred_by_norm.get(norm, [])
                           if c not in claimed]
        if norm_candidates:
            pln = norm_candidates[0]
            claimed.add(pln)
            pairings[ln] = (pln, f"normalized → {pln}", 1.0)
            continue

        # Tier 3: fuzzy on list_name
        unclaimed = [pln for pln in predicted_keys if pln not in claimed]
        best_pln, best_score = _best_fuzzy_match(
            ln, unclaimed, CHOICE_LIST_FUZZY_THRESHOLD,
        )
        if best_pln is not None:
            claimed.add(best_pln)
            pairings[ln] = (
                best_pln, f"fuzzy {best_score:.2f} → {best_pln}", best_score,
            )
            continue

        # Tier 4: value-set overlap (Jaccard on value codes)
        best_overlap_pln = None
        best_overlap_score = 0.0
        for pln in unclaimed:
            p_value_set = set(predicted_lists[pln].keys())
            overlap = _jaccard_overlap(a_value_set, p_value_set)
            if overlap > best_overlap_score:
                best_overlap_pln, best_overlap_score = pln, overlap
        if (best_overlap_pln is not None
                and best_overlap_score >= CHOICE_LIST_JACCARD_THRESHOLD):
            claimed.add(best_overlap_pln)
            pairings[ln] = (
                best_overlap_pln,
                f"value-overlap {best_overlap_score:.2f} → {best_overlap_pln}",
                best_overlap_score,
            )
            continue

        # Tier 5 (miss): capture near-misses across OID and value-overlap.
        # Semantic near-miss removed in Patch 3.2.
        miss_parts = []
        unclaimed_now = [pln for pln in predicted_keys if pln not in claimed]
        if unclaimed_now:
            _, near_oid = _best_fuzzy_match(ln, unclaimed_now, 0.0)
            if near_oid > 0:
                near_cand = max(unclaimed_now,
                    key=lambda c: _fuzzy_score(_norm_oid(ln), _norm_oid(c)))
                miss_parts.append(f"oid {near_cand} ({near_oid:.2f})")

            best_jac = 0.0
            for pln in unclaimed_now:
                p_set = set(predicted_lists[pln].keys())
                jac = _jaccard_overlap(a_value_set, p_set)
                if jac > best_jac:
                    best_jac = jac
            if best_jac > 0:
                miss_parts.append(f"value-overlap {best_jac:.2f}")
        miss_info = ("no match — best: " + ", ".join(miss_parts)) if miss_parts else ""
        pairings[ln] = (None, miss_info, 0.0)

    return pairings


def score_choices(actual_xls: dict, predicted_xls: dict,
                  embed_fn: EmbedFn = None,
                  ) -> tuple[float, list[DiffRow]]:
    """Score choice list match. Caps diffs at 20 per list.

    List-level pairing tiers (see _pair_choice_lists for full ladder):
      1. Exact list_name       3. Fuzzy list_name
      2. Normalized list_name  4. Value-set overlap

    Once paired, value-level comparison runs as before (exact name+label
    match). A future patch may add value-level fuzzy/semantic matching.

    Args:
      embed_fn: ACCEPTED FOR API STABILITY but NOT USED. Semantic Tier 5 was
                disabled in Patch 3.2 because batched embedding of 100+
                choice list strings caused OOM on the trainer container.
                The argument is kept so generate_accuracy_report's call site
                doesn't have to change. To re-enable semantic here in the
                future, the right approach is to chunk the embedding batch.
    """
    diffs   = []
    total   = 0
    matched_score = 0.0   # fractional sum (allows partial credit for fuzzy-label)
    MAX_DIFFS_PER_LIST = 20

    # Aggregate choice lists across all forms
    def _agg(xls):
        lists = {}  # list_name -> {value: label}
        for fid, form in xls.items():
            for ch in form["choices"]:
                ln = ch["list_name"]
                if ln not in lists:
                    lists[ln] = {}
                lists[ln][ch["name"]] = ch["label"]
        return lists

    actual_lists    = _agg(actual_xls)
    predicted_lists = _agg(predicted_xls)

    pairings = _pair_choice_lists(actual_lists, predicted_lists)

    for ln in sorted(actual_lists.keys()):
        a_vals = actual_lists[ln]
        list_diff_count = 0
        paired_pln, tier_label, _score_val = pairings[ln]

        if paired_pln is None:
            # Unpaired — count all values as miss; show pairing info on first
            # diff row only so the cell stays compact.
            total += len(a_vals)
            first = True
            for val, lbl in sorted(a_vals.items()):
                if list_diff_count >= MAX_DIFFS_PER_LIST:
                    diffs.append(DiffRow(
                        "Choices", f"{ln} — (capped)",
                        f"…{len(a_vals)-MAX_DIFFS_PER_LIST}+ more",
                        "See full list",
                        match_info="Capped",
                    ))
                    break
                diffs.append(DiffRow(
                    "Choices", f"{ln} — entire list missing",
                    "— missing —", f"{val}: {lbl}",
                    match_info=tier_label if first else "",
                ))
                first = False
                list_diff_count += 1
            continue

        # List paired — compare values inside it.
        # Patch 4 added value-level matching tiers. Within a paired list:
        #   1. Exact code match              full credit
        #   2. Exact label match (case-i)    full credit, when codes differ
        #   3. Fuzzy label match >= 0.75     partial credit
        #   4. Miss                          no credit
        # Tracking claimed_p prevents a predicted value from being matched
        # to two actual values via different tiers.
        p_vals = predicted_lists[paired_pln]
        claimed_p: set = set()
        first = True

        def _find_label_match(
            target_label: str, p_vals_d: dict, already_claimed: set,
            threshold: float = 0.75,
        ):
            """Find best label match in p_vals_d for target_label.
            Returns (pval, score, tier) or (None, best_score, '')."""
            if not target_label:
                return None, 0.0, ""
            t_norm = target_label.strip().lower()
            if not t_norm:
                return None, 0.0, ""
            # Tier 2: exact label (case-insensitive)
            for pv, pl in p_vals_d.items():
                if pv in already_claimed or not pl:
                    continue
                if pl.strip().lower() == t_norm:
                    return pv, 1.0, "exact-label"
            # Tier 3: fuzzy label
            best_pv, best_sc = None, 0.0
            for pv, pl in p_vals_d.items():
                if pv in already_claimed or not pl:
                    continue
                sc = _fuzzy_score(t_norm, pl.strip().lower())
                if sc > best_sc:
                    best_sc, best_pv = sc, pv
            if best_sc >= threshold:
                return best_pv, best_sc, "fuzzy-label"
            return None, best_sc, ""

        for val in sorted(a_vals.keys()):
            total += 1

            # Tier 1: exact code
            if val in p_vals and val not in claimed_p:
                claimed_p.add(val)
                matched_score += 1.0
                a_lbl = a_vals[val].strip().lower()
                p_lbl = p_vals[val].strip().lower()
                if (a_lbl != p_lbl and a_lbl and p_lbl
                        and list_diff_count < MAX_DIFFS_PER_LIST):
                    diffs.append(DiffRow(
                        "Choices", f"{ln}.{val} — label",
                        p_vals[val], a_vals[val],
                        note="Label differs",
                        match_info=tier_label if first else "",
                    ))
                    first = False
                    list_diff_count += 1
                continue

            # Tier 2/3: label match
            matched_pv, lbl_sc, lbl_tier = _find_label_match(
                a_vals[val], p_vals, claimed_p,
            )
            if matched_pv is not None:
                claimed_p.add(matched_pv)
                matched_score += lbl_sc
                if list_diff_count < MAX_DIFFS_PER_LIST:
                    elem = (f"{ln}.{val}" if ln == paired_pln
                            else f"{ln}.{val} (paired with {paired_pln})")
                    info_first = (tier_label + " | ") if first and tier_label else ""
                    diffs.append(DiffRow(
                        "Choices", elem,
                        f"{matched_pv}: {p_vals[matched_pv]}",
                        f"{val}: {a_vals[val]}",
                        match_info=f"{info_first}{lbl_tier} {lbl_sc:.2f} → {matched_pv}",
                    ))
                    first = False
                    list_diff_count += 1
                continue

            # Tier 4: miss
            if list_diff_count < MAX_DIFFS_PER_LIST:
                elem = (f"{ln}.{val}" if ln == paired_pln
                        else f"{ln}.{val} (paired with {paired_pln})")
                diffs.append(DiffRow(
                    "Choices", elem,
                    "— missing —", f"{val}: {a_vals[val]}",
                    match_info=tier_label if first else "",
                ))
                first = False
                list_diff_count += 1

        # Extras: predicted values not claimed by any tier
        for val in sorted(set(p_vals) - claimed_p):
            if list_diff_count >= MAX_DIFFS_PER_LIST:
                break
            elem = (f"Extra: {ln}.{val}" if ln == paired_pln
                    else f"Extra: {paired_pln}.{val} (paired with {ln})")
            diffs.append(DiffRow(
                "Choices", elem,
                f"{val}: {p_vals[val]}", "— not in actual —",
            ))
            list_diff_count += 1

    return _score(int(round(matched_score)), total), int(round(matched_score)), total, diffs


def score_logic(actual_xls: dict, predicted_xls: dict) -> tuple[float, list[DiffRow]]:
    """
    Score logic match (relevant, constraint, calculation).
    Checks presence, not exact XPath equality (which would be too strict).
    An item gets credit if both actual and predicted have logic (or both don't).
    """
    diffs   = []
    total   = 0
    matched = 0

    LOGIC_COLS = ("relevant", "constraint", "calculation")
    DATA_TYPES = {"text", "integer", "decimal", "date", "datetime",
                  "time", "select_one", "select_multiple", "note", "calculate"}

    def _has_logic(row):
        return any(row.get(c, "").strip() for c in LOGIC_COLS)

    def _logic_summary(row):
        parts = []
        for c in LOGIC_COLS:
            v = row.get(c, "").strip()
            if v:
                parts.append(f"{c}: {v[:60]}{'…' if len(v)>60 else ''}")
        return " | ".join(parts) or "—"

    for fid in sorted(actual_xls.keys()):
        a_survey = {
            r["name"]: r for r in actual_xls[fid]["survey"]
            if r["name"] and r["type"].split(" ")[0].lower() in DATA_TYPES
        }
        p_survey = {}
        if fid in predicted_xls:
            p_survey = {
                r["name"]: r for r in predicted_xls[fid]["survey"]
                if r["name"] and r["type"].split(" ")[0].lower() in DATA_TYPES
            }

        for name in sorted(a_survey.keys()):
            a = a_survey[name]
            if not _has_logic(a) and name not in p_survey:
                continue  # neither has logic, skip

            total += 1
            a_logic = _has_logic(a)
            p_logic = _has_logic(p_survey.get(name, {}))

            if a_logic == p_logic:
                matched += 1
            elif a_logic and not p_logic:
                diffs.append(DiffRow(
                    "Logic", f"{fid}.{name} — missing logic",
                    "— no logic —",
                    _logic_summary(a),
                    "Actual has logic; AI prediction missing",
                ))
            else:
                diffs.append(DiffRow(
                    "Logic", f"{fid}.{name} — extra logic",
                    _logic_summary(p_survey[name]),
                    "— no logic —",
                    "AI added logic not in actual build",
                ))

        # Also show predicted fields that have logic but don't exist in actual.
        # These are AI-generated logic expressions with no counterpart — visible
        # in AI Generated column so reviewers can see what Claude produced.
        for name in sorted(set(p_survey.keys()) - set(a_survey.keys())):
            p = p_survey[name]
            if _has_logic(p):
                diffs.append(DiffRow(
                    "Logic", f"{fid}.{name} — AI-only field with logic",
                    _logic_summary(p),
                    "— field not in actual —",
                    "AI generated logic on a field not in the actual build",
                ))

    return _score(matched, total), matched, total, diffs


# ── XLSX builder ──────────────────────────────────────────────────────────────

def _build_scorecard_sheet(ws, layer_scores: dict, overall: float,
                           study_name: str, generated_date: str):
    """Build Sheet 1: Accuracy Scorecard."""
    _cw(ws, [28, 16, 16, 16, 30])
    row = 1

    # Title banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1,
                value="OpenClinica Study Build Trainer — Accuracy Scorecard")
    c.font = _fn(bold=True, color=WHITE, size=14)
    c.fill = _fl(OC_DARK)
    c.alignment = _al(h="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # Study info row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=f"Study: {study_name}")
    c.font = _fn(size=10, color=WHITE)
    c.fill = _fl(OC_MID)
    c.alignment = _al()
    ws.cell(row=row, column=4).fill = _fl(OC_MID)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    c2 = ws.cell(row=row, column=4, value=f"Generated: {generated_date}")
    c2.font = _fn(size=9, color=WHITE, italic=True)
    c2.fill = _fl(OC_MID)
    c2.alignment = _al(h="right")
    ws.row_dimensions[row].height = 18
    row += 1

    # Overall score band
    row += 1
    # Overall score band — split into label (cols 1-3) and value (cols 4-5)
    # Calculate the formula reference BEFORE writing cells
    score_start_row = row + 4  # actual data rows start after blank + header
    score_end_row   = score_start_row + len(layer_scores) - 1
    overall_cell    = f"=AVERAGE(C{score_start_row}:C{score_end_row})"

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1)
    c.value = "OVERALL ACCURACY SCORE"
    c.font  = _fn(bold=True, color=WHITE, size=13)
    c.fill  = _fl(OC_TEAL)
    c.alignment = _al(h="left")

    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    score_cell = ws.cell(row=row, column=4)
    score_cell.value        = overall_cell
    score_cell.font         = _fn(bold=True, color=WHITE, size=13)
    score_cell.fill         = _fl(OC_TEAL)
    score_cell.alignment    = _al(h="right")
    score_cell.number_format = '0.0"%"'
    ws.row_dimensions[row].height = 26
    row += 1

    row += 1  # blank
    # Column headers
    for col, (hdr, w) in enumerate([
        ("Layer", None), ("Matched", None), ("Score %", None),
        ("Total Elements", None), ("Notes", None),
    ], start=1):
        _hdr_cell(ws, row, col, hdr, bg=OC_MID)
    ws.row_dimensions[row].height = 18
    row += 1

    # Layer rows
    LAYER_NOTES = {
        "Study":          "Study OID and study name match",
        "Events":         "Study event definitions (visits) match",
        "Form Placement": "Form-to-event assignment match",
        "Forms":          "Form definitions match",
        "Items":          "Data items match (name + data type)",
        "Choices":        "Choice list values match",
        "Logic":          "Relevance, constraint, calculation presence match",
    }

    for layer, data in layer_scores.items():
        bg = LAYER_COLORS.get(layer, GREY_L)
        _data_cell(ws, row, 1, layer, bg=bg, bold=True, color=OC_DARK, size=9)
        _data_cell(ws, row, 2, data["matched"], bg=bg, h="center", size=9)
        score_cell = ws.cell(row=row, column=3, value=data["score"] / 100)
        score_cell.font        = _fn(bold=True, size=9,
                                    color=GREEN_C if data["score"] >= 80 else RED_C)
        score_cell.fill        = _fl(bg)
        score_cell.alignment   = _al(h="center")
        score_cell.border      = _bd()
        score_cell.number_format = '0.0%'
        _data_cell(ws, row, 4, data["total"], bg=bg, h="center", size=9)
        _data_cell(ws, row, 5, LAYER_NOTES.get(layer, ""), bg=bg,
                   italic=True, color="555555", size=8)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1  # spacer
    # Legend
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1,
                value="Score ≥ 80% = Good  |  50–79% = Review Needed  |  < 50% = Significant Gaps")
    c.font      = _fn(size=8, italic=True, color="555555")
    c.fill      = _fl(GREY_L)
    c.alignment = _al(h="center")
    ws.row_dimensions[row].height = 14

    ws.freeze_panes = "A5"


def _build_diff_sheet(ws, diffs: list[DiffRow]):
    """Build Sheet 2: Diff Appendix.

    Columns:
      1. Layer
      2. Element / Field
      3. AI Generated
      4. Human Approved
      5. Match Info             ← scorer-written: normalized/fuzzy match details
      6. Notes (Human)          ← human fills in
      7. Convention             ← dropdown: This customer only / All customers / Skip
      8. Claude Questions       ← Claude writes questions on re-submission
      9. Human Response         ← human answers Claude's questions
    """
    NC = 9
    _cw(ws, [18, 30, 32, 32, 28, 28, 22, 34, 34])
    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=1, value="APPENDIX — Build Accuracy Diff Detail")
    c.font      = _fn(bold=True, color=WHITE, size=12)
    c.fill      = _fl(OC_DARK)
    c.alignment = _al(h="center")
    ws.row_dimensions[row].height = 24
    row += 1

    # Subtitle
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=1,
                value=("Each row is a mismatch between AI prediction and human-approved build. "
                       "Match Info (col E) shows whether the scorer matched the row exactly, via "
                       "name normalization, or via fuzzy matching. Fill in Notes + Convention, "
                       "then upload to the Convention Rulebook board and set Submit Trigger → "
                       "Submit for Review. Claude will write questions in column H if any rows "
                       "need clarification — answer in column I and re-submit."))
    c.font      = _fn(size=9, italic=True, color="555555")
    c.fill      = _fl(GREY_L)
    c.alignment = _al()
    ws.row_dimensions[row].height = 20
    row += 1

    row += 1  # blank

    # Column headers
    headers = [
        "Layer",
        "Element / Field",
        "AI Generated",
        "Human Approved",
        "Match Info",
        "Notes (Human)",
        "Convention",
        "Claude Questions",
        "Human Response",
    ]
    for col, hdr in enumerate(headers, start=1):
        bg = OC_MID
        # Highlight the two Q&A columns differently
        if col in (8, 9):
            bg = OC_DARK
        _hdr_cell(ws, row, col, hdr, bg=bg)
    ws.row_dimensions[row].height = 18
    data_start_row = row + 1
    row += 1

    # Column header notes row — sub-header explaining each input column
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1,
                value="← Generated by scorer — do not edit")
    c.font      = _fn(size=7, italic=True, color="888888")
    c.fill      = _fl(GREY_L)
    c.alignment = _al(h="center")

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    c = ws.cell(row=row, column=6,
                value="← Human fills these before submitting")
    c.font      = _fn(size=7, italic=True, color=OC_DARK)
    c.fill      = _fl("FFFDE7")
    c.alignment = _al(h="center")

    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    c = ws.cell(row=row, column=8,
                value="← Q&A between Claude and Human (multi-round)")
    c.font      = _fn(size=7, italic=True, color=WHITE)
    c.fill      = _fl(OC_DARK)
    c.alignment = _al(h="center")
    ws.row_dimensions[row].height = 13
    row += 1

    if not diffs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        c = ws.cell(row=row, column=1, value="✓ No differences found — perfect match!")
        c.font      = _fn(bold=True, color=GREEN_C, size=10)
        c.fill      = _fl(WHITE)
        c.alignment = _al(h="center")
        ws.row_dimensions[row].height = 20
        return

    # Data rows, grouped by layer
    current_layer = None
    for diff in diffs:
        bg = LAYER_COLORS.get(diff.layer, GREY_L)

        if diff.layer != current_layer:
            current_layer = diff.layer
            # Layer group header — spans all 9 columns
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
            c = ws.cell(row=row, column=1, value=f"  {diff.layer.upper()}")
            c.font      = _fn(bold=True, color=WHITE, size=9)
            c.fill      = _fl(OC_DARK)
            c.alignment = _al()
            ws.row_dimensions[row].height = 16
            row += 1

        # Columns 1-5: scorer-generated, read-only visually
        _data_cell(ws, row, 1, diff.layer, bg=bg, bold=True, color=OC_DARK, size=8)
        _data_cell(ws, row, 2, diff.element, bg=bg, size=8)
        _data_cell(ws, row, 3, diff.ai_value, bg=bg, size=8,
                   color=RED_C if "missing" in diff.ai_value.lower() else "000000")
        _data_cell(ws, row, 4, diff.hu_value, bg=bg, size=8,
                   color=RED_C if "missing" in diff.hu_value.lower() else "000000")

        # Column 5: Match Info — scorer-written, italic gray
        _data_cell(ws, row, 5, diff.match_info, bg=bg,
                   italic=True, color="555555", size=8)

        # Column 6: Notes — pale yellow, human fills in
        _data_cell(ws, row, 6, diff.note, bg="FFFDE7" if not diff.note else AMBER,
                   italic=True, color="555555", size=8)

        # Column 7: Convention dropdown — pale yellow, human fills in
        c = ws.cell(row=row, column=7, value="")
        c.font      = _fn(size=8)
        c.fill      = _fl("FFFDE7")
        c.alignment = _al()
        c.border    = _bd()

        # Column 8: Claude Questions — light blue, Claude writes here on re-submission
        c = ws.cell(row=row, column=8, value="")
        c.font      = _fn(size=8, italic=True, color="003366")
        c.fill      = _fl("E8F4FD")   # pale blue — Claude's territory
        c.alignment = _al()
        c.border    = _bd()

        # Column 9: Human Response — pale green, human answers Claude's questions
        c = ws.cell(row=row, column=9, value="")
        c.font      = _fn(size=8)
        c.fill      = _fl("F0FFF0")   # pale green — human response
        c.alignment = _al()
        c.border    = _bd()

        ws.row_dimensions[row].height = 20
        row += 1

    # Convention dropdown validation on column G (was F before adding Match Info)
    dv = DataValidation(
        type="list",
        formula1=CONVENTION_OPTIONS,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid selection",
        error="Choose: 'This customer only', 'All customers', or 'Skip — not a convention'",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"G{data_start_row}:G{row}"

    # Footer
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=1,
                value=(f"Total differences: {len(diffs)}  |  "
                       f"Yellow = human input needed  |  "
                       f"Blue = Claude questions  |  "
                       f"Green = human responses to Claude"))
    c.font      = _fn(bold=True, size=8, color=OC_DARK)
    c.fill      = _fl(OC_LIGHT)
    c.alignment = _al()
    ws.row_dimensions[row].height = 14

    ws.freeze_panes = "A6"


# ── Public entry point ────────────────────────────────────────────────────────

def generate_accuracy_report(
    actual_xml_bytes: bytes,
    actual_xls_bytes: bytes,
    predicted_spec_json: dict,
    predicted_edc_zip_bytes: bytes,
    output_path: str,
    study_name: str = "",
    embed_fn: EmbedFn = None,
) -> dict:
    """
    Generate accuracy report XLSX.

    Args:
      actual_xml_bytes:        ODM XML (human-approved build)
      actual_xls_bytes:        XLSForm ZIP (human-approved forms)
      predicted_spec_json:     Study Spec JSON (from protocol-analysis skill)
      predicted_edc_zip_bytes: EDC Build ZIP (from edc-builder skill)
      output_path:             Where to write the .xlsx file
      study_name:              Display name for header
      embed_fn:                Optional sync embedding function used by the
                               Forms layer's semantic-match tier. Takes a list
                               of strings, returns one float-vector per string.
                               When None, semantic matching is skipped.

    Returns:
      {
        "overall_pct": float,
        "layer_scores": {layer: {"score": float, "matched": int, "total": int}},
        "diff_count": int,
      }
    """
    # 1. Parse inputs
    actual_odm   = _parse_odm(actual_xml_bytes)
    actual_xls   = _parse_xlsform_zip(actual_xls_bytes)
    predicted    = _extract_predicted(predicted_spec_json)
    predicted_xls = _parse_xlsform_zip(predicted_edc_zip_bytes)

    if not study_name:
        study_name = (actual_odm.get("study_name")
                      or predicted.get("study_name")
                      or "Unknown Study")

    # 2. Score each layer
    all_diffs: list[DiffRow] = []

    def _run(fn, *args):
        result = fn(*args)
        score, matched, total, diffs = result
        all_diffs.extend(diffs)
        return score, matched, total

    study_score,  study_matched,  study_total  = _run(score_study, actual_odm, predicted)
    event_score,  event_matched,  event_total  = _run(score_events, actual_odm, predicted)
    place_score,  place_matched,  place_total  = _run(score_form_placement, actual_odm, predicted)
    form_score,   form_matched,   form_total   = _run(score_forms, actual_xls, predicted_xls, embed_fn)
    # Logic runs before Items/Choices so its diffs appear first in the sheet
    # (Items and Choices can generate hundreds of rows that push Logic off the end)
    logic_score,  logic_matched,  logic_total  = _run(score_logic, actual_xls, predicted_xls)
    item_score,   item_matched,   item_total   = _run(score_items, actual_xls, predicted_xls)
    choice_score, choice_matched, choice_total = _run(score_choices, actual_xls, predicted_xls, embed_fn)


    layer_scores = {
        "Study":          {"score": study_score,  "matched": study_matched,  "total": study_total},
        "Events":         {"score": event_score,  "matched": event_matched,  "total": event_total},
        "Form Placement": {"score": place_score,  "matched": place_matched,  "total": place_total},
        "Forms":          {"score": form_score,   "matched": form_matched,   "total": form_total},
        "Items":          {"score": item_score,   "matched": item_matched,   "total": item_total},
        "Choices":        {"score": choice_score, "matched": choice_matched, "total": choice_total},
        "Logic":          {"score": logic_score,  "matched": logic_matched,  "total": logic_total},
    }

    overall = round(sum(d["score"] for d in layer_scores.values()) / len(layer_scores), 1)

    # 3. Build XLSX
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Accuracy Scorecard"
    ws1.sheet_properties.tabColor = OC_TEAL

    ws2 = wb.create_sheet(title="Diff Appendix")
    ws2.sheet_properties.tabColor = OC_ORANGE

    _build_scorecard_sheet(
        ws1, layer_scores, overall, study_name,
        date.today().strftime("%B %d, %Y"),
    )
    _build_diff_sheet(ws2, all_diffs)

    wb.save(output_path)

    return {
        "overall_pct":  overall,
        "layer_scores": {k: {"score": v["score"]} for k, v in layer_scores.items()},
        "diff_count":   len(all_diffs),
    }


# ── Standalone test ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    # Usage: python generate_accuracy_report.py actual.xml actual.zip predicted_spec.json predicted_edc.zip output.xlsx
    if len(sys.argv) != 6:
        print("Usage: generate_accuracy_report.py <actual_xml> <actual_zip> "
              "<predicted_spec_json> <predicted_edc_zip> <output_xlsx>")
        sys.exit(1)

    actual_xml_path, actual_zip_path, spec_json_path, pred_zip_path, out_path = sys.argv[1:]

    result = generate_accuracy_report(
        actual_xml_bytes        = open(actual_xml_path,  "rb").read(),
        actual_xls_bytes        = open(actual_zip_path,  "rb").read(),
        predicted_spec_json     = json.loads(open(spec_json_path).read()),
        predicted_edc_zip_bytes = open(pred_zip_path,    "rb").read(),
        output_path             = out_path,
    )
    print(f"Overall: {result['overall_pct']}%")
    for layer, data in result["layer_scores"].items():
        print(f"  {layer:20s}: {data['score']}%")
    print(f"Diffs: {result['diff_count']}")
    print(f"Written: {out_path}")

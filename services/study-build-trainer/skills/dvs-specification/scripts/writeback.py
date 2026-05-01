"""
writeback.py — DVS to XLSForm Write-Back
Reads an updated DVS xlsx, finds changed rows, and applies changes to
the target XLSForm files. Returns a write-back report.

Usage:
    from writeback import apply_dvs_writeback
    report = apply_dvs_writeback(dvs_path, forms_dir, output_dir)
"""

import os, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column mapping: DVS → XLSForm ─────────────────────────────────────────────
# Maps (Check Type) → which XLSForm survey column receives the Expression value
EXPRESSION_COL_MAP = {
    "Constraint":               "constraint",
    "Calculate + Constraint":   "constraint",
    "Required":                 "required",
    "Relevant":                 "relevant",
    "Derivation / Review Listing": "calculation",
    "Cross-Form Helper":        "calculation",
}

MESSAGE_COL_MAP = {
    "Constraint":               "constraint_message",
    "Calculate + Constraint":   "constraint_message",
    "Required":                 "required_message",
    "Relevant":                 "",   # no message column for relevant
}

# XLSForm columns that write-back is NEVER permitted to touch
PROTECTED_COLS = {
    "type", "name", "label", "bind::oc:itemgroup", "bind::oc:external",
    "appearance", "readonly", "hint", "repeat_count", "image"
}


def _read_dvs_oc4(dvs_path):
    """Read all rows from DVS_OC4 sheet. Returns list of dicts keyed by column header."""
    wb = load_workbook(dvs_path, data_only=True)
    if "DVS_OC4" not in wb.sheetnames:
        raise ValueError(f"No DVS_OC4 sheet found in {dvs_path}")
    ws = wb["DVS_OC4"]
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Header row is row 3 (index 2)
        if i == 2:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if headers is None or not any(c is not None for c in row):
            continue
        if i < 2:
            continue
        row_dict = {}
        for j, h in enumerate(headers):
            if h:
                val = row[j] if j < len(row) else None
                row_dict[h] = str(val).strip() if val is not None else ""
        if row_dict.get("Check ID", "").startswith("DVS-"):
            rows.append(row_dict)
    return rows


def _read_xlsform(form_path):
    """
    Read an XLSForm. Returns:
    { 'settings': {...}, 'survey_headers': [...], 'survey_rows': [{...}], 'choices': [...] }
    Also returns the column index map for the survey sheet.
    """
    wb = load_workbook(form_path, data_only=True)
    result = {"settings": {}, "survey_headers": [], "survey_rows": [], "choices": []}

    # Settings
    if "settings" in wb.sheetnames:
        ws = wb["settings"]
        hdr_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        val_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        for h, v in zip(hdr_row, val_row):
            if h:
                result["settings"][str(h).strip()] = str(v).strip() if v else ""

    # Survey
    if "survey" in wb.sheetnames:
        ws = wb["survey"]
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c else "" for c in row]
                result["survey_headers"] = headers
                continue
            if headers is None:
                continue
            if not any(c is not None for c in row):
                continue
            row_dict = {}
            for j, h in enumerate(headers):
                val = row[j] if j < len(row) else None
                row_dict[h] = str(val).strip() if val is not None else ""
            result["survey_rows"].append(row_dict)

    return result, wb


def _write_xlsform_row(wb, form_path, row_name, updates):
    """
    Apply a dict of {col_name: new_value} updates to the survey row
    where name == row_name. Returns True if row was found and updated.
    """
    if "survey" not in wb.sheetnames:
        return False
    ws = wb["survey"]
    headers = [str(c).strip() if c else "" for c in
               next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    name_col_idx = headers.index("name") if "name" in headers else None
    if name_col_idx is None:
        return False

    # Build header → column letter map
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}

    # Find the row
    target_row = None
    for row in ws.iter_rows(min_row=2):
        name_cell = row[name_col_idx]
        if name_cell.value and str(name_cell.value).strip() == row_name:
            target_row = row
            break

    if target_row is None:
        return False

    # Apply updates
    for col_name, new_val in updates.items():
        if col_name in PROTECTED_COLS:
            continue
        if col_name not in col_map:
            # Column doesn't exist yet — we can't add columns in this pass
            continue
        col_idx = col_map[col_name]
        target_row[col_idx - 1].value = new_val if new_val else None

    return True


def apply_dvs_writeback(dvs_path, forms_dir, output_dir):
    """
    Apply changes from an updated DVS back to XLSForm files.

    Args:
        dvs_path:   Path to the updated DVS xlsx
        forms_dir:  Directory containing the XLSForm .xlsx files
        output_dir: Where to write the updated XLSForms

    Returns:
        dict: write-back report
    """
    report = {
        "dvs_path":        dvs_path,
        "forms_dir":       forms_dir,
        "applied":         datetime.date.today().isoformat(),
        "changes_applied": [],
        "changes_skipped": [],
        "forms_modified":  [],
        "errors":          [],
    }

    # Read DVS
    try:
        dvs_rows = _read_dvs_oc4(dvs_path)
    except Exception as e:
        report["errors"].append(f"Could not read DVS: {e}")
        return report

    # Load all XLSForms in forms_dir
    form_files = {
        os.path.splitext(f)[0]: os.path.join(forms_dir, f)
        for f in os.listdir(forms_dir)
        if f.endswith(".xlsx")
    }

    modified_wbs = {}   # form_id → (wb, path)

    for dvs_row in dvs_rows:
        check_id    = dvs_row.get("Check ID", "")
        status      = dvs_row.get("Status", "")
        check_type  = dvs_row.get("Check Type", "")
        severity    = dvs_row.get("Severity", "")
        target_form = dvs_row.get("Target Form OID", "")
        target_item = dvs_row.get("Target Item OID", "")   # = name in XLSForm
        expression  = dvs_row.get("Expression / Calculation", "")
        message     = dvs_row.get("Constraint / Required / Relevant Message", "")

        if not check_id or not target_form or not target_item:
            report["changes_skipped"].append({
                "check_id": check_id,
                "reason": "Missing Check ID, Target Form OID, or Target Item OID"
            })
            continue

        # Find the form file
        if target_form not in form_files:
            report["changes_skipped"].append({
                "check_id": check_id,
                "reason": f"Form file '{target_form}.xlsx' not found in {forms_dir}"
            })
            continue

        # Load workbook (cache to avoid reloading)
        if target_form not in modified_wbs:
            try:
                wb = load_workbook(form_files[target_form])
                modified_wbs[target_form] = (wb, form_files[target_form])
            except Exception as e:
                report["errors"].append(f"{check_id}: Could not load {target_form}.xlsx — {e}")
                continue

        wb, form_path = modified_wbs[target_form]

        # Build the update dict
        updates = {}

        if status == "Retired":
            # Clear expression and message
            expr_col = EXPRESSION_COL_MAP.get(check_type, "")
            msg_col  = MESSAGE_COL_MAP.get(check_type, "")
            if expr_col:
                updates[expr_col] = ""
            if msg_col:
                updates[msg_col] = ""
            updates["bind::oc:constraint-type"] = ""
        else:
            # Apply expression
            expr_col = EXPRESSION_COL_MAP.get(check_type, "")
            if expr_col and expression:
                updates[expr_col] = expression

            # Apply message
            msg_col = MESSAGE_COL_MAP.get(check_type, "")
            if msg_col and message:
                updates[msg_col] = message

            # Handle severity Hard → bind::oc:constraint-type
            if check_type == "Constraint":
                if severity == "Hard":
                    updates["bind::oc:constraint-type"] = "hard"
                elif severity in ("Soft", "Informational"):
                    updates["bind::oc:constraint-type"] = ""

        if not updates:
            report["changes_skipped"].append({
                "check_id": check_id,
                "reason": "No applicable updates derived from DVS row"
            })
            continue

        # Apply to XLSForm
        success = _write_xlsform_row(wb, form_path, target_item, updates)
        if success:
            if target_form not in report["forms_modified"]:
                report["forms_modified"].append(target_form)
            report["changes_applied"].append({
                "check_id":    check_id,
                "form":        target_form,
                "item":        target_item,
                "updates":     updates,
            })
        else:
            report["changes_skipped"].append({
                "check_id": check_id,
                "reason":   f"Item '{target_item}' not found in {target_form} survey"
            })

    # Save modified workbooks
    os.makedirs(output_dir, exist_ok=True)
    for form_id, (wb, _) in modified_wbs.items():
        if form_id in report["forms_modified"]:
            out_path = os.path.join(output_dir, f"{form_id}.xlsx")
            wb.save(out_path)

    # Summary
    report["summary"] = {
        "dvs_rows_read":      len(dvs_rows),
        "changes_applied":    len(report["changes_applied"]),
        "changes_skipped":    len(report["changes_skipped"]),
        "forms_modified":     len(report["forms_modified"]),
    }

    return report


def build_writeback_report_xlsx(report, output_path):
    """Write the write-back report as a formatted xlsx file."""
    from openpyxl import Workbook
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "SUMMARY"
    ws["A1"] = "DVS Write-Back Report"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1B3A6B")
    ws.merge_cells("A1:D1")

    summary = report.get("summary", {})
    rows = [
        ("DVS File", report.get("dvs_path", "")),
        ("Applied Date", report.get("applied", "")),
        ("DVS Rows Read", summary.get("dvs_rows_read", 0)),
        ("Changes Applied", summary.get("changes_applied", 0)),
        ("Changes Skipped", summary.get("changes_skipped", 0)),
        ("Forms Modified", ", ".join(report.get("forms_modified", []))),
    ]
    for i, (label, val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=9)
        ws.cell(row=i, column=2, value=str(val)).font = Font(size=9)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    # Changes Applied sheet
    ws2 = wb.create_sheet("CHANGES_APPLIED")
    hdrs = ["Check ID", "Form", "Item", "Column Updated", "New Value"]
    for col_i, h in enumerate(hdrs, start=1):
        c = ws2.cell(row=1, column=col_i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="2E6DA4")
    for row_i, change in enumerate(report.get("changes_applied", []), start=2):
        for update_col, new_val in change.get("updates", {}).items():
            ws2.cell(row=row_i, column=1, value=change.get("check_id", ""))
            ws2.cell(row=row_i, column=2, value=change.get("form", ""))
            ws2.cell(row=row_i, column=3, value=change.get("item", ""))
            ws2.cell(row=row_i, column=4, value=update_col)
            ws2.cell(row=row_i, column=5, value=new_val)
    for col in ["A", "B", "C", "D", "E"]:
        ws2.column_dimensions[col].width = 28

    # Skipped sheet
    ws3 = wb.create_sheet("CHANGES_SKIPPED")
    hdrs3 = ["Check ID", "Reason"]
    for col_i, h in enumerate(hdrs3, start=1):
        c = ws3.cell(row=1, column=col_i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="C0392B")
    for row_i, skip in enumerate(report.get("changes_skipped", []), start=2):
        ws3.cell(row=row_i, column=1, value=skip.get("check_id", ""))
        ws3.cell(row=row_i, column=2, value=skip.get("reason", ""))
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 60

    wb.save(output_path)
    print(f"Write-back report: {output_path}")
    return output_path

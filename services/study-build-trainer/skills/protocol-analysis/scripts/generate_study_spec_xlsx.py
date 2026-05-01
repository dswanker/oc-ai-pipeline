"""
generate_xlsx.py — EDC Structure Specification XLSX Generator
Produces a human-editable specification workbook from the EDC structure JSON.

Workbook structure per form:
  - INDEX sheet (workbook-level summary)
  - For each form:
      [FORMID]_survey   — editable XLSForm survey rows
      [FORMID]_choices  — editable choice lists
      [FORMID]_settings — editable settings + metadata tab

When a human edits this file and feeds it back to the EDC structure skill,
the skill reads the changes and regenerates all three outputs (PDF, JSON, XLSX).

Usage:
    from generate_xlsx import build_edc_xlsx
    build_edc_xlsx(data_dict, output_path)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from dep_utils import (
    extract_all_form_dependencies, extract_row_dependencies,
    annotate_survey_with_dependencies, format_deps_short
)

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE_HEX   = "1B3A6B"
MID_BLUE_HEX    = "2E6DA4"
LIGHT_BLUE_HEX  = "D6E4F0"
WHITE_HEX       = "FFFFFF"
GREY_LIGHT_HEX  = "F5F5F5"
GREY_MID_HEX    = "CCCCCC"
AMBER_HEX       = "FFF3CD"
RED_LIGHT_HEX   = "FADBD8"
GREEN_LIGHT_HEX = "D5F5E3"
YELLOW_HEX      = "FFFF99"
ORANGE_HEX      = "FFD580"
TEAL_HEX        = "D0F0F0"

# ── Style helpers ─────────────────────────────────────────────────────────────
def hdr_font(bold=True, color=WHITE_HEX, size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)

def body_font(bold=False, color="1A1A1A", size=8):
    return Font(name="Arial", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=GREY_MID_HEX)
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(horizontal="left"):
    return Alignment(wrap_text=True, vertical="top", horizontal=horizontal)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row_num, col_count, bg=DARK_BLUE_HEX):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = hdr_font(color=WHITE_HEX)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = wrap_align("center")

def style_data_row(ws, row_num, col_count, bg=WHITE_HEX):
    alt = GREY_LIGHT_HEX if row_num % 2 == 0 else WHITE_HEX
    row_bg = bg if bg != WHITE_HEX else alt
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE_HEX):
            cell.fill = fill(row_bg)
        cell.font = body_font()
        cell.border = thin_border()
        cell.alignment = wrap_align()

def status_fill(status):
    s = str(status).upper()
    if "COMPLETE"    in s: return fill(GREEN_LIGHT_HEX)
    if "FLAGGED"     in s: return fill(AMBER_HEX)
    if "PLACEHOLDER" in s: return fill(RED_LIGHT_HEX)
    return fill(WHITE_HEX)

def safe(val):
    return str(val) if val is not None else ""


# ── INDEX sheet ───────────────────────────────────────────────────────────────
def build_index_sheet(wb, data):
    ws = wb.active
    ws.title = "INDEX"

    meta = data.get("study_meta", {})
    forms = data.get("forms", [])

    # Title banner
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = f"EDC STRUCTURE SPECIFICATION — {meta.get('protocol_number','')}  |  Generated: {datetime.date.today().strftime('%d %b %Y')}  |  PENDING HUMAN REVIEW"
    c.font = hdr_font(size=11, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Study meta block
    meta_rows = [
        ("Protocol Number",  meta.get("protocol_number", "")),
        ("Study ID",         meta.get("study_id", "")),
        ("Input Mode",       meta.get("input_mode", "PROTOCOL_ONLY")),
        ("Generated Date",   meta.get("generated_date", "")),
        ("Review Status",    "PENDING HUMAN REVIEW — DO NOT BUILD UNTIL APPROVED"),
        ("Total Forms",      str(len(forms))),
        ("Total Unique CRFs",str(len(forms))),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k).font = hdr_font(color="1A1A1A", bold=True)
        ws.cell(row=i, column=1).fill = fill(LIGHT_BLUE_HEX)
        ws.cell(row=i, column=2, value=v).font = body_font()
        ws.cell(row=i, column=2).fill = fill(WHITE_HEX)
        for col in [1, 2]:
            ws.cell(row=i, column=col).border = thin_border()
            ws.cell(row=i, column=col).alignment = wrap_align()

    # Instructions box
    instr_row = len(meta_rows) + 3
    ws.merge_cells(f"A{instr_row}:L{instr_row}")
    c = ws.cell(row=instr_row, column=1)
    c.value = "HOW TO USE THIS DOCUMENT"
    c.font = hdr_font(color=WHITE_HEX, size=10)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")

    instructions = [
        "1. REVIEW each form tab (named [FORMID]_survey, [FORMID]_choices, [FORMID]_settings)",
        "2. In the survey tab: FLAGGED rows are amber, PLACEHOLDER rows are red — these need your attention",
        "3. To ADD a field: insert a new row, set ACTION = ADD, fill in type, name, label and other columns",
        "4. To DELETE a field: set ACTION = DELETE in column A",
        "5. To MODIFY a field: just edit the cell values. Leave ACTION blank. Optionally explain why in NOTES_FOR_AI (column B)",
        "5. To CHANGE a field property: edit the relevant cell directly (constraint, relevant, label, etc.)",
        "6. To ADD a choice: go to the [FORMID]_choices tab and add a new row",
        "7. ACTION column (col A): blank = keep, DELETE = remove, ADD = new row inserted by reviewer",
        "8. NOTES_FOR_AI column (col B): optional free text — tell Claude why you made a change. Pre-populated with flag reasons for FLAGGED rows.",
        "8. When done: save this file and upload it back to Claude to regenerate all 3 outputs (PDF, JSON, XLSX)",
        "9. The [FORMID]_meta tab is READ-ONLY — it shows technical metadata for reference only",
    ]
    for i, instr in enumerate(instructions, start=instr_row + 1):
        ws.merge_cells(f"A{i}:L{i}")
        c = ws.cell(row=i, column=1)
        c.value = instr
        c.font = body_font(size=8)
        c.fill = fill(GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX)
        c.alignment = wrap_align()
        c.border = thin_border()

    # Form inventory table
    inv_row = instr_row + len(instructions) + 2
    ws.merge_cells(f"A{inv_row}:L{inv_row}")
    c = ws.cell(row=inv_row, column=1)
    c.value = "FORM INVENTORY"
    c.font = hdr_font(color=WHITE_HEX, size=10)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")

    inv_headers = ["#", "Form ID", "Form Title", "Category", "CDASH Domain",
                   "Arm", "Complexity", "Repeating", "ePRO",
                   "Re-uses", "CDASH Alignment", "Library Match", "Survey Tab", "Status"]
    inv_cols = [3, 10, 22, 12, 12, 12, 10, 9, 6, 8, 14, 14, 18, 16]

    hdr_row = inv_row + 1
    for col, (h, w) in enumerate(zip(inv_headers, inv_cols), start=1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = hdr_font(color=WHITE_HEX)
        c.fill = fill(MID_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        set_col_width(ws, get_column_letter(col), w)

    for i, form in enumerate(forms, start=1):
        r = hdr_row + i
        lm = form.get("library_match", {})
        n_flagged = sum(1 for s in form.get("survey", [])
                        if s.get("completion_status") in ("FLAGGED", "PLACEHOLDER"))
        status_str = "✓ Ready" if n_flagged == 0 else f"⚠ {n_flagged} items need review"
        # Category display mapping
        _cat_map = {
            "CDASH_CLINICAL":  "CDASH",
            "CDASH":           "CDASH",
            "CDASH_SAFETY":    "CDASH_SAFETY",
            "CDASH_COMPANION": "CDASH_COMPANION",
            "INFRASTRUCTURE":  "INFRA",
            "CUSTOM":          "CUSTOM",
        }
        _cat_display = _cat_map.get(form.get("form_category", ""),
                                    form.get("form_category", ""))
        # CDASH Alignment — always populated
        cdash_align = form.get("cdash_alignment", "")
        # Library Match — N/A when no library provided
        lm_status = lm.get("status", "")
        lib_match_display = (lm_status if lm_status and lm_status != "PROTOCOL_ONLY"
                             else "N/A — No library provided")
        row_vals = [
            str(i),
            form.get("form_id", ""),
            form.get("form_title", ""),
            _cat_display,
            form.get("cdash_domain", "") or "—",
            form.get("arm_applicability", ""),
            form.get("complexity", ""),
            "Yes" if form.get("has_repeating_group") else "No",
            "Yes" if form.get("is_epro") else "No",
            str(form.get("reuse_count", 0) or 0),
            cdash_align,
            lib_match_display,
            f"{form.get('form_id','')}_survey",
            status_str,
        ]
        for col, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = body_font()
            c.border = thin_border()
            c.alignment = wrap_align()
            bg = GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX
            c.fill = fill(bg)
        # colour status cell (col 14 after CDASH Alignment + Library Match)
        status_cell = ws.cell(row=r, column=14)
        if "⚠" in status_str:
            status_cell.fill = fill(AMBER_HEX)
            status_cell.font = Font(name="Arial", bold=True, color="7D5A00", size=8)
        else:
            status_cell.fill = fill(GREEN_LIGHT_HEX)
            status_cell.font = Font(name="Arial", bold=True, color="1A6B3A", size=8)

    ws.freeze_panes = "A2"


# ── Survey sheet ──────────────────────────────────────────────────────────────
SURVEY_EDITABLE_COLS = [
    ("ACTION",           "Leave blank = keep as-is (edit cells freely) | DELETE = remove row | ADD = new row added by reviewer. No MODIFY value needed — just edit cells and leave ACTION blank; optionally explain in NOTES_FOR_AI.", 12),
    ("NOTES_FOR_AI",     "Optional. Explain why you changed, added, or deleted this row. Claude reads this on the next run.", 32),
    ("type",             "XLSForm field type (e.g. text, integer, select_one NY)", 18),
    ("name",             "Machine-readable field ID (no spaces)", 18),
    ("label",            "User-visible question text", 32),
    ("bind::oc:itemgroup","CDASH domain group (e.g. AE, VS, LB)", 14),
    ("appearance",       "Layout hint (w1-w6, horizontal, minimal, multiline)", 16),
    ("relevant",         "Show/hide condition (XPath expression)", 30),
    ("required",         "yes / true() / expression", 12),
    ("constraint",       "Validation rule (XPath expression)", 30),
    ("constraint_message","Error message shown when constraint fails", 28),
    ("calculation",      "Auto-calculated value (XPath expression)", 30),
    ("readonly",         "yes or blank", 8),
    ("hint",             "Helper text shown below label", 24),
    ("repeat_count",     "Integer or expression (repeating groups only)", 12),
    ("bind::oc:external","External data source: clinicaldata / labranges / tpt", 20),
    ("choice_filter",    "Filter choices based on expression", 24),
    ("DEPENDENCIES",     "Auto-derived: [FormOID].[ItemOID] cross-form references (read-only)", 32),
]

def build_survey_sheet(wb, form):
    form_id = form.get("form_id", "FORM")
    sheet_name = f"{form_id}_survey"[:31]
    ws = wb.create_sheet(title=sheet_name)

    survey = form.get("survey", [])
    n_editable = len(SURVEY_EDITABLE_COLS)

    # Tab colour by status
    n_flagged = sum(1 for r in survey if r.get("completion_status") in ("FLAGGED", "PLACEHOLDER"))
    ws.sheet_properties.tabColor = "E67E22" if n_flagged > 0 else "27AE60"

    # Title banner
    ws.merge_cells(f"A1:{get_column_letter(n_editable)}1")
    c = ws["A1"]
    c.value = f"{form.get('form_title','')}  ({form_id})  —  SURVEY SHEET  |  Arm: {form.get('arm_applicability','')}  |  Visits: {len(form.get('visits_assigned',[]))}  |  Repeating: {'Yes' if form.get('has_repeating_group') else 'No'}  |  ePRO: {'Yes' if form.get('is_epro') else 'No'}"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    # Colour legend
    legend_items = [
        ("GREEN = COMPLETE", GREEN_LIGHT_HEX, "1A6B3A"),
        ("AMBER = FLAGGED — needs review", AMBER_HEX, "7D5A00"),
        ("RED = PLACEHOLDER — must be completed", RED_LIGHT_HEX, "8B1A1A"),
        ("YELLOW = FLAGGED by Claude — see NOTES_FOR_AI for reason", YELLOW_HEX, "5A4000"),
    ]
    for i, (txt, bg, fg) in enumerate(legend_items):
        col = i * 3 + 1
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
        c = ws.cell(row=2, column=col, value=txt)
        c.font = Font(name="Arial", bold=True, size=7, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 14

    # Header row
    for col, (hdr, hint, width) in enumerate(SURVEY_EDITABLE_COLS, start=1):
        c = ws.cell(row=3, column=col, value=hdr)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(DARK_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        c.comment = None
        set_col_width(ws, get_column_letter(col), width)
    ws.row_dimensions[3].height = 16

    # Data rows
    col_keys = [c[0] for c in SURVEY_EDITABLE_COLS]

    # Pre-annotate survey rows with extracted dependencies
    annotate_survey_with_dependencies(survey, form)

    for row_i, row in enumerate(survey, start=4):
        status = row.get("completion_status", "")
        base_fill = status_fill(status)

        for col_i, key in enumerate(col_keys, start=1):
            c = ws.cell(row=row_i, column=col_i)

            if key == "ACTION":
                c.value = ""
                c.fill = fill(WHITE_HEX)
            elif key == "NOTES_FOR_AI":
                # Pre-populate with Claude's flag_reason so the reviewer
                # can see why a row was flagged, then overwrite with their own note
                c.value = row.get("flag_reason", "") or ""
                c.fill = fill(AMBER_HEX) if row.get("flag_reason") else fill(WHITE_HEX)
            elif key == "DEPENDENCIES":
                # Auto-derived — show extracted deps, grey background (read-only)
                row_deps = row.get("dependencies", [])
                c.value = ", ".join(row_deps) if row_deps else ""
                c.fill = fill("E8F4FD") if row_deps else fill(GREY_LIGHT_HEX)
                c.font = body_font(bold=False, color="1B3A6B" if row_deps else "888888")
                c.border = thin_border()
                c.alignment = wrap_align()
                ws.row_dimensions[row_i].height = 14
                continue
            else:
                # Map JSON key to XLSForm column name
                json_key = key.replace("::", "__").replace(":", "_")
                # Try direct key match, then underscored version
                val = row.get(key) or row.get(json_key) or row.get(key.replace("::", "__")) or ""
                c.value = safe(val)
                c.fill = base_fill

            c.font = body_font(bold=(key == "name"))
            c.border = thin_border()
            c.alignment = wrap_align()

        # Row height
        ws.row_dimensions[row_i].height = 14

    # Freeze header
    ws.freeze_panes = "C4"   # Freeze ACTION + NOTES_FOR_AI; scroll starts at col C

    # Add validation for ACTION column
    dv = DataValidation(type="list", formula1='"ADD,DELETE,"', allow_blank=True)
    dv.sqref = f"A4:A{len(survey)+10}"
    ws.add_data_validation(dv)

    # Add validation for type column
    dv_type = DataValidation(
        type="list",
        formula1='"text,integer,decimal,date,select_one,select_multiple,note,calculate,begin group,end group,begin repeat,end repeat"',
        allow_blank=True
    )
    dv_type.sqref = f"C4:C{len(survey)+10}"
    ws.add_data_validation(dv_type)

    return sheet_name


# ── Choices sheet ─────────────────────────────────────────────────────────────
CHOICES_EDITABLE_COLS = [
    ("ACTION",      "Leave blank = keep as-is | DELETE = remove | ADD = new row. Edit cells freely without setting MODIFY.", 10),
    ("NOTES_FOR_AI","Optional. Explain changes to Claude.", 28),
    ("list_name",   "Choice list identifier (no spaces)", 16),
    ("label",       "User-visible option text", 28),
    ("name",        "Machine-readable option value (no spaces)", 20),
    ("source",      "STANDARD or PROTOCOL_SPECIFIC", 18),
    ("filter_column","Column name used for choice_filter", 14),
    ("filter_value", "Value for choice_filter matching", 20),
]

def build_choices_sheet(wb, form):
    form_id = form.get("form_id", "FORM")
    sheet_name = f"{form_id}_choices"[:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = "2E6DA4"

    choices = form.get("choices", [])
    n_cols = len(CHOICES_EDITABLE_COLS)

    # Title
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = f"{form.get('form_title','')}  ({form_id})  —  CHOICES SHEET  |  {len(choices)} choices in {len(set(c.get('list_name','') for c in choices))} lists"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    # Headers
    for col, (hdr, hint, width) in enumerate(CHOICES_EDITABLE_COLS, start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(DARK_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        set_col_width(ws, get_column_letter(col), width)
    ws.row_dimensions[2].height = 16

    # Group choices by list_name with alternating colours per list
    list_colours = [LIGHT_BLUE_HEX, TEAL_HEX, GREY_LIGHT_HEX, WHITE_HEX]
    list_order = []
    for ch in choices:
        ln = ch.get("list_name", "")
        if ln not in list_order:
            list_order.append(ln)

    for row_i, ch in enumerate(choices, start=3):
        ln = ch.get("list_name", "")
        list_idx = list_order.index(ln) % len(list_colours)
        row_bg = list_colours[list_idx]

        # Highlight protocol-specific choices
        if ch.get("source", "") == "PROTOCOL_SPECIFIC":
            row_bg = AMBER_HEX

        vals = [
            "",
            ch.get("list_name", ""),
            ch.get("label", ""),
            ch.get("name", ""),
            ch.get("source", ""),
            ch.get("filter_column", ""),
            ch.get("filter_value", ""),
            "",
        ]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=row_i, column=col_i, value=safe(val))
            c.font = body_font()
            c.fill = fill(row_bg)
            c.border = thin_border()
            c.alignment = wrap_align()
        ws.row_dimensions[row_i].height = 13

    ws.freeze_panes = "B3"

    # Add validation for ACTION
    dv = DataValidation(type="list", formula1='"ADD,DELETE,"', allow_blank=True)
    dv.sqref = f"A3:A{len(choices)+20}"
    ws.add_data_validation(dv)


# ── Settings + Meta sheet ─────────────────────────────────────────────────────
def build_settings_sheet(wb, form):
    form_id = form.get("form_id", "FORM")
    sheet_name = f"{form_id}_settings"[:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = "1B3A6B"

    settings = form.get("settings", {})
    lm = form.get("library_match", {})
    xdeps = form.get("cross_form_dependencies", [])

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"{form.get('form_title','')}  ({form_id})  —  SETTINGS & METADATA"
    c.font = hdr_font(size=10, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── EDITABLE SETTINGS ──
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "EDITABLE — XLSForm Settings"
    c.font = hdr_font(color=WHITE_HEX, size=9)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")

    editable_settings = [
        ("form_title",          settings.get("form_title", ""),   "Human-readable form name"),
        ("form_id",             settings.get("form_id", ""),      "CDASH code or custom ID (no spaces)"),
        ("version",             settings.get("version", "1"),     "Increment when form is updated"),
        ("style",               settings.get("style", "theme-grid"), "Always theme-grid for OpenClinica"),
        ("namespaces",          settings.get("namespaces", ""),   "OpenClinica namespace declaration"),
        ("crossform_references",settings.get("crossform_references", ""), "Leave blank or 'current_event'"),
    ]
    for i, (key, val, hint) in enumerate(editable_settings, start=3):
        ws.cell(row=i, column=1, value=key).font = hdr_font(color="1A1A1A", bold=True, size=8)
        ws.cell(row=i, column=1).fill = fill(LIGHT_BLUE_HEX)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = wrap_align()

        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        c = ws.cell(row=i, column=2, value=val)
        c.font = body_font()
        c.fill = fill(YELLOW_HEX)   # yellow = editable
        c.border = thin_border()
        c.alignment = wrap_align()

        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=6)
        c = ws.cell(row=i, column=5, value=hint)
        c.font = Font(name="Arial", italic=True, size=7, color="666666")
        c.fill = fill(GREY_LIGHT_HEX)
        c.border = thin_border()
        c.alignment = wrap_align()

    # Choice lists are shown in the dedicated [FORMID]_choices tab — not repeated here.
    meta_start_offset = len(editable_settings) + 4

    # ── DEPENDENCIES SUMMARY ──
    form_all_deps = extract_all_form_dependencies(form)
    dep_start = meta_start_offset
    ws.merge_cells(f"A{dep_start}:F{dep_start}")
    c = ws.cell(row=dep_start, column=1)
    c.value = (f"CROSS-FORM DEPENDENCIES — READ-ONLY REFERENCE  "
              f"({len(form_all_deps)} references)  "
              f"| To add a dependency: add a calculate row in the survey tab with "
              f"ACTION=ADD and the XPath in the calculation column")
    c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1B3A6B")
    c.alignment = Alignment(horizontal="left", vertical="center")
    if form_all_deps:
        for d_i, dep in enumerate(form_all_deps, start=dep_start+1):
            ws.merge_cells(f"A{d_i}:F{d_i}")
            c = ws.cell(row=d_i, column=1, value=dep)
            c.font = Font(name="Arial", size=8, color="1B3A6B")
            c.fill = PatternFill("solid", fgColor="EBF5FB" if d_i % 2 == 0 else "FFFFFF")
            c.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )
            c.alignment = Alignment(wrap_text=True, vertical="top")
        meta_start_offset = dep_start + 1 + len(form_all_deps) + 2
    else:
        c_none = ws.cell(row=dep_start+1, column=1, value="No cross-form dependencies identified")
        c_none.font = Font(name="Arial", italic=True, size=8, color="888888")
        ws.merge_cells(f"A{dep_start+1}:F{dep_start+1}")
        meta_start_offset = dep_start + 3

    # ── READ-ONLY METADATA ──
    meta_start = meta_start_offset
    ws.merge_cells(f"A{meta_start}:F{meta_start}")
    c = ws.cell(row=meta_start, column=1)
    c.value = "READ-ONLY — Technical Metadata (for reference)"
    c.font = hdr_font(color=WHITE_HEX, size=9)
    c.fill = fill(GREY_MID_HEX.replace("CC", "88") + "FF"[:0] or "888888")
    c.fill = fill("888888")
    c.alignment = Alignment(horizontal="left", vertical="center")

    metadata = [
        ("form_category",    form.get("form_category", "")),
        ("cdash_domain",     form.get("cdash_domain", "") or "—"),
        ("arm_applicability",form.get("arm_applicability", "")),
        ("complexity",       form.get("complexity", "")),
        ("has_repeating_group", str(form.get("has_repeating_group", False))),
        ("is_epro",          str(form.get("is_epro", False))),
        ("reuse_count",      str(form.get("reuse_count", 0))),
        ("pricing_summary_source", str(form.get("pricing_summary_source", False))),
        ("cdash_alignment",      form.get("cdash_alignment", "—")),
        ("library_match_status", (lm.get("status", "") or "N/A — No library provided")),
        ("library_source_type",  lm.get("source_type", "NONE")),
        ("visits_assigned",  ", ".join(form.get("visits_assigned", []))),
    ]
    for i, (key, val) in enumerate(metadata, start=meta_start + 1):
        ws.cell(row=i, column=1, value=key).font = body_font(bold=True, size=8)
        ws.cell(row=i, column=1).fill = fill(GREY_LIGHT_HEX)
        ws.cell(row=i, column=1).border = thin_border()
        ws.cell(row=i, column=1).alignment = wrap_align()

        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        c = ws.cell(row=i, column=2, value=val)
        c.font = body_font(color="555555", size=8)
        c.fill = fill(GREY_LIGHT_HEX)
        c.border = thin_border()
        c.alignment = wrap_align()

    # ── CROSS-FORM DEPENDENCIES ──
    if xdeps:
        dep_start = meta_start + len(metadata) + 2
        ws.merge_cells(f"A{dep_start}:F{dep_start}")
        c = ws.cell(row=dep_start, column=1)
        c.value = "CROSS-FORM DEPENDENCIES — READ-ONLY. To add a new dependency: insert a calculate row in the survey tab (ACTION=ADD). OID confirmation required for all entries after study configuration."
        c.font = hdr_font(color=WHITE_HEX, size=9)
        c.fill = fill(MID_BLUE_HEX)
        c.alignment = Alignment(horizontal="left", vertical="center")

        dep_hdrs = ["Source Form", "Source Field", "Purpose", "Visit Context", "XPath Pattern", "Status"]
        dep_widths = [12, 14, 26, 14, 30, 28]
        for col, (h, w) in enumerate(zip(dep_hdrs, dep_widths), start=1):
            c = ws.cell(row=dep_start+1, column=col, value=h)
            c.font = hdr_font(color=WHITE_HEX, size=8)
            c.fill = fill(DARK_BLUE_HEX)
            c.border = thin_border()
            c.alignment = wrap_align("center")
            set_col_width(ws, get_column_letter(col), w)

        for i, dep in enumerate(xdeps, start=dep_start+2):
            row_bg = AMBER_HEX  # all deps are flagged
            vals = [
                dep.get("source_form",""),
                dep.get("source_field",""),
                dep.get("purpose",""),
                dep.get("visit_context",""),
                dep.get("xpath_pattern",""),
                dep.get("status","FLAGGED — OID CONFIRMATION REQUIRED"),
            ]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=i, column=col, value=safe(val))
                c.font = body_font(size=7)
                c.fill = fill(row_bg)
                c.border = thin_border()
                c.alignment = wrap_align()

    # Column widths
    set_col_width(ws, "A", 22)
    set_col_width(ws, "B", 28)
    set_col_width(ws, "C", 28)
    set_col_width(ws, "D", 14)
    set_col_width(ws, "E", 18)
    set_col_width(ws, "F", 22)
    ws.freeze_panes = "A3"


# ── Supporting sheets ─────────────────────────────────────────────────────────
def build_conventions_sheet(wb, data):
    """
    Build the CONVENTIONS sheet per
    references/conventions.md §"Surfacing in the Study Specification" → D.

    Reads study_meta.conventions_applied. Inserts a new sheet between
    INDEX and TIMEPOINTS. Skips sheet creation entirely if conventions
    block is absent (legacy data).
    """
    ca = data.get("study_meta", {}).get("conventions_applied", {}) or {}
    if not ca:
        return

    ws = wb.create_sheet(title="CONVENTIONS")
    ws.sheet_properties.tabColor = "27AE60"  # green

    # Banner row
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = ("BUILD CONVENTIONS APPLIED — Defaults from references/conventions.md "
               "applied to this build. Override by editing per-form sheets, or by "
               "adding a study-specific override block to the customer library.")
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Source / version meta strip
    src_row = 2
    ws.cell(row=src_row, column=1, value="Conventions Source:")
    ws.cell(row=src_row, column=2, value=ca.get("source", "references/conventions.md"))
    ws.cell(row=src_row, column=3, value="Version:")
    ws.cell(row=src_row, column=4, value=str(ca.get("version", "1")))
    for col in range(1, 6):
        cell = ws.cell(row=src_row, column=col)
        cell.font = body_font(bold=(col % 2 == 1), size=8)
        cell.fill = fill(LIGHT_BLUE_HEX)
        cell.border = thin_border()
        cell.alignment = wrap_align()

    # Headers row
    hdrs = ["Convention", "Status", "Coverage", "Exemptions / Deviations", "Notes"]
    for col, h in enumerate(hdrs, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=9)
        c.fill = fill(DARK_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
    ws.row_dimensions[3].height = 18

    # Column widths
    set_col_width(ws, "A", 38)
    set_col_width(ws, "B", 14)
    set_col_width(ws, "C", 30)
    set_col_width(ws, "D", 32)
    set_col_width(ws, "E", 38)

    # Read sub-blocks safely
    fdc   = ca.get("future_date_constraint_applied", {}) or {}
    grp   = ca.get("group_wrapping_applied", {}) or {}
    cdash = ca.get("cdash_naming_applied", {}) or {}
    rmc   = ca.get("required_message_coverage", {}) or {}

    icf_applied   = bool(ca.get("icf_form_added_by_default", False))
    fd_const      = int(fdc.get("fields_constrained", 0) or 0)
    fd_exempt     = int(fdc.get("fields_exempted", 0) or 0)
    fd_exempts    = fdc.get("exemptions", []) or []
    grp_wrapped   = int(grp.get("forms_wrapped", 0) or 0)
    grp_name      = grp.get("single_section_group_name", "group0")
    cdash_using   = int(cdash.get("fields_using_cdash", 0) or 0)
    cdash_dev     = int(cdash.get("name_deviations", 0) or 0)
    cdash_devlist = cdash.get("deviations_list", []) or []
    upper_applied = bool(ca.get("uppercase_choice_lists", False))
    rm_required   = int(rmc.get("required_fields", 0) or 0)
    rm_with       = int(rmc.get("fields_with_message", 0) or 0)

    # §7 — common event placement
    cea          = ca.get("common_event_applied", {}) or {}
    cea_event    = cea.get("event_oid", "—")
    cea_forms    = cea.get("forms_in_common_event", []) or []
    cea_excluded = cea.get("forms_excluded_by_override", []) or []
    cea_added    = cea.get("conditional_forms_added", []) or []
    cea_skipped  = cea.get("conditional_forms_skipped", []) or []

    # §8–§19 metrics
    sec  = ca.get("soft_edit_checks_applied", {}) or {}
    sec_strict_req = int(sec.get("strict_required_count", 0) or 0)
    sec_strict_con = int(sec.get("strict_constraint_count", 0) or 0)

    pdr  = ca.get("pdate_for_recall_dates", {}) or {}
    pdr_pdate    = int(pdr.get("pdate_fields", 0) or 0)
    pdr_date     = int(pdr.get("date_fields", 0) or 0)
    pdr_xform    = pdr.get("rule_flagged_crossform_uses", []) or []

    aac  = ca.get("autocomplete_appearance", {}) or {}
    aac_p_elig   = int(aac.get("participate_lists_eligible", 0) or 0)
    aac_p_done   = int(aac.get("participate_lists_with_minimal", 0) or 0)
    aac_s_elig   = int(aac.get("site_lists_eligible", 0) or 0)
    aac_s_done   = int(aac.get("site_lists_with_minimal", 0) or 0)

    ext  = ca.get("external_csv_for_long_lists", {}) or {}
    ext_count    = int(ext.get("lists_exceeded_threshold", 0) or 0)
    ext_csvs     = ext.get("external_csvs_created", []) or []

    icc  = ca.get("item_count_caps", {}) or {}
    icc_site_over   = icc.get("site_forms_over_200", []) or []
    icc_partic_over = icc.get("participate_forms_over_50", []) or []

    bdc  = ca.get("briefdescription_coverage", {}) or {}
    bdc_done = int(bdc.get("applied_count", 0) or 0)
    bdc_total = int(bdc.get("total_data_rows", 0) or 0)
    bdc_missing = int(bdc.get("missing_count", 0) or 0)

    fse  = ca.get("form_style_explicit", {}) or {}
    fse_simple   = int(fse.get("site_simple_single", 0) or 0)
    fse_pages    = int(fse.get("site_simple_pages", 0) or 0)
    fse_grid     = int(fse.get("site_theme_grid", 0) or 0)
    fse_partic   = int(fse.get("participate_simple_pages", 0) or 0)
    fse_missing  = int(fse.get("missing_style", 0) or 0)

    cfr  = ca.get("crossform_references_populated", {}) or {}
    cfr_with_calc = int(cfr.get("forms_with_cross_form_calc", 0) or 0)
    cfr_with_xref = int(cfr.get("forms_with_crossform_references", 0) or 0)

    igk  = ca.get("itemgroup_keep_together", {}) or {}
    igk_records   = int(igk.get("repeating_logical_records", 0) or 0)
    igk_consistent = int(igk.get("repeating_records_consistent", 0) or 0)
    igk_deviations = igk.get("deviations", []) or []

    lik  = ca.get("likert_appearance_rule", {}) or {}
    lik_total      = int(lik.get("likert_fields", 0) or 0)
    lik_compliant  = int(lik.get("likert_compliant", 0) or 0)
    lik_noncompl   = lik.get("likert_non_compliant", []) or []

    vas  = ca.get("vas_appearance_rule", {}) or {}
    vas_total    = int(vas.get("vas_fields", 0) or 0)
    vas_vertical = int(vas.get("vas_vertical", 0) or 0)

    tbl  = ca.get("table_appearance_rule", {}) or {}
    tbl_total      = int(tbl.get("table_fields", 0) or 0)
    tbl_compliant  = int(tbl.get("table_compliant", 0) or 0)

    def _list_text(items):
        if not items:
            return "—"
        names = []
        for it in items[:3]:
            if isinstance(it, dict):
                names.append(f"{it.get('form','?')}.{it.get('field','?')}")
            else:
                names.append(str(it))
        s = ", ".join(names)
        if len(items) > 3:
            s += f" (+{len(items) - 3} more)"
        return s

    if rm_required == 0:
        rm_status, rm_fill = "—", WHITE_HEX
    elif rm_with == rm_required:
        rm_status, rm_fill = "Applied", GREEN_LIGHT_HEX
    else:
        rm_status, rm_fill = "Partial", AMBER_HEX

    rows = [
        ("1. Standalone ICF form (default)",
         "Applied" if icf_applied else "Skipped",
         "form added" if icf_applied else "—",
         "—",
         "Default behaviour per conventions.md §1",
         GREEN_LIGHT_HEX if icf_applied else RED_LIGHT_HEX),
        ("2. Future-date constraint on date fields",
         "Applied" if fd_exempt == 0 else "Partial",
         f"{fd_const} constrained, {fd_exempt} exempted",
         _list_text(fd_exempts),
         ("All date fields constrained" if fd_exempt == 0
          else "Exemptions intentional — see review_flags.constraint_review"),
         GREEN_LIGHT_HEX if fd_exempt == 0 else AMBER_HEX),
        ("3. begin/end group wrapping",
         "Applied",
         f"{grp_wrapped} forms wrapped",
         "—",
         f"Single-section forms use '{grp_name}'",
         GREEN_LIGHT_HEX),
        ("4. CDASH naming convention",
         "Applied" if cdash_dev == 0 else "Partial",
         f"{cdash_using} CDASH names, {cdash_dev} deviations",
         _list_text(cdash_devlist),
         ("No deviations" if cdash_dev == 0
          else "Customer-preferred names carried forward"),
         GREEN_LIGHT_HEX if cdash_dev == 0 else AMBER_HEX),
        ("5. UPPERCASE choice list names",
         "Applied" if upper_applied else "Skipped",
         "applied" if upper_applied else "—",
         "—",
         "All list_name values uppercase",
         GREEN_LIGHT_HEX if upper_applied else RED_LIGHT_HEX),
        ("6. required_message on required fields",
         rm_status,
         f"{rm_with} / {rm_required} required fields",
         "—",
         ("Auto-populated per conventions.md §6" if rm_required > 0
          else "No required fields in this build"),
         rm_fill),
        ("7. Common event with reactive forms",
         ("Applied" if cea_forms and not cea_excluded
          else ("Partial" if cea_forms and cea_excluded
                else "—")),
         (f"{cea_event}: {', '.join(cea_forms)}" if cea_forms else "—"),
         (", ".join(cea_excluded) if cea_excluded
          else (f"+{len(cea_added)} cond added, {len(cea_skipped)} cond skipped"
                if (cea_added or cea_skipped) else "—")),
         ("Override active — see forms_excluded_by_override" if cea_excluded
          else "Default placement per conventions.md §7"),
         (GREEN_LIGHT_HEX if cea_forms and not cea_excluded
          else (AMBER_HEX if cea_forms and cea_excluded
                else WHITE_HEX))),
        ("8. Soft edit checks default",
         ("Partial" if (sec_strict_req or sec_strict_con) else "Applied"),
         (f"{sec_strict_req} strict-req, {sec_strict_con} strict-con"
          if (sec_strict_req or sec_strict_con) else "All soft"),
         "—",
         ("Override(s) per conventions.md §8"
          if (sec_strict_req or sec_strict_con)
          else "All required/constraint rows soft per §8"),
         (AMBER_HEX if (sec_strict_req or sec_strict_con) else GREEN_LIGHT_HEX)),
        ("9. PDate for recall, Date for definite",
         (("Partial" if pdr_xform else "Applied") if (pdr_pdate or pdr_date) else "—"),
         (f"{pdr_pdate} PDate, {pdr_date} Date" if (pdr_pdate or pdr_date) else "No date fields"),
         (_list_text(pdr_xform) if pdr_xform else "—"),
         (f"{len(pdr_xform)} PDate field(s) used in cross-form calc — review"
          if pdr_xform else "Per conventions.md §9"),
         (AMBER_HEX if pdr_xform else (GREEN_LIGHT_HEX if (pdr_pdate or pdr_date) else WHITE_HEX))),
        ("10. Minimal autocomplete on long lists",
         ("Applied" if (aac_p_done == aac_p_elig and aac_s_done == aac_s_elig) else "Partial"),
         f"Participate {aac_p_done}/{aac_p_elig}, Site {aac_s_done}/{aac_s_elig}",
         "—",
         "Threshold: 5+ Participate, 20+ site (§10)",
         (GREEN_LIGHT_HEX if (aac_p_done == aac_p_elig and aac_s_done == aac_s_elig)
          else AMBER_HEX)),
        ("11. External CSV for long choice lists",
         "Applied",
         (f"{ext_count} list(s) externalized" if ext_count else "None needed"),
         (_list_text(ext_csvs) if ext_csvs else "—"),
         "Per conventions.md §11",
         GREEN_LIGHT_HEX),
        ("12. Item-count caps (build-time check)",
         ("Partial" if (icc_site_over or icc_partic_over) else "Applied"),
         (f"{len(icc_site_over)} site over 200, {len(icc_partic_over)} Participate over 50"
          if (icc_site_over or icc_partic_over) else "All within caps"),
         (_list_text([f["form_id"] for f in icc_site_over + icc_partic_over
                       if isinstance(f, dict)])
          if (icc_site_over or icc_partic_over) else "—"),
         ("Forms exceeding cap need review per §12"
          if (icc_site_over or icc_partic_over)
          else "Site ≤200, Participate ≤50 items per form"),
         (AMBER_HEX if (icc_site_over or icc_partic_over) else GREEN_LIGHT_HEX)),
        ("13. bind::oc:briefdescription coverage",
         (("Partial" if bdc_missing else "Applied") if bdc_total else "—"),
         (f"{bdc_done} / {bdc_total} data rows" if bdc_total else "—"),
         "—",
         (f"{bdc_missing} row(s) missing description" if bdc_missing
          else "Auto-populated per conventions.md §13"),
         (AMBER_HEX if bdc_missing else (GREEN_LIGHT_HEX if bdc_total else WHITE_HEX))),
        ("14. Form style declared explicitly",
         ("Partial" if fse_missing else "Applied"),
         f"S-single {fse_simple}, S-pages {fse_pages}, grid {fse_grid}, P-pages {fse_partic}",
         "—",
         (f"{fse_missing} form(s) missing style"
          if fse_missing else "Per conventions.md §14"),
         (AMBER_HEX if fse_missing else GREEN_LIGHT_HEX)),
        ("15. crossform_references on settings sheet",
         (("Applied" if cfr_with_calc == cfr_with_xref else "Partial")
          if cfr_with_calc else "—"),
         (f"{cfr_with_xref} / {cfr_with_calc} forms with cross-form calc"
          if cfr_with_calc else "No cross-form calc rows"),
         "—",
         ("Auto-populated from dependency graph (§15)"
          if cfr_with_calc else "No cross-form calc rows"),
         (GREEN_LIGHT_HEX if cfr_with_calc and cfr_with_calc == cfr_with_xref
          else (AMBER_HEX if cfr_with_calc else WHITE_HEX))),
        ("16. itemgroup keep-together (repeating)",
         (("Partial" if igk_deviations else "Applied") if igk_records else "—"),
         (f"{igk_consistent} / {igk_records} records consistent"
          if igk_records else "No repeating records"),
         (_list_text(igk_deviations) if igk_deviations else "—"),
         (f"{len(igk_deviations)} deviation(s) — review"
          if igk_deviations else "Per conventions.md §16"),
         (AMBER_HEX if igk_deviations else (GREEN_LIGHT_HEX if igk_records else WHITE_HEX))),
        ("17. Likert appearance ≤5 short labels",
         (("Partial" if lik_noncompl else "Applied") if lik_total else "—"),
         (f"{lik_compliant} / {lik_total} compliant" if lik_total else "No Likert fields"),
         (_list_text(lik_noncompl) if lik_noncompl else "—"),
         (f"{len(lik_noncompl)} override(s) carried"
          if lik_noncompl else "Per conventions.md §17"),
         (AMBER_HEX if lik_noncompl else (GREEN_LIGHT_HEX if lik_total else WHITE_HEX))),
        ("18. VAS scales rendered vertically",
         ("Applied" if vas_total else "—"),
         (f"{vas_vertical} / {vas_total} vertical" if vas_total else "No VAS fields"),
         "—",
         ("Per conventions.md §18" if vas_total
          else "Narrow rule — fires only when VAS exists"),
         (GREEN_LIGHT_HEX if vas_total else WHITE_HEX)),
        ("19. Table appearance short labels",
         (("Applied" if tbl_compliant == tbl_total else "Partial")
          if tbl_total else "—"),
         (f"{tbl_compliant} / {tbl_total} compliant"
          if tbl_total else "No table-appearance fields"),
         "—",
         ("Per conventions.md §19" if tbl_total
          else "Narrow rule — fires only when used"),
         (GREEN_LIGHT_HEX if tbl_total and tbl_compliant == tbl_total
          else (AMBER_HEX if tbl_total else WHITE_HEX))),
    ]

    for i, (conv, status, coverage, exempts, notes, status_bg) in enumerate(rows, start=4):
        bg = GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        cells = [conv, status, coverage, exempts, notes]
        for col, val in enumerate(cells, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(size=8, bold=(col == 2))
            c.border = thin_border()
            c.alignment = wrap_align()
            c.fill = fill(status_bg if col == 2 else bg)

    ws.freeze_panes = "A4"


def build_timepoint_sheet(wb, tpt_csv):
    ws = wb.create_sheet(title="TIMEPOINTS")
    ws.sheet_properties.tabColor = "1B3A6B"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"TIMEPOINT CSV — {tpt_csv.get('filename','')}  |  Feed this file to OpenClinica alongside the XLSForms"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    for col, h in enumerate(["event", "timepoint", "NOTES_FOR_AI"], start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(MID_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")

    set_col_width(ws, "A", 24)
    set_col_width(ws, "B", 34)
    set_col_width(ws, "C", 28)

    for i, row in enumerate(tpt_csv.get("rows", []), start=3):
        bg = GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        for col, val in enumerate([row.get("event",""), row.get("timepoint",""), ""], start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = wrap_align()

    ws.freeze_panes = "A3"


def build_labranges_sheet(wb, labranges):
    ws = wb.create_sheet(title="LAB_RANGES")
    ws.sheet_properties.tabColor = "E67E22"

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "LAB RANGES CSV — PLACEHOLDER — Complete all highlighted cells with site-specific values before building"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(MID_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    cols = labranges.get("columns", ["lab_name","test_code","test_name","lower","upper","unit","sex_filter","age_lower","age_upper"])
    widths = [20, 10, 24, 12, 12, 16, 10, 10, 10]

    for col_i, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=2, column=col_i, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(DARK_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        set_col_width(ws, get_column_letter(col_i), w)

    for i, row in enumerate(labranges.get("rows", []), start=3):
        bg = GREY_LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        for col_i, key in enumerate(cols, start=1):
            val = row.get(key, "")
            c = ws.cell(row=i, column=col_i, value=val)
            c.font = body_font()
            c.border = thin_border()
            c.alignment = wrap_align()
            # Highlight placeholders
            if "PLACEHOLDER" in str(val).upper():
                c.fill = fill(RED_LIGHT_HEX)
                c.font = Font(name="Arial", bold=True, size=8, color="8B1A1A")
            else:
                c.fill = fill(bg)

    ws.freeze_panes = "A3"


def build_review_flags_sheet(wb, flags):
    ws = wb.create_sheet(title="REVIEW_FLAGS")
    ws.sheet_properties.tabColor = "C0392B"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "REVIEW FLAGS — All items must be resolved before passing to edc-builder"
    c.font = hdr_font(size=9, color=WHITE_HEX)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    cat_colours = {
        "site_specific":        ("SITE SPECIFIC",         RED_LIGHT_HEX,   "8B1A1A"),
        "oid_confirmation":     ("OID CONFIRMATION",      AMBER_HEX,       "7D5A00"),
        "protocol_ambiguous":   ("PROTOCOL AMBIGUOUS",    AMBER_HEX,       "7D5A00"),
        "constraint_review":    ("CONSTRAINT REVIEW",     AMBER_HEX,       "7D5A00"),
        "choice_list_review":   ("CHOICE LIST REVIEW",    LIGHT_BLUE_HEX,  "1B3A6B"),
        "custom_domain":        ("CUSTOM DOMAIN",         LIGHT_BLUE_HEX,  "1B3A6B"),
        "pdf_mapping_uncertain":("PDF MAPPING UNCERTAIN", AMBER_HEX,       "7D5A00"),
        "name_deviation":       ("NAME DEVIATION",        LIGHT_BLUE_HEX,  "1B3A6B"),
    }

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 22)
    set_col_width(ws, "C", 80)
    set_col_width(ws, "D", 30)

    row = 2
    for key, (label, bg, fg) in cat_colours.items():
        items = flags.get(key, [])
        if not items:
            continue

        # Category header
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1)
        c.value = f"  {label}  ({len(items)} items)"
        c.font = Font(name="Arial", bold=True, size=9, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()
        ws.row_dimensions[row].height = 16
        row += 1

        for item in items:
            ws.cell(row=row, column=1, value="").fill = fill(bg)
            ws.cell(row=row, column=2, value="→").font = body_font(color=fg, bold=True)
            ws.cell(row=row, column=2).fill = fill(bg)
            ws.cell(row=row, column=2).border = thin_border()
            ws.cell(row=row, column=2).alignment = wrap_align("center")

            ws.merge_cells(f"C{row}:D{row}")
            c = ws.cell(row=row, column=3, value=str(item))
            c.font = body_font(size=8)
            c.fill = fill(GREY_LIGHT_HEX if row % 2 == 0 else WHITE_HEX)
            c.border = thin_border()
            c.alignment = wrap_align()
            ws.row_dimensions[row].height = 13
            row += 1

        row += 1  # gap between categories

    ws.freeze_panes = "A2"


# ── Main builder ──────────────────────────────────────────────────────────────

# ── AI Instructions sheet ─────────────────────────────────────────────────────

def build_ai_instructions_sheet(wb, data: dict):
    """
    Build the AI_INSTRUCTIONS sheet — a human-editable tab where the reviewer
    can give Claude structural instructions that apply to the next pipeline run.

    Sections:
      1. Study-Level Instructions  — apply to the whole study
      2. Form-Specific Instructions — apply to named forms
      3. Version History           — append-only audit trail (pipeline manages)

    The pipeline reads this sheet before passing the edited XLSX to Claude and
    injects the instructions as the highest-priority prompt block.
    """
    ws = wb.create_sheet(title="AI_INSTRUCTIONS")
    ws.sheet_properties.tabColor = "FF9800"   # orange — stands out

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = {"A": 14, "B": 72, "C": 18, "D": 32}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ── Title banner ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1,
                value="AI INSTRUCTIONS — Human-to-Claude guidance for next pipeline run")
    c.font = hdr_font(size=11, color=WHITE_HEX)
    c.fill = fill("FF9800")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # Intro text
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1,
                value=("Instructions in this tab are read by the pipeline before Claude "
                       "processes the study. They take priority over all other inputs. "
                       "Add instructions in Sections 1 and 2, then upload this file "
                       "to the 'Edited Study Specification XLSX' column in monday.com "
                       "and re-run."))
    c.font = Font(name="Arial", size=8, italic=True, color="5c4a00")
    c.fill = fill("FFF8E1")
    c.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 2

    # ── Section 1: Study-Level Instructions ──────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="SECTION 1 — STUDY-LEVEL INSTRUCTIONS")
    c.font = hdr_font(color=WHITE_HEX, size=9)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16
    row += 1

    # Section 1 headers
    s1_hdrs = [("PRIORITY", 14), ("INSTRUCTION", 72), ("STATUS", 18), ("NOTES", 32)]
    for col_i, (h, _) in enumerate(s1_hdrs, start=1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(MID_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
    ws.row_dimensions[row].height = 14
    _s1_header_row = row
    row += 1

    # Priority dropdown
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_priority = DataValidation(
        type="list", formula1='"High,Medium,Low"', allow_blank=True)

    # 8 blank instruction rows
    s1_start = row
    example_instructions = [
        ("High",   "Place AE, CM, and DV in their own dedicated visits — do not use a Common event for these forms.", "", ""),
        ("Medium", "Add a DOV (Date of Visit) form to every scheduled visit. Include VISYN (was visit done?) and show all other forms in that visit only when VISYN = Yes.", "", ""),
        ("",       "", "", ""),
        ("",       "", "", ""),
        ("",       "", "", ""),
        ("",       "", "", ""),
        ("",       "", "", ""),
        ("",       "", "", ""),
    ]
    for i, (pri, instr, status, notes) in enumerate(example_instructions):
        r = row + i
        is_example = bool(instr)
        bg = "FFF8E1" if is_example else WHITE_HEX
        vals = [pri, instr, status, notes]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.font = Font(name="Arial", size=8,
                          italic=is_example, color="5c4a00" if is_example else "000000")
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = wrap_align()
        ws.row_dimensions[r].height = 14
    s1_end = row + len(example_instructions) - 1
    row = s1_end + 1

    # Add priority dropdown to section 1
    dv_priority.sqref = f"A{s1_start}:A{s1_end + 20}"
    ws.add_data_validation(dv_priority)

    row += 1  # spacer

    # ── Section 2: Form-Specific Instructions ─────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="SECTION 2 — FORM-SPECIFIC INSTRUCTIONS")
    c.font = hdr_font(color=WHITE_HEX, size=9)
    c.fill = fill(DARK_BLUE_HEX)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16
    row += 1

    s2_hdrs = [("FORM OID", 14), ("INSTRUCTION", 72), ("STATUS", 18), ("NOTES", 32)]
    for col_i, (h, _) in enumerate(s2_hdrs, start=1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(MID_BLUE_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
    ws.row_dimensions[row].height = 14
    row += 1

    # Populate with one row per form as a starting point (blank instruction)
    forms = data.get("forms", [])
    for form in forms:
        fid = form.get("form_id", "")
        c = ws.cell(row=row, column=1, value=fid)
        c.font = body_font(bold=True, size=8)
        c.fill = fill(GREY_LIGHT_HEX)
        c.border = thin_border()
        c.alignment = wrap_align()
        for col_i in [2, 3, 4]:
            c = ws.cell(row=row, column=col_i, value="")
            c.font = body_font(size=8)
            c.fill = fill(WHITE_HEX)
            c.border = thin_border()
            c.alignment = wrap_align()
        ws.row_dimensions[row].height = 13
        row += 1

    # Extra blank rows for new forms
    for _ in range(5):
        for col_i in range(1, 5):
            c = ws.cell(row=row, column=col_i, value="")
            c.border = thin_border()
            c.fill = fill(WHITE_HEX)
        ws.row_dimensions[row].height = 13
        row += 1

    row += 1  # spacer

    # ── Section 3: Version History (append-only) ──────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1,
                value="SECTION 3 — VERSION HISTORY  (pipeline-managed, do not edit)")
    c.font = hdr_font(color=WHITE_HEX, size=9)
    c.fill = fill("888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16
    row += 1

    vh_hdrs = ["VERSION", "DATE", "STUDY INSTRUCTIONS", "FORM INSTRUCTIONS", "SUMMARY"]
    vh_widths = [14, 12, 18, 18, 40]
    for col_i, (h, w) in enumerate(zip(vh_hdrs, vh_widths), start=1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.font = hdr_font(color=WHITE_HEX, size=8)
        c.fill = fill(GREY_MID_HEX)
        c.border = thin_border()
        c.alignment = wrap_align("center")
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[row].height = 14
    row += 1

    # Seed with one "initial generation" row
    study_meta = data.get("study_meta", {})
    version = study_meta.get("generated_date", "")[:10] or "Initial"
    seed_vals = [version, "", "0", "0", "Initial generation — no AI instructions provided"]
    for col_i, val in enumerate(seed_vals, start=1):
        c = ws.cell(row=row, column=col_i, value=val)
        c.font = body_font(size=8, color="555555")
        c.fill = fill(GREY_LIGHT_HEX)
        c.border = thin_border()
        c.alignment = wrap_align()
    ws.row_dimensions[row].height = 13

    ws.freeze_panes = "A2"


def build_edc_xlsx(data: dict, output_path: str):
    wb = Workbook()

    # INDEX sheet (uses default active sheet)
    build_index_sheet(wb, data)

    # AI Instructions sheet — second tab so it's immediately visible
    build_ai_instructions_sheet(wb, data)

    # Supporting sheets
    build_timepoint_sheet(wb, data.get("timepoint_csv", {}))
    build_labranges_sheet(wb, data.get("labranges_csv", {}))
    build_review_flags_sheet(wb, data.get("review_flags", {}))

    # One set of tabs per form
    for form in data.get("forms", []):
        build_survey_sheet(wb, form)
        build_choices_sheet(wb, form)
        build_settings_sheet(wb, form)

    # APPENDIX — CONVENTIONS sheet (last sheet in workbook,
    # per references/conventions.md §"Surfacing" → D)
    build_conventions_sheet(wb, data)

    wb.save(output_path)
    print(f"EDC Structure XLSX written to: {output_path}")


# ── Test run ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import json, sys
    sys.path.insert(0, '/home/claude')
    # Load the full PrTK05 data from the existing test script
    exec(open('/home/claude/prtk05_edc_data.py').read().split('build_edc_pdf')[0])
    build_edc_xlsx(data, "/mnt/user-data/outputs/PrTK05_EDC_Structure.xlsx")

# ── Alias so the function can also be imported by its skill-level name ────
build_study_spec_xlsx = build_edc_xlsx


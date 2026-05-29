"""
build_xlsforms.py — OpenClinica XLSForm Generator
Generates one production-ready .xlsx file per CRF form from the EDC
structure specification.

Each output file is built from form_template.xlsx (must live alongside
this script in the same directory).  The template supplies:
  • Correctly structured settings / choices / survey sheets with OC headers
  • Six reference/instruction tabs (Cross-Form Examples, Custom Annotation
    Examples, Contact Info Examples, eConsent Examples, Hard Checks
    Examples, Offline Forms Examples)
  • bind::oc:external dropdown in the survey sheet

If form_template.xlsx is not found, the script falls back to building
the three functional sheets from scratch (no reference tabs, no dropdowns).

Usage:
    from build_xlsforms import build_all_xlsforms, build_single_xlsform
    results = build_all_xlsforms(spec_data, output_dir)
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os, re, datetime

# Path to the OC form template (must be in the same folder as this script)
_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_SCRIPT_DIR, 'form_template.xlsx')

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1B3A6B"
MID_BLUE   = "2E6DA4"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
GREY_LIGHT = "F5F5F5"
GREY_MID   = "CCCCCC"
AMBER      = "FFF3CD"

# ── Exact column order from OpenClinica template ──────────────────────────────
SURVEY_COLS = [
    "type", "name", "label", "bind::oc:itemgroup", "hint", "appearance",
    "bind::oc:briefdescription", "bind::oc:description", "relevant",
    "required", "required_message", "constraint", "constraint_message",
    "default", "calculation", "trigger", "readonly", "image",
    "repeat_count", "bind::oc:external"
]

CHOICES_COLS = ["list_name", "label", "name", "image"]

SETTINGS_COLS = [
    "form_title", "form_id", "version", "style",
    "crossform_references", "namespaces"
]

# Column widths for survey sheet
SURVEY_WIDTHS = {
    "type": 18, "name": 20, "label": 36, "bind::oc:itemgroup": 14,
    "hint": 24, "appearance": 18, "bind::oc:briefdescription": 20,
    "bind::oc:description": 22, "relevant": 32, "required": 10,
    "required_message": 26, "constraint": 32, "constraint_message": 28,
    "default": 18, "calculation": 32, "trigger": 14, "readonly": 9,
    "image": 12, "repeat_count": 12, "bind::oc:external": 16,
    "choice_filter": 26, "bind::oc:constraint-type": 20,
    "bind::oc:required-type": 18,
}

CHOICES_WIDTHS = {"list_name": 18, "label": 32, "name": 24, "image": 10}
SETTINGS_WIDTHS = {
    "form_title": 28, "form_id": 16, "version": 8, "style": 14,
    "crossform_references": 24, "namespaces": 55
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=9, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color=GREY_MID)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def _header_cell(cell, value):
    cell.value = value
    cell.font = _font(bold=True, color=WHITE, size=9)
    cell.fill = _fill(DARK_BLUE)
    cell.border = _border()
    cell.alignment = _align(h="center", v="center")

def _coerce_cell_value(value):
    """Coerce a value to something openpyxl can write to a cell.
    Lists/tuples become pipe-delimited strings; dicts become JSON;
    anything else is left as-is."""
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return " | ".join(str(v) for v in value)
    if isinstance(value, dict):
        try:
            import json as _json
            return _json.dumps(value)
        except Exception:
            return str(value)
    return value


def _data_cell(cell, value, row_idx=0, flagged=False):
    cell.value = _coerce_cell_value(value)
    cell.font = _font(size=8)
    bg = AMBER if flagged else (GREY_LIGHT if row_idx % 2 == 0 else WHITE)
    cell.fill = _fill(bg)
    cell.border = _border()
    cell.alignment = _align()


# ── XLSX column name → XLSForm column name mapping ────────────────────────────
XLSX_TO_XLSFORM = {
    "bind__oc_itemgroup":        "bind::oc:itemgroup",
    "bind__oc_external":         "bind::oc:external",
    "bind__oc_briefdescription": "bind::oc:briefdescription",
    "bind__oc_description":      "bind::oc:description",
    "bind__oc_constraint_type":  "bind::oc:constraint-type",
    "bind__oc_required_type":    "bind::oc:required-type",
}
# Inverse: given the XLSForm column name, find the JSON key that carries
# its value. Used by _resolve_cell_value when the XLSForm name isn't
# present as a direct key in the row dict.
XLSFORM_TO_JSON = {v: k for k, v in XLSX_TO_XLSFORM.items()}


def _resolve_cell_value(row, col):
    """
    Given a row dict and an XLSForm column name, return the value.
    Tries both the XLSForm column name (e.g. 'bind::oc:itemgroup') AND its
    JSON underscore equivalent (e.g. 'bind__oc_itemgroup'), because
    Claude's extraction emits underscore keys (colons aren't friendly in
    JSON) while the XLSForm spec mandates colon keys.
    """
    # Try exact XLSForm name first
    if col in row and row[col] not in (None, ''):
        return row[col]
    # Fall back to JSON underscore variant
    json_key = XLSFORM_TO_JSON.get(col)
    if json_key and json_key in row:
        return row[json_key]
    return ''

# Columns from spec XLSX to strip (never include in output XLSForms)
STRIP_COLS = {
    "ACTION", "REVIEW_NOTES", "completion_status", "library_source",
    "pdf_original_label", "cdash_standard_name", "cdash_name_deviation",
    "cdash_name_confidence", "flag_reason", "choice_filter_meta",
    "source", "filter_column", "filter_value"
}

# ── Read spec XLSX ─────────────────────────────────────────────────────────────
def read_spec_xlsx(spec_path):
    """
    Read the EDC structure specification XLSX and return a dict of form data.
    Returns: {
        'study_meta': {...},
        'timepoint_csv': {'filename': ..., 'rows': [...]},
        'labranges_csv': {'filename': ..., 'columns': [...], 'rows': [...]},
        'forms': [{settings, choices, survey, form_id, form_title, ...}, ...]
    }
    """
    import openpyxl
    wb = openpyxl.load_workbook(spec_path, data_only=True)
    result = {'study_meta': {}, 'timepoint_csv': {}, 'labranges_csv': {}, 'forms': []}

    # ── INDEX sheet ────────────────────────────────────────────────────────
    if 'INDEX' in wb.sheetnames:
        ws = wb['INDEX']
        meta_map = {}
        for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
            if row[0] and row[1]:
                key = str(row[0]).strip()
                val = str(row[1]).strip() if row[1] else ''
                meta_map[key] = val
        result['study_meta'] = {
            'protocol_number': meta_map.get('Protocol Number', ''),
            'study_id':        meta_map.get('Study ID', ''),
            'input_mode':      meta_map.get('Input Mode', ''),
        }
        # Read form inventory from INDEX
        form_list = []
        reading_forms = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0] == '#':
                reading_forms = True
                continue
            if reading_forms and row and row[0] and str(row[0]).strip().isdigit():
                form_list.append({
                    'num':        str(row[0]).strip(),
                    'form_id':    str(row[1]).strip() if row[1] else '',
                    'form_title': str(row[2]).strip() if row[2] else '',
                    'category':   str(row[3]).strip() if row[3] else '',
                    'cdash':      str(row[4]).strip() if row[4] else '',
                    'arm':        str(row[5]).strip() if row[5] else '',
                    'complexity': str(row[6]).strip() if row[6] else '',
                    'repeating':  str(row[7]).strip() if row[7] else 'No',
                    'epro':       str(row[8]).strip() if row[8] else 'No',
                    'reuse':      str(row[9]).strip() if row[9] else '0',
                    'survey_tab': str(row[10]).strip() if row[10] else '',
                })
        result['_form_list'] = form_list

    # ── TIMEPOINTS sheet ───────────────────────────────────────────────────
    if 'TIMEPOINTS' in wb.sheetnames:
        ws = wb['TIMEPOINTS']
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
            if row[0] and row[1]:
                action = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ''
                if action != 'DELETE':
                    rows.append({'event': str(row[0]).strip(), 'timepoint': str(row[1]).strip()})
        study_id = result['study_meta'].get('study_id', 'study')
        result['timepoint_csv'] = {'filename': f"{study_id}_tpt.csv", 'rows': rows}

    # ── LAB_RANGES sheet ───────────────────────────────────────────────────
    if 'LAB_RANGES' in wb.sheetnames:
        ws = wb['LAB_RANGES']
        headers = []
        data_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i == 0:
                headers = [str(c).strip() for c in row if c is not None]
            else:
                if any(c is not None for c in row[:len(headers)]):
                    row_dict = {}
                    for j, h in enumerate(headers):
                        val = row[j] if j < len(row) else None
                        row_dict[h] = str(val).strip() if val is not None else ''
                    data_rows.append(row_dict)
        result['labranges_csv'] = {
            'filename': 'labranges.csv',
            'columns': headers,
            'rows': data_rows
        }

    # ── Per-form tabs ──────────────────────────────────────────────────────
    form_list = result.get('_form_list', [])
    sheet_names = wb.sheetnames

    for form_meta in form_list:
        form_id = form_meta['form_id']
        survey_tab = form_meta.get('survey_tab', f"{form_id}_survey")
        choices_tab = f"{form_id}_choices"
        settings_tab = f"{form_id}_settings"

        # Try to find tabs (tab names are truncated to 31 chars)
        s_tab = next((s for s in sheet_names if s == survey_tab or
                      s == survey_tab[:31]), None)
        c_tab = next((s for s in sheet_names if s == choices_tab or
                      s == choices_tab[:31]), None)
        st_tab = next((s for s in sheet_names if s == settings_tab or
                       s == settings_tab[:31]), None)

        if not s_tab:
            continue  # form tabs not found, skip

        # Read survey
        survey_rows = []
        extra_cols = []
        ws = wb[s_tab]
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=3, values_only=True)):
            header_row = [str(c).strip() if c else '' for c in row]
        if header_row:
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not any(c is not None for c in row):
                    continue
                action = str(row[0]).strip().upper() if row[0] else ''
                if action == 'DELETE':
                    continue
                row_dict = {}
                for j, h in enumerate(header_row):
                    if h in STRIP_COLS or not h:
                        continue
                    val = row[j] if j < len(row) else None
                    # Remap XLSX col names to XLSForm col names
                    xlsform_key = XLSX_TO_XLSFORM.get(h, h)
                    row_dict[xlsform_key] = str(val).strip() if val is not None else ''
                survey_rows.append(row_dict)
                # Track any extra columns (choice_filter, hard checks etc.)
                for k in row_dict:
                    if k not in SURVEY_COLS and k not in STRIP_COLS and k not in extra_cols:
                        extra_cols.append(k)

        # Read choices
        choices_rows = []
        if c_tab:
            ws = wb[c_tab]
            c_header = None
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                c_header = [str(c).strip() if c else '' for c in row]
            if c_header:
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not any(c is not None for c in row):
                        continue
                    action = str(row[0]).strip().upper() if row[0] else ''
                    if action == 'DELETE':
                        continue
                    row_dict = {}
                    for j, h in enumerate(c_header):
                        if h in STRIP_COLS or not h:
                            continue
                        val = row[j] if j < len(row) else None
                        row_dict[h] = str(val).strip() if val is not None else ''
                    if row_dict.get('list_name') or row_dict.get('name'):
                        choices_rows.append(row_dict)

        # Read settings
        settings_dict = {
            'form_title': form_meta.get('form_title', ''),
            'form_id': form_id,
            'version': '1',
            'style': 'theme-grid',
            'crossform_references': '',
            'namespaces': 'oc="http://openclinica.org/xforms" , OpenClinica="http://openclinica.com/odm"'
        }
        if st_tab:
            ws = wb[st_tab]
            setting_keys = ['form_title', 'form_id', 'version', 'style',
                            'crossform_references', 'namespaces']
            for i, row in enumerate(ws.iter_rows(min_row=3, max_row=8, values_only=True)):
                if i < len(setting_keys) and row[1] is not None:
                    val = str(row[1]).strip()
                    if val:
                        settings_dict[setting_keys[i]] = val

        result['forms'].append({
            'form_id':    form_id,
            'form_title': form_meta.get('form_title', ''),
            'category':   form_meta.get('category', ''),
            'arm':        form_meta.get('arm', 'BOTH'),
            'complexity': form_meta.get('complexity', ''),
            'repeating':  form_meta.get('repeating', 'No'),
            'epro':       form_meta.get('epro', 'No'),
            'reuse':      form_meta.get('reuse', '0'),
            'settings':   settings_dict,
            'choices':    choices_rows,
            'survey':     survey_rows,
            'extra_cols': extra_cols,
            'source_tab': survey_tab,
        })

    return result


# ── Survey row normalization ──────────────────────────────────────────────────
def _normalize_survey_rows(rows):
    """
    Normalize survey rows before writing to XLSForm.

    XLSForm spec (and pyxform validation) requires that end group and
    end repeat rows have a BLANK name. Claude sometimes generates these
    rows with names (e.g. 'AE_REPEAT_END') which causes pyxform to look
    for a matching begin_group by name and fail. This fix also handles
    the OC-8 phantom end group (between begin repeat and end repeat)
    which has no matching begin group by design.
    """
    END_TYPES = {'end group', 'end repeat', 'end_group', 'end_repeat'}
    normalized = []
    for row in rows:
        r = dict(row)
        t = str(r.get('type', '') or '').strip().lower()
        if t in END_TYPES:
            r['name'] = ''
        normalized.append(r)
    return normalized


def _wrap_repeat_group(rows, repeat_group_name, form_id, build_log=None):
    """Wrap the form's data rows in begin_repeat / end_repeat using the
    group name from the study spec (repeat_group_name field).

    Called AFTER _balance_begin_end_tags, which has already stripped the
    OC-8 phantom structure. At that point the survey is either:
      (a) flat — all group/repeat tags stripped
      (b) group-wrapped — begin group / end group preserved

    In case (a): inserts begin_repeat before the first row that has a
    bind::oc:itemgroup value, and end_repeat after the last such row.
    In case (b): renames begin group → begin repeat (using repeat_group_name)
    and end group → end repeat (with empty name).

    Only called when repeat_group_name is a non-empty string.
    """
    has_group = any(
        str(r.get('type', '') or '').strip().lower().replace('_', ' ')
        in ('begin group', 'begin_group')
        for r in rows
    )

    if has_group:
        out = []
        for row in rows:
            t = str(row.get('type', '') or '').strip().lower().replace('_', ' ')
            if t in ('begin group', 'begin_group'):
                new_row = dict(row)
                new_row['type'] = 'begin repeat'
                new_row['name'] = repeat_group_name
                out.append(new_row)
            elif t in ('end group', 'end_group'):
                new_row = dict(row)
                new_row['type'] = 'end repeat'
                new_row['name'] = ''
                out.append(new_row)
            else:
                out.append(row)
        print(f"[edc-builder] {form_id}: converted begin/end group → "
              f"begin/end repeat ({repeat_group_name})", flush=True)
        if build_log is not None:
            build_log.setdefault('repeat_wrap', []).append(
                {'form_id': form_id, 'method': 'group_rename',
                 'repeat_group_name': repeat_group_name})
        return out

    # Flat path — find first and last rows with bind::oc:itemgroup
    first_idx, last_idx = None, None
    for i, row in enumerate(rows):
        if row.get('bind::oc:itemgroup'):
            if first_idx is None:
                first_idx = i
            last_idx = i

    if first_idx is None:
        print(f"[edc-builder] {form_id}: repeat_group_name={repeat_group_name!r} "
              f"set but no rows with bind::oc:itemgroup found — skipping wrap",
              flush=True)
        return rows

    out = (
        rows[:first_idx]
        + [{'type': 'begin repeat', 'name': repeat_group_name}]
        + rows[first_idx:last_idx + 1]
        + [{'type': 'end repeat', 'name': ''}]
        + rows[last_idx + 1:]
    )
    print(f"[edc-builder] {form_id}: wrapped rows {first_idx}–{last_idx} "
          f"in begin/end repeat ({repeat_group_name})", flush=True)
    if build_log is not None:
        build_log.setdefault('repeat_wrap', []).append(
            {'form_id': form_id, 'method': 'flat_wrap',
             'repeat_group_name': repeat_group_name,
             'first_idx': first_idx, 'last_idx': last_idx})
    return out


# ── Begin/end tag pairing balancer ────────────────────────────────────────────
def _balance_begin_end_tags(rows, form_id, build_log=None):
    """Normalize survey block tags for OpenClinica.

    Normalizes begin/end group and repeat tags from OC-8 source library forms.

    OC-8 source CRFs use a phantom pattern — begin group + data + end group +
    empty begin repeat + end repeat — that OC's form-service rejects with
    "Unmatched end statement". This balancer strips the phantom repeat rows
    and preserves only valid begin group / end group pairs.

    Note: begin_repeat / end_repeat ARE valid in OC XLSForms for repeating
    forms, but they are applied AFTER this balancer runs via _wrap_repeat_group,
    which reads repeat_group_name from the study spec. This function's job is
    only to clean up the OC-8 source artifacts.
    """
    def _classify(t):
        u = (t or '').strip().lower().replace('_', ' ')
        if u == 'begin group':  return ('begin', 'group')
        if u == 'begin repeat': return ('begin', 'repeat')
        if u == 'end group':    return ('end',   'group')
        if u == 'end repeat':   return ('end',   'repeat')
        return (None, None)

    stack = []
    corrections = []
    out = []
    _repeat_dropped = False

    for idx, row in enumerate(rows):
        t = str(row.get('type', '') or '')
        action, kind = _classify(t)

        # Repeats are not an XLSForm construct in OC — strip both ends.
        if kind == 'repeat':
            _repeat_dropped = True
            corrections.append(
                f"row {idx + 2}: dropped {t!r} — OC defines repeating "
                f"groups via bind::oc:itemgroup, not begin/end repeat"
            )
            continue

        if action == 'begin':   # begin group
            stack.append('group')
            out.append(dict(row))

        elif action == 'end':   # end group
            if not stack:
                corrections.append(
                    f"row {idx + 2}: dropped orphan {t!r} — no matching "
                    f"begin group on the stack"
                )
                continue
            stack.pop()
            out.append(dict(row))

        else:
            out.append(dict(row))

    # Auto-close any unclosed groups at the tail.
    while stack:
        stack.pop()
        out.append({'type': 'end group', 'name': ''})
        corrections.append(
            "tail: appended 'end group' to close an unclosed begin group"
        )

    if corrections:
        print(f"[edc-builder] balanced begin/end tags for {form_id}:",
              flush=True)
        for c in corrections:
            print(f"  {c}", flush=True)
        if build_log is not None:
            build_log.setdefault('tag_balance_corrections', []).append({
                'form_id': form_id,
                'corrections': corrections,
            })

    # Second-pass group strip DISABLED pending form-design-team review of the
    # correct OC-8 repeating structure (group wrapper + phantom repeat). The
    # _repeat_dropped flag above is still set by the first pass and retained
    # for when this is revisited — re-enable by uncommenting the block below.
    # if _repeat_dropped:
    #     _flat = []
    #     _dropped_groups = 0
    #     for row in out:
    #         _u = (str(row.get('type', '') or '')
    #               .strip().lower().replace('_', ' '))
    #         if _u in ('begin group', 'end group'):
    #             _dropped_groups += 1
    #             print(f"[edc-builder] {form_id}: dropped begin/end group "
    #                   f"(repeat removed, flat field structure used)",
    #                   flush=True)
    #             continue
    #         _flat.append(row)
    #     out = _flat
    #     if _dropped_groups and build_log is not None:
    #         build_log.setdefault('tag_balance_corrections', []).append({
    #             'form_id': form_id,
    #             'corrections': [
    #                 f"dropped {_dropped_groups} begin/end group row(s) "
    #                 f"(repeat removed, flat field structure used)"
    #             ],
    #         })

    return out


# ── Build a single XLSForm .xlsx ───────────────────────────────────────────────
def build_single_xlsform(form_data, output_path, build_log):
    """Build one XLSForm .xlsx file from form_data dict.

    When form_template.xlsx is present alongside this script, the output is
    based on that template so the human receives:
      - The six OC reference/instruction tabs (Cross-Form Examples, etc.)
      - The bind::oc:external dropdown in the survey sheet
      - The official OC header row styling
    If the template is missing, falls back to building the three functional
    sheets from scratch (original behaviour).
    """
    settings = form_data.get('settings', {}) or {}
    choices  = form_data.get('choices', [])
    survey   = form_data.get('survey', [])
    extra_c  = form_data.get('extra_cols', [])
    form_id  = form_data.get('form_id', 'FORM')

    # Apply OpenClinica XLSForm defaults for missing settings fields.
    # OpenClinica's XLSForm validator requires these — without them the
    # form won't load. Claude may omit them in the JSON extraction.
    OC_SETTINGS_DEFAULTS = {
        'form_title':           form_data.get('form_title', form_id),
        'form_id':               form_id,
        'version':               '1',
        'style':                 'theme-grid',
        'crossform_references': '',
        'namespaces':
            'oc="http://openclinica.org/xforms" , '
            'OpenClinica="http://openclinica.com/odm"',
    }
    for k, default in OC_SETTINGS_DEFAULTS.items():
        if not settings.get(k):
            settings[k] = default

    # Build-time version stamp: a unique 12-digit integer per build run, so OC
    # always creates a NEW form version instead of returning a cached rejection
    # (EX) or a deduplicated existing version (SLEEP).
    build_version = int(datetime.datetime.now().strftime("%Y%m%d%H%M"))
    settings['version'] = build_version
    print(f"[edc-builder] {form_id}: version stamp={build_version}", flush=True)

    use_template = os.path.exists(TEMPLATE_PATH)

    # ── Load workbook ──────────────────────────────────────────────────────
    if use_template:
        wb = load_workbook(TEMPLATE_PATH)
        ws_set = wb['settings']
        ws_ch  = wb['choices']
        ws_sv  = wb['survey']

        # Clean up LibreOffice conversion artefacts that cause Excel to show
        # a "repaired content" warning on every open:
        #
        # 1. Invalid sheetView selection elements — LO writes 4 <selection>
        #    pane entries (topRight, bottomLeft×2, bottomRight) for a simple
        #    horizontal freeze. Excel only accepts 2 for a row-only freeze.
        #    Wipe all selections/panes now; freeze_panes='A2' below rewrites
        #    them cleanly.
        #
        # 2. type=None DataValidations — internal LO markers, not real rules.
        #
        # 3. formula2='0' and operator='between' on list validations — invalid
        #    for list type; Excel flags and repairs them.
        # Trim LibreOffice's extra <selection> pane elements from sheetViews.
        # LO writes 4 selections for a simple horizontal freeze; Excel only
        # accepts 1. openpyxl's freeze_panes setter only modifies selection[0]
        # and leaves the extras intact, so we strip them here explicitly.
        # The actual freeze ('A2') is applied later per-sheet.
        for ws in (ws_set, ws_ch, ws_sv):
            ws.sheet_view.selection = ws.sheet_view.selection[:1]

        for ws in (ws_set, ws_ch, ws_sv):
            bad = [dv for dv in ws.data_validations.dataValidation
                   if dv.type is None]
            for dv in bad:
                ws.data_validations.dataValidation.remove(dv)
            for dv in ws.data_validations.dataValidation:
                if dv.type == 'list':
                    dv.formula2 = None
                    dv.operator = None

        # Clear any data rows left in the template (keep row 1 = headers)
        for ws in (ws_set, ws_ch, ws_sv):
            for row_cells in ws.iter_rows(min_row=2):
                for cell in row_cells:
                    cell.value = None
    else:
        wb     = Workbook()
        ws_set = wb.active
        ws_set.title = 'settings'
        ws_ch  = wb.create_sheet('choices')
        ws_sv  = wb.create_sheet('survey')

    # ── Sheet: settings ────────────────────────────────────────────────────
    if not use_template:
        for col_i, col in enumerate(SETTINGS_COLS, start=1):
            _header_cell(ws_set.cell(row=1, column=col_i), col)

    for col_i, col in enumerate(SETTINGS_COLS, start=1):
        val = settings.get(col, '')
        ws_set.cell(row=2, column=col_i).value = val or ''
        ws_set.column_dimensions[get_column_letter(col_i)].width = \
            SETTINGS_WIDTHS.get(col, 20)

    ws_set.row_dimensions[1].height = 16
    ws_set.row_dimensions[2].height = 14
    ws_set.freeze_panes = 'A2'

    # ── Sheet: choices ─────────────────────────────────────────────────────
    choice_extra_cols = []
    for ch in choices:
        for k in ch:
            if k not in CHOICES_COLS and k not in STRIP_COLS and k not in choice_extra_cols:
                choice_extra_cols.append(k)
    all_choice_cols = CHOICES_COLS + choice_extra_cols

    if not use_template:
        for col_i, col in enumerate(all_choice_cols, start=1):
            _header_cell(ws_ch.cell(row=1, column=col_i), col)
    else:
        # Extra columns beyond the template's 4 still need headers written
        for col_i, col in enumerate(all_choice_cols, start=1):
            if col_i > len(CHOICES_COLS):
                _header_cell(ws_ch.cell(row=1, column=col_i), col)

    for col_i in range(1, len(all_choice_cols) + 1):
        col = all_choice_cols[col_i - 1]
        ws_ch.column_dimensions[get_column_letter(col_i)].width = \
            CHOICES_WIDTHS.get(col, 18)

    for row_i, ch in enumerate(choices, start=2):
        for col_i, col in enumerate(all_choice_cols, start=1):
            val = ch.get(col, '')
            _data_cell(ws_ch.cell(row=row_i, column=col_i), val, row_i - 2)

    ws_ch.row_dimensions[1].height = 16
    ws_ch.freeze_panes = 'A2'

    # ── Sheet: survey ──────────────────────────────────────────────────────
    additional_survey_cols = [c for c in extra_c
                               if c not in SURVEY_COLS and c not in STRIP_COLS]
    all_survey_cols = SURVEY_COLS + additional_survey_cols

    if not use_template:
        for col_i, col in enumerate(all_survey_cols, start=1):
            _header_cell(ws_sv.cell(row=1, column=col_i), col)
    else:
        # Extra columns beyond the template's 20 still need headers written
        for col_i, col in enumerate(all_survey_cols, start=1):
            if col_i > len(SURVEY_COLS):
                _header_cell(ws_sv.cell(row=1, column=col_i), col)

    for col_i, col in enumerate(all_survey_cols, start=1):
        ws_sv.column_dimensions[get_column_letter(col_i)].width = \
            SURVEY_WIDTHS.get(col, 16)

    # Strip XLSForm repeat syntax (OC uses bind::oc:itemgroup) and, when a
    # repeat is removed, flatten away the begin/end group wrappers too.
    survey = _balance_begin_end_tags(survey, form_id, build_log)
    _repeat_group_name = (form_data.get('repeat_group_name') or '').strip()
    if _repeat_group_name:
        survey = _wrap_repeat_group(survey, _repeat_group_name, form_id, build_log)
    survey = _normalize_survey_rows(survey)

    placeholders_in_form = []
    for row_i, row in enumerate(survey, start=2):
        has_placeholder = any(
            '[PLACEHOLDER' in str(v).upper() or '[UNIT]' in str(v).upper()
            for v in row.values() if v
        )
        if has_placeholder:
            placeholders_in_form.append(row.get('name', f'row_{row_i}'))

        for col_i, col in enumerate(all_survey_cols, start=1):
            val = _resolve_cell_value(row, col)
            _data_cell(ws_sv.cell(row=row_i, column=col_i), val,
                       row_i - 2, flagged=has_placeholder)
        ws_sv.row_dimensions[row_i].height = 13

    ws_sv.row_dimensions[1].height = 16
    ws_sv.freeze_panes = 'A2'

    # bind::oc:external dropdown — only needed when building from scratch;
    # the template already carries this validation for rows 2:10000.
    if not use_template:
        external_col = len(SURVEY_COLS)   # column 20 = bind::oc:external
        dv = DataValidation(
            type="list",
            formula1='"clinicaldata,contactdata,signature,identifier"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"{get_column_letter(external_col)}2:{get_column_letter(external_col)}10000"
        ws_sv.add_data_validation(dv)

    # Log any placeholders found
    if placeholders_in_form:
        build_log['placeholder_applied'].append({
            'form_id': form_id,
            'fields': placeholders_in_form,
            'note': 'Best-guess values applied — require site-specific completion'
        })

    wb.save(output_path)
    return True


# ── Build all XLSForms ─────────────────────────────────────────────────────────
def build_all_xlsforms(spec_data, output_dir, build_log):
    """
    Build all XLSForm files from spec_data.
    Returns list of (form_id, output_path) tuples.

    Each successfully-built form is also validated using pyxform; results
    are accumulated in build_log['validation_results'] for downstream
    surfacing in BUILD_README and the Build Checklist PDF.
    """
    # Lazy import to keep build_xlsforms.py lean
    try:
        from validate_form import validate_xlsform
    except ImportError:
        validate_xlsform = None

    os.makedirs(output_dir, exist_ok=True)
    built = []
    build_log.setdefault('validation_results', [])

    # Track form_ids to handle variants (same form_id, different designs)
    seen_form_ids = {}

    for form in spec_data.get('forms', []):
        form_id = form.get('form_id', 'FORM')
        source_tab = form.get('source_tab', form_id)

        # Determine output filename
        # If source tab has a suffix (e.g., IE_TRT_survey), use that as filename
        tab_prefix = source_tab.replace('_survey', '').replace('_choices', '')

        if form_id in seen_form_ids:
            # Variant — use tab prefix as filename
            filename = f"{tab_prefix}.xlsx"
        else:
            seen_form_ids[form_id] = True
            # Check if there will be a variant (another form with same form_id)
            same_id_count = sum(1 for f in spec_data['forms'] if f.get('form_id') == form_id)
            if same_id_count > 1:
                filename = f"{tab_prefix}.xlsx"
            else:
                filename = f"{form_id}.xlsx"

        output_path = os.path.join(output_dir, filename)

        try:
            build_single_xlsform(form, output_path, build_log)
            built.append((tab_prefix, output_path))
            build_log['forms_built'].append(form_id)

            # Validate the form we just built
            if validate_xlsform is not None:
                v_result = validate_xlsform(output_path, form_id=tab_prefix)
                build_log['validation_results'].append(v_result)
        except Exception as e:
            build_log['build_errors'].append({
                'form_id': form_id,
                'error': str(e)
            })
            build_log['forms_skipped'].append(form_id)

    return built


# ── Write CSV files ────────────────────────────────────────────────────────────
def write_timepoint_csv(tpt_data, output_path, build_log):
    """Write the timepoint CSV file."""
    rows = tpt_data.get('rows', [])
    with open(output_path, 'w', newline='') as f:
        f.write('event,timepoint\r\n')
        for row in rows:
            event = row.get('event', '').replace(',', '')
            tpt = row.get('timepoint', '').replace(',', '')
            f.write(f'{event},{tpt}\r\n')
    build_log['build_warnings'] = build_log.get('build_warnings', [])
    if not rows:
        build_log['build_warnings'].append('Timepoint CSV has no rows — check TIMEPOINTS sheet')


def write_labranges_csv(lr_data, output_path, build_log):
    """Write the lab ranges CSV file."""
    cols = lr_data.get('columns', [])
    rows = lr_data.get('rows', [])
    placeholders = []

    with open(output_path, 'w', newline='') as f:
        f.write(','.join(cols) + '\r\n')
        for row in rows:
            values = []
            for c in cols:
                val = row.get(c, '')
                if '[PLACEHOLDER' in str(val).upper():
                    placeholders.append(f"{c}: {val}")
                    # Apply best-guess
                    if 'lower' in c.lower():
                        val = '0'
                    elif 'upper' in c.lower():
                        val = '999'
                    elif 'unit' in c.lower():
                        val = '[UNIT]'
                    elif 'lab_name' in c.lower():
                        val = '[LAB_NAME]'
                values.append(str(val).replace(',', ';'))
            f.write(','.join(values) + '\r\n')

    if placeholders:
        build_log['placeholder_applied'].append({
            'form_id': 'labranges.csv',
            'fields': placeholders[:5],
            'note': 'Lab ranges require site-specific values from each participating lab'
        })


if __name__ == '__main__':
    import sys, json
    if len(sys.argv) < 3:
        print("Usage: python build_xlsforms.py <spec_xlsx_path> <output_dir>")
        sys.exit(1)

    spec_path  = sys.argv[1]
    output_dir = sys.argv[2]

    build_log = {
        'forms_built': [], 'forms_skipped': [],
        'placeholder_applied': [], 'oid_placeholders': [],
        'qa_results': [], 'build_warnings': [], 'build_errors': []
    }

    print(f"Reading spec: {spec_path}")
    spec_data = read_spec_xlsx(spec_path)
    print(f"Found {len(spec_data['forms'])} forms")

    forms_dir = os.path.join(output_dir, 'forms')
    csv_dir   = os.path.join(output_dir, 'csv')
    os.makedirs(forms_dir, exist_ok=True)
    os.makedirs(csv_dir,   exist_ok=True)

    built = build_all_xlsforms(spec_data, forms_dir, build_log)
    print(f"Built {len(built)} XLSForms")

    study_id = spec_data['study_meta'].get('study_id', 'study')
    tpt_path = os.path.join(csv_dir, f"{study_id}_tpt.csv")
    write_timepoint_csv(spec_data['timepoint_csv'], tpt_path, build_log)
    print(f"Written: {tpt_path}")

    lr_path = os.path.join(csv_dir, 'labranges.csv')
    write_labranges_csv(spec_data['labranges_csv'], lr_path, build_log)
    print(f"Written: {lr_path}")

    print(f"\nBuild log: {json.dumps(build_log, indent=2)}")

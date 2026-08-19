"""
deterministic_build.py — Chain BX building blocks (no AI calls)

Shared helpers for producing DVS / spec artifacts directly from a
customer-supplied EDC Build ZIP, without running Protocol Analysis.

build_forms_json_from_zip()
    Extracted from main.py's regenerate_dvs endpoint (the openpyxl
    survey/choices reader). No logic changes — just made reusable so
    Chain BX and the existing admin endpoint share one implementation
    instead of two copies drifting apart.

build_event_form_map()
    New. _build_form_event_map() in extract_dvs_from_forms.py normally
    reads form.visits_assigned off a Study Spec JSON that Chain A
    produced. Chain BX has no Chain A run, so this rebuilds the same
    {form_oid: event_oid} shape from whatever's actually available:
      1. Preferred — parse <StudyEventDef><FormRef FormOID=.../></...>
         out of a customer-supplied ODM XML. This is pure structure,
         confirmed present and reliable in real OC4 exports (unlike
         constraint/relevant/calculation logic, which is not).
      2. Fallback — infer from each XLSForm's settings.crossform_references.
         Heuristic, not authoritative; forms that can't be placed this
         way are returned as review_flags-style dicts instead of guessed
         at, matching the existing review_flags pattern in Study Spec
         JSON (site_specific, oid_confirmation, protocol_ambiguous, etc.)
      3. manual_overrides — an optional {form_oid_or_filename: event_oid}
         dict, applied last with highest priority, so a human who's
         inspected an unplaced form can patch the gap without touching
         code or re-running the ODM/settings passes.

Known limitation (confirmed against a real GON001 build): forms with
no cross-form field at all (single-event forms like IE, MH, PE, DS —
nothing in their survey ever references another form or event) carry
no signal for either pass to find. These will always need either the
ODM XML from the SAME build, or a manual_overrides entry.
"""

import io
import os
import xml.etree.ElementTree as ET

import openpyxl


def build_forms_json_from_zip(zip_bytes):
    """Read every .xlsx in an EDC Build ZIP into the forms_json shape
    extract_dvs_data() expects: {"forms": {filename: {survey, choices}}}.

    Mirrors the reader inline in main.py's regenerate_dvs endpoint
    (main.py ~lines 1355-1405) — lifted here unchanged so both call
    sites share one implementation.
    """
    import zipfile

    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    forms_json = {"forms": {}}

    for name in sorted(zf.namelist()):
        if not name.lower().endswith(".xlsx"):
            continue
        try:
            with zf.open(name) as f:
                wb_bytes = f.read()
            wb = openpyxl.load_workbook(io.BytesIO(wb_bytes),
                                         read_only=True, data_only=True)

            def _sheet_rows(sheet_name):
                if sheet_name not in wb.sheetnames:
                    return []
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return []
                headers = [str(h or "").strip() for h in rows[0]]
                out = []
                for r in rows[1:]:
                    row_dict = {headers[i]: r[i]
                                for i in range(len(headers))
                                if i < len(r) and r[i] is not None}
                    if row_dict:
                        out.append(row_dict)
                return out

            forms_json["forms"][os.path.basename(name)] = {
                "survey":  _sheet_rows("survey"),
                "choices": _sheet_rows("choices"),
            }
        except Exception as e:
            print(f"[deterministic-build] skipping {name}: {e}", flush=True)

    return forms_json


# ── ODM namespace (matches odm_reader.py's convention) ────────────────────────
_ODM_NS = {"odm": "http://www.cdisc.org/ns/odm/v1.3"}


def _form_event_map_from_odm(odm_xml_bytes):
    """Parse <StudyEventDef><FormRef FormOID=.../></StudyEventDef> pairs.

    Returns {form_oid: first_event_oid}. Uses the first StudyEventDef
    that references each form — mirrors _build_form_event_map()'s
    "first entry = primary event" convention in extract_dvs_from_forms.py,
    so downstream DVS/UAT code sees the same shape it always has.
    """
    mapping = {}
    try:
        root = ET.fromstring(odm_xml_bytes)
    except ET.ParseError as e:
        print(f"[deterministic-build] ODM parse failed: {e}", flush=True)
        return mapping

    for event_def in root.iter("{http://www.cdisc.org/ns/odm/v1.3}StudyEventDef"):
        event_oid = event_def.get("OID", "")
        if not event_oid:
            continue
        for form_ref in event_def.findall("odm:FormRef", _ODM_NS):
            form_oid = form_ref.get("FormOID", "")
            if form_oid and form_oid not in mapping:
                mapping[form_oid] = event_oid

    return mapping


def _form_event_map_from_settings(forms_json):
    """Fallback: infer {form_oid: event_oid} from each XLSForm's
    settings.crossform_references when no ODM XML is available.

    Heuristic only — crossform_references lists events a form's fields
    read FROM, not necessarily the event the form itself is placed on.
    Forms this can't confidently place are left out of the mapping and
    returned separately in `unplaced` so the caller can flag them
    (matching the existing review_flags pattern) instead of guessing.
    """
    mapping = {}
    unplaced = []

    for filename, form_data in forms_json.get("forms", {}).items():
        survey = form_data.get("survey") or []
        if not survey:
            # Not a form at all (e.g. a checklist .xlsx bundled in the
            # ZIP) — nothing to place, nothing to flag.
            continue

        form_oid = None
        event_oid = None

        for row in survey:
            ext = row.get("bind::oc:external") or row.get("bind__oc_external")
            calc = row.get("calculation") or ""
            if ext == "clinicaldata" and "StudyEventOID='" in calc:
                start = calc.find("StudyEventOID='") + len("StudyEventOID='")
                end = calc.find("'", start)
                if end > start:
                    event_oid = calc[start:end]
            if "FormOID='" in calc and form_oid is None:
                start = calc.find("FormOID='") + len("FormOID='")
                end = calc.find("'", start)
                if end > start:
                    form_oid = calc[start:end]

        key = form_oid or filename.replace(".xlsx", "")
        if event_oid:
            mapping[key] = event_oid
        else:
            unplaced.append({
                "filename": filename,
                "form_oid_guess": key,
                "reason": ("no matching ODM FormRef, and no cross-form "
                           "StudyEventOID reference in this form's own "
                           "survey rows — likely a single-event form "
                           "with no signal to place it automatically"),
                "category": "oid_confirmation",
            })

    return mapping, unplaced


def build_event_form_map(forms_json, odm_xml_bytes=None, manual_overrides=None):
    """Return ({form_oid: event_oid}, unplaced_flags).

    Precedence, highest first:
      1. manual_overrides   — {form_oid_or_filename: event_oid}, human-supplied
      2. ODM XML FormRef    — accurate, pure structure, same-build only
      3. settings heuristic — best-effort fallback

    unplaced_flags is a list of dicts (see _form_event_map_from_settings)
    for any form none of the three passes could place — review_flags
    shaped, not a bare filename list, so a caller can surface *why* a
    form needs manual attention.
    """
    mapping = {}
    if odm_xml_bytes:
        mapping = _form_event_map_from_odm(odm_xml_bytes)

    fallback_map, unplaced = _form_event_map_from_settings(forms_json)

    for form_oid, event_oid in fallback_map.items():
        if form_oid not in mapping:
            mapping[form_oid] = event_oid

    still_unplaced = [
        flag for flag in unplaced
        if flag["form_oid_guess"] not in mapping
        and not any(flag["form_oid_guess"] in k for k in mapping)
    ]

    if manual_overrides:
        placed_filenames = {flag["filename"] for flag in still_unplaced
                             if flag["form_oid_guess"] in manual_overrides
                             or flag["filename"] in manual_overrides}
        for key, event_oid in manual_overrides.items():
            mapping[key] = event_oid
        still_unplaced = [flag for flag in still_unplaced
                           if flag["filename"] not in placed_filenames]

    return mapping, still_unplaced
